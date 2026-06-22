"""
shablon_yaratish.py — Topik kod uchun shablon yaratish
Admin: Shablon yaratish → Topik kod uchun shablon / Import
"""
import os, io, psycopg2, openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from aiogram.types import (
    InlineKeyboardMarkup, InlineKeyboardButton, BufferedInputFile
)

DATABASE_URL = os.getenv("DATABASE_URL")
shablon_state = {}  # user_id -> {step, sinf, fan, mavzular}

def db(): return psycopg2.connect(DATABASE_URL)


# ─── ASOSIY MENYU ────────────────────────────────
async def show_shablon_menu(message, user_id):
    shablon_state[user_id] = {}
    kb = InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(
            text="📋 Topik kod uchun shablon",
            callback_data="sh_topik"
        )],
        [InlineKeyboardButton(
            text="📥 Import qilish",
            callback_data="sh_import"
        )],
    ])
    await message.answer(
        "📚 Shablon yaratish\n\nNimani xohlaysiz?",
        reply_markup=kb
    )


# ─── CALLBACK HANDLER ────────────────────────────
async def handle_shablon_callback(call, user_id):
    data = call.data

    if data == "sh_topik":
        shablon_state[user_id] = {"step": "sinf_fan"}
        await call.message.edit_text(
            "📋 Topik kod uchun shablon\n\n"
            "Sinf va fanni yozing:\n"
            "Masalan: <code>1 Ingliz tili</code>",
            parse_mode="HTML"
        )
        await call.answer()

    elif data == "sh_import":
        shablon_state[user_id] = {"step": "import_wait"}
        await call.message.edit_text(
            "📥 Import qilish\n\n"
            "To'ldirilgan Excel faylni yuboring:"
        )
        await call.answer()


# ─── XABAR HANDLER ───────────────────────────────
async def handle_shablon_message(message, user_id):
    state = shablon_state.get(user_id, {})
    step = state.get("step")

    # Sinf va fan
    if step == "sinf_fan":
        parts = message.text.strip().split(None, 1)
        if len(parts) < 2:
            await message.answer(
                "❌ Noto'g'ri format!\n"
                "Masalan: <code>1 Ingliz tili</code>",
                parse_mode="HTML"
            )
            return
        sinf, fan = parts[0], parts[1]
        shablon_state[user_id] = {
            "step": "mavzular",
            "sinf": sinf,
            "fan": fan
        }
        await message.answer(
            f"✅ {sinf}-sinf, {fan}\n\n"
            "Endi mavzularni yozing:\n"
            "<code>1 / Colours\n"
            "1 / Numbers\n"
            "2 / Animals\n"
            "2 / Family\n"
            "...</code>\n\n"
            "Chorak / Mavzu formatida, har biri yangi qatorda:",
            parse_mode="HTML"
        )

    # Mavzular
    elif step == "mavzular":
        sinf = state.get("sinf", "1")
        fan = state.get("fan", "Fan")
        text = message.text.strip()

        mavzular = []
        for line in text.splitlines():
            line = line.strip()
            if not line:
                continue
            # Format: "1 / Mavzu nomi" yoki "1/ Mavzu" yoki "1 Mavzu"
            if "/" in line:
                parts = line.split("/", 1)
                chorak = parts[0].strip().lstrip("0123456789. ").strip()
                # Chorakni raqam sifatida ol
                chorak_raqam = ""
                for ch in parts[0].strip():
                    if ch.isdigit():
                        chorak_raqam += ch
                mavzu = parts[1].strip()
            else:
                parts = line.split(None, 1)
                chorak_raqam = parts[0].strip() if parts else "1"
                mavzu = parts[1].strip() if len(parts) > 1 else line

            if mavzu and chorak_raqam:
                mavzular.append((chorak_raqam, mavzu))

        if not mavzular:
            await message.answer("❌ Mavzular topilmadi! Qayta yozing.")
            return

        # Excel yaratish
        buf = await _create_shablon(sinf, fan, mavzular)

        fname = f"shablon_{sinf}sinf_{fan.replace(' ', '_')}.xlsx"
        await message.answer_document(
            BufferedInputFile(buf.read(), filename=fname),
            caption=(
                f"✅ Shablon tayyor!\n"
                f"📚 {sinf}-sinf | {fan}\n"
                f"📝 {len(mavzular)} ta mavzu × 2 qator\n"
                f"📊 Jami: {len(mavzular)*2} ta qator\n\n"
                f"Bo'sh ustunlar: Bob, Bo'lim, Kichik mavzu\n"
                f"To'ldirib import qiling ✅"
            )
        )
        shablon_state.pop(user_id, None)

    # Import
    elif step == "import_wait":
        await message.answer(
            "❌ Fayl yuboring! Matn emas."
        )


async def handle_shablon_document(message, user_id, bot):
    """Import: Excel faylni DBga saqlash"""
    state = shablon_state.get(user_id, {})
    if state.get("step") != "import_wait":
        return False

    doc = message.document
    if not doc.file_name.endswith('.xlsx'):
        await message.answer("❌ Faqat .xlsx fayl qabul qilinadi!")
        return True

    file = await bot.get_file(doc.file_id)
    file_bytes = await bot.download_file(file.file_path)

    try:
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes.read()))
        ws = wb.active

        conn = db()
        cur = conn.cursor()
        added = 0
        skipped = 0

        for r in range(2, ws.max_row + 1):
            sinf = ws.cell(r, 1).value
            fan = ws.cell(r, 2).value
            chorak = ws.cell(r, 3).value
            bob = ws.cell(r, 4).value
            bolim = ws.cell(r, 5).value
            mavzu = ws.cell(r, 6).value
            kichik = ws.cell(r, 7).value

            if not sinf or not mavzu:
                continue

            # Topic code yaratish
            # Mavjud oxirgi kodni topamiz
            cur.execute("""
                SELECT topic_code FROM dts_tree
                WHERE grade=%s AND subject_name=%s
                ORDER BY topic_code DESC LIMIT 1
            """, (str(sinf), str(fan) if fan else ""))

            row = cur.fetchone()
            if row:
                # Oxirgi raqamni oshiramiz
                last = row[0]
                parts = last.rsplit('-', 1)
                new_num = str(int(parts[1]) + 1).zfill(3)
                topic_code = f"{parts[0]}-{new_num}"
            else:
                topic_code = f"{sinf}-01-{chorak or 1}-01-01-01-001"

            try:
                cur.execute("""
                    INSERT INTO dts_tree
                    (topic_code, grade, subject_name, quarter,
                     bob_name, bolim_name, mavzu_name, kichik_name,
                     is_deleted)
                    VALUES (%s,%s,%s,%s,%s,%s,%s,%s,FALSE)
                    ON CONFLICT (topic_code) DO NOTHING
                """, (
                    topic_code, str(sinf),
                    str(fan) if fan else "",
                    str(chorak) if chorak else "1",
                    str(bob) if bob else "",
                    str(bolim) if bolim else "",
                    str(mavzu) if mavzu else "",
                    str(kichik) if kichik else "",
                ))
                added += 1
            except Exception:
                skipped += 1

        conn.commit()
        cur.close()
        conn.close()

        await message.answer(
            f"✅ Import tugadi!\n"
            f"➕ Qo'shildi: {added} ta\n"
            f"⏭ O'tkazildi: {skipped} ta"
        )

    except Exception as ex:
        await message.answer(f"❌ Xato: {ex}")

    shablon_state.pop(user_id, None)
    return True


# ─── EXCEL SHABLON YARATISH ───────────────────────
async def _create_shablon(sinf, fan, mavzular):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "DTS_SHABLON"

    headers = ["Sinf", "Fan", "Chorak", "Bob", "Bo'lim", "Mavzu", "Kichik mavzu"]
    header_colors = ["4472C4", "4472C4", "4472C4", "70AD47", "70AD47", "ED7D31", "ED7D31"]

    for col, (h, color) in enumerate(zip(headers, header_colors), 1):
        cell = ws.cell(1, col, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor=color)
        cell.alignment = Alignment(horizontal="center")

    # Har mavzudan 2 qator
    chorak_colors = {
        "1": "DEEAF1", "2": "E2EFDA", "3": "FFF2CC", "4": "FCE4D6"
    }

    row_num = 2
    for chorak, mavzu in mavzular:
        color = chorak_colors.get(str(chorak), "F2F2F2")
        for _ in range(2):  # 2 ta kichik mavzu uchun
            ws.cell(row_num, 1, value=sinf)
            ws.cell(row_num, 2, value=fan)
            ws.cell(row_num, 3, value=chorak)
            # Bob, Bo'lim — bo'sh
            ws.cell(row_num, 6, value=mavzu)
            # Kichik mavzu — bo'sh

            for col in range(1, 8):
                ws.cell(row_num, col).fill = PatternFill("solid", fgColor=color)
                ws.cell(row_num, col).alignment = Alignment(horizontal="left")

            row_num += 1

    # Ustun kengliklari
    for col, width in zip(range(1, 8), [8, 18, 8, 25, 25, 30, 30]):
        ws.column_dimensions[ws.cell(1, col).column_letter].width = width

    # Izoh varag'i
    ws2 = wb.create_sheet("IZOH")
    ws2.cell(1, 1, value="📋 TO'LDIRISH QO'LLANMASI")
    ws2.cell(1, 1).font = Font(bold=True, size=14)
    izohlar = [
        (3, "Sinf", "O'zgartirmang — avtomatik to'ldirilgan"),
        (4, "Fan", "O'zgartirmang — avtomatik to'ldirilgan"),
        (5, "Chorak", "O'zgartirmang — avtomatik to'ldirilgan"),
        (6, "Bob", "To'ldiring: masalan 'Chapter 1. Getting acquainted'"),
        (7, "Bo'lim", "To'ldiring: masalan 'Unit 1. Greetings'"),
        (8, "Mavzu", "O'zgartirmang — mavzu nomi avtomatik"),
        (9, "Kichik mavzu", "To'ldiring: mavzuning kichik qismi"),
    ]
    for r, ustun, izoh in izohlar:
        ws2.cell(r, 1, value=ustun).font = Font(bold=True)
        ws2.cell(r, 2, value=izoh)
    ws2.column_dimensions['A'].width = 15
    ws2.column_dimensions['B'].width = 50

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
