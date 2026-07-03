"""
excel_merger.py — Ikkita Excel faylni aqlli birlashtirish
Shablon + Savol → Birlashtirilgan (topic_code tartibida, bixato)
"""
import openpyxl
from collections import defaultdict
from io import BytesIO

# ── Ustun nomlari (turli fayllar turlicha yozishi mumkin) ──
TOPIC_ALIASES    = ["topic_code","topik_kod","topic","kod","code"]
QUESTION_ALIASES = ["question","savol","вопрос","pytanie"]
OPT_A_ALIASES    = ["option_a","a","variant_a","a)","optiona"]
OPT_B_ALIASES    = ["option_b","b","variant_b","b)","optionb"]
OPT_C_ALIASES    = ["option_c","c","variant_c","c)","optionc"]
OPT_D_ALIASES    = ["option_d","d","variant_d","d)","optiond"]
CORRECT_ALIASES  = ["correct_answer","togri","to'g'ri","correct","javob","answer"]
EXPL_ALIASES     = ["explanation","izoh","izoh/sharh","sharh","explain"]
QTYPE_ALIASES    = ["question_type","tur","type","qtype"]
DIFF_ALIASES     = ["difficulty","qiyinlik","diff","daraja"]
IMAGE_ALIASES    = ["image_url","rasm","image","img","rasm_kodi"]
LANG_ALIASES     = ["language","til","lang"]

def find_col(headers: list, aliases: list) -> int | None:
    """Ustun nomini turli variantlarda qidiradi. 1-indexed."""
    for i, h in enumerate(headers, 1):
        if h and str(h).lower().strip().replace(" ","_") in aliases:
            return i
    return None

def detect_file_role(ws) -> str:
    """
    Fayl roli: 'shablon' yoki 'savol'
    Shablon = question ustuni bo'sh
    Savol   = question ustuni to'liq
    """
    headers = [ws.cell(1,c).value for c in range(1, ws.max_column+1)]
    q_col   = find_col(headers, QUESTION_ALIASES)
    if not q_col:
        return "savol"  # question ustuni yo'q → savol format boshqa
    # Birinchi 5 qatorni tekshirish
    filled = 0
    for r in range(2, min(7, ws.max_row+1)):
        val = ws.cell(r, q_col).value
        if val and str(val).strip():
            filled += 1
    return "savol" if filled >= 3 else "shablon"

def get_headers(ws) -> dict:
    """Barcha ustun nomlarini indeks bilan qaytaradi."""
    raw = [ws.cell(1,c).value for c in range(1, ws.max_column+1)]
    result = {
        "all":     raw,
        "topic":   find_col(raw, TOPIC_ALIASES),
        "question":find_col(raw, QUESTION_ALIASES),
        "opt_a":   find_col(raw, OPT_A_ALIASES),
        "opt_b":   find_col(raw, OPT_B_ALIASES),
        "opt_c":   find_col(raw, OPT_C_ALIASES),
        "opt_d":   find_col(raw, OPT_D_ALIASES),
        "correct": find_col(raw, CORRECT_ALIASES),
        "expl":    find_col(raw, EXPL_ALIASES),
        "qtype":   find_col(raw, QTYPE_ALIASES),
        "diff":    find_col(raw, DIFF_ALIASES),
        "image":   find_col(raw, IMAGE_ALIASES),
        "lang":    find_col(raw, LANG_ALIASES),
    }
    return result

def merge_excel(file1: bytes, file2: bytes) -> tuple[bytes, str]:
    """
    Ikkita Excel faylni birlashtiradi.
    Avtomatik aniqlaydi qaysi shablon, qaysi savol.
    Qaytadi: (natija_bytes, info_matn)
    """
    wb1 = openpyxl.load_workbook(BytesIO(file1), data_only=True)
    wb2 = openpyxl.load_workbook(BytesIO(file2), data_only=True)
    ws1 = wb1.active
    ws2 = wb2.active

    # Rol aniqlash
    role1 = detect_file_role(ws1)
    role2 = detect_file_role(ws2)

    if role1 == "shablon" and role2 == "savol":
        ws_sh, ws_sv = ws1, ws2
    elif role1 == "savol" and role2 == "shablon":
        ws_sh, ws_sv = ws2, ws1
    elif role1 == "savol" and role2 == "savol":
        # Ikkala ham savol — biri manba, biri qabul
        # Qaysi ko'proq to'liq bo'lsa savol
        ws_sv = ws1; ws_sh = ws2
    else:
        # Ikkala ham shablon — birinchisi asosiy
        ws_sh = ws1; ws_sv = ws2

    sh_h = get_headers(ws_sh)
    sv_h = get_headers(ws_sv)

    if not sh_h["topic"] or not sv_h["topic"]:
        return None, "❌ topic_code ustuni topilmadi!"

    # Savol faylidan topic bo'yicha qatorlarni yig'amiz (tartibda)
    sv_by_topic = defaultdict(list)
    for r in range(2, ws_sv.max_row+1):
        tc = ws_sv.cell(r, sv_h["topic"]).value
        if tc and str(tc).strip():
            sv_by_topic[str(tc).strip()].append(r)

    # Natija workbook — shablon strukturasini saqlaymiz
    out_wb = openpyxl.Workbook()
    out_ws = out_wb.active
    out_ws.title = "Birlashtirilgan"

    # Header shablon dan
    out_ws.append(sh_h["all"])

    tc_idx   = defaultdict(int)
    matched  = 0
    skipped  = 0
    total    = 0

    for sh_row in range(2, ws_sh.max_row+1):
        tc = ws_sh.cell(sh_row, sh_h["topic"]).value
        if not tc:
            continue
        tc = str(tc).strip()
        total += 1

        # Shablon qatorini to'liq o'qiymiz
        row_data = [ws_sh.cell(sh_row, c).value
                    for c in range(1, len(sh_h["all"])+1)]

        # Mos savol qatorini tartib bo'yicha olamiz
        sv_list = sv_by_topic.get(tc, [])
        idx     = tc_idx[tc]
        tc_idx[tc] += 1

        if idx < len(sv_list):
            sv_row = sv_list[idx]
            matched += 1

            def sv(key):
                col = sv_h.get(key)
                if col:
                    v = ws_sv.cell(sv_row, col).value
                    return v if v is not None else ""
                return ""

            def sh_col(key):
                return (sh_h.get(key) or 0) - 1  # 0-indexed

            # Savoldan olinadigan maydonlar
            fill_map = {
                "question": sv("question"),
                "opt_a":    sv("opt_a"),
                "opt_b":    sv("opt_b"),
                "opt_c":    sv("opt_c"),
                "opt_d":    sv("opt_d"),
                "correct":  sv("correct"),
                "expl":     sv("expl"),
            }
            # Ixtiyoriy — agar savol faylida bo'lsa yangilash
            if sv("diff"):    fill_map["diff"]  = sv("diff")
            if sv("qtype"):   fill_map["qtype"] = sv("qtype")
            if sv("image") and not (sh_h.get("image") and row_data[sh_col("image")]):
                fill_map["image"] = sv("image")

            # row_data ni yangilaymiz
            for key, val in fill_map.items():
                col_i = sh_col(key)
                if col_i >= 0 and col_i < len(row_data):
                    row_data[col_i] = val
        else:
            skipped += 1

        out_ws.append(row_data)

    # Ustun kengliklarini sozlash
    for col in range(1, len(sh_h["all"])+1):
        letter = openpyxl.utils.get_column_letter(col)
        max_w  = max(
            len(str(out_ws.cell(r, col).value or ""))
            for r in range(1, min(out_ws.max_row+1, 50))
        )
        out_ws.column_dimensions[letter].width = min(max(max_w+2, 12), 50)

    buf = BytesIO()
    out_wb.save(buf)
    buf.seek(0)

    info = (
        f"✅ Birlashtirildi!\n"
        f"📊 Jami: {total} qator\n"
        f"✅ Joylashtirildi: {matched} ta\n"
        f"⏭ O'tkazildi: {skipped} ta\n\n"
        f"📥 Faylni import qiling 👇"
    )
    return buf.read(), info


# ══════════════════════════════════════
# SHABLONNI DB DAN TO'LDIRISH
# ══════════════════════════════════════
def fill_from_db(template_bytes: bytes) -> tuple[bytes, str]:
    """
    Bo'sh shablonni DB dagi generated_tests dan to'ldiradi.
    Tartib: shablon qatori → mos topic_code savoli (navbatma-navbat).
    MUHIM: Shablon satr tartibi O'ZGARMAYDI.
    """
    import psycopg2, os
    from collections import defaultdict

    DATABASE_URL = os.getenv("DATABASE_URL","")

    wb_in = openpyxl.load_workbook(BytesIO(template_bytes), data_only=True)
    ws_in = wb_in.active
    h     = get_headers(ws_in)

    if not h["topic"]:
        return None, "❌ topic_code ustuni topilmadi!"

    # Shablondagi barcha topic_code larni tartibda yig'amiz
    tc_order = defaultdict(list)  # tc → [row_numbers]
    for r in range(2, ws_in.max_row+1):
        tc = ws_in.cell(r, h["topic"]).value
        if tc:
            tc_order[str(tc).strip()].append(r)

    all_tcs = list(tc_order.keys())
    if not all_tcs:
        return None, "❌ topic_code lar topilmadi!"

    # DB dan TARTIBDA savollarni olamiz
    conn = psycopg2.connect(DATABASE_URL); cur = conn.cursor()
    ph = ",".join(["%s"]*len(all_tcs))
    cur.execute(f"""
        SELECT topic_code, question, option_a, option_b, option_c, option_d,
               correct_answer, explanation, question_type, difficulty
        FROM generated_tests
        WHERE topic_code IN ({ph})
        ORDER BY topic_code, id
    """, all_tcs)
    rows = cur.fetchall()
    cur.close(); conn.close()

    # TC bo'yicha guruhlash (tartib saqlanadi)
    db_by_tc = defaultdict(list)
    for row in rows:
        db_by_tc[str(row[0])].append(row)

    # Yangi workbook — shablon strukturasini to'liq saqlaymiz
    wb_out = openpyxl.Workbook()
    ws_out = wb_out.active
    ws_out.title = "To'ldirilgan"
    ws_out.append(h["all"])  # Header o'zgarmaydi

    tc_idx  = defaultdict(int)
    filled  = 0
    empty   = 0

    def ci(key):
        """Column index (0-based) for row_data list."""
        col = h.get(key)
        return (col - 1) if col else -1

    for sh_r in range(2, ws_in.max_row+1):
        tc = ws_in.cell(sh_r, h["topic"]).value
        if not tc:
            continue
        tc = str(tc).strip()

        # Shablon qatorini aynan o'qiymiz (hech narsa o'zgarmaydi)
        row_data = [ws_in.cell(sh_r, c).value
                    for c in range(1, len(h["all"])+1)]

        # DB dan navbatdagi savolni olamiz
        db_list = db_by_tc.get(tc, [])
        idx     = tc_idx[tc]
        tc_idx[tc] += 1

        if idx < len(db_list):
            db_row = db_list[idx]
            # Faqat BO'SH joylarni to'ldirish
            def fill(key, val):
                i = ci(key)
                if i >= 0 and i < len(row_data) and not row_data[i]:
                    row_data[i] = val

            fill("question", db_row[1])
            fill("opt_a",    db_row[2])
            fill("opt_b",    db_row[3])
            fill("opt_c",    db_row[4])
            fill("opt_d",    db_row[5])
            fill("correct",  db_row[6])
            fill("expl",     db_row[7])
            fill("qtype",    db_row[8])
            fill("diff",     db_row[9])
            filled += 1
        else:
            empty += 1

        ws_out.append(row_data)

    # Natija
    buf = BytesIO(); wb_out.save(buf); buf.seek(0)

    has_q = sum(1 for v in db_by_tc.values() for _ in v)
    info  = (
        f"✅ Shablon to'ldirildi!\n"
        f"📊 Jami: {ws_in.max_row-1} qator\n"
        f"✅ To'ldirildi: {filled} ta\n"
        f"⬜ Bo'sh qoldi: {empty} ta"
        + (f"\n\n⚠️ {empty} ta qator uchun DB da savol yo'q." if empty > 0 else "")
        + f"\n\n📥 Faylni import qiling 👇"
    )
    return buf.read(), info
