"""
shablon_yaratish.py — Topik kod uchun shablon yaratish
Admin: Shablon yaratish → Topik kod uchun shablon / Import
"""
import os, io, psycopg2, openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from aiogram.types import (
    InlineKeyboardMarkup, InlineKeyboardButton, BufferedInputFile
)

DATABASE_URL = os.getenv("DATABASE_URL")
shablon_state = {}  # user_id -> {step, sinf, fan, mavzular}

def db(): return psycopg2.connect(DATABASE_URL)


# ─── ASOSIY MENYU ────────────────────────────────
async def show_shablon_menu(message, user_id):
    shablon_state[user_id] = {}
    kb = InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(
            text="📋 Topik kod uchun shablon",
            callback_data="sh_topik"
        )],
        [InlineKeyboardButton(
            text="📥 Import qilish",
            callback_data="sh_import"
        )],
    ])
    await message.answer(
        "📚 Shablon yaratish\n\nNimani xohlaysiz?",
        reply_markup=kb
    )


# ─── CALLBACK HANDLER ────────────────────────────
async def handle_shablon_callback(call, user_id):
    data = call.data

    if data == "sh_topik":
        shablon_state[user_id] = {"step": "sinf_fan"}
        await call.message.edit_text(
            "📋 Topik kod uchun shablon\n\n"
            "Sinf va fanni yozing:\n"
            "Masalan: <code>1 Ingliz tili</code>",
            parse_mode="HTML"
        )
        await call.answer()


    elif data == "sh_import":
        shablon_state[user_id] = {"step": "import_wait"}
        await call.message.edit_text(
            "📥 Import qilish\n\nTo'ldirilgan Excel faylni yuboring:"
        )
        await call.answer()

    elif data.startswith(("shb_cnt_","shb_tp_")) or data in ("sh_download","shb_reset"):
        # Multi-difficulty shablon sozlamalari
        # groups = [{"diff":"oson","type":"single_choice","count":5}, ...]
        st2    = shablon_state.get(user_id, {})
        groups = st2.setdefault("groups", [
            {"diff":"oson",    "type":"single_choice","count":0},
            {"diff":"orta",    "type":"single_choice","count":0},
            {"diff":"qiyin",   "type":"single_choice","count":0},
            {"diff":"murakkab","type":"single_choice","count":0},
        ])

        if data == "shb_reset":
            for g in groups: g["count"] = 0

        elif data.startswith("shb_cnt_"):
            # shb_cnt_oson_5, shb_cnt_orta_10 ...
            parts3 = data[8:].rsplit("_",1)
            diff3, cnt3 = parts3[0], int(parts3[1])
            for g in groups:
                if g["diff"] == diff3: g["count"] = cnt3

        elif data.startswith("shb_tp_"):
            # shb_tp_oson_choice | shb_tp_oson_write
            parts3 = data[7:].rsplit("_",1)
            diff3, tp3 = parts3[0], parts3[1]
            for g in groups:
                if g["diff"] == diff3:
                    g["type"] = "single_choice" if tp3=="choice" else "write_answer"

        elif data == "sh_download":
            sinf     = st2.get("sinf","1")
            fan      = st2.get("fan","Fan")
            mavzular = st2.get("mavzular",[])
            total    = sum(g["count"] for g in groups)
            if total == 0:
                await call.answer("❌ Hech qanday savol tanlanmagan!", show_alert=True); return

            buf = await _create_shablon_multi(sinf, fan, mavzular, groups)
            fname = f"shablon_{sinf}sinf_{fan[:10].replace(' ','_')}.xlsx"
            await call.message.answer_document(
                BufferedInputFile(buf.read(), filename=fname),
                caption=(
                    f"✅ Shablon tayyor!\n🏫 {sinf}-sinf | {fan}\n"
                    f"📚 {len(mavzular)} ta mavzu × {total} ta = {len(mavzular)*total} qator\n"
                    + "\n".join(f"  {g['diff']}: {g['count']} ta ({g['type']})" for g in groups if g['count']>0)
                )
            )
            shablon_state.pop(user_id, None)
            await call.answer("✅"); return

        await call.answer("✅")

        def _sh_multi_kb(groups):
            DIFFS = [("oson","🟢"),("orta","🟡"),("qiyin","🔴"),("murakkab","⚫")]
            rows = []
            for g in groups:
                d = g["diff"]; icon = next(em for df,em in DIFFS if df==d)
                cnt = g["count"]; tp = g["type"]
                rows.append([
                    InlineKeyboardButton(text=f"{icon} {d.capitalize()}", callback_data="noop"),
                    InlineKeyboardButton(text="0" if cnt==0 else f"✅{cnt}", callback_data=f"shb_cnt_{d}_0"),
                    InlineKeyboardButton(text="5"  if cnt!=5  else "✅5",  callback_data=f"shb_cnt_{d}_5"),
                    InlineKeyboardButton(text="10" if cnt!=10 else "✅10", callback_data=f"shb_cnt_{d}_10"),
                    InlineKeyboardButton(text="15" if cnt!=15 else "✅15", callback_data=f"shb_cnt_{d}_15"),
                    InlineKeyboardButton(text="20" if cnt!=20 else "✅20", callback_data=f"shb_cnt_{d}_20"),
                ])
                rows.append([
                    InlineKeyboardButton(
                        text=f"✅ 🔘 Tugmali" if tp=="single_choice" else "🔘 Tugmali",
                        callback_data=f"shb_tp_{d}_choice"
                    ),
                    InlineKeyboardButton(
                        text=f"✅ ✍️ Yozuvli" if tp=="write_answer" else "✍️ Yozuvli",
                        callback_data=f"shb_tp_{d}_write"
                    ),
                ])
            total = sum(g["count"] for g in groups)
            rows.append([InlineKeyboardButton(text=f"🔄 Tozalash", callback_data="shb_reset")])
            rows.append([InlineKeyboardButton(
                text=f"📥 Shablon yuklab olish (jami: {total} ta)",
                callback_data="sh_download"
            )])
            return InlineKeyboardMarkup(inline_keyboard=rows)

        try: await call.message.edit_reply_markup(reply_markup=_sh_multi_kb(groups))
        except: pass



# ─── XABAR HANDLER ───────────────────────────────
async def handle_shablon_message(message, user_id):
    state = shablon_state.get(user_id, {})
    step = state.get("step")

    # Sinf va fan
    if step == "sinf_fan":
        parts = message.text.strip().split(None, 1)
        if len(parts) < 2:
            await message.answer(
                "❌ Noto'g'ri format!\n"
                "Masalan: <code>1 Ingliz tili</code>",
                parse_mode="HTML"
            )
            return
        sinf, fan = parts[0], parts[1]
        shablon_state[user_id] = {
            "step": "mavzular",
            "sinf": sinf,
            "fan": fan
        }
        await message.answer(
            f"✅ {sinf}-sinf, {fan}\n\n"
            "Endi mavzularni yozing:\n"
            "<code>1 / Colours\n"
            "1 / Numbers\n"
            "2 / Animals\n"
            "2 / Family\n"
            "...</code>\n\n"
            "Chorak / Mavzu formatida, har biri yangi qatorda:",
            parse_mode="HTML"
        )

    # Mavzular
    elif step == "mavzular":
        sinf = state.get("sinf", "1")
        fan = state.get("fan", "Fan")
        text = message.text.strip()

        mavzular = []
        for line in text.splitlines():
            line = line.strip()
            if not line:
                continue
            # Format: "1 / Mavzu nomi" yoki "1/ Mavzu" yoki "1 Mavzu"
            if "/" in line:
                parts = line.split("/", 1)
                chorak = parts[0].strip().lstrip("0123456789. ").strip()
                # Chorakni raqam sifatida ol
                chorak_raqam = ""
                for ch in parts[0].strip():
                    if ch.isdigit():
                        chorak_raqam += ch
                mavzu = parts[1].strip()
            else:
                parts = line.split(None, 1)
                chorak_raqam = parts[0].strip() if parts else "1"
                mavzu = parts[1].strip() if len(parts) > 1 else line

            if mavzu and chorak_raqam:
                mavzular.append((chorak_raqam, mavzu))

        if not mavzular:
            await message.answer("❌ Mavzular topilmadi! Qayta yozing.")
            return

        # Multi-difficulty sozlamalar
        shablon_state[user_id]["mavzular"] = mavzular
        shablon_state[user_id]["step"] = "settings"
        groups = shablon_state[user_id].setdefault("groups",[
            {"diff":"oson",    "type":"single_choice","count":0},
            {"diff":"orta",    "type":"single_choice","count":0},
            {"diff":"qiyin",   "type":"single_choice","count":0},
            {"diff":"murakkab","type":"single_choice","count":0},
        ])
        DIFFS = [("oson","🟢"),("orta","🟡"),("qiyin","🔴"),("murakkab","⚫")]
        rows = []
        for g in groups:
            d = g["diff"]; icon = next(em for df,em in DIFFS if df==d)
            cnt = g["count"]; tp = g["type"]
            rows.append([
                InlineKeyboardButton(text=f"{icon} {d.capitalize()}", callback_data="noop"),
                InlineKeyboardButton(text="✅0" if cnt==0  else "0",  callback_data=f"shb_cnt_{d}_0"),
                InlineKeyboardButton(text="✅5" if cnt==5  else "5",  callback_data=f"shb_cnt_{d}_5"),
                InlineKeyboardButton(text="✅10" if cnt==10 else "10", callback_data=f"shb_cnt_{d}_10"),
                InlineKeyboardButton(text="✅15" if cnt==15 else "15", callback_data=f"shb_cnt_{d}_15"),
                InlineKeyboardButton(text="✅20" if cnt==20 else "20", callback_data=f"shb_cnt_{d}_20"),
            ])
            rows.append([
                InlineKeyboardButton(
                    text="✅ 🔘 Tugmali" if tp=="single_choice" else "🔘 Tugmali",
                    callback_data=f"shb_tp_{d}_choice"
                ),
                InlineKeyboardButton(
                    text="✅ ✍️ Yozuvli" if tp=="write_answer" else "✍️ Yozuvli",
                    callback_data=f"shb_tp_{d}_write"
                ),
            ])
        rows.append([InlineKeyboardButton(text="📥 Shablon yuklab olish (0 ta)", callback_data="sh_download")])
        await message.answer(
            f"⚙️ Sozlamalar — har qiyinlikdan nechta?\n📚 {len(mavzular)} ta mavzu\nSon tanlang:",
            reply_markup=InlineKeyboardMarkup(inline_keyboard=rows)
        )

    # Import
    elif step == "import_wait":
        await message.answer(
            "❌ Fayl yuboring! Matn emas."
        )


async def handle_shablon_document(message, user_id, bot):
    """Import: Excel faylni DBga saqlash"""
    state = shablon_state.get(user_id, {})
    if state.get("step") != "import_wait":
        return False

    doc = message.document
    if not doc.file_name.endswith('.xlsx'):
        await message.answer("❌ Faqat .xlsx fayl qabul qilinadi!")
        return True

    file = await bot.get_file(doc.file_id)
    file_bytes = await bot.download_file(file.file_path)

    try:
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes.read()))
        ws = wb.active

        conn = db()
        cur = conn.cursor()
        added = 0
        skipped = 0

        for r in range(2, ws.max_row + 1):
            sinf = ws.cell(r, 1).value
            fan = ws.cell(r, 2).value
            chorak = ws.cell(r, 3).value
            bob = ws.cell(r, 4).value
            bolim = ws.cell(r, 5).value
            mavzu = ws.cell(r, 6).value
            kichik = ws.cell(r, 7).value

            if not sinf or not mavzu:
                continue

            # Topic code yaratish
            # Mavjud oxirgi kodni topamiz
            cur.execute("""
                SELECT topic_code FROM dts_tree
                WHERE grade=%s AND subject_name=%s
                ORDER BY topic_code DESC LIMIT 1
            """, (str(sinf), str(fan) if fan else ""))

            row = cur.fetchone()
            if row:
                # Oxirgi raqamni oshiramiz
                last = row[0]
                parts = last.rsplit('-', 1)
                new_num = str(int(parts[1]) + 1).zfill(3)
                topic_code = f"{parts[0]}-{new_num}"
            else:
                topic_code = f"{sinf}-01-{chorak or 1}-01-01-01-001"

            try:
                cur.execute("""
                    INSERT INTO dts_tree
                    (topic_code, grade, subject_name, quarter,
                     bob_name, bolim_name, mavzu_name, kichik_name,
                     is_deleted)
                    VALUES (%s,%s,%s,%s,%s,%s,%s,%s,FALSE)
                    ON CONFLICT (topic_code) DO NOTHING
                """, (
                    topic_code, str(sinf),
                    str(fan) if fan else "",
                    str(chorak) if chorak else "1",
                    str(bob) if bob else "",
                    str(bolim) if bolim else "",
                    str(mavzu) if mavzu else "",
                    str(kichik) if kichik else "",
                ))
                added += 1
            except Exception:
                skipped += 1

        conn.commit()
        cur.close()
        conn.close()

        await message.answer(
            f"✅ Import tugadi!\n"
            f"➕ Qo'shildi: {added} ta\n"
            f"⏭ O'tkazildi: {skipped} ta"
        )

    except Exception as ex:
        await message.answer(f"❌ Xato: {ex}")

    shablon_state.pop(user_id, None)
    return True


# ─── EXCEL SHABLON YARATISH ───────────────────────
async def _create_shablon(sinf, fan, mavzular, diff='aralash', qtype='single_choice', count=2):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "DTS_SHABLON"

    headers = ["Sinf", "Fan", "Chorak", "Bob", "Bo'lim", "Mavzu", "Kichik mavzu"]
    header_colors = ["4472C4", "4472C4", "4472C4", "70AD47", "70AD47", "ED7D31", "ED7D31"]

    for col, (h, color) in enumerate(zip(headers, header_colors), 1):
        cell = ws.cell(1, col, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor=color)
        cell.alignment = Alignment(horizontal="center")

    # Har mavzudan 2 qator
    chorak_colors = {
        "1": "DEEAF1", "2": "E2EFDA", "3": "FFF2CC", "4": "FCE4D6"
    }

    row_num = 2
    for chorak, mavzu in mavzular:
        color = chorak_colors.get(str(chorak), "F2F2F2")
        for _ in range(count):  # count ta qator
            ws.cell(row_num, 1, value=sinf)
            ws.cell(row_num, 2, value=fan)
            ws.cell(row_num, 3, value=chorak)
            # Bob, Bo'lim — bo'sh
            ws.cell(row_num, 6, value=mavzu)
            # Kichik mavzu — bo'sh

            for col in range(1, 8):
                ws.cell(row_num, col).fill = PatternFill("solid", fgColor=color)
                ws.cell(row_num, col).alignment = Alignment(horizontal="left")

            row_num += 1

    # Ustun kengliklari
    for col, width in zip(range(1, 8), [8, 18, 8, 25, 25, 30, 30]):
        ws.column_dimensions[ws.cell(1, col).column_letter].width = width

    # Izoh varag'i
    ws2 = wb.create_sheet("IZOH")
    ws2.cell(1, 1, value="📋 TO'LDIRISH QO'LLANMASI")
    ws2.cell(1, 1).font = Font(bold=True, size=14)
    izohlar = [
        (3, "Sinf", "O'zgartirmang — avtomatik to'ldirilgan"),
        (4, "Fan", "O'zgartirmang — avtomatik to'ldirilgan"),
        (5, "Chorak", "O'zgartirmang — avtomatik to'ldirilgan"),
        (6, "Bob", "To'ldiring: masalan 'Chapter 1. Getting acquainted'"),
        (7, "Bo'lim", "To'ldiring: masalan 'Unit 1. Greetings'"),
        (8, "Mavzu", "O'zgartirmang — mavzu nomi avtomatik"),
        (9, "Kichik mavzu", "To'ldiring: mavzuning kichik qismi"),
    ]
    for r, ustun, izoh in izohlar:
        ws2.cell(r, 1, value=ustun).font = Font(bold=True)
        ws2.cell(r, 2, value=izoh)
    ws2.column_dimensions['A'].width = 15
    ws2.column_dimensions['B'].width = 50

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf

async def _create_test_shablon_multi(sinf, fan, mavzular, groups):
    """Test savollari shabloni: har group uchun alohida qatorlar."""
    import io, psycopg2, os
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    DATABASE_URL = os.getenv("DATABASE_URL")
    # topic_code → kichik_name mapping
    conn = psycopg2.connect(DATABASE_URL); cur = conn.cursor()
    cur.execute("""
        SELECT DISTINCT topic_code, kichik_name
        FROM dts_tree WHERE is_deleted=FALSE
    """)
    tc_map = {r[0]: r[1] for r in cur.fetchall()}
    cur.close(); conn.close()

    wb = Workbook()
    ws = wb.active
    ws.title = "TEST_SHABLON"

    headers = [
        "topic_code","mavzu_nomi","difficulty","question_type",
        "question","option_a","option_b","option_c","option_d",
        "correct_answer","explanation","image_url","language","time_limit"
    ]
    colors = {
        "oson": "E2EFDA",
        "orta": "FFF2CC",
        "qiyin": "FCE4D6",
        "murakkab": "F2CEEF",
    }
    header_colors = ["4472C4"]*4 + ["DEEAF1"]*10

    for col, (h, hc) in enumerate(zip(headers, header_colors), 1):
        cell = ws.cell(1, col, h)
        cell.font = Font(bold=True, color="FFFFFF" if col<=4 else "000000")
        cell.fill = PatternFill("solid", fgColor="4472C4" if col<=4 else "BDD7EE")
        cell.alignment = Alignment(horizontal="center")

    row_num = 2
    for topic_code, chorak, mavzu_name in mavzular:
        for g in groups:
            if g["count"] == 0: continue
            color = colors.get(g["diff"], "F2F2F2")
            kichik = tc_map.get(topic_code, mavzu_name or topic_code)
            for _ in range(g["count"]):
                ws.cell(row_num, 1, topic_code)
                ws.cell(row_num, 2, kichik)
                ws.cell(row_num, 3, g["diff"])
                ws.cell(row_num, 4, g["type"])
                # 5-14: bo'sh (savol, variantlar...)
                ws.cell(row_num, 13, "uz")  # language default
                ws.cell(row_num, 14, "0")   # time_limit default

                for col in range(1, len(headers)+1):
                    ws.cell(row_num, col).fill = PatternFill("solid", fgColor=color)
                    ws.cell(row_num, col).alignment = Alignment(wrap_text=True)

                row_num += 1

    # Kengliklar
    widths = [30, 25, 10, 15, 50, 20, 20, 20, 20, 15, 40, 20, 8, 10]
    for col, w in enumerate(widths, 1):
        ws.column_dimensions[ws.cell(1, col).column_letter].width = w

    # Izoh varaq
    ws2 = wb.create_sheet("IZOH")
    ws2.cell(1,1,"📋 TO'LDIRISH QO'LLANMASI").font = Font(bold=True, size=13)
    notes = [
        (3,"topic_code","O'zgartirmang"),
        (4,"difficulty","O'zgartirmang: oson/orta/qiyin/murakkab"),
        (5,"question_type","single_choice YOKI write_answer"),
        (6,"question","Savol matnini yozing"),
        (7,"option_a","A variant"),
        (8,"option_b","B variant"),
        (9,"option_c","C variant"),
        (10,"option_d","D variant"),
        (11,"correct_answer","To'g'ri javob: A, B, C yoki D"),
        (12,"explanation","Izoh (ixtiyoriy)"),
        (13,"image_url","Rasm kodi: topic_code-1 ... topic_code-N"),
    ]
    for r, col_name, note in notes:
        ws2.cell(r,1,col_name).font = Font(bold=True)
        ws2.cell(r,2,note)
    ws2.column_dimensions['A'].width = 16
    ws2.column_dimensions['B'].width = 50

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


async def _create_shablon_multi(sinf, fan, mavzular, groups):
    """_create_test_shablon_multi ga yo'naltirish."""
    # mavzular = [(chorak, mavzu)] formatida — topic_code kerak
    # DTS dan topic_code olamiz
    import psycopg2, os
    DATABASE_URL = os.getenv("DATABASE_URL")
    conn = psycopg2.connect(DATABASE_URL); cur = conn.cursor()
    # mavzu_name bo'yicha topic_code topish
    new_mavzular = []
    for chorak, mavzu_name in mavzular:
        cur.execute("""
            SELECT DISTINCT topic_code FROM dts_tree
            WHERE grade=%s AND subject_name=%s AND mavzu_name=%s
              AND is_deleted=FALSE LIMIT 1
        """, (sinf, fan, mavzu_name))
        r = cur.fetchone()
        tc = r[0] if r else mavzu_name
        new_mavzular.append((tc, chorak, mavzu_name))
    cur.close(); conn.close()
    return await _create_test_shablon_multi(sinf, fan, new_mavzular, groups)
