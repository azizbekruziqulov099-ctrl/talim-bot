import os
import io
import pandas as pd
import psycopg2
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from aiogram import F
from aiogram.types import (
    Message, CallbackQuery,
    InlineKeyboardMarkup, InlineKeyboardButton, FSInputFile
)
from aiogram.fsm.context import FSMContext
from aiogram.fsm.state import State, StatesGroup
from loader import dp, bot

DATABASE_URL = os.getenv("DATABASE_URL")
ADMINS = [int(x) for x in os.getenv("ADMINS", "401251407").split(",")]
PAGE_SIZE = 10

class LessonAdminState(StatesGroup):
    waiting_excel = State()

def db():
    return psycopg2.connect(DATABASE_URL)

def kb(buttons):
    return InlineKeyboardMarkup(inline_keyboard=buttons)

def back_btn(cb):
    return [InlineKeyboardButton(text="⬅️ Ortga", callback_data=cb)]

def home_btn():
    return [InlineKeyboardButton(text="🏠 Admin menyu", callback_data="la_home")]

def nav_row(items, page, prefix):
    row = []
    total = (len(items)-1)//PAGE_SIZE+1
    if page > 0:
        row.append(InlineKeyboardButton(text="◀️", callback_data=f"{prefix}|p|{page-1}"))
    row.append(InlineKeyboardButton(text=f"{page+1}/{total}", callback_data="noop"))
    if (page+1)*PAGE_SIZE < len(items):
        row.append(InlineKeyboardButton(text="▶️", callback_data=f"{prefix}|p|{page+1}"))
    return row

# ─── ENTRY ───
@dp.message(F.text == "📝 Dars boshqaruvi")
async def la_entry(message: Message):
    if message.from_user.id not in ADMINS:
        return
    await la_show_grades(message)

async def la_show_grades(target, page=0):
    conn = db(); cur = conn.cursor()
    cur.execute("SELECT DISTINCT grade FROM dts_tree WHERE is_deleted=FALSE ORDER BY grade")
    items = cur.fetchall()
    cur.close(); conn.close()

    chunk   = items[page*PAGE_SIZE:(page+1)*PAGE_SIZE]
    buttons = [[InlineKeyboardButton(text=f"🏫 {r[0]}-sinf", callback_data=f"la_g|{r[0]}")] for r in chunk]
    nav = nav_row(items, page, "la_gs")
    if nav: buttons.append(nav)
    buttons.append(home_btn())

    text = "📝 Dars boshqaruvi\nSinf tanlang:"
    if hasattr(target, 'message'):
        await target.message.edit_text(text, reply_markup=kb(buttons))
    else:
        await target.answer(text, reply_markup=kb(buttons))

@dp.callback_query(F.data.startswith("la_gs|p|"))
async def la_grades_page(call: CallbackQuery):
    page = int(call.data.split("|")[2])
    await call.answer()
    await la_show_grades(call, page)

# ─── SINF → FAN ───
@dp.callback_query(F.data.startswith("la_g|"))
async def la_grade(call: CallbackQuery):
    if call.from_user.id not in ADMINS: return
    await call.answer()
    grade = call.data.split("|")[1]

    conn = db(); cur = conn.cursor()
    cur.execute("""
        SELECT DISTINCT subject_code, subject_name FROM dts_tree
        WHERE grade=%s AND is_deleted=FALSE ORDER BY subject_name
    """, (grade,))
    items = cur.fetchall()
    cur.close(); conn.close()

    buttons = [[InlineKeyboardButton(
        text=f"📘 {name}",
        callback_data=f"la_s|{grade}|{code}"
    )] for code, name in items]
    buttons.append(back_btn("la_back_grades"))
    buttons.append(home_btn())

    await call.message.edit_text(f"📝 {grade}-sinf\nFan tanlang:", reply_markup=kb(buttons))

# ─── FAN → CHORAK ───
@dp.callback_query(F.data.startswith("la_s|"))
async def la_subject(call: CallbackQuery):
    if call.from_user.id not in ADMINS: return
    await call.answer()
    _, grade, scode = call.data.split("|")

    conn = db(); cur = conn.cursor()
    cur.execute("""
        SELECT DISTINCT quarter FROM dts_tree
        WHERE grade=%s AND subject_code=%s AND is_deleted=FALSE ORDER BY quarter
    """, (grade, scode))
    items = cur.fetchall()
    cur.execute("SELECT DISTINCT subject_name FROM dts_tree WHERE grade=%s AND subject_code=%s LIMIT 1", (grade, scode))
    sname = (cur.fetchone() or [scode])[0]
    cur.close(); conn.close()

    buttons = [[InlineKeyboardButton(
        text=f"🗓 {r[0]}-chorak",
        callback_data=f"la_q|{grade}|{scode}|{r[0]}"
    )] for r in items]
    buttons.append(back_btn(f"la_g|{grade}"))
    buttons.append(home_btn())

    await call.message.edit_text(f"📘 {grade}-sinf | {sname}", reply_markup=kb(buttons))

# ─── CHORAK → BOB ───
@dp.callback_query(F.data.startswith("la_q|"))
async def la_quarter(call: CallbackQuery):
    if call.from_user.id not in ADMINS: return
    await call.answer()
    _, grade, scode, quarter = call.data.split("|")

    conn = db(); cur = conn.cursor()
    cur.execute("""
        SELECT DISTINCT bob_code, bob_name FROM dts_tree
        WHERE grade=%s AND subject_code=%s AND quarter=%s AND is_deleted=FALSE
        ORDER BY bob_code
    """, (grade, scode, quarter))
    items = cur.fetchall()
    cur.close(); conn.close()

    buttons = [[InlineKeyboardButton(
        text=f"📖 {name}",
        callback_data=f"la_b|{grade}|{scode}|{quarter}|{code}"
    )] for code, name in items]
    buttons.append(back_btn(f"la_s|{grade}|{scode}"))
    buttons.append(home_btn())

    await call.message.edit_text(f"🗓 {quarter}-chorak", reply_markup=kb(buttons))

# ─── BOB → BO'LIM ───
@dp.callback_query(F.data.startswith("la_b|"))
async def la_bob(call: CallbackQuery):
    if call.from_user.id not in ADMINS: return
    await call.answer()
    _, grade, scode, quarter, bcode = call.data.split("|")

    conn = db(); cur = conn.cursor()
    cur.execute("""
        SELECT DISTINCT bolim_code, bolim_name FROM dts_tree
        WHERE grade=%s AND subject_code=%s AND quarter=%s AND bob_code=%s AND is_deleted=FALSE
        ORDER BY bolim_code
    """, (grade, scode, quarter, bcode))
    items = cur.fetchall()
    cur.execute("SELECT DISTINCT bob_name FROM dts_tree WHERE grade=%s AND bob_code=%s LIMIT 1", (grade, bcode))
    bname = (cur.fetchone() or [bcode])[0]
    cur.close(); conn.close()

    buttons = [[InlineKeyboardButton(
        text=f"📑 {name}",
        callback_data=f"la_bl|{grade}|{scode}|{quarter}|{bcode}|{code}"
    )] for code, name in items]
    buttons.append(back_btn(f"la_q|{grade}|{scode}|{quarter}"))
    buttons.append(home_btn())

    await call.message.edit_text(f"📖 {bname}", reply_markup=kb(buttons))

# ─── BO'LIM → MAVZULAR ───
@dp.callback_query(F.data.startswith("la_bl|"))
async def la_bolim(call: CallbackQuery):
    if call.from_user.id not in ADMINS: return
    await call.answer()
    _, grade, scode, quarter, bcode, blcode = call.data.split("|")

    conn = db(); cur = conn.cursor()
    cur.execute("""
        SELECT DISTINCT mavzu_code, mavzu_name FROM dts_tree
        WHERE grade=%s AND subject_code=%s AND quarter=%s
          AND bob_code=%s AND bolim_code=%s AND is_deleted=FALSE
        ORDER BY mavzu_code
    """, (grade, scode, quarter, bcode, blcode))
    mavzular = cur.fetchall()

    # Har mavzu uchun dars bor/yo'q
    cur.execute("""
        SELECT DISTINCT t.mavzu_code, COUNT(DISTINCT tl.topic_code)
        FROM dts_tree t
        LEFT JOIN teacher_lessons tl ON tl.topic_code = t.topic_code
        WHERE t.grade=%s AND t.subject_code=%s AND t.quarter=%s
          AND t.bob_code=%s AND t.bolim_code=%s AND t.is_deleted=FALSE
        GROUP BY t.mavzu_code
    """, (grade, scode, quarter, bcode, blcode))
    stats = {r[0]: r[1] for r in cur.fetchall()}

    cur.execute("SELECT DISTINCT bolim_name FROM dts_tree WHERE grade=%s AND bolim_code=%s LIMIT 1", (grade, blcode))
    blname = (cur.fetchone() or [blcode])[0]
    cur.close(); conn.close()

    buttons = []
    for mcode, mname in mavzular:
        filled = stats.get(mcode, 0)
        icon   = "✅" if filled > 0 else "❌"
        buttons.append([InlineKeyboardButton(
            text=f"{icon} {mname}",
            callback_data=f"la_m|{grade}|{scode}|{quarter}|{bcode}|{blcode}|{mcode}"
        )])

    total_all  = len(mavzular)
    filled_all = sum(1 for v in stats.values() if v > 0)

    buttons.append(back_btn(f"la_b|{grade}|{scode}|{quarter}|{bcode}"))
    buttons.append(home_btn())

    await call.message.edit_text(
        f"📑 {blname}\nMavzular: {filled_all}/{total_all}\n✅ Dars bor  ❌ Bo'sh",
        reply_markup=kb(buttons)
    )

# ─── MAVZU → KICHIK MAVZULAR ───
@dp.callback_query(F.data.startswith("la_m|"))
async def la_mavzu(call: CallbackQuery):
    if call.from_user.id not in ADMINS: return
    await call.answer()
    _, grade, scode, quarter, bcode, blcode, mcode = call.data.split("|")

    conn = db(); cur = conn.cursor()
    cur.execute("""
        SELECT DISTINCT kichik_code, kichik_name, topic_code,
               subject_name, mavzu_name
        FROM dts_tree
        WHERE grade=%s AND subject_code=%s AND quarter=%s
          AND bob_code=%s AND bolim_code=%s AND mavzu_code=%s
          AND is_deleted=FALSE
        ORDER BY kichik_code
    """, (grade, scode, quarter, bcode, blcode, mcode))
    rows = cur.fetchall()

    if not rows:
        await call.message.edit_text("❌ Kichik mavzular topilmadi")
        cur.close(); conn.close()
        return

    topic_codes = [r[2] for r in rows]
    cur.execute("SELECT topic_code FROM teacher_lessons WHERE topic_code = ANY(%s)", (topic_codes,))
    existing = {r[0] for r in cur.fetchall()}
    cur.close(); conn.close()

    sname  = rows[0][3]
    mname  = rows[0][4]
    filled = len([r for r in rows if r[2] in existing])
    total  = len(rows)

    buttons = []
    for kcode, kname, topic_code, *_ in rows:
        icon = "✅" if topic_code in existing else "❌"
        if topic_code in existing:
            buttons.append([InlineKeyboardButton(
                text=f"✅ {kname}",
                callback_data=f"la_lesson|{topic_code}"
            )])
        else:
            buttons.append([InlineKeyboardButton(
                text=f"❌ {kname}",
                callback_data="noop"
            )])

    from storage import user_state as us
    if not isinstance(us.get(call.from_user.id), dict):
        us[call.from_user.id] = {}
    us[call.from_user.id]["la_meta"] = f"{grade}|{scode}|{quarter}|{bcode}|{blcode}|{mcode}"

    buttons.append([InlineKeyboardButton(text="📥 Shablon yuklab ol", callback_data="la_tmpl")])
    buttons.append([InlineKeyboardButton(text="📤 Import qilish",     callback_data="la_imp")])
    buttons.append(back_btn(f"la_bl|{grade}|{scode}|{quarter}|{bcode}|{blcode}"))
    buttons.append(home_btn())

    await call.message.edit_text(
        f"📝 {grade}-sinf | {sname}\n"
        f"📖 {mname}\n"
        f"━━━━━━━━━━━━━━\n"
        f"Kichik mavzular: {filled}/{total}\n"
        f"✅ Dars bor  ❌ Bo'sh",
        reply_markup=kb(buttons)
    )

# ─── SHABLON ───
@dp.callback_query(F.data == "la_tmpl")
async def la_template(call: CallbackQuery):
    if call.from_user.id not in ADMINS: return
    await call.answer("📥 Tayyor bo'lmoqda...")

    from storage import user_state as us
    meta = us.get(call.from_user.id, {}).get("la_meta", "")
    if not meta:
        await call.message.answer("❌ Mavzu tanlanmagan, qaytadan tanlang")
        return

    _, grade, scode, quarter, bcode, blcode, mcode = ("x|" + meta).split("|")

    conn = db(); cur = conn.cursor()
    cur.execute("""
        SELECT DISTINCT kichik_code, kichik_name, topic_code, subject_name, mavzu_name
        FROM dts_tree
        WHERE grade=%s AND subject_code=%s AND quarter=%s
          AND bob_code=%s AND bolim_code=%s AND mavzu_code=%s
          AND is_deleted=FALSE
        ORDER BY kichik_code
    """, (grade, scode, quarter, bcode, blcode, mcode))
    rows = cur.fetchall()

    topic_codes = [r[2] for r in rows]
    cur.execute("""
        SELECT topic_code, intro, part_1, part_2, part_3, part_4,
               simple_1, simple_2, example_1, example_2,
               exercise_1, exercise_2, summary
        FROM teacher_lessons WHERE topic_code = ANY(%s)
    """, (topic_codes,))
    lessons = {r[0]: r for r in cur.fetchall()}
    cur.close(); conn.close()

    sname = rows[0][3] if rows else ""
    mname = rows[0][4] if rows else ""

    buf = make_excel(rows, lessons, grade, sname, mname)

    from aiogram.types import BufferedInputFile
    await call.message.answer_document(
        BufferedInputFile(buf.read(), filename=f"lesson_{grade}.xlsx"),
        caption=(
            f"📥 Shablon\n🏫 {grade}-sinf | {sname}\n📖 {mname}\n\n"
            f"✅ Yashil — dars bor\n⬜ Oq — to'ldiring\n\n"
            f"To'ldirib yuborish uchun 👇"
        ),
        reply_markup=kb([[InlineKeyboardButton(
            text="📤 Import qilish",
            callback_data="la_imp"
        )]])
    )

# ─── IMPORT ───
@dp.callback_query(F.data == "la_imp")
async def la_import_prompt(call: CallbackQuery, state: FSMContext):
    if call.from_user.id not in ADMINS: return
    await call.answer()
    await state.set_state(LessonAdminState.waiting_excel)
    await call.message.answer("📤 To'ldirilgan Excel faylini yuboring:")

@dp.message(LessonAdminState.waiting_excel)
async def la_import_excel(message: Message, state: FSMContext):
    if not message.document:
        await message.answer("❌ Excel fayl yuboring")
        return

    file = await bot.get_file(message.document.file_id)
    buf  = io.BytesIO()
    await bot.download_file(file.file_path, buf)
    buf.seek(0)

    try:
        df = pd.read_excel(buf, sheet_name="DARSLAR", dtype=str, header=1)
    except Exception as e:
        await message.answer(f"❌ Excel o'qib bo'lmadi:\n{e}")
        await state.clear()
        return

    conn = db(); cur = conn.cursor()
    added = updated = skipped = 0

    def v(row, col):
        val = row.get(col, "")
        return "" if str(val) in ("nan","None","") else str(val).strip()

    for _, row in df.iterrows():
        tc    = v(row, "topic_code")
        intro = v(row, "intro")
        if not tc or not intro:
            skipped += 1
            continue

        fields = (intro, v(row,"part_1"), v(row,"part_2"), v(row,"part_3"), v(row,"part_4"),
                  v(row,"simple_1"), v(row,"simple_2"), v(row,"example_1"), v(row,"example_2"),
                  v(row,"exercise_1"), v(row,"exercise_2"), v(row,"summary"),
                  v(row,"simple_3"), v(row,"simple_4"))

        cur.execute("SELECT id FROM teacher_lessons WHERE topic_code=%s", (tc,))
        if cur.fetchone():
            cur.execute("""
                UPDATE teacher_lessons SET
                    intro=%s,part_1=%s,part_2=%s,part_3=%s,part_4=%s,
                    simple_1=%s,simple_2=%s,example_1=%s,example_2=%s,
                    exercise_1=%s,exercise_2=%s,summary=%s,
                    simple_3=%s,simple_4=%s
                WHERE topic_code=%s
            """, (*fields, tc))
            updated += 1
        else:
            cur.execute("""
                INSERT INTO teacher_lessons
                (topic_code,intro,part_1,part_2,part_3,part_4,
                 simple_1,simple_2,example_1,example_2,exercise_1,exercise_2,summary,
                 simple_3,simple_4)
                VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
            """, (tc, *fields))
            added += 1

    conn.commit(); cur.close(); conn.close()
    await state.clear()
    await message.answer(f"✅ Import tugadi!\n➕ Yangi: {added}\n✏️ Yangilandi: {updated}\n⏭ O'tkazildi: {skipped}")

@dp.callback_query(F.data.startswith("la_lesson|"))
async def la_lesson_detail(call: CallbackQuery):
    if call.from_user.id not in ADMINS: return
    await call.answer()

    topic_code = call.data.replace("la_lesson|", "")

    conn = db(); cur = conn.cursor()
    cur.execute("""
        SELECT topic_code, intro, part_1, part_2, part_3, part_4, summary
        FROM teacher_lessons WHERE topic_code=%s
    """, (topic_code,))
    lesson = cur.fetchone()
    cur.close(); conn.close()

    if not lesson:
        await call.message.answer("❌ Dars topilmadi")
        return

    intro_preview = str(lesson[1] or "")[:100] + "..." if lesson[1] and len(str(lesson[1])) > 100 else str(lesson[1] or "")

    await call.message.answer(
        f"📝 Dars: `{topic_code}`\n\n"
        f"📖 Kirish:\n{intro_preview}\n\n"
        f"Nima qilmoqchisiz?",
        parse_mode="Markdown",
        reply_markup=kb([
            [InlineKeyboardButton(text="👁 O'quvchi ko'rinishi", callback_data=f"la_preview|{topic_code}")],
            [InlineKeyboardButton(text="✏️ Tahrirlash (shablon)", callback_data=f"la_edit|{topic_code}")],
            [InlineKeyboardButton(text="🗑 O'chirish", callback_data=f"la_delete|{topic_code}")],
            [InlineKeyboardButton(text="⬅️ Ortga", callback_data="noop")]
        ])
    )


@dp.callback_query(F.data.startswith("la_preview|"))
async def la_preview_lesson(call: CallbackQuery):
    """Darsni o'quvchi ko'rinishida ko'rsatish"""
    if call.from_user.id not in ADMINS: return
    await call.answer()

    topic_code = call.data.replace("la_preview|", "")

    conn = db(); cur = conn.cursor()
    cur.execute("""
        SELECT intro, part_1, part_2, part_3, part_4,
               simple_1, simple_2, simple_3, simple_4,
               example_1, example_2, exercise_1, exercise_2, summary
        FROM teacher_lessons WHERE topic_code=%s
    """, (topic_code,))
    row = cur.fetchone()

    # DTS daraxtidan sinf, fan, mavzu nomi
    cur.execute("""
        SELECT grade, subject_name, mavzu_name, kichik_name
        FROM dts_tree WHERE topic_code=%s LIMIT 1
    """, (topic_code,))
    tree = cur.fetchone()
    cur.close(); conn.close()

    if not row:
        await call.message.answer("❌ Dars topilmadi")
        return

    grade   = tree[0] if tree else "?"
    subj    = tree[1] if tree else ""
    mavzu   = tree[2] if tree else ""
    kichik  = tree[3] if tree else topic_code

    parts_map = {
        "📖 Kirish":          row[0],
        "📘 1-qism":          row[1],
        "📗 2-qism":          row[2],
        "📙 3-qism":          row[3],
        "📕 4-qism":          row[4],
        "💡 Sodda 1":         row[5],
        "💡 Sodda 2":         row[6],
        "💡 Sodda 3":         row[7],
        "💡 Sodda 4":         row[8],
        "📌 Misol 1":         row[9],
        "📌 Misol 2":         row[10],
        "✏️ Mashq 1":         row[11],
        "✏️ Mashq 2":         row[12],
        "📝 Xulosa":          row[13],
    }

    header = (
        f"👤 O'quvchi ko'rinishi (preview)\n"
        f"━━━━━━━━━━━━━━\n"
        f"🎓 {grade}-sinf | 📚 {subj}\n"
        f"📍 {mavzu} → {kichik}\n"
        f"━━━━━━━━━━━━━━"
    )
    await call.message.answer(header)

    step = 1
    total = sum(1 for v in parts_map.values() if v and str(v).strip())
    for label, content in parts_map.items():
        if not content or not str(content).strip():
            continue
        text = (
            f"━━━━━━━━━━━━━━\n"
            f"{label} | {step}/{total}\n\n"
            f"{content}"
        )
        await call.message.answer(text[:4096])
        step += 1

    await call.message.answer(
        "✅ Dars oxiri — o'quvchi shu ko'rinishda ko'radi.",
        reply_markup=kb([[
            InlineKeyboardButton(text="⬅️ Ortga", callback_data=f"la_lesson|{topic_code}")
        ]])
    )


@dp.callback_query(F.data.startswith("la_edit|"))
async def la_edit_lesson(call: CallbackQuery):
    if call.from_user.id not in ADMINS: return
    await call.answer("📥 Shablon tayyorlanmoqda...")

    topic_code = call.data.replace("la_edit|", "")

    conn = db(); cur = conn.cursor()
    cur.execute("""
        SELECT kichik_code, kichik_name, topic_code, subject_name, mavzu_name
        FROM dts_tree WHERE topic_code=%s LIMIT 1
    """, (topic_code,))
    tree_row = cur.fetchone()

    cur.execute("""
        SELECT topic_code, intro, part_1, part_2, part_3, part_4,
               simple_1, simple_2, example_1, example_2,
               exercise_1, exercise_2, summary
        FROM teacher_lessons WHERE topic_code=%s
    """, (topic_code,))
    lesson = cur.fetchone()
    cur.close(); conn.close()

    if not tree_row or not lesson:
        await call.message.answer("❌ Ma'lumot topilmadi")
        return

    grade        = "—"
    subject_name = tree_row[3] or ""
    mavzu_name   = tree_row[4] or topic_code

    rows    = [tree_row]
    lessons = {topic_code: lesson}

    buf = make_excel(rows, lessons, grade, subject_name, mavzu_name)

    from aiogram.types import BufferedInputFile
    await call.message.answer_document(
        BufferedInputFile(buf.read(), filename=f"edit_{topic_code}.xlsx"),
        caption=(
            f"✏️ Tahrirlash shabloni\n\n"
            f"🔑 {topic_code}\n\n"
            f"To'ldirib import qiling 👇"
        ),
        reply_markup=kb([[InlineKeyboardButton(
            text="📤 Import qilish",
            callback_data="la_imp"
        )]])
    )


@dp.callback_query(F.data.startswith("la_delete|"))
async def la_delete_confirm(call: CallbackQuery):
    if call.from_user.id not in ADMINS: return
    await call.answer()

    topic_code = call.data.replace("la_delete|", "")

    await call.message.answer(
        f"⚠️ Haqiqatan ham o'chirasizmi?\n\n"
        f"🔑 `{topic_code}`",
        parse_mode="Markdown",
        reply_markup=kb([
            [
                InlineKeyboardButton(text="🗑 Ha, o'chir", callback_data=f"la_delete_yes|{topic_code}"),
                InlineKeyboardButton(text="❌ Bekor", callback_data="noop")
            ]
        ])
    )


@dp.callback_query(F.data.startswith("la_delete_yes|"))
async def la_delete_yes(call: CallbackQuery):
    if call.from_user.id not in ADMINS: return
    await call.answer()

    topic_code = call.data.replace("la_delete_yes|", "")

    conn = db(); cur = conn.cursor()
    cur.execute("DELETE FROM teacher_lessons WHERE topic_code=%s", (topic_code,))
    conn.commit()
    cur.close(); conn.close()

    await call.message.edit_text(f"✅ O'chirildi: `{topic_code}`", parse_mode="Markdown")
@dp.callback_query(F.data == "la_back_grades")
async def la_back_grades(call: CallbackQuery):
    await call.answer()
    await la_show_grades(call)

@dp.callback_query(F.data == "la_home")
async def la_home(call: CallbackQuery):
    await call.answer()
    await call.message.delete()

@dp.callback_query(F.data == "noop")
async def noop(call: CallbackQuery):
    await call.answer()

# ─── EXCEL SHABLON ───
def make_excel(rows, lessons, grade, subject_name, mavzu_name):
    import io
    wb = Workbook()
    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    HEADER = "1F4E79"; EXISTS = "E2EFDA"; EMPTY = "FFFFFF"; FIXED = "D6E4F0"

    columns = [
        ("topic_code",16),("grade",8),("subject",18),("mavzu",22),
        ("intro",45),("part_1",45),("part_2",45),("part_3",45),("part_4",45),
        ("simple_1",38),("simple_2",38),("simple_3",38),("simple_4",38),
        ("example_1",38),("example_2",38),
        ("exercise_1",38),("exercise_2",38),("summary",45),
    ]
    editable = {"intro","part_1","part_2","part_3","part_4",
                "simple_1","simple_2","simple_3","simple_4",
                "example_1","example_2","exercise_1","exercise_2","summary"}

    ws = wb.active; ws.title = "DARSLAR"
    ws.merge_cells(f"A1:{get_column_letter(len(columns))}1")
    t = ws["A1"]
    t.value = f"📝 {grade}-sinf | {subject_name} | {mavzu_name}"
    t.font = Font(bold=True, color="FFFFFF", name="Arial", size=12)
    t.fill = PatternFill("solid", start_color=HEADER)
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    for col, (name, width) in enumerate(columns, 1):
        c = ws.cell(row=2, column=col, value=name)
        c.font = Font(bold=True, color="FFFFFF", name="Arial", size=9)
        c.fill = PatternFill("solid", start_color=HEADER)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = border
        ws.column_dimensions[get_column_letter(col)].width = width
    ws.row_dimensions[2].height = 25

    col_map = {name: i+1 for i, (name, _) in enumerate(columns)}

    for row_idx, (kcode, kname, topic_code, *_) in enumerate(rows, 3):
        lesson = lessons.get(topic_code)
        lm = {}
        if lesson:
            keys = ["topic_code","intro","part_1","part_2","part_3","part_4",
                    "simple_1","simple_2","example_1","example_2","exercise_1","exercise_2","summary"]
            lm = {k: (lesson[i] or "") for i, k in enumerate(keys)}

        fixed = {"topic_code": topic_code, "grade": grade, "subject": subject_name, "mavzu": kname}

        for cname, cidx in col_map.items():
            if cname in fixed:
                val = fixed[cname]; bg = FIXED
            elif cname in editable:
                val = lm.get(cname, ""); bg = EXISTS if lesson else EMPTY
            else:
                val = ""; bg = EMPTY
            c = ws.cell(row=row_idx, column=cidx, value=val)
            c.font = Font(name="Arial", size=10)
            c.fill = PatternFill("solid", start_color=bg)
            c.alignment = Alignment(vertical="top", wrap_text=True)
            c.border = border
        ws.row_dimensions[row_idx].height = 80

    # Namuna sheet
    ws2 = wb.create_sheet("NAMUNA")
    ws2.merge_cells(f"A1:{get_column_letter(len(columns))}1")
    n = ws2["A1"]
    n.value = "✅ NAMUNA"; n.font = Font(bold=True, color="FFFFFF", name="Arial", size=12)
    n.fill = PatternFill("solid", start_color="375623")
    n.alignment = Alignment(horizontal="center", vertical="center")
    for col, (name, width) in enumerate(columns, 1):
        c = ws2.cell(row=2, column=col, value=name)
        c.font = Font(bold=True, color="FFFFFF", name="Arial", size=9)
        c.fill = PatternFill("solid", start_color=HEADER)
        c.border = border
        ws2.column_dimensions[get_column_letter(col)].width = width
    sample = ["TEST_001",grade,subject_name,"Namuna mavzu",
              "🌟 Kirish matni...","📖 1-qism...","📌 2-qism...","","",
              "💡 Sodda izoh...","","🎬 Misol...","","","",
              "✅ Xulosa: 1)... 2)..."]
    for col, val in enumerate(sample, 1):
        c = ws2.cell(row=3, column=col, value=val)
        c.font = Font(name="Arial", size=10)
        c.fill = PatternFill("solid", start_color="FFFFFF")
        c.alignment = Alignment(vertical="top", wrap_text=True)
        c.border = border
    ws2.row_dimensions[3].height = 100

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
