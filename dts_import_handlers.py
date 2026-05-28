from aiogram.types import Message
from difflib import SequenceMatcher
from openpyxl import load_workbook
from openpyxl import Workbook
from Talim import dp
import psycopg2
import os
import re
from aiogram import F
from aiogram.fsm.context import (
    FSMContext
)
from difflib import SequenceMatcher
from Talim import bot
from aiogram.types import FSInputFile
from aiogram.types import (

    Message,
    CallbackQuery,

    InlineKeyboardMarkup,
    InlineKeyboardButton

)
from aiogram.fsm.state import (
    State,
    StatesGroup
)
class DTSImportState(

    StatesGroup

):

    waiting_excel = State()

DATABASE_URL = os.getenv("DATABASE_URL")

dts_import_cache = {}

def normalize_text(text):

    if text is None:
        return ""

    text = str(text)

    text = text.lower()

    text = text.replace("ʻ", "'")
    text = text.replace("`", "'")
    text = text.replace("ʼ", "'")

    text = text.replace("–", "-")
    text = text.replace("—", "-")

    text = text.replace("_", " ")

    text = re.sub(
        r"\s+",
        " ",
        text
    )

    return text.strip()

def normalize_grade(

    grade

):

    grade = normalize_text(
        grade
    )

    grade = grade.replace(
        "-sinf",
        ""
    )

    grade = grade.replace(
        "sinf",
        ""
    )

    grade = grade.replace(
        " ",
        ""
    )

    return grade

def validate_grade(

    grade

):

    if len(grade) < 1:
        return False

    return True

def get_grade_code(

    grade

):

    grade = normalize_grade(
        grade
    )

    if not validate_grade(
        grade
    ):

        raise Exception(
            "Noto‘g‘ri level"
        )

    return grade


def get_subject_code(

    cur,

    grade,

    subject_name

):

    grade = normalize_text(
        grade
    )

    subject_name = normalize_text(
        subject_name
    ).upper()

    cur.execute("""
    SELECT subject_code
    FROM dts_tree
    WHERE grade=%s
    AND subject_name=%s
    LIMIT 1
    """, (
        grade,
        subject_name
    ))

    row = cur.fetchone()

    if row:

        return row[0]

    cur.execute("""
    SELECT MAX(
        CAST(subject_code AS INTEGER)
    )
    FROM dts_tree
    WHERE grade=%s
    """, (
        grade,
    ))

    last_code = cur.fetchone()[0]

    next_code = str(
        (last_code or 0) + 1
    ).zfill(2)

    return next_code
def normalize_quarter(

    quarter

):

    quarter = normalize_text(
        quarter
    )

    quarter = quarter.replace(
        "chorak",
        ""
    )

    quarter = quarter.replace(
        "-",
        ""
    )

    quarter = quarter.replace(
        " ",
        ""
    )

    return quarter

def validate_quarter(

    quarter

):

    if len(quarter) < 1:
        return False

    return True

def get_quarter_code(

    quarter

):

    quarter = normalize_quarter(
        quarter
    )

    if not validate_quarter(
        quarter
    ):

        raise Exception(
            "Noto‘g‘ri chorak"
        )

    return quarter

def get_bob_code(

    cur,

    grade,

    subject_code,

    quarter_code,

    bob_name

):

    bob_name = normalize_text(
        bob_name
    )

    cur.execute("""
    SELECT bob_code
    FROM dts_tree
    WHERE grade=%s
    AND subject_code=%s
    AND quarter=%s
    AND bob_name=%s
    LIMIT 1
    """, (
        grade,
        subject_code,
        quarter_code,
        bob_name
    ))

    row = cur.fetchone()

    if row:

        return row[0]

    cur.execute("""
    SELECT MAX(
        CAST(bob_code AS INTEGER)
    )
    FROM dts_tree
    WHERE grade=%s
    AND subject_code=%s
    AND quarter=%s
    """, (
        grade,
        subject_code,
        quarter_code
    ))

    last_code = cur.fetchone()[0]

    next_code = str(
        (last_code or 0) + 1
    ).zfill(2)

    return next_code

def get_bolim_code(

    cur,

    grade,

    subject_code,

    quarter_code,

    bob_code,

    bolim_name

):

    grade = get_grade_code(
        grade
    )

    quarter_code = get_quarter_code(
        quarter_code
    )

    bolim_name = normalize_text(
        bolim_name
    )

    cur.execute("""

    SELECT bolim_code
    FROM dts_tree
    WHERE grade=%s
    AND subject_code=%s
    AND quarter=%s
    AND bob_code=%s
    AND bolim_name=%s
    LIMIT 1

    """, (

        grade,
        subject_code,
        quarter_code,
        bob_code,
        bolim_name

    ))

    row = cur.fetchone()

    if row:

        return row[0]

    cur.execute("""

    SELECT MAX(
        CAST(bolim_code AS INTEGER)
    )
    FROM dts_tree
    WHERE grade=%s
    AND subject_code=%s
    AND quarter=%s
    AND bob_code=%s

    """, (

        grade,
        subject_code,
        quarter_code,
        bob_code

    ))

    last_code = cur.fetchone()[0]

    if last_code is None:

        return "01"

    new_code = int(last_code) + 1

    return str(new_code).zfill(2)

def get_mavzu_code(

    cur,

    grade,

    subject_code,

    quarter_code,

    bob_code,

    bolim_code,

    mavzu_name

):

    grade = get_grade_code(
        grade
    )

    quarter_code = get_quarter_code(
        quarter_code
    )

    mavzu_name = normalize_text(
        mavzu_name
    )

    cur.execute("""
    SELECT mavzu_code
    FROM dts_tree
    WHERE grade=%s
    AND subject_code=%s
    AND quarter=%s
    AND bob_code=%s
    AND bolim_code=%s
    AND mavzu_name=%s
    LIMIT 1
    """, (
        grade,
        subject_code,
        quarter_code,
        bob_code,
        bolim_code,
        mavzu_name
    ))

    row = cur.fetchone()

    if row:

        return row[0]

    cur.execute("""
    SELECT MAX(
        CAST(mavzu_code AS INTEGER)
    )
    FROM dts_tree
    WHERE grade=%s
    AND subject_code=%s
    AND quarter=%s
    AND bob_code=%s
    AND bolim_code=%s
    """, (
        grade,
        subject_code,
        quarter_code,
        bob_code,
        bolim_code
    ))

    last_code = cur.fetchone()[0]

    next_code = str(
        (last_code or 0) + 1
    ).zfill(2)

    return next_code

def get_kichik_code(

    cur,

    grade,

    subject_code,

    quarter_code,

    bob_code,

    bolim_code,

    mavzu_code,

    kichik_name

):

    grade = get_grade_code(
        grade
    )

    quarter_code = get_quarter_code(
        quarter_code
    )

    kichik_name = normalize_text(
        kichik_name
    )

    cur.execute("""
    SELECT kichik_code
    FROM dts_tree
    WHERE grade=%s
    AND subject_code=%s
    AND quarter=%s
    AND bob_code=%s
    AND bolim_code=%s
    AND mavzu_code=%s
    AND kichik_name=%s
    LIMIT 1
    """, (
        grade,
        subject_code,
        quarter_code,
        bob_code,
        bolim_code,
        mavzu_code,
        kichik_name
    ))

    row = cur.fetchone()

    if row:

        return row[0]

    cur.execute("""
    SELECT MAX(
        CAST(kichik_code AS INTEGER)
    )
    FROM dts_tree
    WHERE grade=%s
    AND subject_code=%s
    AND quarter=%s
    AND bob_code=%s
    AND bolim_code=%s
    AND mavzu_code=%s
    """, (
        grade,
        subject_code,
        quarter_code,
        bob_code,
        bolim_code,
        mavzu_code
    ))

    last_code = cur.fetchone()[0]

    next_code = str(
        (last_code or 0) + 1
    ).zfill(3)

    return next_code

def build_topic_code(

    grade,

    subject_code,

    quarter_code,

    bob_code,

    bolim_code,

    mavzu_code,

    kichik_code

):

    grade = get_grade_code(
        grade
    )

    quarter_code = get_quarter_code(
        quarter_code
    )

    return (

        f"{grade}-"

        f"{subject_code}-"

        f"{quarter_code}-"

        f"{bob_code}-"

        f"{bolim_code}-"

        f"{mavzu_code}-"

        f"{kichik_code}"

    )

def insert_row(

    cur,
    row

):

    grade = get_grade_code(
        row[0]
    )

    subject_name = normalize_text(
        row[1]
    ).upper()

    quarter_code = get_quarter_code(
        row[2]
    )

    bob_name = normalize_text(
        row[3]
    )

    bolim_name = normalize_text(
        row[4]
    )

    mavzu_name = normalize_text(
        row[5]
    )

    kichik_name = normalize_text(
        row[6]
    )

    subject_code = get_subject_code(

        cur,

        grade,

        subject_name

    )

    bob_code = get_bob_code(

        cur,

        grade,

        subject_code,

        quarter_code,

        bob_name

    )

    bolim_code = get_bolim_code(

        cur,

        grade,

        subject_code,

        quarter_code,

        bob_code,

        bolim_name

    )

    mavzu_code = get_mavzu_code(

        cur,

        grade,

        subject_code,

        quarter_code,

        bob_code,

        bolim_code,

        mavzu_name

    )

    kichik_code = get_kichik_code(

        cur,

        grade,

        subject_code,

        quarter_code,

        bob_code,

        bolim_code,

        mavzu_code,

        kichik_name

    )

    topic_code = build_topic_code(

        grade,

        subject_code,

        quarter_code,

        bob_code,

        bolim_code,

        mavzu_code,

        kichik_code

    )

    cur.execute("""
    SELECT 1
    FROM dts_tree
    WHERE topic_code=%s
    LIMIT 1
    """, (
        topic_code,
    ))

    exists = cur.fetchone()

    if exists:

        return {

            "status": "exists",

            "topic_code": topic_code

        }

    cur.execute("""
    INSERT INTO dts_tree (

        topic_code,

        grade,

        subject_code,
        subject_name,

        quarter,

        bob_code,
        bob_name,

        bolim_code,
        bolim_name,

        mavzu_code,
        mavzu_name,

        kichik_code,
        kichik_name

    )
    VALUES (

        %s,

        %s,

        %s,
        %s,

        %s,

        %s,
        %s,

        %s,
        %s,

        %s,
        %s,

        %s,
        %s

    )
    """, (

        topic_code,

        grade,

        subject_code,
        subject_name,

        quarter_code,

        bob_code,
        bob_name,

        bolim_code,
        bolim_name,

        mavzu_code,
        mavzu_name,

        kichik_code,
        kichik_name

    ))

    return {

        "status": "inserted",

        "topic_code": topic_code

    }
def analyze_import(

    cur,
    rows

):

    error_rows = []
    duplicate_rows = []
    existing_rows = []
    valid_rows = []

    seen = set()

    for i, row in enumerate(

        rows,
        start=2

    ):

        try:

            if len(row) < 7:

                error_rows.append({

                    "row_no": i,

                    "reason": (
                        "Ustunlar soni yetarli emas"
                    )

                })

                continue

            grade = get_grade_code(
                row[0]
            )

            subject_name = normalize_text(
                row[1]
            ).upper()

            quarter_code = get_quarter_code(
                row[2]
            )

            bob_name = normalize_text(
                row[3]
            )

            bolim_name = normalize_text(
                row[4]
            )

            mavzu_name = normalize_text(
                row[5]
            )

            kichik_name = normalize_text(
                row[6]
            )

            if not all([

                grade,

                subject_name,

                quarter_code,

                bob_name,

                bolim_name,

                mavzu_name,

                kichik_name

            ]):

                error_rows.append({

                    "row_no": i,

                    "reason": (
                        "Bo‘sh ustun bor"
                    )

                })

                continue

            subject_code = get_subject_code(

                cur,

                grade,

                subject_name

            )

            bob_code = get_bob_code(

                cur,

                grade,

                subject_code,

                quarter_code,

                bob_name

            )

            bolim_code = get_bolim_code(

                cur,

                grade,

                subject_code,

                quarter_code,

                bob_code,

                bolim_name

            )

            mavzu_code = get_mavzu_code(

                cur,

                grade,

                subject_code,

                quarter_code,

                bob_code,

                bolim_code,

                mavzu_name

            )

            kichik_code = get_kichik_code(

                cur,

                grade,

                subject_code,

                quarter_code,

                bob_code,

                bolim_code,

                mavzu_code,

                kichik_name

            )

            topic_code = build_topic_code(

                grade,

                subject_code,

                quarter_code,

                bob_code,

                bolim_code,

                mavzu_code,

                kichik_code

            )

            tree_key = (

                grade,

                subject_name,

                quarter_code,

                bob_name,

                bolim_name,

                mavzu_name,

                kichik_name

            )

            if tree_key in seen:

                duplicate_rows.append({

                    "row_no": i,

                    "reason": (
                        "Excel ichida takroriy"
                    )

                })

                continue

            seen.add(tree_key)

            cur.execute("""
            SELECT 1
            FROM dts_tree
            WHERE topic_code=%s
            LIMIT 1
            """, (
                topic_code,
            ))

            exists = cur.fetchone()

            if exists:

                existing_rows.append({

                    "row_no": i,

                    "topic_code": topic_code,

                    "reason": (
                        "Bazada mavjud"
                    )

                })

                continue

            valid_rows.append({

                "row_no": i,

                "row": row,

                "topic_code": topic_code

            })

        except Exception as e:

            error_rows.append({

                "row_no": i,

                "reason": str(e)

            })

    return {

        "error_rows": error_rows,

        "duplicate_rows": duplicate_rows,
        
        "existing_rows": existing_rows,

        "valid_rows": valid_rows

    }

def confirm_import(

    cur,
    valid_rows

):

    inserted_count = 0

    inserted_rows = []
    failed_rows = []

    for item in valid_rows:

        try:

            row = item["row"]
            row_no = item.get("row_no")
            topic_code = item.get("topic_code")

            result = insert_row(
                cur,
                row
            )

            if result.get("status") == "inserted":

                inserted_count += 1

                inserted_rows.append({

                    "row_no": row_no,
                    "topic_code": topic_code

                })

            else:

                failed_rows.append({

                    "row_no": row_no,
                    "topic_code": topic_code,
                    "reason": result.get(
                        "reason",
                        "Insert bajarilmadi"
                    )

                })

        except Exception as e:

            failed_rows.append({

                "row_no": item.get("row_no"),
                "topic_code": item.get("topic_code"),
                "reason": str(e)

            })

    return {

        "inserted_count": inserted_count,

        "inserted_rows": inserted_rows,

        "failed_rows": failed_rows

    }

async def dts_navigator(

    call: CallbackQuery,

    page=0

):

    conn = psycopg2.connect(
        DATABASE_URL
    )

    cur = conn.cursor()

    cur.execute("""
    SELECT DISTINCT grade
    FROM dts_tree
    WHERE is_deleted=FALSE
    ORDER BY grade
    """)

    grades = cur.fetchall()

    page_size = 10

    start = page * page_size

    end = start + page_size

    current_items = grades[
        start:end
    ]

    buttons = []

    for (grade,) in current_items:

        buttons.append([

            InlineKeyboardButton(

                text=f"🏫 {grade}-sinf",

                callback_data=(
                    f"dts_grade_{grade}"
                )

            )

        ])

    nav_buttons = []

    if page > 0:

        nav_buttons.append(

            InlineKeyboardButton(

                text="⬅️",

                callback_data=(
                    f"dts_nav_{page - 1}"
                )

            )

        )

    total_pages = (
        len(grades) - 1
    ) // page_size + 1

    nav_buttons.append(

        InlineKeyboardButton(

            text=(
                f"{page + 1}/{total_pages}"
            ),

            callback_data="ignore"

        )

    )

    if end < len(grades):

        nav_buttons.append(

            InlineKeyboardButton(

                text="➡️",

                callback_data=(
                    f"dts_nav_{page + 1}"
                )

            )

        )

    buttons.append(nav_buttons)

    kb = InlineKeyboardMarkup(
        inline_keyboard=buttons
    )

    await call.message.edit_text(

        "📚 DTS Navigator",

        reply_markup=kb

    )

async def dts_grade(

    call: CallbackQuery

):

    grade = call.data.replace(
        "dts_grade_",
        ""
    )

    conn = psycopg2.connect(
        DATABASE_URL
    )

    cur = conn.cursor()

    cur.execute("""
    SELECT DISTINCT
        subject_code,
        subject_name
    FROM dts_tree
    WHERE grade=%s
    AND is_deleted=FALSE
    ORDER BY subject_code
    """, (
        grade,
    ))

    subjects = cur.fetchall()

    buttons = []

    for code, name in subjects:

        buttons.append([

            InlineKeyboardButton(

                text=f"📚 {name}",

                callback_data=(
                    f"dts_subject_"
                    f"{grade}_"
                    f"{code}"
                )

            )

        ])

    buttons.append([

        InlineKeyboardButton(

            text="⬅️ Orqaga",

            callback_data="dts_navigator"

        )

    ])

    kb = InlineKeyboardMarkup(
        inline_keyboard=buttons
    )

    await call.message.edit_text(

        f"🏫 {grade}-sinf fanlari",

        reply_markup=kb

    )

async def dts_subject(

    call: CallbackQuery

):
    (
    _,
    _,
    grade,
    subject_code 
    )= call.data.split("_")

    conn = psycopg2.connect(
        DATABASE_URL
    )

    cur = conn.cursor()

    cur.execute("""
    SELECT DISTINCT
        quarter
    FROM dts_tree
    WHERE grade=%s
    AND subject_code=%s
    AND is_deleted=FALSE
    ORDER BY quarter
    """, (
        grade,
        subject_code
    ))

    quarters = cur.fetchall()

    cur.execute("""
    SELECT subject_name
    FROM dts_tree
    WHERE subject_code=%s
    LIMIT 1
    """, (
        subject_code,
    ))

    subject_name = cur.fetchone()[0]

    buttons = []

    for (quarter,) in quarters:

        buttons.append([

            InlineKeyboardButton(

                text=f"🗓 {quarter}",

                callback_data=(

                    f"dts_quarter_"
                    f"{grade}_"
                    f"{subject_code}_"
                    f"{quarter}"

                )

            )

        ])

    buttons.append([

        InlineKeyboardButton(

            text="⬅️ Orqaga",

            callback_data=(
                f"dts_grade_{grade}"
            )

        )

    ])

    kb = InlineKeyboardMarkup(
        inline_keyboard=buttons
    )

    await call.message.edit_text(

        f"📚 {subject_name}",

        reply_markup=kb

    )
async def dts_quarter(

    call: CallbackQuery

):

    (
        _,
        _,
        grade,
        subject_code,
        quarter

    ) = call.data.split("_")

    conn = psycopg2.connect(
        DATABASE_URL
    )

    cur = conn.cursor()

    cur.execute("""
    SELECT DISTINCT
        bob_code,
        bob_name
    FROM dts_tree
    WHERE grade=%s
    AND subject_code=%s
    AND quarter=%s
    AND is_deleted=FALSE
    ORDER BY bob_code
    """, (
        grade,
        subject_code,
        quarter
    ))

    rows = cur.fetchall()

    buttons = []

    for code, name in rows:

        buttons.append([

            InlineKeyboardButton(

                text=f"📖 {name}",

                callback_data=(

                    f"dts_bob_"
                    f"{grade}_"
                    f"{subject_code}_"
                    f"{quarter}_"
                    f"{code}"

                )

            )

        ])

    buttons.append([

        InlineKeyboardButton(

            text="⬅️ Orqaga",

            callback_data=(

                f"dts_subject_"
                f"{grade}_"
                f"{subject_code}"

            )

        )

    ])

    kb = InlineKeyboardMarkup(
        inline_keyboard=buttons
    )

    await call.message.edit_text(

        f"🗓 {quarter}",

        reply_markup=kb

    )

async def dts_bob(

    call: CallbackQuery

):

    (
        _,
        _,
        grade,
        subject_code,
        quarter,
        bob_code

    ) = call.data.split("_")

    conn = psycopg2.connect(
        DATABASE_URL
    )

    cur = conn.cursor()

    cur.execute("""
    SELECT DISTINCT
        bolim_code,
        bolim_name
    FROM dts_tree
    WHERE grade=%s
    AND subject_code=%s
    AND quarter=%s
    AND bob_code=%s
    AND is_deleted=FALSE
    ORDER BY bolim_code
    """, (
        grade,
        subject_code,
        quarter,
        bob_code
    ))

    rows = cur.fetchall()

    buttons = []

    for code, name in rows:

        buttons.append([

            InlineKeyboardButton(

                text=f"📑 {name}",

                callback_data=(

                    f"dts_bolim_"
                    f"{grade}_"
                    f"{subject_code}_"
                    f"{quarter}_"
                    f"{bob_code}_"
                    f"{code}"

                )

            )

        ])

    buttons.append([

        InlineKeyboardButton(

            text="⬅️ Orqaga",

            callback_data=(

                f"dts_quarter_"
                f"{grade}_"
                f"{subject_code}_"
                f"{quarter}"

            )

        )

    ])

    kb = InlineKeyboardMarkup(
        inline_keyboard=buttons
    )

    await call.message.edit_text(

        "📖 Boblar",

        reply_markup=kb

    )

async def dts_bolim(

    call: CallbackQuery

):

    (
        _,
        _,
        grade,
        subject_code,
        quarter,
        bob_code,
        bolim_code

    ) = call.data.split("_")

    conn = psycopg2.connect(
        DATABASE_URL
    )

    cur = conn.cursor()

    cur.execute("""
    SELECT DISTINCT
        mavzu_code,
        mavzu_name
    FROM dts_tree
    WHERE grade=%s
    AND subject_code=%s
    AND quarter=%s
    AND bob_code=%s
    AND bolim_code=%s
    AND is_deleted=FALSE
    ORDER BY mavzu_code
    """, (
        grade,
        subject_code,
        quarter,
        bob_code,
        bolim_code
    ))

    rows = cur.fetchall()

    buttons = []

    for code, name in rows:

        buttons.append([

            InlineKeyboardButton(

                text=f"📝 {name}",

                callback_data=(

                    f"dts_mavzu_"
                    f"{grade}_"
                    f"{subject_code}_"
                    f"{quarter}_"
                    f"{bob_code}_"
                    f"{bolim_code}_"
                    f"{code}"

                )

            )

        ])

    buttons.append([

        InlineKeyboardButton(

            text="⬅️ Orqaga",

            callback_data=(

                f"dts_bob_"
                f"{grade}_"
                f"{subject_code}_"
                f"{quarter}_"
                f"{bob_code}"

            )

        )

    ])

    kb = InlineKeyboardMarkup(
        inline_keyboard=buttons
    )

    await call.message.edit_text(

        "📑 Bo‘limlar",

        reply_markup=kb

    )

async def dts_mavzu(

    call: CallbackQuery

):

    (
        _,
        _,
        grade,
        subject_code,
        quarter,
        bob_code,
        bolim_code,
        mavzu_code

    ) = call.data.split("_")

    conn = psycopg2.connect(
        DATABASE_URL
    )

    cur = conn.cursor()

    cur.execute("""
    SELECT DISTINCT
        kichik_code,
        kichik_name
    FROM dts_tree
    WHERE grade=%s
    AND subject_code=%s
    AND quarter=%s
    AND bob_code=%s
    AND bolim_code=%s
    AND mavzu_code=%s
    AND is_deleted=FALSE
    ORDER BY kichik_code
    """, (
        grade,
        subject_code,
        quarter,
        bob_code,
        bolim_code,
        mavzu_code
    ))

    rows = cur.fetchall()

    buttons = []

    for code, name in rows:

        buttons.append([

            InlineKeyboardButton(

                text=f"🔹 {name}",

                callback_data=(

                    f"dts_small_"
                    f"{grade}_"
                    f"{subject_code}_"
                    f"{quarter}_"
                    f"{bob_code}_"
                    f"{bolim_code}_"
                    f"{mavzu_code}_"
                    f"{code}"

                )

            )

        ])

    buttons.append([

        InlineKeyboardButton(

            text="⬅️ Orqaga",

            callback_data=(

                f"dts_bolim_"
                f"{grade}_"
                f"{subject_code}_"
                f"{quarter}_"
                f"{bob_code}_"
                f"{bolim_code}"

            )

        )

    ])

    kb = InlineKeyboardMarkup(
        inline_keyboard=buttons
    )

    await call.message.edit_text(

        "📝 Mavzular",

        reply_markup=kb

    )

async def dts_small(

    call: CallbackQuery

):

    (
        _,
        _,
        grade,
        subject_code,
        quarter,
        bob_code,
        bolim_code,
        mavzu_code,
        kichik_code

    ) = call.data.split("_")

    conn = psycopg2.connect(
        DATABASE_URL
    )

    cur = conn.cursor()

    cur.execute("""
    SELECT

        topic_code,

        subject_name,

        bob_name,

        bolim_name,

        mavzu_name,

        kichik_name

    FROM dts_tree
    WHERE grade=%s
    AND subject_code=%s
    AND quarter=%s
    AND bob_code=%s
    AND bolim_code=%s
    AND mavzu_code=%s
    AND kichik_code=%s
    AND is_deleted=FALSE
    LIMIT 1
    """, (
        grade,
        subject_code,
        quarter,
        bob_code,
        bolim_code,
        mavzu_code,
        kichik_code
    ))

    row = cur.fetchone()

    if not row:

        await call.answer(
            "Topilmadi"
        )

        return

    (
        topic_code,

        subject_name,

        bob_name,

        bolim_name,

        mavzu_name,

        kichik_name

    ) = row

    text = f"""

📚 DTS Ma’lumotlari

━━━━━━━━━━━━━━━

🏫 Sinf:
{grade}

📖 Fan:
{subject_name}

🗓 Chorak:
{quarter}

📘 Bob:
{bob_name}

📑 Bo‘lim:
{bolim_name}

📝 Mavzu:
{mavzu_name}

🔹 Kichik mavzu:
{kichik_name}

━━━━━━━━━━━━━━━

🆔 Topic Code:
{topic_code}

"""

    kb = InlineKeyboardMarkup(
        inline_keyboard=[

            [

                InlineKeyboardButton(

                    text="⬅️ Orqaga",

                    callback_data=(

                        f"dts_mavzu_"
                        f"{grade}_"
                        f"{subject_code}_"
                        f"{quarter}_"
                        f"{bob_code}_"
                        f"{bolim_code}_"
                        f"{mavzu_code}"

                    )

                )

            ]

        ]
    )

    await call.message.edit_text(

        text,

        reply_markup=kb

    )




async def dts_bob(

    call: CallbackQuery,

    page=0

):

    _, _, subject_code, quarter, bob_code = (
        call.data.split("_")
    )

    conn = psycopg2.connect(
        DATABASE_URL
    )

    cur = conn.cursor()

    cur.execute("""
    SELECT DISTINCT
        bolim_code,
        bolim_name
    FROM dts_tree
    WHERE subject_code=%s
    AND quarter=%s
    AND bob_code=%s
    AND is_deleted=FALSE
    ORDER BY bolim_code
    """, (
        subject_code,
        quarter,
        bob_code
    ))

    bolimlar = cur.fetchall()

    page_size = 10

    start = page * page_size

    end = start + page_size

    current_items = bolimlar[
        start:end
    ]

    buttons = []

    for code, name in current_items:

        buttons.append([

            InlineKeyboardButton(

                text=f"📑 {name} [{code}]",

                callback_data=(

                    f"dts_bolim_"

                    f"{subject_code}_"

                    f"{quarter}_"

                    f"{bob_code}_"

                    f"{code}"

                )

            )

        ])

    nav_buttons = []

    if page > 0:

        nav_buttons.append(

            InlineKeyboardButton(

                text="⬅️ Oldingi",

                callback_data=(

                    f"dts_bobpage_"

                    f"{subject_code}_"

                    f"{quarter}_"

                    f"{bob_code}_"

                    f"{page - 1}"

                )

            )

        )

    total_pages = (
        len(bolimlar) - 1
    ) // page_size + 1

    nav_buttons.append(

        InlineKeyboardButton(

            text=(
                f"{page + 1}/{total_pages}"
            ),

            callback_data="ignore"

        )

    )

    if end < len(bolimlar):

        nav_buttons.append(

            InlineKeyboardButton(

                text="Keyingi ➡️",

                callback_data=(

                    f"dts_bobpage_"

                    f"{subject_code}_"

                    f"{quarter}_"

                    f"{bob_code}_"

                    f"{page + 1}"

                )

            )

        )

    buttons.append(nav_buttons)

    buttons.append([

        InlineKeyboardButton(

            text="⬅️ Orqaga",

            callback_data=(

                f"dts_quarter_"

                f"{subject_code}_"

                f"{quarter}"

            )

        )

    ])

    kb = InlineKeyboardMarkup(
        inline_keyboard=buttons
    )

    await call.message.edit_text(

        f"📖 Bob [{bob_code}]",

        reply_markup=kb

    )

async def dts_bolim(

    call: CallbackQuery,

    page=0

):

    (
        _,
        _,
        subject_code,
        quarter,
        bob_code,
        bolim_code

    ) = call.data.split("_")

    conn = psycopg2.connect(
        DATABASE_URL
    )

    cur = conn.cursor()

    cur.execute("""
    SELECT DISTINCT
        mavzu_code,
        mavzu_name
    FROM dts_tree
    WHERE subject_code=%s
    AND quarter=%s
    AND bob_code=%s
    AND bolim_code=%s
    AND is_deleted=FALSE
    ORDER BY mavzu_code
    """, (
        subject_code,
        quarter,
        bob_code,
        bolim_code
    ))

    mavzular = cur.fetchall()

    page_size = 10

    start = page * page_size

    end = start + page_size

    current_items = mavzular[
        start:end
    ]

    buttons = []

    for code, name in current_items:

        buttons.append([

            InlineKeyboardButton(

                text=f"📝 {name} [{code}]",

                callback_data=(

                    f"dts_mavzu_"

                    f"{subject_code}_"

                    f"{quarter}_"

                    f"{bob_code}_"

                    f"{bolim_code}_"

                    f"{code}"

                )

            )

        ])

    nav_buttons = []

    if page > 0:

        nav_buttons.append(

            InlineKeyboardButton(

                text="⬅️ Oldingi",

                callback_data=(

                    f"dts_bolimpage_"

                    f"{subject_code}_"

                    f"{quarter}_"

                    f"{bob_code}_"

                    f"{bolim_code}_"

                    f"{page - 1}"

                )

            )

        )

    total_pages = (
        len(mavzular) - 1
    ) // page_size + 1

    nav_buttons.append(

        InlineKeyboardButton(

            text=(
                f"{page + 1}/{total_pages}"
            ),

            callback_data="ignore"

        )

    )

    if end < len(mavzular):

        nav_buttons.append(

            InlineKeyboardButton(

                text="Keyingi ➡️",

                callback_data=(

                    f"dts_bolimpage_"

                    f"{subject_code}_"

                    f"{quarter}_"

                    f"{bob_code}_"

                    f"{bolim_code}_"

                    f"{page + 1}"

                )

            )

        )

    buttons.append(nav_buttons)

    buttons.append([

        InlineKeyboardButton(

            text="⬅️ Orqaga",

            callback_data=(

                f"dts_bob_"

                f"{subject_code}_"

                f"{quarter}_"

                f"{bob_code}"

            )

        )

    ])

    kb = InlineKeyboardMarkup(
        inline_keyboard=buttons
    )

    await call.message.edit_text(

        f"📑 Bo‘lim [{bolim_code}]",

        reply_markup=kb

    )

async def dts_mavzu(

    call: CallbackQuery

):

    (
        _,
        _,
        subject_code,
        quarter,
        bob_code,
        bolim_code,
        mavzu_code

    ) = call.data.split("_")

    conn = psycopg2.connect(
        DATABASE_URL
    )

    cur = conn.cursor()

    cur.execute("""
    SELECT DISTINCT
        kichik_code,
        kichik_name,
        topic_code
    FROM dts_tree
    WHERE subject_code=%s
    AND quarter=%s
    AND bob_code=%s
    AND bolim_code=%s
    AND mavzu_code=%s
    AND is_deleted=FALSE
    ORDER BY kichik_code
    """, (
        subject_code,
        quarter,
        bob_code,
        bolim_code,
        mavzu_code
    ))

    kichiklar = cur.fetchall()

    buttons = []

    for code, name, topic_code in kichiklar:

        buttons.append([

            InlineKeyboardButton(

                text=f"🔹 {name} [{code}]",

                callback_data=(
                    f"dts_small_"
                    f"{topic_code}"
                )

            )

        ])

    buttons.append([

        InlineKeyboardButton(

            text="⬅️ Orqaga",

            callback_data=(
                f"dts_bolim_"
                f"{subject_code}_"
                f"{quarter}_"
                f"{bob_code}_"
                f"{bolim_code}"
            )

        )

    ])

    kb = InlineKeyboardMarkup(
        inline_keyboard=buttons
    )

    await call.message.edit_text(

        f"📝 Mavzu [{mavzu_code}]",

        reply_markup=kb

    )

@dp.callback_query(
    lambda c: c.data.startswith(
        "dts_bolimpage_"
    )
)
async def dts_bolim_page(

    call: CallbackQuery

):

    (
        _,
        _,
        subject_code,
        quarter,
        bob_code,
        bolim_code,
        page

    ) = call.data.split("_")

    await dts_bolim(
        call,
        int(page)
    )

async def dts_small(
    call: CallbackQuery
):

    topic_code = call.data.replace(
        "dts_small_",
        ""
    )

    conn = psycopg2.connect(
        DATABASE_URL
    )

    cur = conn.cursor()

    cur.execute("""

    SELECT
        topic_code,
        grade,
        subject_name,
        subject_code,
        quarter,
        bob_name,
        bob_code,
        bolim_name,
        bolim_code,
        mavzu_name,
        mavzu_code,
        kichik_name,
        kichik_code
    FROM dts_tree
    WHERE topic_code=%s
    AND is_deleted=FALSE
    LIMIT 1

    """, (

        topic_code,

    ))

    row = cur.fetchone()

    if not row:

        await call.answer(
            "Topilmadi"
        )

        return

    (
        topic_code,
        grade,
        subject_name,
        subject_code,
        quarter,
        bob_name,
        bob_code,
        bolim_name,
        bolim_code,
        mavzu_name,
        mavzu_code,
        kichik_name,
        kichik_code

    ) = row

    text = f"""
📚 DTS Ma’lumotlari
━━━━━━━━━━━━━━━

🏫 Sinf:
{grade}

📖 Fan:
{subject_name}

🗓 Chorak:
{quarter}

📘 Bob:
{bob_name}

📑 Bo‘lim:
{bolim_name}

📝 Mavzu:
{mavzu_name}

🔹 Kichik mavzu:
{kichik_name}

━━━━━━━━━━━━━━━

🆔 Topic Code:
{topic_code}
"""

    kb = InlineKeyboardMarkup(

        inline_keyboard=[

            [

                InlineKeyboardButton(

                    text="⬅️ Orqaga",

                    callback_data=(

                        f"dts_mavzu_"
                        f"{subject_code}_"
                        f"{quarter}_"
                        f"{bob_code}_"
                        f"{bolim_code}_"
                        f"{mavzu_code}"

                    )

                )

            ]

        ]

    )

    await call.message.edit_text(

        text,
        reply_markup=kb

    )

    cur.close()
    conn.close()

async def dts_menu(

    message

):

    kb = InlineKeyboardMarkup(
        inline_keyboard=[

            [

                InlineKeyboardButton(

                    text="📥 Import DTS",

                    callback_data="dts_import"

                )

            ],

            [

                InlineKeyboardButton(

                    text="🧭 DTS Navigator",

                    callback_data="dts_navigator"

                )

            ],

            [

                InlineKeyboardButton(

                    text="🔎 DTS Qidiruv",

                    callback_data="dts_search"

                )

            ],

            [

                InlineKeyboardButton(

                    text="📤 Export DTS",

                    callback_data="dts_export"

                )

            ]

        ]
    )

    await message.answer(

        "📚 DTS Boshqaruv Paneli",

        reply_markup=kb

    )

async def dts_import(

    call: CallbackQuery,

    state: FSMContext

):

    await state.set_state(
        DTSImportState.waiting_excel
    )

    kb = InlineKeyboardMarkup(
        inline_keyboard=[

            [

                InlineKeyboardButton(

                    text="⬅️ Orqaga",

                    callback_data="dts_menu"

                )

            ]

        ]
    )
    await call.message.edit_text(

        """

    📥 DTS Excel Import

    ━━━━━━━━━━━━━━━

    Excel fayl yuboring.

    Kerakli ustunlar:

    1. Sinf
    2. Fan
    3. Chorak
    4. Bob
    5. Bo‘lim
    6. Mavzu
    7. Kichik mavzu

    ━━━━━━━━━━━━━━━

    """,

        reply_markup=kb

    )

@dp.message(
    DTSImportState.waiting_excel,
    F.document
)
async def dts_excel_import(
    message: Message,
    state: FSMContext
):
    current_state = await state.get_state()

    if current_state != (
        DTSImportState.waiting_excel.state
    ):

        return

    document = message.document

    if not document:
        await message.answer(
            "❌ Excel yuboring"
        )
        return

    os.makedirs(
        "temp",
        exist_ok=True
    )

    file_path = (
        f"temp/{document.file_name}"
    )

    await bot.download(
        document,
        destination=file_path
    )

    await message.answer(
        "📥 Fayl yuklab olindi"
    )

    wb = load_workbook(
        file_path
    )

    ws = wb.active

    rows = list(
        ws.iter_rows(
            min_row=2,
            values_only=True
        )
    )

    try:

        conn = psycopg2.connect(
            DATABASE_URL
        )

        cur = conn.cursor()

    except Exception as e:

        await message.answer(
            f"❌ DB ERROR:\n{e}"
        )

        return

    result = analyze_import(
        cur,
        rows
    )

    user_id = message.from_user.id

    dts_import_cache[user_id] = result

    text = f"""
📥 DTS Import Tahlili

━━━━━━━━━━━━━━━

📄 Jami qator:
{len(rows)}

✅ Import qilinadi:
{len(result["valid_rows"])}

⚠️ Excel ichida takroriy:
{len(result["duplicate_rows"])}

📦 Bazada mavjud:
{len(result["existing_rows"])}

❌ Xatolar:
{len(result["error_rows"])}

━━━━━━━━━━━━━━━
"""

    kb = InlineKeyboardMarkup(
        inline_keyboard=[

            [

                InlineKeyboardButton(
                    text="✅ Import qilish",
                    callback_data="dts_confirm_import"
                )

            ],

            [

                InlineKeyboardButton(
                    text="📊 Muammolar tahlili",
                    callback_data="dts_problems"
                )

            ],

            [

                InlineKeyboardButton(
                    text="📥 Xatolarni yuklab olish",
                    callback_data="dts_download_errors"
                )

            ],

            [

                InlineKeyboardButton(
                    text="❌ Bekor qilish",
                    callback_data="dts_cancel_import"
                )

            ],

            [

                InlineKeyboardButton(
                    text="⬅️ Menu",
                    callback_data="dts_menu"
                )

            ]

        ]
    )

    await message.answer(
        text,
        reply_markup=kb
    )

    cur.close()
    conn.close()

    await state.clear()

@dp.callback_query(
    lambda c: c.data == "dts_confirm_import"
)
async def dts_confirm_import(

    call: CallbackQuery

):

    await call.answer()

    user_id = call.from_user.id

    cache = dts_import_cache.get(
        user_id
    )

    if not cache:

        await call.answer(
            "❌ Cache topilmadi"
        )

        return

    valid_rows = cache[
        "valid_rows"
    ]

    try:

        conn = psycopg2.connect(
            DATABASE_URL
        )

        cur = conn.cursor()

        result = confirm_import(
            cur,
            valid_rows
        )

        conn.commit()

        inserted_count = result[
            "inserted_count"
        ]

        failed_rows = result[
            "failed_rows"
        ]

        text = f"""
✅ DTS import tugadi

━━━━━━━━━━━━━━━

📥 Qo‘shildi:
{inserted_count}

❌ Xatolar:
{len(failed_rows)}

━━━━━━━━━━━━━━━
"""

        await call.message.edit_text(
            text
        )

        dts_import_cache.pop(
            user_id,
            None
        )

    except Exception as e:

        await call.message.answer(
            f"❌ Import xatosi:\n{e}"
        )

    finally:

        cur.close()
        conn.close()

@dp.callback_query(
    lambda c: c.data == "dts_problems"
)
async def dts_problems(
    call: CallbackQuery
):

    await call.answer()

    user_id = call.from_user.id

    cache = dts_import_cache.get(
        user_id
    )

    if not cache:

        await call.answer(
            "❌ Cache topilmadi"
        )

        return

    text = "📊 DTS Import Muammolari\n\n"

    for row in cache["duplicate_rows"]:

        text += (
            f"⚠️ Takroriy\n"
            f"Qator: {row['row_no']}\n"
            f"Code: {row['topic_code']}\n"
            f"{row['reason']}\n\n"
        )

    for row in cache["existing_rows"]:

        text += (
            f"📦 Bazada mavjud\n"
            f"Qator: {row['row_no']}\n"
            f"Code: {row['topic_code']}\n"
            f"{row['reason']}\n\n"
        )

    for row in cache["error_rows"]:

        text += (
            f"❌ Xato\n"
            f"Qator: {row['row_no']}\n"
            f"{row['reason']}\n\n"
        )

    if text == "📊 DTS Import Muammolari\n\n":

        text += "✅ Muammo yo‘q"

    await call.message.answer(
        text[:4000]
    )

@dp.callback_query(
    lambda c: c.data == "dts_download_errors"
)
async def dts_download_errors(
    call: CallbackQuery
):

    await call.answer()

    user_id = call.from_user.id

    cache = dts_import_cache.get(
        user_id
    )

    if not cache:

        await call.answer(
            "❌ Cache topilmadi"
        )

        return

    text = "DTS IMPORT XATOLARI\n\n"

    for row in cache["duplicate_rows"]:

        text += (
            f"TAKRORIY\n"
            f"Qator: {row['row_no']}\n"
            f"Code: {row['topic_code']}\n"
            f"{row['reason']}\n\n"
        )

    for row in cache["existing_rows"]:

        text += (
            f"BAZADA MAVJUD\n"
            f"Qator: {row['row_no']}\n"
            f"Code: {row['topic_code']}\n"
            f"{row['reason']}\n\n"
        )

    for row in cache["error_rows"]:

        text += (
            f"XATO\n"
            f"Qator: {row['row_no']}\n"
            f"{row['reason']}\n\n"
        )

    if text == "DTS IMPORT XATOLARI\n\n":

        text += "Xatolar yo‘q"

    os.makedirs(
        "temp",
        exist_ok=True
    )

    file_path = (
        f"temp/errors_{user_id}.txt"
    )

    with open(
        file_path,
        "w",
        encoding="utf-8"
    ) as f:

        f.write(text)

    await call.message.answer_document(
        FSInputFile(file_path)
    )

@dp.callback_query(
    lambda c: c.data == "dts_cancel_import"
)
async def dts_cancel_import(
    call: CallbackQuery
):

    await call.answer()

    user_id = call.from_user.id

    if user_id in dts_import_cache:

        del dts_import_cache[user_id]

    await call.message.edit_text(
        "❌ DTS import bekor qilindi"
    )
