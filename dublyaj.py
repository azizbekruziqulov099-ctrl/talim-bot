"""dublyaj.py — Video/audio DUBLYAJ qilish.

Bosqichlar:
  1. Video yuklanadi (video_admin.yukla orqali)
  2. Videodan audio ajratiladi (ffmpeg)
  3. Audio matnga aylantiriladi (faster-whisper, mahalliy, internet shart emas)
  4. Matn tarjima qilinadi (Gemini API)
  5. Tarjima qilingan matndan yangi ovoz yaratiladi (edge-tts)
  6. Yangi ovoz videoga ulanadi, asl ovoz almashtiriladi (ffmpeg)

ESLATMA: bu LIP-SYNC emas — faqat ovoz almashtiriladi, lab harakati
asl tilga mos qoladi. Whisper modeli birinchi ishlatilganda yuklab olinadi
(internet kerak, keyin keshlanadi — qayta yuklanmaydi).
"""
import os
import re
import io
import json
import subprocess
import urllib.request
import urllib.error
import psycopg2

DATABASE_URL = os.getenv("DATABASE_URL", "")


def _db():
    return psycopg2.connect(DATABASE_URL)


def klon_ovoz_jadval():
    try:
        c = _db(); cr = c.cursor()
        cr.execute("""CREATE TABLE IF NOT EXISTS klon_ovoz(
            user_id BIGINT PRIMARY KEY,
            voice_id TEXT NOT NULL,
            yaratildi TIMESTAMP DEFAULT NOW()
        )""")
        c.commit(); cr.close(); c.close()
    except Exception as e:
        print(f"[dublyaj] klon_ovoz_jadval: {e}")


def klon_ovoz_saqla(user_id, voice_id):
    klon_ovoz_jadval()
    try:
        c = _db(); cr = c.cursor()
        cr.execute("""INSERT INTO klon_ovoz(user_id, voice_id) VALUES(%s,%s)
            ON CONFLICT(user_id) DO UPDATE SET voice_id=EXCLUDED.voice_id,
            yaratildi=NOW()""", (user_id, voice_id))
        c.commit(); cr.close(); c.close()
        return True
    except Exception as e:
        print(f"[dublyaj] klon_ovoz_saqla: {e}")
        return False


def klon_ovoz_ol(user_id):
    """Bu foydalanuvchining klonlangan ovoz_id'sini qaytaradi, yo'q bo'lsa None."""
    klon_ovoz_jadval()
    try:
        c = _db(); cr = c.cursor()
        cr.execute("SELECT voice_id FROM klon_ovoz WHERE user_id=%s", (user_id,))
        r = cr.fetchone(); cr.close(); c.close()
        return r[0] if r else None
    except Exception as e:
        print(f"[dublyaj] klon_ovoz_ol: {e}")
        return None

MAX_VAQT_SONIYA = 240
WHISPER_MODEL_OLCHAMI = "base"     # tiny/base/small — CPU uchun muvozanat

# Til kodi -> (ko'rsatiladigan nom, ayol ovoz, erkak ovoz)
TILLAR = {
    "en": ("🇬🇧 Ingliz",    "en-US-JennyNeural",    "en-US-GuyNeural"),
    "ru": ("🇷🇺 Rus",       "ru-RU-SvetlanaNeural", "ru-RU-DmitryNeural"),
    "fr": ("🇫🇷 Fransuz",   "fr-FR-DeniseNeural",   "fr-FR-HenriNeural"),
    "es": ("🇪🇸 Ispan",     "es-ES-ElviraNeural",   "es-ES-AlvaroNeural"),
    "de": ("🇩🇪 Nemis",     "de-DE-KatjaNeural",    "de-DE-ConradNeural"),
    "tr": ("🇹🇷 Turk",      "tr-TR-EmelNeural",     "tr-TR-AhmetNeural"),
    "ar": ("🇸🇦 Arab",      "ar-SA-ZariyahNeural",  "ar-SA-HamedNeural"),
    "zh": ("🇨🇳 Xitoy",     "zh-CN-XiaoxiaoNeural", "zh-CN-YunxiNeural"),
    "ko": ("🇰🇷 Koreys",    "ko-KR-SunHiNeural",    "ko-KR-InJoonNeural"),
    "ja": ("🇯🇵 Yapon",     "ja-JP-NanamiNeural",   "ja-JP-KeitaNeural"),
    "kk": ("🇰🇿 Qozoq",     "kk-KZ-AigulNeural",    "kk-KZ-DauletNeural"),
    "uz": ("🇺🇿 O'zbek",    "uz-UZ-MadinaNeural",   "uz-UZ-SardorNeural"),
}


def til_nomi(kod):
    return TILLAR.get(kod, (kod, None, None))[0]


def til_royxati_tugmalar(callback_prefix):
    """[[InlineKeyboardButton...]] — chaqiruvchi Talim.py o'zi InlineKeyboardButton import qiladi."""
    return [(kod, nom) for kod, (nom, _, _) in TILLAR.items()]


# ═══════════════ NLLB — MUSTAQIL TARJIMA (Gemini'ga UMUMAN bog'liq emas) ═══════════════
# Meta'ning ochiq kodli NLLB-200 modeli — serverning o'zida ishlaydi.
# Hech qanday tashqi API, kvota yoki internet (birinchi yuklashdan keyin) kerak emas.
# DIQQAT: og'ir — ~2-4GB RAM/disk talab qiladi, birinchi chaqiruvda modelni
# yuklab oladi (internet kerak, keyin keshlanadi).

NLLB_MODEL_NOMI = "facebook/nllb-200-distilled-600M"

# Oddiy til kodidan FLORES-200 kodiga (NLLB shu formatni talab qiladi)
FLORES_KOD = {
    "uz": "uzn_Latn", "en": "eng_Latn", "ru": "rus_Cyrl", "fr": "fra_Latn",
    "es": "spa_Latn", "de": "deu_Latn", "tr": "tur_Latn", "ar": "arb_Arab",
    "zh": "zho_Hans", "ko": "kor_Hang", "ja": "jpn_Jpan", "kk": "kaz_Cyrl",
}

_NLLB_MODEL = None
_NLLB_TOKENIZER = None


def _nllb_ol():
    """NLLB modelini bir marta yuklaydi (og'ir), keyin keshdan foydalanadi."""
    global _NLLB_MODEL, _NLLB_TOKENIZER
    if _NLLB_MODEL is None:
        from transformers import AutoModelForSeq2SeqLM, AutoTokenizer
        _NLLB_TOKENIZER = AutoTokenizer.from_pretrained(NLLB_MODEL_NOMI)
        _NLLB_MODEL = AutoModelForSeq2SeqLM.from_pretrained(NLLB_MODEL_NOMI)
    return _NLLB_MODEL, _NLLB_TOKENIZER


def _flores_kodga_ot(oddiy_kod):
    """Oddiy til kodini (uz, en...) FLORES-200 kodiga o'giradi.
    Topilmasa — inglizchaga taxmin qiladi."""
    return FLORES_KOD.get(str(oddiy_kod or "").lower(), "eng_Latn")


def _nllb_forced_id(tokenizer, flores_kod):
    """transformers versiyalari orasida API farqi bor — ikkalasini ham sinaymiz."""
    try:
        return tokenizer.convert_tokens_to_ids(flores_kod)
    except Exception:
        pass
    try:
        return tokenizer.lang_code_to_id[flores_kod]
    except Exception:
        return tokenizer.convert_tokens_to_ids("eng_Latn")


def nllb_tarjima_qil_segmentlar(segmentlar, manba_til_oddiy, maqsad_til_kod):
    """NLLB orqali har segmentni tarjima qiladi — Gemini'ga bog'liq EMAS.
    ([(boshlanish, tugash, tarjima), ...], xato)"""
    try:
        model, tokenizer = _nllb_ol()
        manba_flores = _flores_kodga_ot(manba_til_oddiy)
        maqsad_flores = _flores_kodga_ot(maqsad_til_kod)
        tokenizer.src_lang = manba_flores
        forced_id = _nllb_forced_id(tokenizer, maqsad_flores)

        natija = []
        for boshi, tugash, matn in segmentlar:
            matn = matn.strip()
            if not matn:
                natija.append((boshi, tugash, matn))
                continue
            inputs = tokenizer(matn, return_tensors="pt", truncation=True, max_length=256)
            tokenlar = model.generate(**inputs, forced_bos_token_id=forced_id, max_length=256)
            tarjima = tokenizer.batch_decode(tokenlar, skip_special_tokens=True)[0]
            natija.append((boshi, tugash, tarjima))
        return (natija, None)
    except Exception as e:
        return (None, str(e)[:300])


# ═══════════════ GOOGLE CLOUD TRANSLATION API ═══════════════
# Gemini'dan BUTUNLAY BOSHQA xizmat/kvota. Harfi bo'yicha to'lanadi,
# "daqiqasiga necha so'rov" chegarasi yo'q. Oyiga 500,000 harf bepul.
# Yengil (og'ir kutubxona kerak emas), tez, sifatli.

# ═══════════════ QO'LDA TARJIMA — EXCEL JADVAL ═══════════════
# Avtomatik tarjima sifatidan qat'iy nazar — foydalanuvchi o'zi tarjima
# qiladi. Bot faqat vaqt-tuzilmani (segmentlarni) jadvalga chiqaradi va
# to'ldirilgan jadvalni qayta o'qiydi.

def segmentlar_excelga(segmentlar, chiqish_yol):
    """Segmentlarni Excel jadvaliga chiqaradi — Tarjima ustuni bo'sh,
    foydalanuvchi to'ldiradi. (muvaffaqiyat, xato)"""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Tarjima"

        sarlavhalar = ["#", "Boshlanish (s)", "Tugash (s)", "Original matn",
                       "Tarjima (shu yerga yozing)"]
        boshliq_fon = PatternFill(start_color="4A5568", end_color="4A5568", fill_type="solid")
        for col, sarlavha in enumerate(sarlavhalar, 1):
            c = ws.cell(row=1, column=col, value=sarlavha)
            c.font = Font(name="Arial", bold=True, color="FFFFFF")
            c.fill = boshliq_fon

        sariq = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
        for i, (boshi, tugash, matn) in enumerate(segmentlar, start=2):
            ws.cell(row=i, column=1, value=i - 1).font = Font(name="Arial")
            ws.cell(row=i, column=2, value=round(boshi, 2)).font = Font(name="Arial")
            ws.cell(row=i, column=3, value=round(tugash, 2)).font = Font(name="Arial")
            ws.cell(row=i, column=4, value=matn).font = Font(name="Arial")
            tj = ws.cell(row=i, column=5, value="")
            tj.font = Font(name="Arial")
            tj.fill = sariq

        ws.column_dimensions["A"].width = 5
        ws.column_dimensions["B"].width = 14
        ws.column_dimensions["C"].width = 14
        ws.column_dimensions["D"].width = 40
        ws.column_dimensions["E"].width = 40
        ws.freeze_panes = "A2"

        wb.save(chiqish_yol)
        if not os.path.exists(chiqish_yol):
            return (False, "Excel fayli yaratilmadi")
        return (True, None)
    except Exception as e:
        return (False, str(e)[:300])


def excel_dan_segmentlar(excel_yol, asl_segmentlar):
    """Foydalanuvchi to'ldirgan Excel'dan tarjimalarni o'qib, segmentlarga
    joylashtiradi. Qator tartibi eksport qilingandagi bilan bir xil deb
    hisoblanadi. ([(boshlanish, tugash, tarjima), ...], xato)"""
    try:
        import openpyxl
        wb = openpyxl.load_workbook(excel_yol, data_only=True)
        ws = wb.active

        natija = []
        bosh_tarjima = 0
        for i, (boshi, tugash, asl) in enumerate(asl_segmentlar):
            satr = i + 2   # 1-qator sarlavha
            xujayra = ws.cell(row=satr, column=5).value
            tarjima = str(xujayra).strip() if xujayra else ""
            if tarjima:
                bosh_tarjima += 1
            natija.append((boshi, tugash, tarjima or asl))

        if bosh_tarjima == 0:
            return (None, "Tarjima ustuni bo'sh — hech qanday gap to'ldirilmagan")
        return (natija, None)
    except Exception as e:
        return (None, str(e)[:300])



def google_tarjima_qil_segmentlar(segmentlar, manba_til_oddiy, maqsad_til_kod):
    """Google Cloud Translation API orqali BARCHA segmentlarni BITTA
    so'rovda tarjima qiladi. ([(boshlanish, tugash, tarjima), ...], xato)"""
    import html as _html

    api_key = os.getenv("GOOGLE_TRANSLATE_API_KEY", "")
    if not api_key:
        return (None, "GOOGLE_TRANSLATE_API_KEY sozlanmagan (Railway muhitida yo'q)")

    # Bo'sh matnli segmentlarni API'ga yubormaymiz, keyin joyiga qaytaramiz
    yuboriladigan = [(i, s[2]) for i, s in enumerate(segmentlar) if s[2].strip()]
    if not yuboriladigan:
        return (list(segmentlar), None)

    try:
        url = f"https://translation.googleapis.com/language/translate/v2?key={api_key}"
        body = json.dumps({
            "q": [m for _, m in yuboriladigan],
            "source": manba_til_oddiy,
            "target": maqsad_til_kod,
            "format": "text",
        }).encode("utf-8")
        req = urllib.request.Request(url, data=body,
                                     headers={"Content-Type": "application/json"})
        with urllib.request.urlopen(req, timeout=30) as resp:
            natija = json.loads(resp.read().decode("utf-8"))
        tarjimalar = [_html.unescape(t["translatedText"])
                     for t in natija["data"]["translations"]]

        indeks_dan_tarjima = {idx: tarj for (idx, _), tarj in zip(yuboriladigan, tarjimalar)}
        yakuniy = []
        for i, (boshi, tugash, asl) in enumerate(segmentlar):
            yakuniy.append((boshi, tugash, indeks_dan_tarjima.get(i, asl)))
        return (yakuniy, None)
    except urllib.error.HTTPError as e:
        xabar = e.read().decode("utf-8", errors="ignore")[:300]
        return (None, f"Google Translate xato ({e.code}): {xabar}")
    except Exception as e:
        return (None, str(e)[:300])


# ═══════════════ 2. AUDIO AJRATISH ═══════════════

def video_dan_audio_ajrat(video_yol, audio_yol):
    """ffmpeg orqali video faylidan audio ajratib oladi. (muvaffaqiyat, xato)"""
    try:
        natija = subprocess.run(
            ["ffmpeg", "-y", "-i", video_yol, "-vn", "-acodec", "libmp3lame",
             "-q:a", "2", audio_yol],
            capture_output=True, text=True, timeout=120)
        if natija.returncode != 0:
            return (False, (natija.stderr or "")[-300:])
        if not os.path.exists(audio_yol) or os.path.getsize(audio_yol) == 0:
            return (False, "Audio fayl yaratilmadi")
        return (True, None)
    except Exception as e:
        return (False, str(e)[:300])


def video_davomiyligi(video_yol):
    """Videoning umumiy uzunligini soniyada qaytaradi (ffprobe)."""
    try:
        natija = subprocess.run(
            ["ffprobe", "-v", "error", "-show_entries", "format=duration",
             "-of", "csv=p=0", video_yol],
            capture_output=True, text=True, timeout=15)
        return float(natija.stdout.strip())
    except Exception:
        return 0.0


# ═══════════════ 3. MATNGA AYLANTIRISH (Whisper) ═══════════════

_WHISPER_MODEL = None

def _whisper_ol():
    global _WHISPER_MODEL
    if _WHISPER_MODEL is None:
        from faster_whisper import WhisperModel
        _WHISPER_MODEL = WhisperModel(WHISPER_MODEL_OLCHAMI, device="cpu", compute_type="int8")
    return _WHISPER_MODEL


def matnga_aylantir(audio_yol):
    """Audiodan SEGMENTLAR ro'yxatini qaytaradi — har biri o'z vaqti bilan.
    ([(boshlanish_son, tugash_son, matn), ...], til_kodi, xato)
    Vaqt-moslashtirilgan dublyaj uchun shart — har gap o'z vaqtida gapirilishi kerak."""
    try:
        model = _whisper_ol()
        segments, info = model.transcribe(audio_yol, beam_size=5)
        natija = []
        for s in segments:
            matn = s.text.strip()
            if matn:
                natija.append((round(s.start, 2), round(s.end, 2), matn))
        if not natija:
            return (None, None, "Nutq aniqlanmadi (audio jim yoki tushunarsiz)")
        return (natija, info.language, None)
    except Exception as e:
        return (None, None, str(e)[:300])


def matn_yigindisi(segmentlar):
    """Segmentlarni bitta matnga birlashtiradi — admin ko'rsatish uchun."""
    return " ".join(s[2] for s in segmentlar)


# ═══════════════ 4. TARJIMA (Gemini) ═══════════════

def tarjima_qil(matn, maqsad_til_kod):
    """Gemini API orqali tarjima qiladi (oddiy, bitta matn). (tarjima, xato)"""
    api_key = os.getenv("GEMINI_API_KEY", "")
    if not api_key:
        return (None, "GEMINI_API_KEY sozlanmagan (Railway muhitida yo'q)")

    maqsad_nom = til_nomi(maqsad_til_kod)
    prompt = (f"Quyidagi matnni {maqsad_nom} tiliga tarjima qil. "
              f"FAQAT tarjima qilingan matnni qaytar — izoh, tirnoq, "
              f"boshqa hech narsa qo'shma:\n\n{matn}")

    try:
        url = ("https://generativelanguage.googleapis.com/v1beta/"
               f"models/gemini-2.5-flash:generateContent?key={api_key}")
        body = json.dumps({"contents": [{"parts": [{"text": prompt}]}]}).encode("utf-8")
        req = urllib.request.Request(url, data=body,
                                     headers={"Content-Type": "application/json"})
        with urllib.request.urlopen(req, timeout=30) as resp:
            natija = json.loads(resp.read().decode("utf-8"))
        tarjima = natija["candidates"][0]["content"]["parts"][0]["text"].strip()
        tarjima = tarjima.strip('"\'')
        return (tarjima, None)
    except urllib.error.HTTPError as e:
        return (None, f"Gemini xato ({e.code}): {e.read()[:200]}")
    except Exception as e:
        return (None, str(e)[:300])


def tarjima_qil_segmentlar(segmentlar, maqsad_til_kod):
    """Barcha segmentlarni BITTA Gemini chaqiruvida tarjima qiladi,
    tartibni saqlab. ([(boshlanish, tugash, tarjima), ...], xato)"""
    api_key = os.getenv("GEMINI_API_KEY", "")
    if not api_key:
        return (None, "GEMINI_API_KEY sozlanmagan (Railway muhitida yo'q)")

    maqsad_nom = til_nomi(maqsad_til_kod)
    royxat = [{"id": i, "matn": s[2]} for i, s in enumerate(segmentlar)]
    prompt = (
        f"Quyida JSON ro'yxat bor — har birida \"id\" va \"matn\" maydoni bor.\n"
        f"Har bir \"matn\"ni {maqsad_nom} tiliga tarjima qil.\n\n"
        f"JAVOBNI FAQAT shu JSON formatda qaytar (boshqa hech narsa yozma, "
        f"``` belgilarisiz):\n"
        f'[{{"id": 0, "tarjima": "..."}}, {{"id": 1, "tarjima": "..."}}, ...]\n\n'
        f"Matnlar:\n{json.dumps(royxat, ensure_ascii=False)}"
    )

    try:
        url = ("https://generativelanguage.googleapis.com/v1beta/"
               f"models/gemini-2.5-flash:generateContent?key={api_key}")
        body = json.dumps({"contents": [{"parts": [{"text": prompt}]}]}).encode("utf-8")
        req = urllib.request.Request(url, data=body,
                                     headers={"Content-Type": "application/json"})
        with urllib.request.urlopen(req, timeout=60) as resp:
            natija = json.loads(resp.read().decode("utf-8"))
        javob_matn = natija["candidates"][0]["content"]["parts"][0]["text"].strip()
        javob_matn = re.sub(r"^```(json)?|```$", "", javob_matn.strip(), flags=re.M).strip()
        tarjimalar = json.loads(javob_matn)

        id_dan_tarjima = {t["id"]: t["tarjima"] for t in tarjimalar}
        natija_royxat = []
        for i, (boshi, tugash, asl) in enumerate(segmentlar):
            tarj = id_dan_tarjima.get(i, asl)   # topilmasa aslini qoldiradi
            natija_royxat.append((boshi, tugash, tarj))
        return (natija_royxat, None)
    except urllib.error.HTTPError as e:
        return (None, f"Gemini xato ({e.code}): {e.read()[:200]}")
    except (json.JSONDecodeError, KeyError) as e:
        return (None, f"Gemini javobini o'qib bo'lmadi: {e}")
    except Exception as e:
        return (None, str(e)[:300])


# ═══════════════ 5. YANGI OVOZ (edge-tts — jinsga qarab, umumiy ovoz) ═══════════════

async def matndan_ovoz_fayl(matn, ovoz_nomi, chiqish_yol):
    """edge-tts orqali istalgan tilda ovoz yaratib, faylga yozadi. (muvaffaqiyat, xato)"""
    try:
        import edge_tts
        com = edge_tts.Communicate(matn, ovoz_nomi)
        await com.save(chiqish_yol)
        if not os.path.exists(chiqish_yol) or os.path.getsize(chiqish_yol) == 0:
            return (False, "Ovoz fayli yaratilmadi")
        return (True, None)
    except Exception as e:
        return (False, str(e)[:300])


# ═══════════════ 5b. OVOZ KLONLASH (ElevenLabs) ═══════════════
# Foydalanuvchi o'z ovozidan namuna beradi -> shu ovozda dublyaj qilinadi.

def elevenlabs_ovoz_klonla(namuna_audio_yol, ovoz_nomi, user_id):
    """Foydalanuvchi ovoz namunasidan klonlangan ovoz yaratadi.
    (voice_id, xato)"""
    api_key = os.getenv("ELEVENLABS_API_KEY", "")
    if not api_key:
        return (None, "ELEVENLABS_API_KEY sozlanmagan (Railway muhitida yo'q)")
    try:
        import urllib.request as _ur

        chegara = f"----samtm{user_id}"
        with open(namuna_audio_yol, "rb") as f:
            audio_bayt = f.read()

        qismlar = []
        qismlar.append(f"--{chegara}\r\n"
                       f'Content-Disposition: form-data; name="name"\r\n\r\n'
                       f"{ovoz_nomi}\r\n".encode("utf-8"))
        qismlar.append(f"--{chegara}\r\n"
                       f'Content-Disposition: form-data; name="files"; filename="namuna.ogg"\r\n'
                       f"Content-Type: audio/ogg\r\n\r\n".encode("utf-8"))
        qismlar.append(audio_bayt)
        qismlar.append(f"\r\n--{chegara}--\r\n".encode("utf-8"))
        tana = b"".join(qismlar)

        req = _ur.Request(
            "https://api.elevenlabs.io/v1/voices/add",
            data=tana,
            headers={
                "xi-api-key": api_key,
                "Content-Type": f"multipart/form-data; boundary={chegara}",
            })
        with _ur.urlopen(req, timeout=60) as resp:
            natija = json.loads(resp.read().decode("utf-8"))
        voice_id = natija.get("voice_id")
        if not voice_id:
            return (None, f"voice_id qaytmadi: {natija}")
        return (voice_id, None)
    except urllib.error.HTTPError as e:
        xabar = e.read().decode("utf-8", errors="ignore")[:300]
        return (None, f"ElevenLabs xato ({e.code}): {xabar}")
    except Exception as e:
        return (None, str(e)[:300])


async def elevenlabs_ovoz_fayl(matn, voice_id, chiqish_yol, barqarorlik=0.5, jonlilik=0.5):
    """Klonlangan ovozda matndan audio yaratadi.
    barqarorlik (stability): 0=juda ekspressiv/o'zgaruvchan, 1=juda barqaror/monoton
    jonlilik (style): 0=neytral, 1=kuchli uslub/ohang
    (muvaffaqiyat, xato)"""
    api_key = os.getenv("ELEVENLABS_API_KEY", "")
    if not api_key:
        return (False, "ELEVENLABS_API_KEY sozlanmagan (Railway muhitida yo'q)")
    try:
        url = f"https://api.elevenlabs.io/v1/text-to-speech/{voice_id}"
        body = json.dumps({
            "text": matn,
            "model_id": "eleven_multilingual_v2",
            "voice_settings": {
                "stability": barqarorlik,
                "similarity_boost": 0.75,
                "style": jonlilik,
                "use_speaker_boost": True,
            }
        }).encode("utf-8")
        req = urllib.request.Request(url, data=body, headers={
            "xi-api-key": api_key,
            "Content-Type": "application/json",
            "Accept": "audio/mpeg",
        })
        with urllib.request.urlopen(req, timeout=45) as resp:
            audio_bayt = resp.read()
        with open(chiqish_yol, "wb") as f:
            f.write(audio_bayt)
        if not os.path.exists(chiqish_yol) or os.path.getsize(chiqish_yol) == 0:
            return (False, "Ovoz fayli yaratilmadi")
        return (True, None)
    except urllib.error.HTTPError as e:
        xabar = e.read().decode("utf-8", errors="ignore")[:300]
        return (False, f"ElevenLabs xato ({e.code}): {xabar}")
    except Exception as e:
        return (False, str(e)[:300])


# ═══════════════ 6. VIDEOGA ULASH (ffmpeg) ═══════════════

def videoga_ulash(video_yol, yangi_audio_yol, chiqish_yol):
    """Videoning audiosini yangisiga almashtiradi — video kadrlari o'zgarmaydi.
    (muvaffaqiyat, xato)"""
    try:
        natija = subprocess.run(
            ["ffmpeg", "-y", "-i", video_yol, "-i", yangi_audio_yol,
             "-c:v", "copy", "-map", "0:v:0", "-map", "1:a:0",
             "-shortest", chiqish_yol],
            capture_output=True, text=True, timeout=120)
        if natija.returncode != 0:
            return (False, (natija.stderr or "")[-300:])
        if not os.path.exists(chiqish_yol) or os.path.getsize(chiqish_yol) == 0:
            return (False, "Yakuniy video yaratilmadi")
        return (True, None)
    except Exception as e:
        return (False, str(e)[:300])


# ═══════════════ VAQT-MOSLASHTIRILGAN YIG'ISH ═══════════════

def _davomiylik_ol(audio_yol):
    """ffprobe orqali audio faylning soniyadagi uzunligini oladi."""
    try:
        natija = subprocess.run(
            ["ffprobe", "-v", "error", "-show_entries", "format=duration",
             "-of", "csv=p=0", audio_yol],
            capture_output=True, text=True, timeout=15)
        return float(natija.stdout.strip())
    except Exception:
        return 0.0


def _tezlikni_moslash(kirish_yol, chiqish_yol, tezlik):
    """ffmpeg atempo orqali audio tezligini o'zgartiradi (ohang o'zgarmaydi).
    atempo 0.5-2.0 oralig'ida ishlaydi — chegaradan tashqarisini zanjirlaymiz."""
    tezlik = max(0.5, min(2.0, tezlik))
    try:
        natija = subprocess.run(
            ["ffmpeg", "-y", "-i", kirish_yol, "-filter:a", f"atempo={tezlik}",
             chiqish_yol], capture_output=True, text=True, timeout=60)
        return natija.returncode == 0 and os.path.exists(chiqish_yol)
    except Exception:
        return False


def _sukunat_yarat(soniya, chiqish_yol):
    """Berilgan uzunlikda sukunat (jimlik) audio fayl yaratadi.
    Faqat asl fon olinmagan/muvaffaqiyatsiz bo'lganda ZAXIRA sifatida ishlatiladi."""
    if soniya <= 0:
        return False
    try:
        natija = subprocess.run(
            ["ffmpeg", "-y", "-f", "lavfi", "-i", "anullsrc=r=24000:cl=mono",
             "-t", str(round(soniya, 2)), chiqish_yol],
            capture_output=True, text=True, timeout=30)
        return natija.returncode == 0 and os.path.exists(chiqish_yol)
    except Exception:
        return False


def _asl_fon_ol(asl_audio_yol, boshlanish, davomiylik, chiqish_yol):
    """ASL videodagi shu vaqt oralig'ini kesib oladi — bu 'gapirmagan'
    joylarga tabiiy fon (xona shovqini, muhit tovushi) sifatida ishlatiladi,
    mutlaq raqamli sukunat o'rniga. Bu ovozni ANCHA tabiiyroq qiladi."""
    if davomiylik <= 0 or not asl_audio_yol or not os.path.exists(asl_audio_yol):
        return False
    try:
        natija = subprocess.run(
            ["ffmpeg", "-y", "-i", asl_audio_yol,
             "-ss", str(max(0, round(boshlanish, 2))),
             "-t", str(round(davomiylik, 2)),
             "-acodec", "libmp3lame", chiqish_yol],
            capture_output=True, text=True, timeout=30)
        return (natija.returncode == 0 and os.path.exists(chiqish_yol)
                and os.path.getsize(chiqish_yol) > 0)
    except Exception:
        return False


async def segmentlardan_ovoz_yigindisi(segmentlar_tarjima, ovoz_nomi, papka, jami_davomiylik,
                                       dvigatel="edge", tezlik_moljal=1.0,
                                       barqarorlik=0.5, jonlilik=0.5, asl_audio_yol=None):
    """Har tarjima segmentidan OVOZ yaratadi, ORIGINAL VAQTIGA moslab
    (kerak bo'lsa tezlikni o'zgartirib), keyin BITTA uzluksiz audio faylga
    yig'adi — video davomiyligiga teng. (chiqish_yol, xato)

    segmentlar_tarjima: [(boshlanish, tugash, tarjima_matni), ...]
    dvigatel: "edge" (umumiy ovoz, ovoz_nomi=edge-tts nomi) yoki
              "elevenlabs" (klonlangan ovoz, ovoz_nomi=voice_id)
    tezlik_moljal: 1.0=oddiy, <1.0=sekinroq, >1.0=tezroq.
        DIQQAT: 1.0 dan farqli qiymat videoga MUKAMMAL sinxronlikni
        biroz buzishi mumkin — tezlik ustunligi beriladi.
    asl_audio_yol: berilsa, gapirmagan joylar ASL VIDEODAGI fon bilan
        to'ldiriladi (mutlaq sukunat o'rniga) — ovoz TABIIYROQ chiqadi.
        Berilmasa yoki kesib ololmasa — sukunatga tushadi (zaxira).
    """
    parchalar = []          # concat ro'yxatiga kiradigan fayllar
    oxirgi_joy = 0.0         # oxirgi qo'yilgan parchaning tugash vaqti

    def _bushliqni_toldir(boshlanish, davomiylik, nom):
        """Asl fon bilan (bo'lsa) yoki sukunat bilan (zaxira) to'ldiradi.
        Muvaffaqiyatli bo'lsa fayl yo'lini, aks holda None qaytaradi."""
        yol = os.path.join(papka, f"{nom}.mp3")
        if asl_audio_yol and _asl_fon_ol(asl_audio_yol, boshlanish, davomiylik, yol):
            return yol
        yol2 = os.path.join(papka, f"{nom}_suk.mp3")
        if _sukunat_yarat(davomiylik, yol2):
            return yol2
        return None

    for idx, (boshi, tugash, matn) in enumerate(segmentlar_tarjima):
        if not matn.strip():
            continue

        # 1) Boshlanishgacha bo'lgan bo'shliqni ASL FON bilan to'ldiramiz
        #    (mutlaq sukunat emas — shu sabab tabiiyroq chiqadi)
        bushliq = boshi - oxirgi_joy
        if bushliq > 0.05:
            suk_yol = _bushliqni_toldir(oxirgi_joy, bushliq, f"fon_{idx}")
            if suk_yol:
                parchalar.append(suk_yol)

        # 2) Ushbu segment uchun ovoz yaratamiz (oddiy tezlikda)
        xom_yol = os.path.join(papka, f"seg_{idx}_xom.mp3")
        try:
            if dvigatel == "elevenlabs":
                ok, _ = await elevenlabs_ovoz_fayl(matn, ovoz_nomi, xom_yol,
                                                    barqarorlik, jonlilik)
                if not ok:
                    continue
            else:
                import edge_tts
                com = edge_tts.Communicate(matn, ovoz_nomi)
                await com.save(xom_yol)
        except Exception as e:
            print(f"[dublyaj] segment {idx} ovoz xatosi: {e}")
            continue
        if not os.path.exists(xom_yol) or os.path.getsize(xom_yol) == 0:
            continue

        # 3) Uzunlikni ORIGINAL segment davomiyligiga moslaymiz
        #    (foydalanuvchi tezlik moljali bilan birga)
        maqsad_davomiylik = max(0.3, (tugash - boshi) / max(0.5, min(2.0, tezlik_moljal)))
        haqiqiy_davomiylik = _davomiylik_ol(xom_yol)
        yakuniy_yol = os.path.join(papka, f"seg_{idx}.mp3")

        if haqiqiy_davomiylik > 0.05:
            tezlik = haqiqiy_davomiylik / maqsad_davomiylik
            if 0.97 <= tezlik <= 1.03:
                # deyarli bir xil — tezlikni o'zgartirish shart emas
                os.replace(xom_yol, yakuniy_yol)
            elif _tezlikni_moslash(xom_yol, yakuniy_yol, tezlik):
                pass   # muvaffaqiyatli moslashtirildi
            else:
                os.replace(xom_yol, yakuniy_yol)   # moslay olmadik — asl holicha qoldiramiz
        else:
            os.replace(xom_yol, yakuniy_yol)

        parchalar.append(yakuniy_yol)
        oxirgi_joy = boshi + _davomiylik_ol(yakuniy_yol)

    # 4) Video oxirigacha ASL FON bilan to'ldiramiz
    if jami_davomiylik - oxirgi_joy > 0.05:
        suk_oxir = _bushliqni_toldir(oxirgi_joy, jami_davomiylik - oxirgi_joy, "fon_oxir")
        if suk_oxir:
            parchalar.append(suk_oxir)

    if not parchalar:
        return (None, "Hech qanday ovoz segmenti yaratilmadi")

    # 5) Barchasini BITTA audio faylga ketma-ket ulaymiz
    royxat_yol = os.path.join(papka, "concat_royxat.txt")
    with open(royxat_yol, "w", encoding="utf-8") as f:
        for p in parchalar:
            f.write(f"file '{os.path.abspath(p)}'\n")

    yakuniy_audio = os.path.join(papka, "yigindi_audio.mp3")
    try:
        natija = subprocess.run(
            ["ffmpeg", "-y", "-f", "concat", "-safe", "0", "-i", royxat_yol,
             "-c:a", "libmp3lame", "-q:a", "2", yakuniy_audio],
            capture_output=True, text=True, timeout=120)
        if natija.returncode != 0:
            return (None, (natija.stderr or "")[-300:])
        if not os.path.exists(yakuniy_audio):
            return (None, "Yig'indi audio yaratilmadi")
        return (yakuniy_audio, None)
    except Exception as e:
        return (None, str(e)[:300])
