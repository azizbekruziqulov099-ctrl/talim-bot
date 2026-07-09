try:
    from auto_trainer import auto_train_scheduler, train_all_profiles
except Exception as _ate:
    print(f"auto_trainer import xato: {_ate}")
    async def auto_train_scheduler(): pass
    async def train_all_profiles(*a, **k): return {}
from admin_handlers import *
from learning import *
from generator_handlers import *
from test_engine import *
from register import *
from learning import continue_learning
import asyncio
from aiogram.types import ReplyKeyboardRemove
from aiogram import Bot, Dispatcher, types
from urllib.parse import quote
from aiogram.filters import *
from dts_import_handlers import *
from ai_generatori import *
from storage import user_state, temp_user, registration_message, reg_kbd_message, admin_state
from keyboards import get_main_keyboard
from loader import dp, bot
from test_generator import save_test
from aiogram import F
import pandas as pd
from aiogram.fsm.context import FSMContext
from aiogram.types import (
    ReplyKeyboardMarkup,
    KeyboardButton,
    InlineKeyboardMarkup,
    InlineKeyboardButton
)
from learning import (
    lesson_next,
    lesson_prev,
    lesson_tts,
    lesson_tts_help,
    lesson_back_main,
    lesson_consolidation_test,
    lesson_test_answer,
    send_test_question,
    open_teacher_lesson,
    continue_learning
)
import lesson_admin
import student_dashboard
import json
import random
import time
import edge_tts
from aiogram.types import FSInputFile
import psycopg2
try:
    from db_pool import db as _db_pool, release as _db_release, get_user_cached, invalidate_user
    _pool_available = True
except:
    _db_pool=None; _db_release=None; _pool_available=False

def _get_db_conn():
    """Connection pool dan ulanish olish."""
    if _pool_available and _db_pool:
        try: return _db_pool()
        except: pass
    return psycopg2.connect(DATABASE_URL)
import re
import os
import subprocess
from openpyxl import Workbook
from openpyxl.styles import Font

with open("regions.json", "r", encoding="utf-8") as f:
    REGIONS = json.load(f)

ADMINS = [401251407]

# ═══ MULTI-ACCOUNT: effektiv ID ═══
# Har akkauntga alohida ichki ID: idx=0 -> telegram_id, idx=N -> telegram_id*1000+N
_EFF_CACHE = {}   # {telegram_id: eff_uid}

def _eff_uid(telegram_id):
    """Aktiv akkauntning ichki ID si.
    user_accounts.uid ustunidan o'qiladi — telefonga bog'liq emas.
    Eski qatorlarda uid bo'lmasa eski qoida bilan hisoblanadi."""
    if telegram_id in _EFF_CACHE:
        return _EFF_CACHE[telegram_id]
    eff = telegram_id
    try:
        conn = _get_db_conn(); cur = conn.cursor()
        nomzod = None
        try:
            cur.execute("""SELECT uid, account_index FROM user_accounts
                WHERE telegram_id=%s AND is_active=TRUE LIMIT 1""", (telegram_id,))
            r = cur.fetchone()
            if r:
                _uid, _idx = r
                if _uid:
                    nomzod = int(_uid)
                elif _idx:
                    nomzod = telegram_id * 1000 + int(_idx)
        except Exception:
            conn.rollback()
            cur.execute("""SELECT account_index FROM user_accounts
                WHERE telegram_id=%s AND is_active=TRUE LIMIT 1""", (telegram_id,))
            r = cur.fetchone()
            if r and r[0]:
                nomzod = telegram_id * 1000 + int(r[0])

        if nomzod and nomzod != telegram_id:
            cur.execute("SELECT 1 FROM users WHERE user_id=%s LIMIT 1", (nomzod,))
            if cur.fetchone():
                eff = nomzod
            else:
                print(f"[eff_uid] {nomzod} da users yo'q -> {telegram_id}")
        cur.close(); conn.close()
    except Exception as e:
        print(f"[eff_uid] {e}")
    _EFF_CACHE[telegram_id] = eff
    return eff


# uid -> telegram_id keshi (xabar yuborish uchun)
_TG_CACHE = {}

def _tg_id(uid):
    """Ichki ID dan haqiqiy telegram ID.
    Akkaunt ko'chirilgan bo'lishi mumkin — DB dan qidiramiz."""
    if uid in _TG_CACHE:
        return _TG_CACHE[uid]
    tg = uid // 1000 if uid > 10_000_000_000 else uid
    try:
        conn = _get_db_conn(); cur = conn.cursor()
        cur.execute("""SELECT telegram_id FROM user_accounts WHERE uid=%s
            ORDER BY is_active DESC LIMIT 1""", (uid,))
        r = cur.fetchone()
        if r and r[0]:
            tg = int(r[0])
        cur.close(); conn.close()
    except Exception:
        pass
    _TG_CACHE[uid] = tg
    return tg

def _eff_clear(telegram_id):
    """Akkaunt almashganda keshni tozalaymiz."""
    _EFF_CACHE.pop(telegram_id, None)
    _TG_CACHE.clear()

def _is_admin(uid):
    return _tg_id(uid) in ADMINS

def _get_user_qisqa(uid):
    """(ism, rol, sinf)"""
    try:
        c = _get_db_conn(); cr = c.cursor()
        cr.execute("SELECT full_name, role, class FROM users WHERE user_id=%s", (uid,))
        r = cr.fetchone(); cr.close(); c.close()
        return r if r else (None, None, None)
    except Exception:
        return (None, None, None)

def _mavzu_hash(nom):
    """Mavzu nomining qisqa belgisi — callback_data ga sig'ishi uchun."""
    import hashlib
    return hashlib.md5((nom or "").strip().encode()).hexdigest()[:6]

# Excel dan o'qilgan rasm tavsiflari: {image_id: description_en}
_RASM_TAVSIF = {}
# Limit tugagan foydalanuvchilar (ertaga davom etish uchun)
_IMG_PENDING = {}
# O'chirish tasdig'i kutayotgan so'rovlar: {uid: (shart, args, izoh, soni)}
_ochir_soqi = {}

DATABASE_URL = os.getenv("DATABASE_URL")
API_TOKEN = os.getenv("BOT_TOKEN")

conn = _get_db_conn()
cur = conn.cursor()

user_locks = {}
state_history = {}
generator_process = None
topic_stats_state = {}

wb = Workbook()

# README
ws1 = wb.active
ws1.title = "README"

ws1["A1"] = "DIFFICULTY"
ws1["A1"].font = Font(bold=True)

ws1["A2"] = "oson"
ws1["A3"] = "o'rta"
ws1["A4"] = "qiyin"
ws1["A5"] = "murakkab"

ws1["C1"] = "QUESTION_TYPE"
ws1["C1"].font = Font(bold=True)

ws1["C2"] = "single_choice"
ws1["C3"] = "multiple_choice"
ws1["C4"] = "true_false"
ws1["C5"] = "write_answer"
ws1["C6"] = "image_question"

ws1["E1"] = "LIFE_LEVEL"
ws1["E1"].font = Font(bold=True)

ws1["E2"] = "0"
ws1["E3"] = "1"
ws1["E4"] = "2"
ws1["E5"] = "3"
ws1["E6"] = "4"

ws1["G1"] = "LANGUAGE"
ws1["G1"].font = Font(bold=True)

ws1["G2"] = "uz"
ws1["G3"] = "ru"
ws1["G4"] = "en"

# TESTLAR
ws2 = wb.create_sheet("TESTLAR")

headers = [
    "topic_code",
    "difficulty",
    "situation",
    "question",
    "option_a",
    "option_b",
    "option_c",
    "option_d",
    "correct_answer",
    "explanation",
    "question_type",
    "is_latex",
    "image_url",
    "audio_text",
    "language",
    "life_level",
    "age_group",
    "time_limit"
]

for col_num, header in enumerate(headers, start=1):
    cell = ws2.cell(row=1, column=col_num)
    cell.value = header
    cell.font = Font(bold=True)

# NAMUNA
ws3 = wb.create_sheet("NAMUNA")

ws3.append(headers)

ws3.append([
    "_1_01_01_01_01_001",
    "oson",
    "oddiy",
    "What animal gives us milk?",
    "Cow",
    "Cat",
    "Dog",
    "Horse",
    "Cow",
    "A cow gives milk",
    "single_choice",
    False,
    "",
    "",
    "uz",
    1,
    "10-11",
    60
])

wb.save("test_import_template.xlsx")

print("test_import_template.xlsx yaratildi")

# BUTTON ID (faqat shu bilan ishlaymiz)
BTN_SURVEY = "survey"
BTN_STATS = "stats"

BTN_MY = "my_stats"
BTN_GLOBAL = "global_stats"

BACK = "🔙 Ortga"
HOME = "🏠 Bosh menyu"
HOME2 = "🏠 Bosh ekran"

SUBJECTS_BY_LEVEL = {

    "👶 Boshlang‘ich": [
        "Matematika",
        "Ona tili",
        "O‘qish",
        "Tabiiy fan"
    ],

    "📘 O‘rta": [
        "Algebra",
        "Geometriya",
        "Fizika",
        "Kimyo",
        "Biologiya",
        "Tarix",
        "Geografiya"
    ],

    "🎓 Yuqori": [
        "Algebra",
        "Geometriya",
        "Fizika",
        "Kimyo",
        "Biologiya",
        "Huquq",
        "Iqtisod",
        "Informatika"
    ]
}

SUBJECTS_BY_CLASS = {

    "0-sinf": [
        ["🔤 O‘zbekcha so‘zlar","🇬🇧 English Kids"],
        ["🇷🇺 Русский детям","🔢 Sonlar olami"],
        ["🎨 Ranglar","🐶 Hayvonlar"],
        ["🧩 Jumboqlar","👀 Diqqat"],
        ["🎯 Mantiqiy o‘yinlar"]
    ]
}

# 1-4 sinf
PRIMARY_SUBJECTS = [
    ["Matematika", "Ona tili", "O‘qish"],
    ["Ingliz tili", "Tabiiy fan", "Tarbiya"],
    ["Musiqa", "Rasm"]
]

# 5-6 sinf
MIDDLE_SUBJECTS = [
    ["Matematika", "Ona tili", "Adabiyot"],
    ["Ingliz tili", "Rus tili", "Tarix"],
    ["Biologiya", "Geografiya", "Informatika"],
    ["Texnologiya"]
]

# 7-9 sinf
UPPER_SUBJECTS = [
    ["Algebra", "Geometriya", "Fizika"],
    ["Kimyo", "Biologiya", "Tarix"],
    ["Geografiya", "Informatika", "Ingliz tili"],
    ["Ona tili", "Adabiyot"]
]

# 10-11 sinf
HIGH_SUBJECTS = [
    ["Algebra", "Geometriya", "Fizika"],
    ["Kimyo", "Biologiya", "Tarix"],
    ["Huquq", "Iqtisod", "Geografiya"],
    ["Informatika", "Ingliz tili", "Ona tili"],
    ["Adabiyot"]
]

ZERO_TEST_TYPES = [
    "🔤 Harflar",
    "📖 So‘zlar",
    "🖼 Rasmli o‘yin",
    "🎵 Eshit va top",
    "🎁 Aralash"
]

def set_state(user_id, state):

    user_state[user_id] = state

    if user_id not in state_history:
        state_history[user_id] = []

    state_history[user_id].append(state)

def base_keyboard(extra=[]):

    keyboard = []
    row = []

    for i, item in enumerate(extra, start=1):

        row.append(KeyboardButton(text=item))

        # har 3 tadan keyin yangi qator
        if i % 4 == 0:
            keyboard.append(row)
            row = []

    # qolganlari
    if row:
        keyboard.append(row)

    # pastki tugmalar
    keyboard.append([KeyboardButton(text=BACK)])
    keyboard.append([KeyboardButton(text=HOME)])

    return ReplyKeyboardMarkup(
        keyboard=keyboard,
        resize_keyboard=True
    )

def make_keyboard(items):
    return ReplyKeyboardMarkup(
        keyboard=[[KeyboardButton(text=i)] for i in items],
        resize_keyboard=True
    )

def question_nav_keyboard(test):

    return ReplyKeyboardMarkup(
        keyboard=[
            [
                KeyboardButton(text="➡️"),
                KeyboardButton(text=FINISH)
            ]
        ],
        resize_keyboard=True
    )

def get_stats_keyboard():
    return ReplyKeyboardMarkup(
        keyboard=[
            [KeyboardButton(text="📚 DTS")],
            [KeyboardButton(text="📈 Umumiy statistika")],
            [KeyboardButton(text=BACK)],
            [KeyboardButton(text=HOME)]
        ],
        resize_keyboard=True
    )

def check_survey(user_id):

    conn = _get_db_conn()
    cur = conn.cursor()

    cur.execute("""
    SELECT survey_done
    FROM users
    WHERE user_id=%s
    """, (user_id,))

    row = cur.fetchone()

    conn.close()

    if not row:
        return False

    return row[0] == 1

def init_db():
    conn = _get_db_conn()
    cur = conn.cursor()

    # user_id INTEGER -> BIGINT (Telegram ID lari INTEGER chegarasidan oshib ketadi)
    try:
        cur.execute("ALTER TABLE users ALTER COLUMN user_id TYPE BIGINT")
        cur.execute("ALTER TABLE survey_answers ALTER COLUMN user_id TYPE BIGINT")
        conn.commit()
    except Exception:
        conn.rollback()
    # Books jadvali
    try:
        cur.execute("""CREATE TABLE IF NOT EXISTS books (
            id SERIAL PRIMARY KEY, title TEXT, fan TEXT, sinf TEXT,
            muallif TEXT, total_pages INT DEFAULT 0,
            created_at TIMESTAMP DEFAULT NOW()
        )""")
        conn.commit()
    except: conn.rollback()
    for _add_col in [
        "ALTER TABLE books ADD COLUMN IF NOT EXISTS total_pages INT DEFAULT 0",
        "ALTER TABLE books ADD COLUMN IF NOT EXISTS fan TEXT",
        "ALTER TABLE books ADD COLUMN IF NOT EXISTS sinf TEXT",
        "ALTER TABLE books ADD COLUMN IF NOT EXISTS muallif TEXT",
    ]:
        try: cur.execute(_add_col); conn.commit()
        except: conn.rollback()
    # book_pages, book_exercises
    try:
        cur.execute("""CREATE TABLE IF NOT EXISTS book_pages (
            id SERIAL PRIMARY KEY, book_id INT, page_num INT,
            section_name TEXT, full_text TEXT, exercise_count INT DEFAULT 0,
            UNIQUE(book_id,page_num)
        )""")
        cur.execute("""CREATE TABLE IF NOT EXISTS book_exercises (
            id SERIAL PRIMARY KEY, book_id INT, page_num INT,
            mavzu TEXT, fan TEXT, sinf TEXT,
            ex_type TEXT DEFAULT 'misol', savol TEXT, qiyinlik TEXT DEFAULT 'orta'
        )""")
        conn.commit()
    except: conn.rollback()

    for _col_sql in [
        'ALTER TABLE users ADD COLUMN IF NOT EXISTS kindergarten TEXT',
        'ALTER TABLE users ADD COLUMN IF NOT EXISTS "group" TEXT',
    ]:
        try: cur.execute(_col_sql); conn.commit()
        except Exception: conn.rollback()

    cur.execute("""
    CREATE TABLE IF NOT EXISTS survey_answers (
        id SERIAL PRIMARY KEY,
        user_id INTEGER,
        survey_id INTEGER,
        answer TEXT
    )
    """)

    cur.execute("""
    CREATE TABLE IF NOT EXISTS images(
        id SERIAL PRIMARY KEY,
        name TEXT UNIQUE,
        file_id TEXT
    )
    """)

    # USERS
    cur.execute("""
    CREATE TABLE IF NOT EXISTS users (
        id SERIAL PRIMARY KEY,
        user_id INTEGER,
        full_name TEXT,
        role TEXT,
        region TEXT,
        district TEXT,
        school TEXT,
        class TEXT,
        survey_done INTEGER DEFAULT 0
    )
    """)

    # SURVEY
    cur.execute("""
    CREATE TABLE IF NOT EXISTS surveys (
        id SERIAL PRIMARY KEY,
        role TEXT,
        question TEXT,
        q_type TEXT,
        a TEXT,
        b TEXT,
        c TEXT,
        d TEXT
    )
    """)

    cur.execute("""
    CREATE TABLE IF NOT EXISTS results (
        id SERIAL PRIMARY KEY,
        user_id INTEGER,
        class TEXT,
        subject TEXT,
        score INTEGER,
        total INTEGER,
        region TEXT,
        district TEXT,
        school TEXT,
        role TEXT
    )
    """)

    # subject ustuni — eski DB uchun
    cur.execute("""
        ALTER TABLE users 
        ADD COLUMN IF NOT EXISTS subject TEXT
    """)

    # Dars tarixi — takrorlash va statistika uchun
    cur.execute("""
    CREATE TABLE IF NOT EXISTS lesson_history (
        id          SERIAL PRIMARY KEY,
        user_id     INTEGER NOT NULL,
        topic_code  TEXT NOT NULL,
        mavzu       TEXT,
        fan         TEXT,
        score       INTEGER DEFAULT 0,
        total       INTEGER DEFAULT 0,
        learned_at  TIMESTAMP DEFAULT NOW()
    )
    """)

    cur.execute("""
        CREATE INDEX IF NOT EXISTS idx_lesson_history_user
        ON lesson_history(user_id, learned_at DESC)
    """)

    cur.execute("""
    CREATE TABLE IF NOT EXISTS error_log (
        id         SERIAL PRIMARY KEY,
        user_id    BIGINT,
        username   TEXT,
        error_text TEXT,
        created_at TIMESTAMP DEFAULT NOW(),
        is_read    BOOLEAN DEFAULT FALSE
    )
    """)

    # ===== Yetishmayotgan 10 ta jadval (auto-create, idempotent) =====
    cur.execute("""
    ALTER TABLE users ADD COLUMN IF NOT EXISTS gender     TEXT
    """)
    cur.execute("""
    ALTER TABLE users ADD COLUMN IF NOT EXISTS birth_date DATE
    """)

    cur.execute("""
    CREATE TABLE IF NOT EXISTS dts_tree (
        id SERIAL PRIMARY KEY,
        topic_code TEXT UNIQUE,
        grade TEXT, subject_code TEXT, subject_name TEXT, quarter TEXT,
        bob_code TEXT, bob_name TEXT, bolim_code TEXT, bolim_name TEXT,
        mavzu_code TEXT, mavzu_name TEXT, kichik_code TEXT, kichik_name TEXT,
        is_deleted BOOLEAN DEFAULT FALSE
    )
    """)
    cur.execute("CREATE INDEX IF NOT EXISTS idx_dts_tree_grade_subj ON dts_tree (grade, subject_code, quarter)")
    cur.execute("CREATE INDEX IF NOT EXISTS idx_dts_tree_active ON dts_tree (is_deleted)")

    cur.execute("""
    CREATE TABLE IF NOT EXISTS generated_tests (
        id SERIAL PRIMARY KEY,
        topic_code TEXT, difficulty TEXT, situation TEXT, question TEXT,
        option_a TEXT, option_b TEXT, option_c TEXT, option_d TEXT,
        correct_answer TEXT, explanation TEXT, question_type TEXT,
        is_latex BOOLEAN DEFAULT FALSE, image_url TEXT, audio_text TEXT,
        language TEXT, life_level TEXT, age_group TEXT, time_limit INTEGER
    )
    """)
    # Rasm file_id ustuni (eski DB uchun)
    cur.execute("ALTER TABLE generated_tests ADD COLUMN IF NOT EXISTS image_file_id TEXT")
    # Kitob va bilim jadvallari
    for _tbl in [
        "CREATE TABLE IF NOT EXISTS books (id SERIAL PRIMARY KEY, title TEXT, fan TEXT, sinf TEXT, muallif TEXT, file_id TEXT, created_at TIMESTAMP DEFAULT NOW())",
        "CREATE TABLE IF NOT EXISTS book_sections (id SERIAL PRIMARY KEY, book_id INT, title TEXT, from_page INT, to_page INT, tartib INT)",
        "CREATE TABLE IF NOT EXISTS book_chunks (id SERIAL PRIMARY KEY, section_id INT, book_id INT, matn TEXT, latex TEXT, chunk_type TEXT, page_num INT, keywords TEXT)",
        "CREATE TABLE IF NOT EXISTS knowledge_facts (id SERIAL PRIMARY KEY, mavzu TEXT, fan TEXT, sinf TEXT, fact_type TEXT, savol TEXT, javob TEXT, izoh TEXT, yosh_5_7 TEXT, yosh_8_11 TEXT, yosh_12plus TEXT, source_ai TEXT, quality INT DEFAULT 5, book_id INT, chunk_text TEXT, keywords TEXT, created_at TIMESTAMP DEFAULT NOW())",
        "CREATE TABLE IF NOT EXISTS weak_topics (id SERIAL PRIMARY KEY, mavzu TEXT, fan TEXT, error_count INT DEFAULT 1, last_error TIMESTAMP DEFAULT NOW(), UNIQUE(mavzu, fan))",
        "CREATE TABLE IF NOT EXISTS books (id SERIAL PRIMARY KEY, title TEXT, fan TEXT, sinf TEXT, muallif TEXT, total_pages INT, created_at TIMESTAMP DEFAULT NOW())",
        "CREATE TABLE IF NOT EXISTS book_pages (id SERIAL PRIMARY KEY, book_id INT, page_num INT, section_name TEXT, full_text TEXT, exercise_count INT DEFAULT 0, UNIQUE(book_id,page_num))",
        "CREATE TABLE IF NOT EXISTS book_exercises (id SERIAL PRIMARY KEY, book_id INT, page_num INT, mavzu TEXT, fan TEXT, sinf TEXT, ex_type TEXT DEFAULT 'misol', savol TEXT, qiyinlik TEXT DEFAULT 'orta')",
        "CREATE TABLE IF NOT EXISTS rasm_queue (id SERIAL PRIMARY KEY, user_id BIGINT, kod TEXT, done BOOLEAN DEFAULT FALSE, created_at TIMESTAMP DEFAULT NOW())",
        "CREATE TABLE IF NOT EXISTS answer_feedback (id SERIAL PRIMARY KEY, question TEXT, answer_given TEXT, correct_answer TEXT, was_correct BOOLEAN, user_id BIGINT, created_at TIMESTAMP DEFAULT NOW())",
        "CREATE TABLE IF NOT EXISTS book_exercises (id SERIAL PRIMARY KEY, book_id INT, section_id INT, mavzu TEXT, fan TEXT, sinf TEXT, ex_type TEXT, savol TEXT, yechim TEXT, javob TEXT, formula TEXT, qiyinlik TEXT DEFAULT 'orta', source_ai TEXT, created_at TIMESTAMP DEFAULT NOW())",
    ]:
        try: cur.execute(_tbl); conn.commit()
        except: conn.rollback()
    cur.execute("""
    CREATE TABLE IF NOT EXISTS test_corrections (
        id SERIAL PRIMARY KEY,
        test_id INTEGER,
        topic_code TEXT,
        question TEXT,
        user_id BIGINT,
        comment TEXT,
        status TEXT DEFAULT 'new',
        created_at TIMESTAMP DEFAULT NOW()
    )
    """)
    cur.execute("CREATE INDEX IF NOT EXISTS idx_gen_tests_topic ON generated_tests (topic_code)")
    cur.execute("CREATE INDEX IF NOT EXISTS idx_gen_tests_topic_diff ON generated_tests (topic_code, difficulty)")

    cur.execute("""
    CREATE TABLE IF NOT EXISTS teacher_lessons (
        id SERIAL PRIMARY KEY,
        topic_code TEXT UNIQUE,
        intro      TEXT, image_intro TEXT,
        part_1     TEXT, image_1 TEXT,
        part_2     TEXT, image_2 TEXT,
        part_3     TEXT, image_3 TEXT,
        part_4     TEXT, image_4 TEXT,
        part_5     TEXT, image_5 TEXT,
        part_6     TEXT, image_6 TEXT,
        part_7     TEXT, image_7 TEXT,
        simple_1   TEXT, simple_2 TEXT, simple_3 TEXT, simple_4 TEXT,
        simple_5   TEXT, simple_6 TEXT, simple_7 TEXT,
        example_1  TEXT, example_2 TEXT, example_3 TEXT,
        example_4  TEXT, example_5 TEXT,
        summary    TEXT
    )
    """)

    # Eski jadvalga yangi ustunlar qo'shish (mavjud ma'lumotlarga tegmaydi)
    for _col in [
        "image_intro TEXT", "image_1 TEXT", "image_2 TEXT", "image_3 TEXT",
        "image_4 TEXT",     "image_5 TEXT", "image_6 TEXT", "image_7 TEXT",
        "part_5 TEXT",  "part_6 TEXT",  "part_7 TEXT",
        "simple_5 TEXT","simple_6 TEXT","simple_7 TEXT",
        "example_3 TEXT","example_4 TEXT","example_5 TEXT",
        "image_e_1 TEXT","image_e_2 TEXT","image_e_3 TEXT",
        "image_e_4 TEXT","image_e_5 TEXT",
    ]:
        try:
            cur.execute(f"ALTER TABLE teacher_lessons ADD COLUMN IF NOT EXISTS {_col}")
            conn.commit()
        except Exception:
            conn.rollback()

    cur.execute("""
    CREATE TABLE IF NOT EXISTS learned_topics (
        id SERIAL PRIMARY KEY,
        user_id BIGINT NOT NULL,
        topic_code TEXT NOT NULL,
        score INTEGER DEFAULT 0,
        repeat_count INTEGER DEFAULT 0,
        learned_at TIMESTAMP DEFAULT NOW(),
        next_repeat DATE,
        UNIQUE (user_id, topic_code)
    )
    """)
    cur.execute("CREATE INDEX IF NOT EXISTS idx_learned_user ON learned_topics (user_id)")
    cur.execute("CREATE INDEX IF NOT EXISTS idx_learned_repeat ON learned_topics (next_repeat)")

    cur.execute("""
    CREATE TABLE IF NOT EXISTS user_progress (
        user_id BIGINT PRIMARY KEY,
        xp INTEGER DEFAULT 0,
        streak INTEGER DEFAULT 0,
        last_active DATE
    )
    """)

    cur.execute("""
    CREATE TABLE IF NOT EXISTS lesson_progress (
        id SERIAL PRIMARY KEY,
        user_id BIGINT NOT NULL UNIQUE,
        topic_code TEXT,
        current_step INTEGER DEFAULT 0,
        completed BOOLEAN DEFAULT FALSE
    )
    """)
    cur.execute("CREATE INDEX IF NOT EXISTS idx_lesson_prog_user ON lesson_progress (user_id)")
    # Agar eski jadval UNIQUE siz bo'lsa — qo'shamiz
    try:
        cur.execute("ALTER TABLE lesson_progress ADD CONSTRAINT lesson_progress_user_id_key UNIQUE (user_id)")
        conn.commit()
    except Exception:
        conn.rollback()

    cur.execute("""
    CREATE TABLE IF NOT EXISTS achievements (
        id SERIAL PRIMARY KEY,
        user_id BIGINT NOT NULL,
        badge_code TEXT NOT NULL,
        earned_at TIMESTAMP DEFAULT NOW(),
        UNIQUE (user_id, badge_code)
    )
    """)

    cur.execute("""
    CREATE TABLE IF NOT EXISTS exams (
        id SERIAL PRIMARY KEY,
        title TEXT, grade TEXT, exam_date DATE,
        is_mandatory BOOLEAN DEFAULT FALSE,
        created_by BIGINT
    )
    """)

    cur.execute("""
    CREATE TABLE IF NOT EXISTS exam_results (
        id SERIAL PRIMARY KEY,
        user_id BIGINT NOT NULL,
        exam_id INTEGER REFERENCES exams(id) ON DELETE CASCADE,
        status TEXT DEFAULT 'pending',
        UNIQUE (user_id, exam_id)
    )
    """)
    cur.execute("CREATE INDEX IF NOT EXISTS idx_exam_results_user ON exam_results (user_id)")

    cur.execute("""
    CREATE TABLE IF NOT EXISTS topic_generation (
        topic_code TEXT PRIMARY KEY,
        current_count INTEGER DEFAULT 0,
        target_count INTEGER DEFAULT 10,
        last_generated_at TIMESTAMP
    )
    """)
    # ===== Yetishmayotgan jadvallar tugadi =====

    conn.commit()
    conn.close()

def get_grades():

    conn = _get_db_conn()
    cur = conn.cursor()

    cur.execute("""
        SELECT grade FROM (
            SELECT DISTINCT grade FROM dts_tree
        ) _g
        ORDER BY
            CASE WHEN grade ~ '^[0-9]+$' THEN grade::int ELSE 9999 END,
            grade
    """)

    rows = cur.fetchall()

    cur.close()
    conn.close()

    return [row[0] for row in rows]

def get_subjects_by_grade(grade):

    conn = _get_db_conn()
    cur = conn.cursor()

    cur.execute("""
        SELECT DISTINCT subject_name
        FROM dts_tree
        WHERE grade = %s
        ORDER BY subject_name
    """, (grade,))

    rows = cur.fetchall()

    cur.close()
    conn.close()

    return [row[0] for row in rows]

def get_topics_by_grade_subject(
    grade,
    subject_name
):

    conn = _get_db_conn()
    cur = conn.cursor()

    cur.execute("""
        SELECT
            topic_code,
            kichik_name
        FROM dts_tree
        WHERE grade = %s
        AND subject_name = %s
        ORDER BY topic_code
    """, (
        grade,
        subject_name
    ))

    rows = cur.fetchall()

    cur.close()
    conn.close()

    return rows

def get_topic_name(topic_code):

    conn = _get_db_conn()
    cur = conn.cursor()

    cur.execute("""
        SELECT
            grade,
            subject_name,
            quarter,
            bob_name,
            bolim_name,
            mavzu_name,
            kichik_name
        FROM dts_tree
        WHERE topic_code=%s
        LIMIT 1
    """, (topic_code,))

    row = cur.fetchone()

    cur.close()
    conn.close()

    return row

def get_topic_test_count(topic_code):

    conn = _get_db_conn()
    cur = conn.cursor()

    cur.execute("""
        SELECT COUNT(*)
        FROM generated_tests
        WHERE topic_code=%s
    """, (topic_code,))

    count = cur.fetchone()[0]

    cur.close()
    conn.close()

    return count

def get_topic_statistics(topic_code):

    conn = _get_db_conn()
    cur = conn.cursor()

    cur.execute("""
        SELECT

            COUNT(*) as total,

            SUM(
                CASE
                WHEN difficulty='oson'
                THEN 1
                ELSE 0
                END
            ),

            SUM(
                CASE
                WHEN difficulty='o''rta'
                THEN 1
                ELSE 0
                END
            ),

            SUM(
                CASE
                WHEN difficulty='qiyin'
                THEN 1
                ELSE 0
                END
            ),

            SUM(
                CASE
                WHEN difficulty='murakkab'
                THEN 1
                ELSE 0
                END
            ),

            SUM(
                CASE
                WHEN question_type='single_choice'
                THEN 1
                ELSE 0
                END
            ),

            SUM(
                CASE
                WHEN question_type='multiple_choice'
                THEN 1
                ELSE 0
                END
            ),

            SUM(
                CASE
                WHEN question_type='true_false'
                THEN 1
                ELSE 0
                END
            ),

            SUM(
                CASE
                WHEN question_type='write_answer'
                THEN 1
                ELSE 0
                END
            )

        FROM generated_tests
        WHERE topic_code=%s
    """, (topic_code,))

    row = cur.fetchone()

    cur.close()
    conn.close()

    return row

async def show_topics_page(
    message,
    user_id
):

    data = topic_stats_state[user_id]

    topics = data["topics"]
    page = data["page"]

    start = page * 10
    end = start + 10

    current_topics = topics[start:end]

    keyboard = []

    text = "📚 MAVZULAR\n\n"

    for row in current_topics:

        topic_code = row[0]
        kichik_name = row[1]

        count = get_topic_test_count(
            topic_code
        )

        keyboard.append([
            KeyboardButton(
                text=f"{kichik_name} ({count})"
            )
        ])

        topic_stats_state[user_id][
            f"topic_{kichik_name} ({count})"
        ] = topic_code

    keyboard.append([
        KeyboardButton(text="⬅️ Oldingi"),
        KeyboardButton(text="➡️ Keyingi")
    ])

    keyboard.append([
        KeyboardButton(text="🔍 Mavzu qidirish")
    ])

    keyboard.append([
        KeyboardButton(text="🔙 Ortga")
    ])

    await message.answer(
        text,
        reply_markup=ReplyKeyboardMarkup(
            keyboard=keyboard,
            resize_keyboard=True
        )
    )
    
@dp.message(F.photo)
async def save_image(message: types.Message):
    if message.from_user.id not in ADMINS:
        return

    # ── Collage rasm split handler ──
    # Caption: "split:TC:start:rows:cols" yoki "split:TC:start" (default 5x6)
    # Misol: "split:3-02-3:1:5:6" → TC=3-02-3, start=1, 5 qator, 6 ustun
    uid_sp = message.from_user.id

    # ── Collage rasm split handler ──
    _col_cap = (message.caption or "").strip()
    if message.photo or message.document or (message.caption and message.caption.startswith('split:')):
        # rasm_queue da kutayotgan kodlar bormi?
        try:
            conn_rq2=_get_db_conn();cur_rq2=conn_rq2.cursor()
            cur_rq2.execute("""CREATE TABLE IF NOT EXISTS rasm_queue
                (id SERIAL PRIMARY KEY, user_id BIGINT, kod TEXT,
                 done BOOLEAN DEFAULT FALSE, created_at TIMESTAMP DEFAULT NOW())""")
            conn_rq2.commit()
            cur_rq2.execute("SELECT COUNT(*) FROM rasm_queue WHERE user_id=%s AND done=FALSE",(uid_sp,))
            pending=(cur_rq2.fetchone() or [0])[0]
            cur_rq2.close();conn_rq2.close()
        except: pending=0
    if pending>0:
        ROWS_sp = 5; COLS_sp = 6
        total_sp = 30

        # rasm_queue dan keyingi 30 ta kodni olish
        conn_rq3=_get_db_conn();cur_rq3=conn_rq3.cursor()
        cur_rq3.execute("""SELECT id,kod FROM rasm_queue
            WHERE user_id=%s AND done=FALSE ORDER BY id LIMIT 30""",(uid_sp,))
        rows_rq=cur_rq3.fetchall()
        cur_rq3.close();conn_rq3.close()

        if not rows_rq:
            await message.answer("✅ Barcha rasmlar allaqachon saqlandi!")
            return

        batch_ids_rq = [r[0] for r in rows_rq]
        batch_kodlar = [r[1] for r in rows_rq]
        ROWS_sp = 5; COLS_sp = 6
        total_sp = len(batch_kodlar)
        status_sp = await message.answer(f"✂️ Qirqilmoqda... {len(batch_kodlar)} ta")

        try:
            from PIL import Image as PILImage
            from io import BytesIO
            from aiogram.types import BufferedInputFile

            # Rasmni yuklab olish
            photo = message.photo[-1]
            buf_p = BytesIO()
            await message.bot.download(photo.file_id, destination=buf_p)
            buf_p.seek(0)
            img = PILImage.open(buf_p)
            W, H = img.size

            COLS, ROWS = COLS_sp, ROWS_sp
            cw, ch = W // COLS, H // ROWS
            saved_sp = 0

            for i, kod in enumerate(batch_kodlar):
                r, c = i // COLS, i % COLS
                x1, y1 = c * cw, r * ch
                x2, y2 = x1 + cw, y1 + ch
                cell = img.crop((x1, y1, x2, y2))
                buf_c = BytesIO()
                cell.save(buf_c, format="PNG")
                buf_c.seek(0)
                try:
                    sent_sp = await message.bot.send_photo(
                        message.chat.id,
                        BufferedInputFile(buf_c.read(), f"{kod}.png"),
                        caption=f"🖼 {kod}"
                    )
                    fid_sp = sent_sp.photo[-1].file_id
                    conn_sp = _get_db_conn(); cur_sp = conn_sp.cursor()
                    cur_sp.execute("""INSERT INTO images(name,file_id) VALUES(%s,%s)
                        ON CONFLICT(name) DO UPDATE SET file_id=EXCLUDED.file_id""",(kod,fid_sp))
                    cur_sp.execute("UPDATE rasm_queue SET done=TRUE WHERE user_id=%s AND kod=%s",(uid_sp,kod))
                    conn_sp.commit(); cur_sp.close(); conn_sp.close()
                    saved_sp += 1
                except Exception as _e:
                    print(f"cell {kod}: {_e}")

            conn_rq4=_get_db_conn();cur_rq4=conn_rq4.cursor()
            cur_rq4.execute("SELECT COUNT(*) FROM rasm_queue WHERE user_id=%s AND done=FALSE",(uid_sp,))
            remaining=(cur_rq4.fetchone() or [0])[0]
            cur_rq4.close();conn_rq4.close()

            if remaining > 0:
                await status_sp.edit_text(
                    f"✅ {saved_sp} ta saqlandi!\n"
                    f"📊 Qolgan: {remaining} ta\n\n"
                    f"Keyingi collage yuboring 👇"
                )
            else:
                await status_sp.edit_text(
                    f"🎉 Hammasi tayyor!\n"
                    f"✅ {saved_sp} ta + avvalgisi = {new_idx} ta jami saqlandi!"
                )
                admin_state.pop(f"{user_id}_rasm_kodlar", None)
                admin_state.pop(f"{user_id}_rasm_idx", None)

        except Exception as _e:
            await status_sp.edit_text(f"❌ Split xato: {_e}")
        return

    caption = (message.caption or "").strip()

    # Agar caption "split:TOPIC_CODE:rows:cols:prefix" formatida bo'lsa
    # Misol: split:1-01-1-01-01-01-001:1:6:p  → -p-1..-p-6
    #        split:1-01-1-01-01-01-001:1:5:e  → -e-1..-e-5
    # msplit:QATOR:USTUN:TC1,TC2,TC3,...
    # Misol: msplit:5:2:TC1,TC2,TC3,TC4,TC5 → har TC dan 10 ta panel
    if caption.lower().startswith("msplit:"):
        parts = caption[7:].strip().split(":")
        rows   = int(parts[0]) if parts else 5
        cols   = int(parts[1]) if len(parts)>1 else 4
        tcs    = [t.strip() for t in parts[2].split(",") if t.strip()] if len(parts)>2 else []
        prefix = parts[3].strip() if len(parts)>3 else ""
        if not tcs:
            await message.answer("❌ TC kodlar yo'q! Format: msplit:5:4:TC1,TC2,..."); return
        panel_per_tc = rows * cols
        total = panel_per_tc * len(tcs)
        sep = f"-{prefix}-" if prefix else "-"
        await message.answer(
            f"⏳ msplit: {rows}×{cols}={panel_per_tc} panel × {len(tcs)} TC = {total} bo'lak\n"
            f"TC lar: {', '.join(tcs[:3])}{'...' if len(tcs)>3 else ''}"
        )
        buf = await message.bot.download(message.photo[-1].file_id)
        from PIL import Image as PILImage
        from io import BytesIO
        buf.seek(0)
        img = PILImage.open(buf)
        total_cols = cols
        total_rows = rows * len(tcs)
        w, h = img.size
        pw = w // total_cols
        ph = h // total_rows

        conn2 = _get_db_conn(); cur2 = conn2.cursor()
        saved = 0
        for tc_idx, tc in enumerate(tcs):
            for r in range(rows):
                for c in range(cols):
                    n = r * cols + c + 1
                    global_row = tc_idx * rows + r
                    x1, y1 = c*pw, global_row*ph
                    x2, y2 = x1+pw, y1+ph
                    piece = img.crop((x1,y1,x2,y2))
                    pb = BytesIO(); piece.save(pb, format="JPEG", quality=90); pb.seek(0)
                    from aiogram.types import BufferedInputFile
                    sent = await message.answer_photo(
                        BufferedInputFile(pb.read(), f"{tc}{sep}{n}.jpg"),
                        caption=f"{tc}{sep}{n}"
                    )
                    fid = sent.photo[-1].file_id
                    name = f"{tc}{sep}{n}"
                    cur2.execute("""
                        INSERT INTO images(name, file_id)
                        VALUES(%s,%s)
                        ON CONFLICT(name) DO UPDATE SET file_id=EXCLUDED.file_id
                    """, (name, fid))
                    saved += 1
        conn2.commit(); cur2.close(); conn2.close()
        await message.answer(f"✅ {saved} ta rasm saqlandi!\n{len(tcs)} ta mavzu × {panel_per_tc} panel")
        return

    if caption.lower().startswith("split:"):
        parts  = caption[6:].strip().split(":")
        topic_code = parts[0].strip()
        rows   = int(parts[1]) if len(parts) > 1 else 1
        cols   = int(parts[2]) if len(parts) > 2 else 6
        prefix = parts[3].strip() if len(parts) > 3 else ""
        total  = rows * cols
        sep    = f"-{prefix}-" if prefix else "-"
        await message.answer(
            f"⏳ Rasm {rows}×{cols}={total} ga bo'linmoqda...\n"
            f"📌 {topic_code}{sep}1 ... {topic_code}{sep}{total}"
        )

        # Rasmni yuklab olish
        import io
        from PIL import Image

        file = await message.bot.get_file(message.photo[-1].file_id)
        buf = io.BytesIO()
        await message.bot.download_file(file.file_path, buf)
        buf.seek(0)
        img = Image.open(buf)
        W, H = img.size

        # 5 qator x 8 ustun = 40 ta (yoki berilgan o'lcham)
        cell_w = W // cols
        cell_h = H // rows

        conn = _get_db_conn()
        cur  = conn.cursor()
        saved = 0

        for row in range(rows):
            for col in range(cols):
                n = row * cols + col + 1
                if n > total:
                    break

                # Kesish
                x1 = col * cell_w
                y1 = row * cell_h
                x2 = x1 + cell_w
                y2 = y1 + cell_h
                piece = img.crop((x1, y1, x2, y2))

                # Telegram ga yuborish va file_id olish
                piece_buf = io.BytesIO()
                piece.save(piece_buf, format="JPEG")
                piece_buf.seek(0)

                from aiogram.types import BufferedInputFile
                name_preview = f"{topic_code}-{prefix}-{n}"
                sent = await message.answer_photo(
                    BufferedInputFile(piece_buf.read(), filename=f"{name_preview}.jpg"),
                    caption=name_preview
                )
                file_id = sent.photo[-1].file_id

                # DBga saqlash
                sep2 = f"-{prefix}-" if prefix else "-"
                name = f"{topic_code}{sep2}{n}"
                cur.execute("""
                    INSERT INTO images(name, file_id)
                    VALUES(%s,%s)
                    ON CONFLICT (name) DO UPDATE SET file_id=EXCLUDED.file_id
                """, (name, file_id))
                saved += 1

        conn.commit()
        cur.close(); conn.close()
        sep3 = f"-{prefix}-" if prefix else "-"
        await message.answer(
            f"✅ {saved} ta rasm saqlandi!\n"
            f"📁 {topic_code}{sep3}1 ... {topic_code}{sep3}{total}"
        )
        return

    # Oddiy rasm saqlash
    if not caption:
        await message.answer("Rasm nomini captionga yozing\n\nYoki 40 ga bo'lish uchun:\nsplit:TOPIC_CODE")
        return

    name = caption
    file_id = message.photo[-1].file_id
    conn = _get_db_conn()
    cur = conn.cursor()
    cur.execute("""
    INSERT INTO images(name, file_id)
    VALUES(%s,%s)
    ON CONFLICT (name)
    DO UPDATE SET file_id = EXCLUDED.file_id
    """, (name, file_id))
    conn.commit()
    cur.close(); conn.close()
    await message.answer(f"✅ Saqlandi: {name}")

# ====== START ======
def _save_error_log(uid_sp, username, error_text):
    """Xatoni bazaga yozadi (sinxron)."""
    try:
        conn_ = _get_db_conn()
        cur_  = conn_.cursor()
        cur_.execute(
            "INSERT INTO error_log(uid_sp, username, error_text) VALUES(%s,%s,%s)",
            (uid_sp, username, error_text[:1000])
        )
        conn_.commit(); cur_.close(); conn_.close()
    except Exception:
        pass

def _get_unread_errors():
    """O'qilmagan xatolar sonini qaytaradi."""
    try:
        conn_ = _get_db_conn()
        cur_  = conn_.cursor()
        cur_.execute("SELECT COUNT(*) FROM error_log WHERE is_read=FALSE")
        n = cur_.fetchone()[0]
        cur_.close(); conn_.close()
        return n
    except Exception:
        return 0

async def _notify_admins(error_text, user_id, username):
    """Adminlarga yangi xato haqida xabar yuboradi."""
    try:
        n = _get_unread_errors()
        txt = (
            f"🆘 Yangi xato #{n}\n"
            f"👤 User: {username} ({user_id})\n"
            f"❌ {error_text[:300]}"
        )
        for admin_id in ADMINS:
            try:
                await bot.send_message(admin_id, txt)
            except Exception:
                pass
    except Exception:
        pass

async def _error_and_home(source, user_id, err, label="Xato"):
    """Xato xabarini ko'rsatib, 2 soniyada bosh menyuga qaytaradi."""
    import traceback, asyncio
    tb = traceback.format_exc()
    short = str(err)[:200]

    # Username olish
    try:
        uname = source.from_user.username or str(user_id) if hasattr(source, "from_user") else str(user_id)
    except Exception:
        uname = str(user_id)

    # Bazaga saqlash
    _save_error_log(user_id, uname, f"{label}: {short}\n{tb[:500]}")

    # Foydalanuvchiga xabar
    try:
        msg_fn = source.answer if hasattr(source, "answer") else source.message.answer
        await msg_fn(f"⚠️ Xatolik yuz berdi.\n\nBosh menyuga qaytilmoqda...")
    except Exception:
        pass

    await asyncio.sleep(2)

    # Bosh menyuga qaytarish
    try:
        conn_ = _get_db_conn()
        cur_  = conn_.cursor()
        cur_.execute("SELECT role FROM users WHERE user_id=%s", (uid_sp,))
        row_  = cur_.fetchone()
        cur_.close(); conn_.close()
        role_ = row_[0] if row_ else "🧒 O'quvchi"
        if _is_admin(user_id): role_ = "Admin"
    except Exception:
        role_ = "🧒 O'quvchi"

    n_err = _get_unread_errors() if _is_admin(user_id) else 0
    kb_ = get_main_keyboard(role_, unread_errors=n_err)
    try:
        msg_fn = source.answer if hasattr(source, "answer") else source.message.answer
        await msg_fn("🏠 Bosh menyu", reply_markup=kb_)
    except Exception:
        pass

    # Adminlarga xabardorlik
    await _notify_admins(f"{label}: {short}", user_id, uname)
    print(f"[ERROR] user={user_id} | {label}: {short}\n{tb}")

from aiogram.filters import Command

@dp.message(Command("menu"))
@dp.message(Command("cancel"))
@dp.message(Command("stop"))
async def cmd_menu(message: types.Message, state: FSMContext):
    """Har qanday holatda bosh menyuga qaytish."""
    uid = message.from_user.id
    try: await state.clear()
    except: pass
    user_state.pop(uid, None)
    from storage import lesson_state as _ls, temp_user as _tu
    from storage import registration_message as _rm
    _ls.pop(uid, None); _tu.pop(uid, None); _rm.pop(uid, None)
    from test_engine import test_sessions, start_test
    if uid in test_sessions:
        s_ = test_sessions.pop(uid, {})
        if s_.get("timer_task"):
            try: s_["timer_task"].cancel()
            except: pass
    try:
        conn = _get_db_conn(); cur = conn.cursor()
        cur.execute("SELECT role FROM users WHERE user_id=%s", (uid,))
        row = cur.fetchone(); cur.close(); conn.close()
        role = row[0] if row else "🧒 O'quvchi"
    except: role = "🧒 O'quvchi"
    if uid in ADMINS: role = "Admin"
    await message.answer(
        "🏠 Bosh menyu",
        reply_markup=get_main_keyboard(role)
    )

@dp.message(CommandStart())
async def start(message: types.Message, state: FSMContext):
    # /start — har qanday holatda barcha state tozalanadi
    _tg = message.from_user.id
    _eff_clear(_tg)                 # aktiv akkaunt o'zgargan bo'lishi mumkin
    uid = _eff_uid(_tg)
    try: await state.clear()
    except: pass
    user_state.pop(uid, None)
    from storage import lesson_state as _ls, temp_user as _tu
    from storage import registration_message as _rm, reg_kbd_message as _rkm
    _ls.pop(uid, None)
    _tu.pop(uid, None)
    _rm.pop(uid, None)
    _rkm.pop(uid, None)
    from test_engine import test_sessions
    if uid in test_sessions:
        s_ = test_sessions.pop(uid, {})
        if s_.get("timer_task"):
            try: s_["timer_task"].cancel()
            except: pass

    conn = _get_db_conn()
    cur = conn.cursor()

    cur.execute(
        "SELECT role, full_name, class FROM users WHERE user_id=%s",
        (uid,)
    )

    user = cur.fetchone()

    conn.close()

    # RO'YXATDAN O'TGAN FOYDALANUVCHI
    if user:

        role, full_name, grade = user

        if _tg in ADMINS:
            role = "Admin"

        if role == "Admin":
            _n_err = _get_unread_errors()
            await message.answer(
                f"👋 Qaytganingiz bilan!\n🎭 Rol: {role}"
                + (f"\n🆘 {_n_err} ta o'qilmagan xato bor!" if _n_err else ""),
                reply_markup=get_main_keyboard(role, unread_errors=_n_err)
            )
            return

        # O'quvchi uchun yoshga mos kutib olish
        if "quvchi" in role.lower() or role.strip() in ("🧒 O'quvchi", "🧒O'quvchi", "O'quvchi"):

            from progress import update_streak
            from student_dashboard import build_dashboard

            update_streak(uid)

            try:
                text, keyboard = await build_dashboard(uid)
            except Exception as _de:
                import traceback
                print(f"build_dashboard ERROR: {traceback.format_exc()}")
                text = "👋 Xush kelibsiz!"
                keyboard = None

            # Majburiy imtihon bormi?
            from progress import get_pending_exams
            pending   = get_pending_exams(uid)
            mandatory = [e for e in pending if e[3]]

            if mandatory:
                exam = mandatory[0]
                await message.answer(
                    f"{text}\n\n"
                    f"🚨 MAJBURIY IMTIHON:\n"
                    f"📘 {exam[1]}\n"
                    f"📅 Sana: {exam[2]}",
                    reply_markup=InlineKeyboardMarkup(
                        inline_keyboard=[
                            [InlineKeyboardButton(
                                text="▶️ Imtihonni boshlash",
                                callback_data=f"exam_start_{exam[0]}"
                            )],
                            [InlineKeyboardButton(
                                text="⏭ Keyinroq",
                                callback_data="exam_later"
                            )]
                        ]
                    )
                )
                await message.answer("👇", reply_markup=get_main_keyboard("🧒 O'quvchi"))
            else:
                # Dashboard inline keyboard bilan
                if keyboard:
                    await message.answer(text, parse_mode="HTML", reply_markup=keyboard)
                else:
                    await message.answer(text, parse_mode="HTML")
                # O'quvchi reply keyboard — har doim yuboriladi
                _student_kb = get_main_keyboard("🧒 O'quvchi")
                await message.answer("👇", reply_markup=_student_kb)
        else:
            await message.answer(
                f"👋 Qaytganingiz bilan!\n🎭 Rol: {role}",
                reply_markup=get_main_keyboard(role)
            )

        return

    # YANGI FOYDALANUVCHI — inline registratsiya
    from register import start_registration
    await start_registration(message)

# Test import uchun vaqtincha fayl yo'llari (user_id -> path)
test_import_files = {}

async def prepare_test_import(message, user_id):
    """Excel ni yuklab, tekshirib, tasdiq so'raydi (darrov import qilmaydi)."""
    file = await bot.get_file(message.document.file_id)
    path = f"temp_import_{user_id}.xlsx"
    await bot.download_file(file.file_path, path)

    try:
        _xls = pd.ExcelFile(path)
        _sheet = "TESTLAR" if "TESTLAR" in _xls.sheet_names else _xls.sheet_names[0]
        df = pd.read_excel(path, sheet_name=_sheet)
    except Exception as _e:
        await message.answer(f"❌ Excel o'qib bo'lmadi: {_e}")
        admin_state[user_id] = None
        return

    if "topic_code" not in df.columns:
        await message.answer(
            "❌ Excel ustunlari mos emas.\n"
            "Birinchi qatorda 'topic_code, difficulty, question ...' ustunlari bo'lishi kerak."
        )
        admin_state[user_id] = None
        return

    valid = 0
    for _, r in df.iterrows():
        if not pd.isna(r.get("topic_code")) and not pd.isna(r.get("question")):
            valid += 1

    test_import_files[user_id] = path
    admin_state[user_id] = "test_import_confirm"

    await message.answer(
        f"📋 Faylda {valid} ta savol topildi.\n\nImport qilaylikmi?",
        reply_markup=ReplyKeyboardMarkup(
            keyboard=[[
                KeyboardButton(text="✅ Ha, import qil"),
                KeyboardButton(text="❌ Bekor"),
            ]],
            resize_keyboard=True
        )
    )

async def import_tests_excel(target, path, user_id):

    # Varaq nomi har qanday bo'lsa ishlasin: avval TESTLAR, bo'lmasa birinchi varaq
    try:
        _xls = pd.ExcelFile(path)
        _sheet = "TESTLAR" if "TESTLAR" in _xls.sheet_names else _xls.sheet_names[0]
        df = pd.read_excel(path, sheet_name=_sheet)
    except Exception as _e:
        await target.answer(f"❌ Excel o'qib bo'lmadi: {_e}")
        return

    # ── RASM TAVSIFLARI (varaq nomi har xil bo'lishi mumkin) ──
    global _RASM_TAVSIF
    _RASM_TAVSIF = {}

    # 1) TESTLAR varag'ining o'zida tavsif ustuni bormi? (eng ishonchli)
    try:
        _tc = next((c for c in df.columns
                    if "tavsif" in str(c).lower() or "description" in str(c).lower()), None)
        if _tc is not None and "image_url" in df.columns:
            for _, _r in df.iterrows():
                _k = str(_r.get("image_url") or "").strip()
                _v = str(_r.get(_tc) or "").strip()
                if _k and _v and _k.lower() != "nan" and _v.lower() != "nan":
                    _RASM_TAVSIF[_k] = _v
            if _RASM_TAVSIF:
                print(f"[rasm_tavsif] TESTLAR '{_tc}' ustunidan {len(_RASM_TAVSIF)} ta")
    except Exception as _e:
        print(f"[rasm_tavsif] ustun xato: {_e}")

    # 2) Topilmasa — boshqa varaqlarni mazmuni bo'yicha qidiramiz
    if not _RASM_TAVSIF:
        # TESTLAR dagi haqiqiy rasm kodlari (moslikni tekshirish uchun)
        try:
            _kodlar = set(str(v).strip() for v in df.get("image_url", pd.Series([])).dropna())
        except Exception:
            _kodlar = set()
        try:
            for _sn in _xls.sheet_names:
                if _sn == _sheet:
                    continue
                try:
                    _dfr = pd.read_excel(path, sheet_name=_sn)
                except Exception:
                    continue
                if _dfr.shape[1] < 2 or len(_dfr) == 0:
                    continue
                _cols = [str(c).lower() for c in _dfr.columns]
                # id ustuni: qiymatlari haqiqiy rasm kodlariga mos kelishi SHART
                _idc = None
                for _c in _dfr.columns:
                    _vals = set(str(x).strip() for x in _dfr[_c].dropna().head(50))
                    if _kodlar and len(_vals & _kodlar) >= max(1, len(_vals) // 2):
                        _idc = _c; break
                if _idc is None:
                    continue
                # tavsif ustuni: desc/tavsif/_en/prompt, yoki eng uzun matnli
                _dsc = next((c for c, l in zip(_dfr.columns, _cols)
                             if "desc" in l or "tavsif" in l or "_en" in l or "prompt" in l), None)
                if _dsc is None:
                    _best, _blen = None, 0
                    for _c in _dfr.columns:
                        if _c == _idc: continue
                        try:
                            _ml = _dfr[_c].dropna().astype(str).str.len().mean()
                        except Exception:
                            _ml = 0
                        if _ml > _blen: _best, _blen = _c, _ml
                    if _blen > 10: _dsc = _best
                if _dsc is None:
                    continue
                for _, _r in _dfr.iterrows():
                    _k = str(_r[_idc]).strip()
                    _v = str(_r[_dsc]).strip()
                    if _k and _v and _k.lower() != "nan" and _v.lower() != "nan":
                        _RASM_TAVSIF[_k] = _v
                if _RASM_TAVSIF:
                    print(f"[rasm_tavsif] '{_sn}' varaqdan {len(_RASM_TAVSIF)} ta")
                    break
        except Exception as _e:
            print(f"[rasm_tavsif] varaq xato: {_e}")

    if not _RASM_TAVSIF:
        print("[rasm_tavsif] tavsif topilmadi")
    else:
        # DB ga saqlaymiz (bot qayta ishga tushsa ham yo'qolmasin)
        try:
            _c = _get_db_conn(); _cr = _c.cursor()
            _cr.execute("""CREATE TABLE IF NOT EXISTS rasm_tavsif(
                image_id TEXT PRIMARY KEY, description TEXT)""")

            # Eski tavsiflarni olamiz — o'zgarganini aniqlash uchun
            _cr.execute("SELECT image_id, description FROM rasm_tavsif")
            _eski = {r[0]: (r[1] or "") for r in _cr.fetchall()}

            _ozgargan = []
            for _k, _v in _RASM_TAVSIF.items():
                if _k in _eski and _eski[_k].strip() != (_v or "").strip():
                    _ozgargan.append(_k)
                _cr.execute("""INSERT INTO rasm_tavsif(image_id,description) VALUES(%s,%s)
                    ON CONFLICT(image_id) DO UPDATE SET description=EXCLUDED.description""",
                    (_k, _v))

            # Tavsif o'zgargan bo'lsa — ESKI RASM MOS EMAS, o'chiramiz
            if _ozgargan:
                _cr.execute("DELETE FROM images WHERE name = ANY(%s)", (_ozgargan,))
                _cr.execute("""UPDATE generated_tests SET image_file_id=NULL
                    WHERE image_url = ANY(%s)""", (_ozgargan,))
                print(f"[rasm_tavsif] ⚠️ {len(_ozgargan)} ta tavsif o'zgardi -> rasm qayta chiziladi")
                print(f"[rasm_tavsif] o'zgarganlar: {_ozgargan[:5]}")

            _c.commit(); _cr.close(); _c.close()
            print(f"[rasm_tavsif] DB ga {len(_RASM_TAVSIF)} ta saqlandi")

            if _ozgargan:
                try:
                    await target.answer(
                        f"🔄 {len(_ozgargan)} ta rasm tavsifi o'zgargan.\n"
                        f"Eski rasmlar o'chirildi — qaytadan chiziladi."
                    )
                except Exception: pass
        except Exception as _e:
            print(f"[rasm_tavsif] DB xato: {_e}")

    if "topic_code" not in df.columns:
        await target.answer(
            "❌ Excel ustunlari mos emas.\n"
            "Birinchi qatorda 'topic_code, difficulty, question ...' ustunlari bo'lishi kerak."
        )
        return

    success = 0
    duplicates = 0
    errors = 0
    error_rows = []

    # Fayldagi barcha topik kodlarni olish
    _all_tcs = df["topic_code"].dropna().astype(str).str.strip()
    _all_tcs = [tc for tc in _all_tcs.unique() if tc not in ("","nan")]

    # Har topik uchun eskiyi o'chiramiz (qayta import)
    if _all_tcs:
        import psycopg2 as _pg0
        _c0 = _pg0.connect(os.getenv("DATABASE_URL")); _cu0 = _c0.cursor()
        for _tc in _all_tcs:
            _cu0.execute("DELETE FROM generated_tests WHERE topic_code=%s", (_tc,))
        _c0.commit(); _cu0.close(); _c0.close()

    for index, row in df.iterrows():

        try:
            # Bo'sh qatorni o'tkazib yuborish
            _tc = row.get("topic_code","")
            _q  = row.get("question","")
            if pd.isna(_tc) or pd.isna(_q) or str(_tc).strip() in ("","nan") or str(_q).strip() in ("","nan"):
                continue

            test_data = {
                "topic_code": str(row["topic_code"]).strip(),
                "difficulty": str(row["difficulty"]).strip(),
                "situation": "" if pd.isna(row["situation"]) else str(row["situation"]),
                "question": str(row["question"]).strip(),
                "option_a": "" if pd.isna(row["option_a"]) else str(row["option_a"]),
                "option_b": "" if pd.isna(row["option_b"]) else str(row["option_b"]),
                "option_c": "" if pd.isna(row["option_c"]) else str(row["option_c"]),
                "option_d": "" if pd.isna(row["option_d"]) else str(row["option_d"]),
                "correct_answer": str(row["correct_answer"]).strip(),
                "explanation": "" if pd.isna(row["explanation"]) else str(row["explanation"]),
                "question_type": str(row["question_type"]).strip(),
                "is_latex": False if pd.isna(row["is_latex"]) else (row["is_latex"] not in (0, 0.0, False, "False", "false", "0")),
                "image_url": None if pd.isna(row["image_url"]) else str(row["image_url"]),
                "audio_text": None if pd.isna(row["audio_text"]) else str(row["audio_text"]),
                "language": "uz" if pd.isna(row["language"]) else str(row["language"]),
                "life_level": 1 if pd.isna(row["life_level"]) else int(row["life_level"]),
                "age_group": None if pd.isna(row["age_group"]) else str(row["age_group"]),
                "time_limit": 60 if pd.isna(row["time_limit"]) else int(row["time_limit"])
            }

            import psycopg2 as _pg
            _conn = _pg.connect(os.getenv("DATABASE_URL"))
            _cur = _conn.cursor()

            # Avval to'liq tekshiruv — duplikatmi? (question_type ham hisobga olinadi)
            _cur.execute("""
                SELECT COUNT(*) FROM generated_tests
                WHERE topic_code=%s AND question=%s
                  AND option_a=%s AND option_b=%s
                  AND option_c=%s AND option_d=%s
                  AND correct_answer=%s
                  AND COALESCE(question_type,'single_choice')=COALESCE(%s,'single_choice')
            """, (
                test_data["topic_code"], test_data["question"],
                test_data["option_a"], test_data["option_b"],
                test_data["option_c"], test_data["option_d"],
                test_data["correct_answer"],
                test_data["question_type"],
            ))
            already = _cur.fetchone()[0]

            if already > 0:
                result = "duplicate"
            else:
                try:
                    _cur.execute("""
                        INSERT INTO generated_tests
                        (topic_code, difficulty, situation, question,
                         option_a, option_b, option_c, option_d,
                         correct_answer, explanation, question_type,
                         is_latex, image_url, audio_text, language,
                         life_level, age_group, time_limit)
                        VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,
                                %s::boolean,%s,%s,%s,%s,%s,%s)
                    """, (
                        test_data["topic_code"], test_data["difficulty"],
                        test_data["situation"], test_data["question"],
                        test_data["option_a"], test_data["option_b"],
                        test_data["option_c"], test_data["option_d"],
                        test_data["correct_answer"], test_data["explanation"],
                        test_data["question_type"],
                        test_data["is_latex"],
                        test_data["image_url"], test_data["audio_text"],
                        test_data["language"], test_data["life_level"],
                        test_data["age_group"], test_data["time_limit"],
                    ))
                    _conn.commit()
                    result = "saved"
                except Exception as _ex:
                    _conn.rollback()
                    result = f"error: {_ex}"

            _cur.close(); _conn.close()

            if result == "saved":
                success += 1

            elif result == "duplicate":

                duplicates += 1

                error_rows.append({
                    "row_number": index + 2,
                    "question": row.get("question"),
                    "error": "Duplikat yoki o'xshash savol"
                })

            else:
                errors += 1

        except Exception as e:

            errors += 1

            error_rows.append({
                "row_number": index + 2,
                "question": row.get("question"),
                "error": str(e)
            })

    if error_rows:

        error_file = "import_errors.xlsx"

        df_errors = pd.DataFrame(error_rows)

        df_errors.to_excel(
            error_file,
            index=False
        )

    await target.answer(
        f"✅ Import tugadi\n\n"
        f"📥 Saqlandi: {success}\n"
        f"⚠️ Duplikat: {duplicates}\n"
        f"❌ Xato: {errors}"
    )

    if error_rows:

        await target.answer_document(
            FSInputFile("import_errors.xlsx"),
            caption="📋 Import xatolari hisoboti"
        )

    admin_state[user_id] = None

    # ═══ RASMLARNI AVTOMATIK CHIZISH ═══
    asyncio.create_task(_auto_generate_images(user_id))


async def _auto_generate_images(user_id, resume=False):
    """image_url kodiga mos tavsif bo'yicha rasm chizadi.
    Limit tugasa to'xtaydi va 'davom etish' tugmasini beradi."""
    import asyncio as _a
    print(f"[auto_img] boshlandi (resume={resume})")

    conn = _get_db_conn(); cur = conn.cursor()
    try:
        cur.execute("ALTER TABLE generated_tests ADD COLUMN IF NOT EXISTS image_file_id TEXT")
        conn.commit()
    except Exception as e:
        print(f"[auto_img] alter: {e}"); conn.rollback()
    try:
        cur.execute("""
            SELECT id, image_url
            FROM generated_tests
            WHERE image_url IS NOT NULL AND image_url <> ''
              AND (image_file_id IS NULL OR image_file_id = '')
            ORDER BY id
        """)
        rows = cur.fetchall()
    except Exception as e:
        print(f"[auto_img] select: {e}")
        cur.close(); conn.close()
        return
    cur.close(); conn.close()

    if not rows:
        try: await bot.send_message(_tg_id(user_id), "✅ Barcha rasmlar chizilgan!")
        except Exception: pass
        return

    # Tavsiflar xotirada yo'q bo'lsa DB dan yuklaymiz
    if not _RASM_TAVSIF:
        try:
            c1=_get_db_conn(); cr1=c1.cursor()
            cr1.execute("SELECT image_id, description FROM rasm_tavsif")
            for _k,_v in cr1.fetchall():
                _RASM_TAVSIF[_k]=_v
            cr1.close(); c1.close()
            print(f"[auto_img] DB dan {len(_RASM_TAVSIF)} ta tavsif")
        except Exception as e:
            print(f"[auto_img] tavsif DB: {e}")

    total = len(rows)
    status = await bot.send_message(_tg_id(user_id), f"🎨 Rasm chizish\n📊 Qolgan: {total} ta")
    ok = 0; fail = 0; skip = 0; limit_hit = False; cf_tugadi = False

    from rasim_generator import generate_cf_flux_ex, generate_pollinations, generate_together_flux
    from aiogram.types import BufferedInputFile

    for idx, (tid, img_code) in enumerate(rows, 1):
        img_code = (img_code or "").strip()
        if img_code.startswith("http") or img_code.startswith("\\") or img_code.startswith("$"):
            skip += 1; continue

        tavsif = _RASM_TAVSIF.get(img_code, "")
        if not tavsif:
            print(f"[auto_img] {tid} tavsif yo'q: {img_code}")
            skip += 1; continue

        # Allaqachon chizilganmi?
        try:
            c0=_get_db_conn(); cr0=c0.cursor()
            cr0.execute("SELECT file_id FROM images WHERE name=%s LIMIT 1",(img_code,))
            ex=cr0.fetchone()
            if ex and ex[0]:
                cr0.execute("UPDATE generated_tests SET image_file_id=%s WHERE id=%s",(ex[0],tid))
                c0.commit(); cr0.close(); c0.close()
                ok += 1; continue
            cr0.close(); c0.close()
        except Exception: pass

        try:
            await status.edit_text(
                f"🎨 Chizilmoqda {idx}/{total}\n"
                f"{'🌻 Pollinations' if cf_tugadi else '☁️ Cloudflare'}\n"
                f"✅ {ok}   ❌ {fail}   ⏭ {skip}\n\n"
                f"📝 {tavsif[:60]}"
            )
        except Exception: pass

        prompt = (f"{tavsif}, colorful cartoon illustration for children, "
                  f"clean simple shapes, bright colors, white background, "
                  f"if any person appears they are modern present-day Uzbek Central Asian "
                  f"people in clean contemporary clothes, cheerful 2020s setting, "
                  f"no vintage, no traditional costumes, no rural poverty")

        img = None; err = None
        if not cf_tugadi:
            for _try in (1, 2):
                img, err = await generate_cf_flux_ex(prompt, steps=8)
                if img or err == "limit":
                    break
                if _try == 1:
                    print(f"[auto_img] {tid} qayta urinamiz (10s)")
                    await _a.sleep(10)
            if err == "limit":
                cf_tugadi = True
                print("[auto_img] Cloudflare limiti tugadi -> Pollinations")
                try:
                    await bot.send_message(
                        _tg_id(user_id),
                        "⏸ Cloudflare kunlik limiti tugadi.\n"
                        "🌻 Pollinations (bepul, cheksiz) bilan davom etmoqda..."
                    )
                except Exception: pass

        # Cloudflare tugagan bo'lsa yoki ishlamasa — Pollinations
        if not img:
            for _try in (1, 2):
                img = await generate_pollinations(prompt)
                if img: break
                if _try == 1: await _a.sleep(8)

        # Oxirgi zaxira
        if not img:
            try: img = await generate_together_flux(prompt, steps=8)
            except Exception: img = None

        if img:
            try:
                sent = await bot.send_photo(
                    chat_id=_tg_id(user_id),
                    photo=BufferedInputFile(img, f"{img_code}.png"),
                    caption=f"🖼 {idx}/{total} · <code>{img_code}</code>\n📝 {tavsif[:70]}",
                    parse_mode="HTML",
                    disable_notification=True
                )
                fid = sent.photo[-1].file_id
                c2=_get_db_conn(); cr2=c2.cursor()
                cr2.execute("UPDATE generated_tests SET image_file_id=%s WHERE id=%s",(fid,tid))
                try:
                    cr2.execute("INSERT INTO images(name,file_id) VALUES(%s,%s) ON CONFLICT DO NOTHING",
                                (img_code,fid))
                except Exception: pass
                c2.commit(); cr2.close(); c2.close()
                ok += 1
            except Exception as e:
                print(f"[auto_img] {tid} yuborish: {e}")
                fail += 1
        else:
            fail += 1
            print(f"[auto_img] {tid} chizilmadi")

        await _a.sleep(1)

    # ═══ YAKUN ═══
    manba = "🌻 Pollinations" if cf_tugadi else "☁️ Cloudflare"
    if fail > 0:
        _IMG_PENDING[user_id] = True
        try:
            await status.edit_text(
                f"✅ <b>Rasm chizish tugadi</b>\n\n"
                f"✅ Chizildi: {ok}\n"
                f"⏭ O'tkazildi: {skip}\n"
                f"❌ Chizilmadi: {fail}\n\n"
                f"Chizilmaganlarini qayta urinib ko'ramizmi?",
                parse_mode="HTML",
                reply_markup=InlineKeyboardMarkup(inline_keyboard=[[
                    InlineKeyboardButton(text="🔄 Qayta urinish", callback_data="rsmres")
                ]])
            )
        except Exception: pass
    else:
        _IMG_PENDING.pop(user_id, None)
        try:
            await status.edit_text(
                f"🎉 <b>Barcha rasmlar tayyor!</b>\n\n"
                f"✅ Chizildi: {ok}\n"
                f"⏭ O'tkazildi: {skip}\n"
                f"🖼 Manba: {manba}",
                parse_mode="HTML"
            )
        except Exception: pass
    print(f"[auto_img] tugadi ok={ok} fail={fail} skip={skip} cf_tugadi={cf_tugadi}")


async def _daily_image_resume():
    """Har kuni limit tiklangach (00:10 UTC) rasm chizishni davom ettiradi."""
    import asyncio as _a
    from datetime import datetime, timezone
    while True:
        try:
            now = datetime.now(timezone.utc)
            secs = ((23 - now.hour) * 3600 + (59 - now.minute) * 60) + 600
            await _a.sleep(max(secs, 60))

            # Chizilmagan rasm bormi?
            try:
                c=_get_db_conn(); cr=c.cursor()
                cr.execute("""SELECT COUNT(*) FROM generated_tests
                    WHERE image_url IS NOT NULL AND image_url<>''
                      AND (image_file_id IS NULL OR image_file_id='')""")
                qolgan=(cr.fetchone() or [0])[0]
                cr.close(); c.close()
            except Exception:
                qolgan = 0
            if qolgan == 0:
                continue

            for uid in ADMINS:
                try:
                    await bot.send_message(
                        uid,
                        f"🌅 <b>Limit tiklandi</b>\n\n"
                        f"⏳ Chizilmagan rasm: {qolgan} ta\n\nDavom etamizmi?",
                        parse_mode="HTML",
                        reply_markup=InlineKeyboardMarkup(inline_keyboard=[[
                            InlineKeyboardButton(text="▶️ Ha, davom et", callback_data="rsmres"),
                            InlineKeyboardButton(text="⏸ Keyinroq", callback_data="rsmlat"),
                        ]])
                    )
                except Exception as e:
                    print(f"[daily_img] {uid}: {e}")
        except Exception as e:
            print(f"[daily_img] {e}")
            await _a.sleep(3600)

@dp.message()
async def handle_all(
    message: Message,
    state: FSMContext
):
    user_id = _eff_uid(message.from_user.id)
    try:
        await _handle_all_inner(message, state, user_id)
    except Exception as _e:
        await _error_and_home(message, user_id, _e, "Xatolik")

async def _rq_save(source, user_id, name, rol, sinf):
    """Tez kirish — foydalanuvchini saqlash (multi-account)."""
    tg = _tg_id(user_id)          # haqiqiy telegram ID (user_accounts uchun)
    conn2=_get_db_conn(); cur2=conn2.cursor()
    rol_uz = {"student":"O'quvchi","teacher":"O'qituvchi","parent":"Ota-ona"}.get(rol,rol)
    sinf_txt = f"{sinf}-sinf" if sinf else ""
    print(f"[rq_save] user={user_id} name={name} rol={rol_uz}")
    try:
        cur2.execute("UPDATE users SET full_name=%s,role=%s,class=%s WHERE user_id=%s",
                    (name,rol_uz,sinf_txt,user_id))
        if cur2.rowcount==0:
            cur2.execute("INSERT INTO users(user_id,full_name,role,class) VALUES(%s,%s,%s,%s)",
                        (user_id,name,rol_uz,sinf_txt))
        cur2.execute("SELECT MAX(account_index) FROM user_accounts WHERE telegram_id=%s",(tg,))
        max_idx=(cur2.fetchone() or [None])[0]
        new_idx = 0 if max_idx is None else max_idx + 1
        print(f"[rq_save] max_idx={max_idx} new_idx={new_idx}")
        cur2.execute("UPDATE user_accounts SET is_active=FALSE WHERE telegram_id=%s",(tg,))
        try:
            cur2.execute("""
                INSERT INTO user_accounts(telegram_id,account_index,uid,full_name,role,class,is_active)
                VALUES(%s,%s,%s,%s,%s,%s,TRUE)
                ON CONFLICT(telegram_id,account_index) DO UPDATE
                SET uid=EXCLUDED.uid,full_name=EXCLUDED.full_name,role=EXCLUDED.role,
                    class=EXCLUDED.class,is_active=TRUE
            """, (tg, new_idx, user_id, name, rol_uz, sinf_txt))
        except Exception:
            conn2.rollback()
            cur2.execute("""
                INSERT INTO user_accounts(telegram_id,account_index,full_name,role,class,is_active)
                VALUES(%s,%s,%s,%s,%s,TRUE)
                ON CONFLICT(telegram_id,account_index) DO UPDATE
                SET full_name=EXCLUDED.full_name,role=EXCLUDED.role,class=EXCLUDED.class,is_active=TRUE
            """, (tg, new_idx, name, rol_uz, sinf_txt))
        conn2.commit()
        print(f"[rq_save] ✅ saqlandi index={new_idx}")
    except Exception as e:
        import traceback; traceback.print_exc()
        print(f"[rq_save] ❌ XATO: {e}")
        conn2.rollback()
    cur2.close(); conn2.close()
    _eff_clear(tg)                 # kesh yangilansin
    user_state.pop(user_id, None)
    temp_user.pop(user_id, None)
    kb = get_main_keyboard(rol_uz)
    welcome=f"✅ Xush kelibsiz, {name}!\n🎯 {rol_uz} {sinf_txt}"
    # CallbackQuery bo'lsa .message.answer, Message bo'lsa .answer
    if hasattr(source, "message") and source.message is not None:
        # CallbackQuery
        await source.message.answer(welcome, reply_markup=kb)
    else:
        # Message
        await source.answer(welcome, reply_markup=kb)

async def _save_quick_user(call, user_id):
    """rq_sinf callbackdan saqlash."""
    data = temp_user.get(user_id, {})
    name = data.get("full_name","Foydalanuvchi")
    rol  = data.get("role","student")
    sinf = data.get("class")
    await _rq_save(call, user_id, name, rol, sinf)

async def _handle_all_inner(message: Message, state: FSMContext, user_id: int):

    # ═══ /start SINONIMLARI ═══
    # "menyu", "bosh", "01" — /start bilan bir xil ishlaydi
    if message.text:
        _t = message.text.strip().lower().rstrip("!.?")
        if _t in ("menyu", "menu", "bosh", "boshi", "01", "start",
                  "меню", "бош", "бошига"):
            # Yozuvli test javobi bo'lsa — tegmaymiz
            from test_engine import test_sessions as _tsx
            if not (user_id in _tsx and user_state.get(user_id) == "text_answer"):
                await start(message, state)
                return

    # ═══ 🎓 TO'GARAK TESTLARI (bosh menyu) ═══
    if message.text and message.text.strip() in ("🎓 To'garak testlari", "🎓 Togarak testlari",
                                                  "/togarak_test", "/tt"):
        from togarak import get_student_togaraklar
        tgs = get_student_togaraklar(user_id)
        if not tgs:
            await message.answer(
                "📚 Siz hech qaysi to'garakka a'zo emassiz.\n\n"
                "«📚 To'garaklar» bo'limidan qo'shiling.")
            return
        if len(tgs) == 1:
            import togarak_test as _tt
            tgid = tgs[0]["id"]
            ro_yxat = _tt.mavzular(tgid)
            if not ro_yxat:
                await message.answer("❌ Bu to'garak fani bo'yicha test yo'q.")
                return
            temp_user[f"tt_sel:{user_id}:{tgid}"] = []
            temp_user[f"tt_sah:{user_id}:{tgid}"] = 0
            ochiq = sum(1 for m in ro_yxat if m[3])
            yopiq = len(ro_yxat) - ochiq
            s = f"🎓 {tgs[0]['nomi']}\n📚 {ochiq} mavzu ochiq"
            if yopiq: s += f" · 🔒 {yopiq} yopiq"
            s += "\n\nMavzularni belgilang:"
            await message.answer(s, reply_markup=_tt.mavzu_kb(tgid, ro_yxat, set(), 0))
            return
        rows = [[InlineKeyboardButton(text=f"🎓 {t['nomi'][:32]}",
                                      callback_data=f"tt_tg:{t['id']}")] for t in tgs]
        await message.answer("🎓 To'garak tanlang:",
                             reply_markup=InlineKeyboardMarkup(inline_keyboard=rows))
        return

    # ═══ 📊 IMTIHON (o'qituvchi) ═══
    if message.text and message.text.strip() in ("📊 Imtihon", "/imtihon"):
        from togarak import get_teacher_togaraklar
        tgs = get_teacher_togaraklar(user_id)
        if not tgs:
            await message.answer("❌ Sizda to'garak yo'q.")
            return
        rows = [[InlineKeyboardButton(text=f"📚 {t['nomi'][:32]}",
                                      callback_data=f"im_menu:{t['id']}")] for t in tgs]
        await message.answer("📊 To'garak tanlang:",
                             reply_markup=InlineKeyboardMarkup(inline_keyboard=rows))
        return

    # ═══════════════════════════════════════════
    # 👨‍👩‍👧 OTA-ONA PANELI (matn tugmalari)
    # handler_parent.py routeri handle_all dan keyin ulanadi va
    # hech qachon ishlamaydi — shuning uchun shu yerda ushlaymiz.
    # ═══════════════════════════════════════════
    if message.text and not message.text.startswith("/"):
        _mt = message.text.strip().lower()
        _ota_tugma = None
        if "farzand" in _mt and "qo'sh" not in _mt and "qosh" not in _mt:
            _ota_tugma = "asosiy"
        elif "yo'qlama" in _mt or "yoqlama" in _mt or "davomat" in _mt:
            _ota_tugma = "yoqlama"
        elif "uy vazifa" in _mt or _mt in ("📝 uy", "uy"):
            _ota_tugma = "vazifa"
        elif "imtihon" in _mt and not _mt.startswith("/"):
            _ota_tugma = "imtihon"
        elif "nazorat" in _mt:
            _ota_tugma = "nazorat"
        elif "baho" in _mt and "qo'y" not in _mt:
            _ota_tugma = "baho"

        if _ota_tugma:
            _rol = (_get_user_qisqa(user_id)[1] or "").lower()
            _ota_mi = ("ota" in _rol or "ona" in _rol or "parent" in _rol)
            # Ota-ona paneli FAQAT ota-onaga. Boshqalar o'z handlerlariga o'tadi.
            if not _ota_mi:
                _ota_tugma = None

        if _ota_tugma:
            import ota_ona as _oo
            lst = _oo.farzandlar(user_id)
            if not lst:
                await message.answer(
                    "👨‍👩‍👧 <b>Sizda farzand ulanmagan</b>\n\n"
                    "Farzandingizni ulash uchun:",
                    parse_mode="HTML",
                    reply_markup=InlineKeyboardMarkup(inline_keyboard=[[
                        InlineKeyboardButton(text="➕ Farzand qo'shish", callback_data="fk_add")]]))
                return
            if len(lst) == 1:
                await _ota_ekran(message, user_id, lst[0][0], _ota_tugma)
                return
            rows = [[InlineKeyboardButton(text=f"👤 {ism[:24]} {sinf}",
                     callback_data=f"op_{_ota_tugma}:{cid}")] for cid, ism, sinf in lst]
            await message.answer("👨‍👩‍👧 Farzandni tanlang:",
                                 reply_markup=InlineKeyboardMarkup(inline_keyboard=rows))
            return

    # ═══ AKKAUNT: ko'chirish kodi kiritildi ═══
    if user_state.get(user_id) == "ak_kod_kirit" and message.text:
        user_state.pop(user_id, None)
        import akkaunt as _ak
        kod = message.text.strip().upper()
        if len(kod) != 8:
            await message.answer("❌ Kod 8 belgi bo'lishi kerak.")
            return
        tg = message.from_user.id
        ok, xabar = _ak.kochirish_bajar(kod, tg)
        _eff_clear(tg)
        if ok:
            await message.answer(
                f"{xabar}\n\n"
                f"Akkaunt shu telefonga ulandi va aktiv qilindi.\n"
                f"/start bosing.",
                reply_markup=get_main_keyboard(""))
        else:
            await message.answer(xabar)
        return

    # ═══ IMTIHON: nom yozildi ═══
    if str(admin_state.get(user_id) or "").startswith("im_nom:") and message.text:
        _q = str(admin_state[user_id]).split(":")
        tgid = int(_q[1]); turi = _q[2]
        admin_state.pop(user_id, None)
        nomi = message.text.strip()[:60]

        import baholash as _bh
        iid = _bh.imtihon_yarat(tgid, user_id, nomi, turi)
        if not iid:
            await message.answer("❌ Imtihon yaratilmadi."); return

        if turi == "yozma":
            await message.answer(
                f"✅ <b>{nomi}</b> yaratildi\n\nEndi baho qo'ying:",
                parse_mode="HTML",
                reply_markup=InlineKeyboardMarkup(inline_keyboard=[[
                    InlineKeyboardButton(text="✍️ Baho qo'yish", callback_data=f"im_bal:{iid}")]]))
        else:
            admin_state[user_id] = f"im_soni:{tgid}:{iid}"
            await message.answer(
                f"✅ <b>{nomi}</b> yaratildi\n\nNechta savol bo'lsin? (5–50)",
                parse_mode="HTML")
        return

    # ═══ IMTIHON: savol soni ═══
    if str(admin_state.get(user_id) or "").startswith("im_soni:") and message.text:
        _q = str(admin_state[user_id]).split(":")
        tgid = int(_q[1]); iid = int(_q[2])
        admin_state.pop(user_id, None)
        try:
            soni = max(5, min(50, int(message.text.strip())))
        except Exception:
            await message.answer("❌ Faqat raqam (5–50)")
            admin_state[user_id] = f"im_soni:{tgid}:{iid}"
            return
        c = _get_db_conn(); cr = c.cursor()
        cr.execute("UPDATE togarak_imtihon SET savol_soni=%s WHERE id=%s", (soni, iid))
        c.commit(); cr.close(); c.close()
        await message.answer(
            f"✅ {soni} ta savol\n\nSavollar qayerdan olinsin?",
            reply_markup=InlineKeyboardMarkup(inline_keyboard=[
                [InlineKeyboardButton(text="📗 Mavzu tanlayman",
                                      callback_data=f"im_manba:{tgid}:{iid}:tanla")],
                [InlineKeyboardButton(text="🎲 O'tilgan mavzulardan random",
                                      callback_data=f"im_manba:{tgid}:{iid}:random")],
            ]))
        return

    # ═══ IMTIHON: yozma foiz ═══
    if str(admin_state.get(user_id) or "").startswith("im_foiz:") and message.text:
        _q = str(admin_state[user_id]).split(":")
        iid = int(_q[1]); uid2 = int(_q[2])
        admin_state.pop(user_id, None)
        try:
            foiz = float(message.text.strip().replace(",", "."))
        except Exception:
            await message.answer("❌ 1 dan 100 gacha raqam yozing")
            admin_state[user_id] = f"im_foiz:{iid}:{uid2}"
            return
        if not (0 <= foiz <= 100):
            await message.answer("❌ 1 dan 100 gacha bo'lsin")
            admin_state[user_id] = f"im_foiz:{iid}:{uid2}"
            return

        import baholash as _bh
        _bh.baho_qoy(iid, uid2, foiz, manba="teacher")
        imt = _bh.imtihon_ol(iid)
        d = _bh.daraja(foiz)

        c = _get_db_conn(); cr = c.cursor()
        cr.execute("SELECT full_name FROM users WHERE user_id=%s", (uid2,))
        ism = (cr.fetchone() or ["O'quvchi"])[0]
        cr.execute("SELECT parent_id FROM parent_child WHERE child_id=%s", (uid2,))
        otalar = [r[0] for r in cr.fetchall()]
        cr.close(); c.close()

        await message.answer(
            f"✅ {ism} — <b>{foiz:.0f}%</b> {d}", parse_mode="HTML",
            reply_markup=InlineKeyboardMarkup(inline_keyboard=[[
                InlineKeyboardButton(text="✍️ Yana baho qo'yish", callback_data=f"im_bal:{iid}"),
            ]]))

        xabar = (f"📊 <b>Imtihon natijasi</b>\n\n"
                 f"📝 {imt['nomi'] if imt else '—'}\n"
                 f"🎯 Baho: <b>{foiz:.0f}%</b>\n🎖 {d}")
        for kim in [uid2] + otalar:
            try: await bot.send_message(_tg_id(kim), xabar, parse_mode="HTML")
            except Exception: pass
        return

    # ═══ RASM HOLATI (admin) ═══
    if message.text and message.text.strip() == "/rasmlar":
        if not _is_admin(user_id):
            return
        c = _get_db_conn(); cr = c.cursor()
        try:
            cr.execute("""SELECT
                COUNT(*) FILTER (WHERE image_url IS NOT NULL AND image_url<>''),
                COUNT(*) FILTER (WHERE image_file_id IS NOT NULL AND image_file_id<>''),
                COUNT(*) FROM generated_tests""")
            kod, chizilgan, jami = cr.fetchone()
            cr.execute("SELECT COUNT(*) FROM images")
            saqlangan = (cr.fetchone() or [0])[0]
            cr.execute("""SELECT topic_code, image_url FROM generated_tests
                WHERE image_url IS NOT NULL AND image_url<>''
                  AND (image_file_id IS NULL OR image_file_id='') LIMIT 5""")
            kutayotgan = cr.fetchall()
            cr.close(); c.close()

            t = ["🖼 <b>Rasm holati</b>", ""]
            t.append(f"📊 Jami test: {jami}")
            t.append(f"🔖 Rasm kodi bor: {kod}")
            t.append(f"✅ Rasm chizilgan: {chizilgan}")
            t.append(f"⏳ Chizilmagan: {kod - chizilgan}")
            t.append(f"💾 images jadvalida: {saqlangan}")
            if kutayotgan:
                t.append("\n<b>Chizilmaganlar:</b>")
                for tc, iu in kutayotgan:
                    t.append(f"   <code>{iu}</code>")
                t.append("\n<i>Excel qayta yuklang — chiziladi.</i>")
            elif kod == 0:
                t.append("\n<i>Hech qaysi testda rasm kodi yo'q.</i>")
            else:
                t.append("\n✅ <i>Barcha rasmlar tayyor.</i>")
            await message.answer("\n".join(t), parse_mode="HTML")
        except Exception as e:
            try: cr.close(); c.close()
            except Exception: pass
            await message.answer(f"❌ {e}")
        return

    # ═══ TESTLARNI O'CHIRISH (admin) ═══
    if message.text and message.text.startswith("/ochir"):
        if not _is_admin(user_id):
            return
        arg = message.text[6:].strip()
        if not arg:
            await message.answer(
                "🗑 <b>Testlarni o'chirish</b>\n"
                "<i>Faqat savollar o'chadi. Topic kodlar va mavzu tuzilmasi qoladi.</i>\n\n"
                "<code>/ochir kod 5-01-1-01-01-01-001</code>\n"
                "   shu koddagi testlar\n\n"
                "<code>/ochir boshi 5-01-1</code>\n"
                "   shu bilan boshlanadigan kodlardagi testlar\n\n"
                "<code>/ochir sinf 5</code>\n"
                "   5-sinfning barcha testlari\n\n"
                "<code>/ochir sinf 5 MATEMATIKA</code>\n"
                "   5-sinf matematika testlari\n\n"
                "<code>/ochir rasm</code>\n"
                "   barcha rasmlar (testlar qoladi)\n\n"
                "<i>Avval nechta o'chishini ko'rsatadi, tasdiqlaysiz.</i>",
                parse_mode="HTML")
            return

        q = arg.split()
        rejim = q[0].lower()
        c = _get_db_conn(); cr = c.cursor()
        try:
            if rejim == "kod" and len(q) > 1:
                shart, args = "topic_code=%s", [q[1]]
                izoh = f"kod: {q[1]}"
            elif rejim == "boshi" and len(q) > 1:
                shart, args = "topic_code LIKE %s", [q[1] + "%"]
                izoh = f"'{q[1]}' bilan boshlanadigan"
            elif rejim == "sinf" and len(q) > 1:
                sub = " ".join(q[2:]) if len(q) > 2 else None
                if sub:
                    shart = """topic_code IN (SELECT topic_code FROM dts_tree
                        WHERE grade::TEXT=%s AND subject_name ILIKE %s AND is_deleted=FALSE)"""
                    args = [q[1], f"%{sub}%"]
                    izoh = f"{q[1]}-sinf · {sub}"
                else:
                    shart = """topic_code IN (SELECT topic_code FROM dts_tree
                        WHERE grade::TEXT=%s AND is_deleted=FALSE)"""
                    args = [q[1]]
                    izoh = f"{q[1]}-sinf (barcha fan)"
            elif rejim == "rasm":
                cr.execute("""SELECT COUNT(*) FROM generated_tests
                    WHERE image_file_id IS NOT NULL AND image_file_id<>''""")
                n = (cr.fetchone() or [0])[0]
                cr.close(); c.close()
                if n == 0:
                    await message.answer("ℹ️ O'chiriladigan rasm yo'q."); return
                await message.answer(
                    f"🗑 <b>{n} ta rasm</b> o'chiriladi.\nTestlar qoladi.\n\nTasdiqlaysizmi?",
                    parse_mode="HTML",
                    reply_markup=InlineKeyboardMarkup(inline_keyboard=[[
                        InlineKeyboardButton(text="✅ Ha", callback_data="ochrasm_yes"),
                        InlineKeyboardButton(text="❌ Yo'q", callback_data="ochrasm_no"),
                    ]]))
                return
            else:
                cr.close(); c.close()
                await message.answer("❌ Noto'g'ri format. <code>/ochir</code> yozing.",
                                     parse_mode="HTML")
                return

            cr.execute(f"SELECT COUNT(*) FROM generated_tests WHERE {shart}", tuple(args))
            n = (cr.fetchone() or [0])[0]
            cr.execute(f"SELECT DISTINCT topic_code FROM generated_tests WHERE {shart} LIMIT 5",
                       tuple(args))
            namuna = [r[0] for r in cr.fetchall()]
            cr.close(); c.close()

            if n == 0:
                await message.answer(f"ℹ️ '{izoh}' bo'yicha test topilmadi."); return

            # Bu testlarda nechta rasm bor?
            c2 = _get_db_conn(); cr2 = c2.cursor()
            cr2.execute(f"""SELECT COUNT(DISTINCT image_url) FROM generated_tests
                WHERE {shart} AND image_url IS NOT NULL AND image_url<>''""", tuple(args))
            rasm_soni = (cr2.fetchone() or [0])[0]
            cr2.close(); c2.close()

            _t = [f"🗑 <b>{n} ta test</b> o'chiriladi", f"📍 {izoh}", ""]
            _t.append("Kodlar:")
            for k in namuna:
                _t.append(f"   <code>{k}</code>")
            _t.append("")
            _t.append("✅ Topic kodlar va mavzu tuzilmasi <b>qoladi</b>")
            if rasm_soni:
                _t.append(f"🖼 Bu testlarda <b>{rasm_soni} ta rasm</b> bor")
                _t.append("   Rasmlar ham o'chsinmi?")
            _t.append("⚠️ Qaytarib bo'lmaydi")

            _ochir_soqi[user_id] = (shart, args, izoh, n)

            tugmalar = [[
                InlineKeyboardButton(text=f"✅ Faqat {n} ta test", callback_data="ochir_yes"),
                InlineKeyboardButton(text="❌ Bekor", callback_data="ochir_no"),
            ]]
            if rasm_soni:
                tugmalar.insert(0, [InlineKeyboardButton(
                    text=f"🗑 Test + {rasm_soni} ta rasm", callback_data="ochir_yes_rasm")])

            await message.answer("\n".join(_t), parse_mode="HTML",
                reply_markup=InlineKeyboardMarkup(inline_keyboard=tugmalar))
        except Exception as e:
            try: cr.close(); c.close()
            except Exception: pass
            await message.answer(f"❌ Xato: {e}")
        return

    # ═══ MAVZU TASHXISI (admin) ═══
    if message.text and message.text.startswith("/mavzu"):
        if not _is_admin(user_id):
            return
        arg = message.text[6:].strip()
        if not arg:
            await message.answer(
                "🔍 <b>Mavzu tashxisi</b>\n\n"
                "<code>/mavzu 5</code> — 5-sinf mavzulari\n"
                "<code>/mavzu 5 01-01</code> — aniq mavzu",
                parse_mode="HTML")
            return
        qism = arg.split()
        gr = qism[0]
        mc = qism[1] if len(qism) > 1 else None
        c = _get_db_conn(); cr = c.cursor()
        try:
            if mc:
                # Aniq mavzu: topic kodlar + kichik mavzular + savollar
                cr.execute("""SELECT topic_code, mavzu_name, kichik_name, subject_name,
                       (SELECT COUNT(*) FROM generated_tests g WHERE g.topic_code=d.topic_code)
                    FROM dts_tree d WHERE mavzu_code=%s AND grade::TEXT=%s AND is_deleted=FALSE
                    ORDER BY topic_code LIMIT 30""", (mc, gr))
                rows = cr.fetchall()
                t = [f"🔍 <b>{gr}-sinf · {mc}</b>\n"]
                for tc, mn, kn, sn, n in rows:
                    t.append(f"<code>{tc}</code> · {n} test")
                    t.append(f"   📗 {mn}")
                    if kn: t.append(f"   └ {kn}")
                jami = sum(r[4] for r in rows)
                t.append(f"\n<b>Jami: {len(rows)} topic, {jami} test</b>")

                # Haqiqiy savollar — mos keladimi?
                kodlar = [r[0] for r in rows]
                if kodlar:
                    cr.execute("""SELECT topic_code, LEFT(question,50)
                        FROM generated_tests WHERE topic_code=ANY(%s) LIMIT 5""", (kodlar,))
                    t.append("\n📝 <b>Savollar:</b>")
                    for tc, q in cr.fetchall():
                        t.append(f"<code>{tc}</code>")
                        t.append(f"   {q}")
            else:
                cr.execute("""SELECT d.mavzu_code, d.mavzu_name, d.subject_name,
                       COUNT(DISTINCT d.topic_code), COUNT(g.id)
                    FROM dts_tree d LEFT JOIN generated_tests g ON g.topic_code=d.topic_code
                    WHERE d.grade::TEXT=%s AND d.is_deleted=FALSE AND d.mavzu_code IS NOT NULL
                    GROUP BY d.mavzu_code, d.mavzu_name, d.subject_name
                    ORDER BY d.mavzu_code LIMIT 25""", (gr,))
                rows = cr.fetchall()
                t = [f"🔍 <b>{gr}-sinf mavzulari</b>\n"]
                for mcode, mn, sn, ntopic, ntest in rows:
                    t.append(f"<code>{mcode}</code> · {(mn or '—')[:28]}")
                    t.append(f"   {ntopic} topic · {ntest} test")
            cr.close(); c.close()
            await message.answer("\n".join(t)[:4000], parse_mode="HTML")
        except Exception as e:
            cr.close(); c.close()
            await message.answer(f"❌ {e}")
        return

    # ═══ OVOZ SINOVI (admin) ═══
    if message.text and message.text.startswith("/ovoz"):
        if not _is_admin(user_id):
            return
        matn = message.text[5:].strip()
        if not matn:
            await message.answer(
                "🎧 <b>Ovoz sinovi</b>\n\n"
                "Matn yozing:\n<code>/ovoz To'g'ri javob: to'qqiz. San'at va a'zo.</code>\n\n"
                "4 variant yuboriladi — qaysi biri yaxshi eshitilsa ayting.",
                parse_mode="HTML")
            return
        try:
            import ovoz as _ov
        except Exception as e:
            await message.answer(f"❌ ovoz.py topilmadi: {e}")
            return

        kut = await message.answer("🎧 Tayyorlanmoqda...")
        try:
            await message.answer(
                f"📝 <b>Xom matn:</b>\n<code>{matn[:300]}</code>\n\n"
                f"🇺🇿 <b>Kirill:</b>\n<code>{_ov.tayyorla(matn)[:300]}</code>\n\n"
                f"🔤 <b>Lotin:</b>\n<code>{_ov.tayyorla(matn, kirillga=False)[:300]}</code>",
                parse_mode="HTML")
        except Exception as e:
            await message.answer(f"⚠️ Matn tayyorlash: {e}")

        from aiogram.types import BufferedInputFile
        n = 0
        JINS_NOM = {"qiz": "👧 Qiz", "ogil": "👦 O'g'il"}
        for jins in ("qiz", "ogil"):
            for kir in (True, False):
                alifbo = "Kirill" if kir else "Lotin"
                nom = JINS_NOM[jins] + " · " + alifbo
                try:
                    audio = await _ov.ovoz_yarat(matn, jins=jins, yosh=10, kirillga=kir)
                    if audio:
                        n += 1
                        await message.answer_voice(
                            BufferedInputFile(audio, f"ovoz_{n}.mp3"),
                            caption=f"{n}. {nom}")
                except Exception as e:
                    await message.answer(f"❌ {nom}: {e}")
        try: await kut.delete()
        except Exception: pass
        if n == 0:
            await message.answer("❌ Hech qanday ovoz yaratilmadi. Log ni tekshiring.")
        else:
            await message.answer(f"✅ {n} ta variant. Qaysi biri yaxshi?")
        return

    # Test paytida yozilgan xabarni o'chirish
    from test_engine import test_sessions
    if user_id in test_sessions:
        if user_state.get(user_id) != "text_answer":
            try:
                await message.delete()
            except Exception:
                pass
            return

    # ── YANGI AKKAUNT: ism yozildi → rol tanlash ──
    if user_state.get(user_id) == "nacc_name" and message.text:
        name = message.text.strip()
        user_state.pop(user_id, None)
        rows2 = InlineKeyboardMarkup(inline_keyboard=[
            [InlineKeyboardButton(text="🧒 O'quvchi",   callback_data=f"nacc_rol:student|{name}")],
            [InlineKeyboardButton(text="👨‍🏫 O'qituvchi", callback_data=f"nacc_rol:teacher|{name}")],
            [InlineKeyboardButton(text="👨‍👩‍👧 Ota-ona",    callback_data=f"nacc_rol:parent|{name}")],
        ])
        await message.answer(f"✅ {name}\n\nRolni tanlang:", reply_markup=rows2)
        return

    # ── TO'GARAK STATE HANDLERS (yuqoriga ko'chirilgan) ──
    # Tez kirish ism (yuqoriga ko'chirildi)
    if str(user_state.get(user_id) or "").startswith("rq_name:") and message.text:
        rol = str(user_state[user_id]).split(":")[1]
        name = message.text.strip()
        print(f"[rq_name] user={user_id} rol={rol} name={name}")
        user_state.pop(user_id, None)
        if user_id not in temp_user or not isinstance(temp_user.get(user_id),dict):
            temp_user[user_id] = {}
        temp_user[user_id]["role"] = rol
        if rol == "student":
            # Sinf tanlash
            temp_user[user_id]["full_name"] = name
            rows2 = [[InlineKeyboardButton(text=f"{i}-sinf", callback_data=f"rq_sinf:{i}") for i in range(j, j+4) if i<=11] for j in range(1, 12, 4)]
            await message.answer(
                f"✅ {name}\n\nSinfni tanlang:",
                reply_markup=InlineKeyboardMarkup(inline_keyboard=rows2)
            )
        else:
            # O'qituvchi/ota-ona — to'g'ridan saqlash
            await _rq_save(message, user_id, name, rol, None)
        return

    if str(admin_state.get(user_id) or "").startswith("tg_kun_vaqt:") and message.text:
        parts3=str(admin_state[user_id]).split(":")
        tgid3=int(parts3[1]); sana3=parts3[2]; reja_id3=int(parts3[3])
        admin_state.pop(user_id,None)
        vaqt=message.text.strip()
        from datetime import datetime
        KUNLAR=["Dushanba","Seshanba","Chorshanba","Payshanba","Juma","Shanba","Yakshanba"]
        d=datetime.strptime(sana3,"%Y-%m-%d").date()
        conn2=_get_db_conn();cur2=conn2.cursor()
        try:
            cur2.execute("UPDATE togarak_reja SET dars_sana=%s, dars_kuni=%s, dars_vaqt=%s WHERE id=%s",
                        (sana3, KUNLAR[d.weekday()], vaqt, reja_id3))
            conn2.commit()
        except Exception as e:
            conn2.rollback(); print(f"kun_vaqt: {e}")
            try:
                cur2.execute("UPDATE togarak_reja SET dars_sana=%s, dars_vaqt=%s WHERE id=%s",
                            (sana3, vaqt, reja_id3))
                conn2.commit()
            except Exception as e2: conn2.rollback(); print(f"kun_vaqt2: {e2}")
        cur2.close(); conn2.close()
        await message.answer(f"✅ {KUNLAR[d.weekday()]} {d.strftime('%d.%m')} — {vaqt}\nDars belgilandi!")
        return

    if str(admin_state.get(user_id) or "").startswith("tg_reja_vaqt_save:") and message.text:
        parts3=str(admin_state[user_id]).split(":"); tgid3,reja_id3,kun_id3=int(parts3[1]),int(parts3[2]),int(parts3[3])
        admin_state.pop(user_id,None)
        KUNLAR=["Dushanba","Seshanba","Chorshanba","Payshanba","Juma","Shanba"]
        vaqt=message.text.strip()
        conn2=_get_db_conn();cur2=conn2.cursor()
        try:
            cur2.execute("UPDATE togarak_reja SET dars_kuni=%s, dars_vaqt=%s WHERE id=%s",
                        (KUNLAR[kun_id3], vaqt, reja_id3))
            # Jadval jadvaliga ham qo'shamiz
            cur2.execute("DELETE FROM togarak_jadval WHERE togarak_id=(SELECT togarak_id FROM togarak_reja WHERE id=%s) AND kun_id=%s",(reja_id3,kun_id3))
            cur2.execute("""SELECT togarak_id FROM togarak_reja WHERE id=%s""",(reja_id3,))
            tg_id_r=(cur2.fetchone() or [tgid3])[0]
            cur2.execute("INSERT INTO togarak_jadval(togarak_id,kun_id,kun_nomi,boshlanish) VALUES(%s,%s,%s,%s)",
                        (tg_id_r,kun_id3,KUNLAR[kun_id3],vaqt))
            conn2.commit()
        except Exception as e: conn2.rollback(); print(f"reja_vaqt: {e}")
        cur2.close(); conn2.close()
        await message.answer(f"✅ {KUNLAR[kun_id3]}: {vaqt} — belgilandi!")
        return

    if str(admin_state.get(user_id) or "").startswith("tg_jadval_vaqt:") and message.text:
        parts3=str(admin_state[user_id]).split(":"); tgid3,kun_id3=int(parts3[1]),int(parts3[2])
        admin_state.pop(user_id,None)
        KUNLAR=["Dushanba","Seshanba","Chorshanba","Payshanba","Juma","Shanba"]
        vaqt=message.text.strip()
        bosh,tug="","" 
        if "-" in vaqt: bosh,tug=vaqt.split("-",1); bosh,tug=bosh.strip(),tug.strip()
        else:
            bosh=vaqt.strip()
            # Avto +2 soat
            try:
                h,m=bosh.split(":"); h=int(h)+2
                if h>=24: h-=24
                tug=f"{h:02d}:{m}"
            except: tug=""
        conn2=_get_db_conn();cur2=conn2.cursor()
        try:
            cur2.execute("DELETE FROM togarak_jadval WHERE togarak_id=%s AND kun_id=%s",(tgid3,kun_id3))
            cur2.execute("INSERT INTO togarak_jadval(togarak_id,kun_id,kun_nomi,boshlanish,tugash) VALUES(%s,%s,%s,%s,%s)",(tgid3,kun_id3,KUNLAR[kun_id3],bosh,tug))
            conn2.commit()
        except Exception as e: conn2.rollback(); print(f"jadval: {e}")
        cur2.execute("SELECT kun_id,boshlanish,tugash FROM togarak_jadval WHERE togarak_id=%s",(tgid3,))
        mavjud={r[0]:(r[1],r[2]) for r in cur2.fetchall()}
        cur2.close(); conn2.close()
        txt="⚙️ Dars kunlarini sozlash\n"+"─"*20+"\n\n"
        for i,k in enumerate(KUNLAR):
            if i in mavjud:
                v=mavjud[i]; vt=f"{v[0]}-{v[1]}" if v[1] else v[0]
                txt+=f"✅ {k}: {vt}\n"
            else: txt+=f"⚪ {k}: —\n"
        rows2=[]
        for i,k in enumerate(KUNLAR):
            if i in mavjud:
                v=mavjud[i]; vt=f"{v[0]}-{v[1]}" if v[1] else v[0]
                rows2.append([InlineKeyboardButton(text=f"✅ {k} — {vt}",callback_data=f"tg_jadval_kun:{tgid3}:{i}"),InlineKeyboardButton(text="🗑",callback_data=f"tg_jadval_del:{tgid3}:{i}")])
            else:
                rows2.append([InlineKeyboardButton(text=f"⚪ {k} — vaqt qo'shish",callback_data=f"tg_jadval_kun:{tgid3}:{i}")])
        rows2.append([InlineKeyboardButton(text="✅ Tayyor — jadvalga",callback_data=f"tg_reja:{tgid3}:0")])
        tug_txt=f" (tugash: {tug})" if tug else ""
        await message.answer(f"✅ {KUNLAR[kun_id3]}: {bosh}{tug_txt} belgilandi!\n\n"+txt,reply_markup=InlineKeyboardMarkup(inline_keyboard=rows2))
        return

    if str(admin_state.get(user_id) or "").startswith("tg_new_parol:") and message.text:
        tgid=int(str(admin_state[user_id]).split(":")[1])
        admin_state.pop(user_id,None)
        yangi=message.text.strip()
        if len(yangi)<4: await message.answer("❌ Kamida 4 belgi!"); return
        conn2=_get_db_conn();cur2=conn2.cursor()
        cur2.execute("UPDATE togaraklar SET parol=%s WHERE id=%s AND teacher_id=%s",(yangi,tgid,user_id))
        conn2.commit(); cur2.close(); conn2.close()
        await message.answer(f"✅ Parol yangilandi!\n\n🔑 Yangi parol: <code>{yangi}</code>",parse_mode="HTML")
        return



    # Dars paytida yozilgan xabarni avtomatik o'chirish
    if isinstance(user_state.get(user_id), dict):
        if user_state[user_id].get("board_message_id"):
            try:
                await message.delete()
            except Exception:
                pass

    if user_state.get(user_id) == "user_rasm" and message.text:
        tavsif = message.text.strip()
        user_state[user_id] = None

        # Uzunlik tekshirish
        words = tavsif.split()
        if len(words) < 2:
            await message.answer("❌ Juda qisqa! Kamida 2 so'z yozing.\nMasalan: «qishki manzara, bolalar»")
            user_state[user_id] = "user_rasm"
            return
        if len(words) > 10:
            await message.answer(f"❌ Juda uzun ({len(words)} so'z)! Maksimal 10 so'z.\nQisqaroq yozing.")
            user_state[user_id] = "user_rasm"
            return

        # Limit qayta tekshirish
        conn2 = _get_db_conn(); cur2 = conn2.cursor()
        try:
            cur2.execute("SELECT COUNT(*) FROM images WHERE name LIKE %s AND created_at >= CURRENT_DATE",
                        (f"user_{user_id}_%",))
            today_count = cur2.fetchone()[0]
        except: today_count = 0
        cur2.close(); conn2.close()

        if today_count >= 2:
            await message.answer("⏰ Kunlik limit tugadi (2 ta). Ertaga qaytib keling!")
            return

        status_u = await message.answer(f"🎨 Chizilmoqda... «{tavsif[:50]}»")

        async def do_user_rasm():
            try:
                from rasim_generator import generate_smart
                img, prompt = await generate_smart(
                    tavsif, "ta'lim", "", "multik",
                    is_admin=(_is_admin(user_id))
                )
                if img:
                    from aiogram.types import BufferedInputFile
                    fname = f"user_{user_id}_{int(__import__('time').time())}"
                    sent = await message.answer_photo(
                        BufferedInputFile(img, f"{fname}.png"),
                        caption=f"🎨 {tavsif[:60]}\n\n📝 <i>{(prompt or '')[:800]}</i>",
                        parse_mode="HTML"
                    )
                    fid = sent.photo[-1].file_id
                    try:
                        conn3=_get_db_conn();cur3=conn3.cursor()
                        cur3.execute("INSERT INTO images(name,file_id) VALUES(%s,%s) ON CONFLICT DO NOTHING",
                                    (fname,fid))
                        conn3.commit();cur3.close();conn3.close()
                    except: pass
                    qolgan = 1 - (today_count)
                    msg = f"✅ Rasm tayyor!"
                    if qolgan > 0:
                        msg += f"\n📊 Bugun yana {qolgan} ta yaratish mumkin"
                    else:
                        msg += "\n⏰ Bugungi limit tugadi"
                    await status_u.edit_text(msg)
                else:
                    await status_u.edit_text(
                        "❌ Rasm yaratilmadi.\n\n"
                        "Sabab: DALL-E API ishlamayapti.\n"
                        "Boshqacharoq tavsif bilan qayta urinib ko'ring."
                    )
            except Exception as e:
                await status_u.edit_text(f"❌ Xato: {e}")

        asyncio.create_task(do_user_rasm())
        return

    # ── Report comment handler ──
    if str(admin_state.get(user_id) or "").startswith("edit_test_field:") and message.text:
        parts3 = admin_state[user_id].split(":")
        tid3 = int(parts3[1]); field3 = parts3[2]
        val3  = message.text.strip()
        admin_state.pop(user_id, None)
        conn2=_get_db_conn();cur2=conn2.cursor()
        cur2.execute(f"UPDATE generated_tests SET {field3}=%s WHERE id=%s",(val3,tid3))
        conn2.commit();cur2.close();conn2.close()
        await message.answer(f"✅ Test #{tid3} yangilandi!\n🔑 {field3} = {val3[:60]}")
        return

    if str(user_state.get(user_id) or "").startswith("report_comment:") and message.text:
        cur_idx = int(str(user_state[user_id]).split(":")[1])
        comment = message.text.strip()
        user_state[user_id] = None

        from test_engine import test_sessions
        st2 = test_sessions.get(user_id) or {}
        tests = st2.get("questions", [])
        tc = st2.get("topic_code","")

        if cur_idx < len(tests):
            q = tests[cur_idx][0]
            conn2 = _get_db_conn(); cur2 = conn2.cursor()
            cur2.execute("SELECT id FROM generated_tests WHERE question=%s LIMIT 1", (q,))
            row2 = cur2.fetchone(); tid = row2[0] if row2 else None
            try:
                cur2.execute("""INSERT INTO test_corrections(test_id,topic_code,question,user_id,comment,status)
                    VALUES(%s,%s,%s,%s,%s,'new')
                    ON CONFLICT DO NOTHING""", (tid, tc, q[:200], user_id, comment))
                conn2.commit()
            except: pass
            cur2.close(); conn2.close()
            # Admin xabar
            for aid in ADMINS:
                try:
                    await bot.send_message(aid,
                        f"✏️ Xato test bildirish\n"
                        f"👤 {user_id}\n"
                        f"📝 {q[:100]}\n"
                        f"💬 {comment}",
                        reply_markup=InlineKeyboardMarkup(inline_keyboard=[[
                            InlineKeyboardButton(
                                text="🔧 Shu testni tuzatish",
                                callback_data=f"admin_fix_test:{tid if tid else 0}"
                            )
                        ]])
                    )
                except: pass
        await message.answer("✅ Rahmat! Xabaringiz adminlarga yuborildi.\n⏩ Test davom etmoqda...")
        from test_engine import test_sessions, _advance
        st2 = test_sessions.get(user_id)
        if st2 and not st2.get("answered"):
            try: await _advance(user_id)
            except Exception as _te: print(f"test davom: {_te}")
        return

    if str(admin_state.get(user_id) or "").startswith("kitob_edit_text:") and message.text:
        parts3=str(admin_state[user_id]).split(":")
        book_id3,page3=int(parts3[1]),int(parts3[2])
        admin_state.pop(user_id,None)
        from kitob_bazasi import extract_exercises
        new_text=message.text.strip()
        exs=extract_exercises(new_text)
        conn2=_get_db_conn();cur2=conn2.cursor()
        cur2.execute("UPDATE book_pages SET full_text=%s,exercise_count=%s WHERE book_id=%s AND page_num=%s",
                    (new_text,len(exs),book_id3,page3))
        cur2.execute("DELETE FROM book_exercises WHERE book_id=%s AND page_num=%s",(book_id3,page3))
        for ex in exs:
            cur2.execute("INSERT INTO book_exercises(book_id,page_num,savol) VALUES(%s,%s,%s)",
                        (book_id3,page3,ex[:1000]))
        conn2.commit(); cur2.close(); conn2.close()
        await message.answer(f"✅ Bet {page3} yangilandi! ({len(exs)} misol)")
        return

    if str(admin_state.get(user_id) or "").startswith("kitob_search:") and message.text:
        book_id2=int(str(admin_state[user_id]).split(":")[1])
        query=message.text.strip()
        admin_state.pop(user_id,None)
        from kitob_bazasi import search_book
        results=[r for r in search_book(query) if r["book_id"]==book_id2]
        if not results:
            await message.answer(f"❌ '{query}' topilmadi"); return
        rows2=[[InlineKeyboardButton(text=f"📖 Bet {r['page']}",callback_data=f"kitob_bet:{book_id2}:{r['page']}")] for r in results[:10]]
        txt=f"🔍 '{query}' — {len(results)} ta bet:\n\n"
        for r in results[:5]:
            txt+=f"📄 Bet {r['page']}: {r['text'][:80]}...\n\n"
        await message.answer(txt[:2000],reply_markup=InlineKeyboardMarkup(inline_keyboard=rows2))
        return

    # Tez kirish — ism yozish


    # ══ TO'GARAK YARATISH STATE ══
    if user_state.get(user_id) == "tg_create_nomi" and message.text:
        temp_user[user_id]["tg_nomi"] = message.text.strip()
        user_state[user_id] = "tg_create_fan"
        await message.answer("📖 Fan nomini yozing (Matematika, Ingliz tili...):")
        return

    if user_state.get(user_id) == "tg_create_fan" and message.text:
        temp_user[user_id]["tg_fan"] = message.text.strip()
        user_state[user_id] = "tg_create_parol"
        await message.answer("🔑 Parol o'rnating (harf va raqam, 4 ta va ko'p):")
        return

    if user_state.get(user_id) == "tg_create_parol" and message.text:
        parol = message.text.strip()
        from togarak import check_parol
        if not check_parol(parol):
            await message.answer("❌ Parol kamida 4 belgi bo'lsin!"); return
        temp_user[user_id]["tg_parol"] = parol
        user_state[user_id] = "tg_create_oylik"
        await message.answer("💰 Oylik to'lov miqdorini yozing (so'mda, 0 = bepul):")
        return

    if user_state.get(user_id) == "tg_create_oylik" and message.text:
        try: summa = int(message.text.strip().replace(" ","").replace(",",""))
        except: summa = 0
        temp_user[user_id]["tg_summa"] = summa
        user_state[user_id] = "tg_create_sana"
        await message.answer("📅 Oylik to'lov kunini yozing (1-28):")
        return

    if user_state.get(user_id) == "tg_create_sana" and message.text:
        try: sana = max(1, min(28, int(message.text.strip())))
        except: sana = 1
        d = temp_user.get(user_id, {})
        from togarak import create_togarak
        tid = create_togarak(
            teacher_id=user_id,
            nomi=d.get("tg_nomi","To'garak"),
            fan=d.get("tg_fan",""),
            parol=d.get("tg_parol","0000"),
            oylik_summa=d.get("tg_summa",0),
            oylik_sana=sana
        )
        user_state.pop(user_id,None)
        await message.answer(
            f"✅ To'garak yaratildi!\n"
            f"📚 {d.get('tg_nomi')}\n"
            f"🔑 Parol: {d.get('tg_parol')}\n"
            f"🔢 ID: {tid}\n\n"
            f"O'quvchilarga ID va parolni bering!"
        )
        return

    # O'quvchi to'garakka qo'shilish
    if user_state.get(user_id) == "stg_join_id" and message.text:
        try: tgid = int(message.text.strip())
        except:
            await message.answer("❌ ID raqam bo'lishi kerak!"); return
        user_state.pop(user_id, None)
        from togarak import send_join_request
        res = send_join_request(tgid, user_id)
        if res["ok"]:
            # O'qituvchiga xabar yuborish
            try:
                conn2=_get_db_conn();cur2=conn2.cursor()
                cur2.execute("SELECT full_name,class FROM users WHERE user_id=%s",(user_id,))
                u=cur2.fetchone(); cur2.close(); conn2.close()
                uname = f"{u[0]} ({u[1] or ''})" if u else str(user_id)
                await message.bot.send_message(
                    res["teacher_id"],
                    f"📨 Yangi so'rov!\n👤 {uname}\n📚 {res['togarak_nomi']}\n\nQabul qilasizmi?",
                    reply_markup=InlineKeyboardMarkup(inline_keyboard=[[
                        InlineKeyboardButton(text="✅ Qabul", callback_data=f"tg_req_approve:{user_id}|{tgid}"),
                        InlineKeyboardButton(text="❌ Rad",   callback_data=f"tg_req_reject:{user_id}|{tgid}"),
                    ]])
                )
            except Exception as e:
                print(f"teacher notify: {e}")
            await message.answer(f"✅ So'rov yuborildi!\n⏳ O'qituvchi tasdiqlashini kuting.")
        else:
            await message.answer(res["msg"])
        return


    # Shaxsiy xabar yuborish
    if str(admin_state.get(user_id) or "").startswith("tg_send_pm:") and message.text:
        parts3=str(admin_state[user_id]).split(":")
        tgid3,uid3=int(parts3[1]),int(parts3[2])
        admin_state.pop(user_id,None)
        matn=message.text.strip()
        from togarak import send_group_message
        send_group_message(tgid3, user_id, matn, receiver_id=uid3)
        # Qabul qiluvchiga yuborish
        conn2=_get_db_conn();cur2=conn2.cursor()
        cur2.execute("SELECT full_name FROM users WHERE user_id=%s",(user_id,))
        sender_name=(cur2.fetchone() or ["?"])[0]; cur2.close(); conn2.close()
        try:
            await message.bot.send_message(uid3, f"💬 {sender_name}:\n{matn}")
        except: pass
        await message.answer("✅ Yuborildi!")
        return

    if str(admin_state.get(user_id) or "").startswith("tg_reja_add_manual:") and message.text:
        tgid=int(str(admin_state[user_id]).split(":")[1])
        admin_state.pop(user_id,None)
        from togarak import add_to_reja
        add_to_reja(tgid, message.text.strip(), "maxsus")
        await message.answer(f"✅ '{message.text.strip()}' rejaga qo'shildi!")
        return

    if str(admin_state.get(user_id) or "").startswith("hw_submit:") and message.text:
        parts3=str(admin_state[user_id]).split(":")
        hw_id3,tgid3=int(parts3[1]),int(parts3[2])
        admin_state.pop(user_id,None)
        from features import submit_homework
        submit_homework(hw_id3,user_id,message.text.strip())
        # O'qituvchiga xabar
        conn2=_get_db_conn();cur2=conn2.cursor()
        cur2.execute("SELECT teacher_id,mavzu FROM homework WHERE id=%s",(hw_id3,))
        hw2=cur2.fetchone()
        cur2.execute("SELECT full_name FROM users WHERE user_id=%s",(user_id,))
        uname=(cur2.fetchone() or ["?"])[0]; cur2.close(); conn2.close()
        if hw2:
            try: await message.bot.send_message(hw2[0],
                f"📝 Yangi topshiriq!\n👤 {uname}\n📌 {hw2[1]}\n\n{message.text[:200]}")
            except: pass
        await message.answer("✅ Vazifa topshirildi!")
        return

    if str(admin_state.get(user_id) or "").startswith("hw_new:") and message.text:
        parts3=str(admin_state[user_id]).split(":")
        tgid3,step=int(parts3[1]),parts3[2]
        if step=="mavzu":
            temp_user[user_id]["hw_mavzu"]=message.text.strip()
            admin_state[user_id]=f"hw_new:{tgid3}:topshiriq"
            await message.answer("📝 Topshiriq matnini yozing:")
        elif step=="topshiriq":
            temp_user[user_id]["hw_topshiriq"]=message.text.strip()
            admin_state[user_id]=f"hw_new:{tgid3}:deadline"
            await message.answer("📅 Deadline sanasini yozing (DD.MM.YYYY) yoki 'yoq' deng:")
        elif step=="deadline":
            admin_state.pop(user_id,None)
            mavzu=temp_user[user_id].get("hw_mavzu","")
            topshiriq=temp_user[user_id].get("hw_topshiriq","")
            deadline=None
            if message.text.strip().lower() not in ("yoq","yo'q","-"):
                try:
                    from datetime import datetime
                    deadline=datetime.strptime(message.text.strip(),"%d.%m.%Y").date()
                except: pass
            from features import add_homework
            from togarak import get_group_members, send_group_message
            hw_id=add_homework(tgid3,user_id,mavzu,topshiriq,deadline)
            # A'zolarga xabar
            members=get_group_members(tgid3)
            dl_txt=str(deadline) if deadline else "muddatsiz"
            for uid3 in members:
                try: await message.bot.send_message(uid3,
                    f"📝 Yangi uyga vazifa!\n📌 {mavzu}\n{topshiriq}\n📅 {dl_txt}")
                except: pass
            await message.answer(f"✅ Vazifa qo'shildi va {len(members)} ta o'quvchiga yuborildi!")
        return

    # Guruhga xabar yuborish
    if str(admin_state.get(user_id) or "").startswith("tg_send_msg:") and message.text:
        parts3=str(admin_state[user_id]).split(":")
        tgid3=int(parts3[1]); mode=parts3[2]
        admin_state.pop(user_id,None)
        from togarak import get_group_members, send_group_message
        matn=message.text.strip()
        members=get_group_members(tgid3)
        sent=0
        for uid3 in members:
            if uid3==user_id: continue
            try:
                await message.bot.send_message(uid3, f"📢 To'garak xabari:\n\n{matn}")
                sent+=1
            except: pass
        send_group_message(tgid3, user_id, matn)
        await message.answer(f"✅ Xabar {sent} ta a'zoga yuborildi!")
        return

    # To'garak o'chirish tasdiqlash
    if str(user_state.get(user_id) or "").startswith("tg_del_parol:") and message.text:
        tgid = int(str(user_state[user_id]).split(":")[1])
        user_state.pop(user_id,None)
        conn2=_get_db_conn();cur2=conn2.cursor()
        cur2.execute("SELECT parol,nomi FROM togaraklar WHERE id=%s AND teacher_id=%s",(tgid,user_id))
        row2=cur2.fetchone(); cur2.close(); conn2.close()
        if not row2:
            await message.answer("❌ Topilmadi!"); return
        if message.text.strip() != row2[0]:
            await message.answer("❌ Parol noto'g'ri! To'garak o'chirilmadi."); return
        from togarak import delete_togarak
        delete_togarak(tgid,user_id)
        await message.answer(f"✅ \"{row2[1]}\" to'garagi o'chirildi.")
        return

    if str(admin_state.get(user_id) or "").startswith("tg_set_nomi:") and message.text:
        tgid=int(str(admin_state[user_id]).split(":")[1])
        admin_state.pop(user_id,None)
        nomi=message.text.strip()
        if len(nomi)<3: await message.answer("❌ Nom kamida 3 belgi!"); return
        conn2=_get_db_conn();cur2=conn2.cursor()
        cur2.execute("UPDATE togaraklar SET nomi=%s WHERE id=%s AND teacher_id=%s",(nomi,tgid,user_id))
        conn2.commit(); cur2.close(); conn2.close()
        await message.answer(f"✅ To'garak nomi: {nomi}")
        return

    if str(admin_state.get(user_id) or "").startswith("tg_set_summa:") and message.text:
        tgid=int(str(admin_state[user_id]).split(":")[1])
        admin_state.pop(user_id,None)
        try: summa=int(message.text.strip().replace(" ","").replace(",",""))
        except: await message.answer("❌ Faqat raqam yozing!"); return
        conn2=_get_db_conn();cur2=conn2.cursor()
        cur2.execute("UPDATE togaraklar SET oylik_summa=%s WHERE id=%s AND teacher_id=%s",(summa,tgid,user_id))
        conn2.commit(); cur2.close(); conn2.close()
        await message.answer(f"✅ Oylik summa: {summa:,} so'm")
        return

    if str(user_state.get(user_id) or "").startswith("tg_del_confirm:") and message.text:
        tgid = int(str(user_state[user_id]).split(":")[1])
        user_state.pop(user_id,None)
        conn2=_get_db_conn();cur2=conn2.cursor()
        cur2.execute("SELECT parol FROM togaraklar WHERE id=%s AND teacher_id=%s",(tgid,user_id))
        row2=cur2.fetchone(); cur2.close(); conn2.close()
        if not row2:
            await message.answer("❌ Topilmadi!"); return
        if message.text.strip() != row2[0]:
            await message.answer("❌ Parol noto'g'ri!"); return
        from togarak import delete_togarak
        delete_togarak(tgid,user_id)
        await message.answer("✅ To'garak o'chirildi!")
        return

    # Kabinet o'zgartirish
    if str(user_state.get(user_id) or "").startswith("kb_change_") and message.text:
        field = str(user_state[user_id]).replace("kb_change_","")
        val = message.text.strip()
        user_state.pop(user_id, None)
        conn2=_get_db_conn();cur2=conn2.cursor()
        col_map = {"name":"full_name","bdate":"birth_date","school":"school","role":"role","class":"class"}
        col = col_map.get(field)
        # Rol uchun normallashtirish
        if field == "role":
            rl = val.lower()
            if "quvchi" in rl or "student" in rl: val = "O'quvchi"
            elif "qituvchi" in rl or "teacher" in rl: val = "O'qituvchi"
            elif "ota" in rl or "ona" in rl or "parent" in rl: val = "Ota-ona"
        if col:
            cur2.execute(f"UPDATE users SET {col}=%s WHERE user_id=%s",(val,user_id))
            # user_accounts da ham yangilash
            try:
                cur2.execute(f"UPDATE user_accounts SET {col}=%s WHERE telegram_id=%s AND is_active=TRUE",(val,user_id))
            except: pass
            conn2.commit()
        cur2.close(); conn2.close()
        labels = {"name":"Ism","bdate":"Tug'ilgan sana","school":"Maktab","role":"Rol","class":"Sinf"}
        # Rol o'zgarganda menyuni yangilash
        if field == "role":
            await message.answer(f"✅ Rol yangilandi: {val}", reply_markup=get_main_keyboard(val))
        else:
            await message.answer(f"✅ {labels.get(field,field)} yangilandi: {val}")
        return

    if user_state.get(user_id) == "parent_link_id" and message.text:
        user_state.pop(user_id, None)
        import ota_ona as _oo
        kod = message.text.strip()
        if not (kod.isdigit() and len(kod) == 6):
            await message.answer(
                "❌ Kod 6 xonali raqam bo'lishi kerak.\n\n"
                "Farzandingiz botda «👤 Kabinet → 🔗 Ota-onani ulash» dan kod oladi.")
            return

        child_id = _oo.kod_tekshir(kod)
        if not child_id:
            await message.answer(
                "❌ Kod noto'g'ri yoki muddati o'tgan.\n\n"
                "Farzandingizdan yangi kod so'rang (15 daqiqa amal qiladi).")
            return

        if child_id == user_id:
            await message.answer("❌ O'zingizni farzand qilib ulay olmaysiz.")
            return

        if _oo.bogliqmi(user_id, child_id):
            await message.answer("ℹ️ Bu farzand allaqachon ulangan.")
            _oo.kod_ochir(kod)
            return

        ism_c, _, sinf_c = _oo.kim(child_id)
        ism_p, _, _ = _oo.kim(user_id)

        # Farzandga tasdiq so'rovi
        try:
            await bot.send_message(
                _tg_id(child_id),
                f"🔔 <b>Ulanish so'rovi</b>\n\n"
                f"👤 {ism_p or 'Foydalanuvchi'}\n"
                f"sizni farzandi sifatida ulamoqchi.\n\n"
                f"U sizning baholaringiz, davomatingiz va\n"
                f"imtihon natijalaringizni ko'radi.\n\n"
                f"Tasdiqlaysizmi?",
                parse_mode="HTML",
                reply_markup=InlineKeyboardMarkup(inline_keyboard=[[
                    InlineKeyboardButton(text="✅ Ha, bu mening ota-onam",
                                         callback_data=f"fk_ok:{user_id}:{kod}"),
                ], [
                    InlineKeyboardButton(text="❌ Yo'q, men tanimayman",
                                         callback_data=f"fk_no:{user_id}:{kod}"),
                ]]))
        except Exception as e:
            print(f"[fk] farzandga yuborilmadi: {e}")
            await message.answer("❌ Farzandingizga xabar yuborib bo'lmadi.")
            return

        await message.answer(
            f"📨 So'rov yuborildi.\n\n"
            f"👤 {ism_c or 'Farzand'} {sinf_c or ''}\n"
            f"⏳ U tasdiqlashini kuting.")
        return

    if str(admin_state.get(user_id) or "").startswith("parent_send_msg:") and message.text:
        teacher_id=int(str(admin_state[user_id]).split(":")[1])
        admin_state.pop(user_id, None)
        conn2=_get_db_conn();cur2=conn2.cursor()
        cur2.execute("SELECT full_name FROM users WHERE user_id=%s",(user_id,))
        sender=(cur2.fetchone() or ["Ota-ona"])[0]; cur2.close(); conn2.close()
        try:
            await message.bot.send_message(teacher_id,
                f"📨 Ota-ona xabari ({sender}):\n{message.text}")
            await message.answer("✅ O'qituvchiga yuborildi!")
        except:
            await message.answer("❌ O'qituvchiga yuborib bo'lmadi.")
        return

    # Kitob o'chirish — parol tasdiqlash
    if str(admin_state.get(user_id) or "").startswith("kitob_del_confirm:") and message.text:
        book_id2=int(str(admin_state[user_id]).split(":")[1])
        entered=message.text.strip()
        conn2=_get_db_conn();cur2=conn2.cursor()
        cur2.execute("SELECT parol,title FROM books WHERE id=%s",(book_id2,))
        row2=cur2.fetchone(); cur2.close(); conn2.close()
        if not row2:
            await message.answer("❌ Kitob topilmadi!"); admin_state.pop(user_id,None); return
        parol2,title2=row2
        if entered != (parol2 or "0000"):
            await message.answer(f"❌ Parol noto'g'ri! Kitob o'chirilmadi."); admin_state.pop(user_id,None); return
        conn2=_get_db_conn();cur2=conn2.cursor()
        cur2.execute("DELETE FROM book_exercises WHERE book_id=%s",(book_id2,))
        cur2.execute("DELETE FROM book_pages WHERE book_id=%s",(book_id2,))
        cur2.execute("DELETE FROM books WHERE id=%s",(book_id2,))
        conn2.commit();cur2.close();conn2.close()
        admin_state.pop(user_id,None)
        await message.answer(f"✅ '{title2}' o'chirildi!")
        return

    # Kitob paroli o'rnatish
    if str(admin_state.get(user_id) or "").startswith("kitob_set_parol:") and message.text:
        book_id2=int(str(admin_state[user_id]).split(":")[1])
        parol2=message.text.strip()
        if not (parol2.isdigit() and len(parol2)==4):
            await message.answer("❌ 4 xonali raqam yozing!"); return
        conn2=_get_db_conn();cur2=conn2.cursor()
        cur2.execute("UPDATE books SET parol=%s WHERE id=%s",(parol2,book_id2))
        conn2.commit();cur2.close();conn2.close()
        admin_state.pop(user_id,None)
        await message.answer(f"✅ Parol {parol2} o'rnatildi!")
        return

    # Kitob bet navigatsiya
    if str(admin_state.get(user_id) or "").startswith("kitob_goto:") and message.text:
        try: page2 = int(message.text.strip())
        except:
            await message.answer("❌ Faqat raqam yozing!"); return
        book_id2 = int(str(admin_state[user_id]).split(":")[1])
        admin_state.pop(user_id, None)
        from kitob_bazasi import get_page, render_page_as_image
        pg = get_page(book_id2, page2)
        if not pg:
            await message.answer(f"❌ Bet {page2} topilmadi!"); return
        conn2=_get_db_conn();cur2=conn2.cursor()
        cur2.execute("SELECT total_pages FROM books WHERE id=%s",(book_id2,))
        tot2=(cur2.fetchone() or [0])[0]; cur2.close(); conn2.close()
        nav=[]
        if page2>1: nav.append(InlineKeyboardButton(text="◀️",callback_data=f"kitob_bet:{book_id2}:{page2-1}"))
        nav.append(InlineKeyboardButton(text=f"📄 {page2}/{tot2}",callback_data=f"kitob_goto:{book_id2}"))
        if page2<tot2: nav.append(InlineKeyboardButton(text="▶️",callback_data=f"kitob_bet:{book_id2}:{page2+1}"))
        kb2=InlineKeyboardMarkup(inline_keyboard=[nav,[
            InlineKeyboardButton(text="📝 Matn",callback_data=f"kitob_matn:{book_id2}:{page2}"),
            InlineKeyboardButton(text="🎯 Misollar",callback_data=f"kitob_test:{book_id2}:{page2}")
        ]])
        img2=await render_page_as_image(pg["text"],page2)
        caption2=f"📖 Bet {page2}" + (f" — {pg['section']}" if pg.get("section") else "")
        if img2:
            from aiogram.types import BufferedInputFile
            await message.answer_photo(BufferedInputFile(img2,f"bet_{page2}.png"),caption=caption2,reply_markup=kb2)
        else:
            await message.answer(pg["text"][:800],reply_markup=kb2)
        return

    # Qo'lda terish — ma'lumotlar
    if admin_state.get(user_id) == "kitob_qolda_info" and message.text:
        parts3 = message.text.strip().split("|")
        title3 = parts3[0].strip() if len(parts3)>0 else "Kitob"
        fan3   = parts3[1].strip() if len(parts3)>1 else "Fan"
        sinf3  = parts3[2].strip() if len(parts3)>2 else "1"
        mual3  = parts3[3].strip() if len(parts3)>3 else ""
        admin_state.pop(user_id, None)
        conn2=_get_db_conn();cur2=conn2.cursor()
        cur2.execute("INSERT INTO books(title,fan,sinf,muallif,total_pages) VALUES(%s,%s,%s,%s,0) RETURNING id",
                    (title3,fan3,sinf3,mual3))
        book_id3=cur2.fetchone()[0]; conn2.commit(); cur2.close(); conn2.close()
        admin_state[user_id] = f"kitob_qolda_bet:{book_id3}:1"
        await message.answer(
            f"✅ {title3} yaratildi!\n\n"
            "📝 Bet 1 matnini yozing:\n"
            "Tugash: <code>tugat</code>",
            parse_mode="HTML"
        )
        return

    # Qo'lda terish — betlar
    if str(admin_state.get(user_id) or "").startswith("kitob_qolda_bet:") and message.text:
        parts3 = str(admin_state[user_id]).split(":")
        book_id3, page_num3 = int(parts3[1]), int(parts3[2])
        text3 = message.text.strip()
        # Menu tugmalarini o'tkazib yuborish
        menu_items = {"📚 Kitoblar ▾","🚀 Mavzu tayyorla","📋 Shablonlar",
                     "🧠 Bilimlar ▾","📊 Hisobotlar & Xatolar","👥 Foydalanuvchilar",
                     "🖼 Rasmlar boshqaruvi","🎨 AI Rasm yaratish","⚙️ Akkaunt sozlamalari",
                     "📖 Darslar holati","🧭 DTS topik boshqaruvi"}
        if text3 in menu_items or text3.startswith("/"):
            return
        if text3.lower() in ("tugat","stop","done"):
            conn2=_get_db_conn();cur2=conn2.cursor()
            cur2.execute("UPDATE books SET total_pages=%s WHERE id=%s",(page_num3-1,book_id3))
            conn2.commit(); cur2.close(); conn2.close()
            admin_state.pop(user_id,None)
            await message.answer(f"✅ Kitob tayyor!\n📄 {page_num3-1} bet\n🔑 ID: {book_id3}")
            return
        from kitob_bazasi import extract_exercises
        exercises = extract_exercises(text3)
        conn2=_get_db_conn();cur2=conn2.cursor()
        cur2.execute("""INSERT INTO book_pages(book_id,page_num,full_text,exercise_count)
            VALUES(%s,%s,%s,%s) ON CONFLICT(book_id,page_num) DO UPDATE SET full_text=EXCLUDED.full_text""",
            (book_id3,page_num3,text3,len(exercises)))
        for ex in exercises:
            cur2.execute("INSERT INTO book_exercises(book_id,page_num,savol) VALUES(%s,%s,%s) ON CONFLICT DO NOTHING",
                        (book_id3,page_num3,ex[:1000]))
        cur2.execute("UPDATE books SET total_pages=GREATEST(total_pages,%s) WHERE id=%s",(page_num3,book_id3))
        conn2.commit(); cur2.close(); conn2.close()
        admin_state[user_id] = f"kitob_qolda_bet:{book_id3}:{page_num3+1}"
        # A4 rasm sifatida ko'rsat
        from kitob_bazasi import render_page_as_image
        img3 = await render_page_as_image(text3, page_num3)
        kb3 = InlineKeyboardMarkup(inline_keyboard=[[
            InlineKeyboardButton(text=f"📝 Bet {page_num3+1} yozing", callback_data=f"kitob_next_bet:{book_id3}:{page_num3+1}"),
            InlineKeyboardButton(text="✅ Kitobni tugatish", callback_data=f"kitob_qolda_tugat:{book_id3}:{page_num3}"),
        ]])
        caption3 = f"✅ Bet {page_num3} saqlandi ({len(exercises)} misol)"
        if img3:
            from aiogram.types import BufferedInputFile
            await message.answer_photo(BufferedInputFile(img3,f"bet_{page_num3}.png"), caption=caption3, reply_markup=kb3)
        else:
            await message.answer(caption3, reply_markup=kb3)
        return

    # ═══ BRAIN ═══
    if (user_id not in ADMINS
            and message.text
            and not message.text.startswith("/")
            and user_state.get(user_id) == "ai_mode"):  # FAQAT ai_mode da ishlaydi
        _skip = {
            "⚙️ Sozlamalar","🎯 Bugungi reja","📚 Bilimni mustahkamlash",
            "🧪 Bilimni sinash","📈 Rivojlanishim","🌍 Hamjamiyat","👤 Kabinet",
            "🤖 Yordamchi","👇","✅ Ha, import qil","❌ Bekor",
            "🔙 Ortga","🏠 Bosh menyu","🏠 Bosh ekran",
            "▶️ O'rganishni boshlash","🔊 O'qib berish",
        }
        if message.text not in _skip:
            try:
                from brain import process_message as _b
                _cn = _get_db_conn(); _cu = _cn.cursor()
                _cu.execute("SELECT class FROM users WHERE user_id=%s",(user_id,))
                _gr = _cu.fetchone(); _grade = str(_gr[0]) if _gr else None
                _cu.close(); _cn.close()
                _res = await _b(message.text, user_id, _grade)
                if _res.get("message"):
                    await message.answer(_res["message"])
                if _res.get("action") == "START_TEST" and _res.get("topic"):
                    _c2 = _get_db_conn(); _cu2 = _c2.cursor()
                    _cu2.execute("""SELECT question,option_a,option_b,option_c,option_d,
                           correct_answer,explanation,question_type,is_latex,
                           COALESCE(NULLIF(image_file_id,''), image_url) AS image_url,audio_text,language,time_limit
                           FROM generated_tests WHERE topic_code=%s
                           ORDER BY RANDOM() LIMIT 20""",
                        (_res["topic"]["topic_code"],))
                    _t = _cu2.fetchall(); _cu2.close(); _c2.close()
                    if _t: await start_test(user_id, _t, message)
                elif _res.get("action") == "START_LESSON" and _res.get("topic"):
                    await open_teacher_lesson(message,
                        topic_code=_res["topic"]["topic_code"], _user_id=user_id)
                elif _res.get("action") == "SHOW_STATS":
                    await continue_learning(message)
            except Exception as _be:
                print(f"brain: {_be}")
            return
    # ════════════════════════════════

    if message.text == "⚙️ Sozlamalar":
        from student_settings import show_settings
        await show_settings(message, user_id)
        return

    if message.text == "📚 Bilimni mustahkamlash":
        conn2 = _get_db_conn(); cur2 = conn2.cursor()
        cur2.execute("SELECT class FROM users WHERE user_id=%s", (user_id,))
        _gr = cur2.fetchone()
        _my_grade = str(_gr[0]) if _gr and _gr[0] else "1"

        # Dars bor fanlar (o'z sinfi + CEFR kabi raqamsiz sinflar)
        cur2.execute("""
            SELECT DISTINCT d.subject_name,
                   COUNT(tl.topic_code) as cnt,
                   d.grade
            FROM teacher_lessons tl
            JOIN dts_tree d ON d.topic_code = tl.topic_code
            WHERE (d.grade = %s OR d.grade !~ '^[0-9]+$')
              AND d.is_deleted = FALSE
            GROUP BY d.subject_name, d.grade
            ORDER BY d.grade, d.subject_name
        """, (_my_grade,))
        subjects = cur2.fetchall()
        cur2.close(); conn2.close()

        if not subjects:
            await message.answer("📚 Hali darslar yozilmagan. Admin darslar qo'shishi kerak!")
            return

        rows = []
        for subj, cnt, grade in subjects:
            lbl = f"{grade}-sinf" if str(grade).isdigit() else str(grade)
            icon = "⭐" if str(grade) == str(_my_grade) else "🌐"
            rows.append([InlineKeyboardButton(
                text=f"{icon} {subj} ({cnt} ta dars)",
                callback_data=f"mustah_subj:{grade}:{subj}"
            )])

        await message.answer(
            "📚 Bilimni mustahkamlash\nFan tanlang:",
            reply_markup=InlineKeyboardMarkup(inline_keyboard=rows)
        )
        return

    if message.text == "📚 To'garaklar":
        conn2=_get_db_conn();cur2=conn2.cursor()
        cur2.execute("SELECT role FROM users WHERE user_id=%s",(user_id,))
        row2=cur2.fetchone(); cur2.close(); conn2.close()
        role2 = str(row2[0] if row2 else "")
        from togarak import get_teacher_togaraklar, get_student_togaraklar, togarak_list_kb

        if "qituvchi" in role2:
            # O'QITUVCHI — o'z to'garaklari + yaratish
            tgs = get_teacher_togaraklar(user_id)
            rows2 = [[InlineKeyboardButton(
                text=f"📚 {t['nomi']} ({t['azolar']}/{t['max']})",
                callback_data=f"tg_info:{t['id']}"
            )] for t in tgs]
            rows2.append([InlineKeyboardButton(text="➕ Yangi to'garak ochish", callback_data="tg_yangi")])
            await message.answer(
                f"📚 Mening to'garaklarim ({len(tgs)} ta):",
                reply_markup=InlineKeyboardMarkup(inline_keyboard=rows2)
            )
        else:
            # O'QUVCHI — a'zo bo'lgan to'garaklar + izlash
            tgs = get_student_togaraklar(user_id)
            rows2 = [[InlineKeyboardButton(
                text=f"📚 {t['nomi']} | 👨‍🏫 {t['teacher']}",
                callback_data=f"stg_info:{t['id']}"
            )] for t in tgs]
            rows2.append([InlineKeyboardButton(text="🔍 To'garak izlash (ID bilan)", callback_data="stg_join")])
            txt = f"📚 Mening to'garaklarim ({len(tgs)} ta):"
            if not tgs:
                txt = "📚 Siz hali hech qaysi to'garakka a'zo emassiz.\n\n🔍 To'garak ID va parolini o'qituvchingizdan oling!"
            await message.answer(txt, reply_markup=InlineKeyboardMarkup(inline_keyboard=rows2))
        return

    # Ota-ona message handlers → handler_parent.py

    if message.text == "🎨 Rasm chizdir":
        conn2 = _get_db_conn(); cur2 = conn2.cursor()
        try:
            cur2.execute("SELECT COUNT(*) FROM images WHERE name LIKE %s AND created_at >= CURRENT_DATE",
                        (f"user_{user_id}_%",))
            today_count = cur2.fetchone()[0]
        except: today_count = 0
        cur2.close(); conn2.close()

        if today_count >= 2:
            await message.answer(
                f"⏰ Kunlik limit: 2 ta rasm\n"
                f"✅ Bugun: {today_count} ta\n\n"
                "Ertaga yana 2 ta rasm yaratish mumkin!"
            )
            return

        qolgan = 2 - today_count
        user_state[user_id] = "user_rasm"
        await message.answer(
            f"🎨 Rasm yaratish\n\n"
            f"📊 Bugun qolgan: {qolgan} ta\n\n"
            "Nimani chizishni xohlaysiz?\n"
            "O'zbek tilida yozing:\n\n"
            "Masalan:\n"
            "• «3 ta qizil olma»\n"
            "• «maktabda dars»\n"
            "• «qishki manzara»"
        )
        return

    if message.text == "/id":
        await message.answer(f"🆔 Sizning Telegram ID ingiz:\n<code>{_tg_id(user_id)}</code>\n\nOta-onangizga yuboring!", parse_mode="HTML")
        return

    if message.text == "/stop":
        user_state.pop(user_id, None)
        await message.answer("✅ AI yordamchi o'chirildi.")
        return

    if message.text == "🤖 Yordamchi":
        # Rolni aniqlash
        conn2=_get_db_conn();cur2=conn2.cursor()
        cur2.execute("SELECT role FROM users WHERE user_id=%s",(user_id,))
        row2=cur2.fetchone(); cur2.close(); conn2.close()
        role2 = str(row2[0] if row2 else "")
        is_admin = _is_admin(user_id)

        if is_admin:
            txt = ("🤖 Admin AI Yordamchi\n\n"
                   "Nima qila olaman:\n"
                   "• Foydalanuvchilar statistikasi\n"
                   "• Test yaratish maslahatlar\n"
                   "• DTS mavzu tahlili\n"
                   "• Excel/shablon yordam\n\n"
                   "Savolingizni yozing 👇")
        elif "qituvchi" in role2:
            txt = ("🤖 O'qituvchi AI Yordamchi\n\n"
                   "Nima qila olaman:\n"
                   "• Dars rejalari tuzish\n"
                   "• Savollar yaratish\n"
                   "• Mavzuni tushuntirish\n"
                   "• To'garak rejalash\n\n"
                   "Savolingizni yozing 👇")
        else:
            txt = ("🤖 O'quvchi AI Yordamchi\n\n"
                   "Nima qila olaman:\n"
                   "• Mavzuni tushuntiraman\n"
                   "• Misol yechamiz\n"
                   "• Test beraman\n"
                   "• Kitobdan misol beraman\n\n"
                   "Savolingizni yozing 👇")

        user_state[user_id] = "ai_mode"
        await message.answer(txt, reply_markup=InlineKeyboardMarkup(inline_keyboard=[[
            InlineKeyboardButton(text="❌ AI ni o'chirish", callback_data="ai_stop")
        ]]))
        return

    if message.text == "🧪 Bilimni sinash":
        conn2 = _get_db_conn(); cur2 = conn2.cursor()
        cur2.execute("SELECT class FROM users WHERE user_id=%s", (user_id,))
        _gr = cur2.fetchone()
        _my_grade = str(_gr[0]) if _gr and _gr[0] else "1"

        # Barcha fanlar (sinf filtrsiz)
        cur2.execute("""
            SELECT DISTINCT d.subject_name,
                   COUNT(DISTINCT g.topic_code) as cnt,
                   d.grade
            FROM generated_tests g
            JOIN dts_tree d ON d.topic_code = g.topic_code
            WHERE d.is_deleted = FALSE
            GROUP BY d.subject_name, d.grade
            ORDER BY d.grade, d.subject_name
        """)
        subjects = cur2.fetchall()
        cur2.close(); conn2.close()

        if not subjects:
            await message.answer("🧪 Hali test mavjud emas. Admin testlar qo'shishi kerak!")
            return

        rows = []
        for subj, cnt, grade in subjects:
            icon = "⭐" if str(grade) == str(_my_grade) else "🌐"
            rows.append([InlineKeyboardButton(
                text=f"{icon} {subj} ({cnt} ta mavzu)",
                callback_data=f"sinash_subj:{grade}:{subj}"
            )])
        rows.append([InlineKeyboardButton(text="⚡ Tezkor (aralash 20ta)", callback_data="tset_start_quick")])

        await message.answer(
            "🧪 Bilimni sinash\nFan tanlang:",
            reply_markup=InlineKeyboardMarkup(inline_keyboard=rows)
        )
        return

    if message.text == "🏠 Bosh ekran":
        user_state[user_id] = None
        # Eski xabarlarni o'chirish
        try:
            for i in range(message.message_id - 1, message.message_id - 20, -1):
                try:
                    await bot.delete_message(message.chat.id, i)
                except Exception:
                    pass
        except Exception:
            pass
        try:
            from student_dashboard import build_dashboard
            text, kb = await build_dashboard(user_id)
            await message.answer(text, parse_mode="HTML", reply_markup=kb)
        except Exception:
            pass
        conn2 = _get_db_conn()
        cur2  = conn2.cursor()
        cur2.execute("SELECT role FROM users WHERE user_id=%s", (user_id,))
        row = cur2.fetchone()
        cur2.close(); conn2.close()
        await message.answer("🏠 Asosiy menyu", reply_markup=get_main_keyboard(row[0] if row else "🧒 O'quvchi"))
        return

    if message.text == "🎯 Bugungi reja":
        await continue_learning(message)
        return

    if message.text == "🔊 O'qib berish":

        await read_current_page(
            message.from_user.id,
            message,
            user_state
        )

        return

    if message.text == "▶️ O'rganishni boshlash":

        await open_teacher_lesson(message)

        return

    if message.text == "📈 Rivojlanishim":
        from togarak import get_student_togaraklar, get_student_progress, get_togarak_progress
        tgs = get_student_togaraklar(user_id)

        txt = "📈 Rivojlanishim\n" + "─"*20 + "\n\n"
        rows2 = []

        if not tgs:
            txt += "📚 Hali hech qaysi to'garakka a'zo emassiz.\n\n"
            txt += "To'garakka qo'shilish uchun \"📚 To'garaklar\" bo'limiga o'ting."
        else:
            txt += "To'garaklaringiz:\n\n"
            for t in tgs:
                prog = get_togarak_progress(t["id"])
                sp = get_student_progress(t["id"], user_id)
                bdaraja = "⭐" if sp["avg_baho"]>=4.5 else ("👍" if sp["avg_baho"]>=3.5 else ("📖" if sp["avg_baho"]>0 else "—"))
                txt += f"📚 <b>{t['nomi']}</b>\n"
                txt += f"   {bdaraja} Baho: {sp['avg_baho'] or '—'} · 📋 {sp['yoqlama_pct']}% · 📊 {prog['pct']}%\n\n"
                rows2.append([
                    InlineKeyboardButton(text=f"📂 {t['nomi']}", callback_data=f"stg_info:{t['id']}"),
                ])

        await message.answer(txt[:3000], parse_mode="HTML", reply_markup=InlineKeyboardMarkup(inline_keyboard=rows2))
        return

    if message.text == "🌍 Hamjamiyat":
        await student_community(message)
        return

    if message.text == "👤 Kabinet":
        conn2=_get_db_conn();cur2=conn2.cursor()
        cur2.execute("SELECT full_name,role,class,school,birth_date FROM users WHERE user_id=%s",(user_id,))
        u=cur2.fetchone()
        # Akkauntlar soni
        cur2.execute("SELECT COUNT(*) FROM user_accounts WHERE telegram_id=%s",(_tg_id(user_id),))
        acc_count=(cur2.fetchone() or [0])[0]
        cur2.close();conn2.close()
        if not u:
            await message.answer("❌ Ma'lumot topilmadi. /start bosing."); return
        ism,rol,sinf,maktab,tugilgan=u
        txt=(f"👤 <b>Kabinet</b>\n"
             f"─────────────\n"
             f"📛 Ism: {ism or '—'}\n"
             f"🎭 Rol: {rol or '—'}\n")
        if sinf: txt+=f"🎓 Sinf: {sinf}\n"
        if maktab: txt+=f"🏫 Maktab: {maktab}\n"
        if tugilgan: txt+=f"🎂 Tug'ilgan: {tugilgan}\n"
        rows2=[
            [InlineKeyboardButton(text="✏️ Ismni o'zgartirish",callback_data="kb_change_name"),
             InlineKeyboardButton(text="🎭 Rolni o'zgartirish",callback_data="kb_change_role")],
        ]
        _rol = str(rol or "").lower()
        if "ota" in _rol or "ona" in _rol or "parent" in _rol:
            rows2.append([InlineKeyboardButton(text="👨‍👩‍👧 Farzandlarim",
                                               callback_data="fk_list")])
            rows2.append([InlineKeyboardButton(text="➕ Farzand qo'shish",
                                               callback_data="fk_add")])
        else:
            rows2.append([InlineKeyboardButton(text="🔗 Ota-onani ulash",
                                               callback_data="fk_mine")])
        rows2 += [
            [InlineKeyboardButton(text=f"🔄 Akkaunt almashtirish ({acc_count})",callback_data="kb_switch_acc")],
            [InlineKeyboardButton(text="📱 Akkauntlar / ko'chirish", callback_data="ak_menu")],
        ]
        try:
            import akkaunt as _ak0
            if _ak0.joy_bormi(_tg_id(user_id)):
                rows2.append([InlineKeyboardButton(text="➕ Yangi akkaunt",callback_data="kb_new_acc")])
            else:
                rows2.append([InlineKeyboardButton(
                    text=f"🚫 Limit: {_ak0.MAX_AKKAUNT} ta akkaunt", callback_data="ak_menu")])
        except Exception:
            rows2.append([InlineKeyboardButton(text="➕ Yangi akkaunt",callback_data="kb_new_acc")])
        await message.answer(txt,parse_mode="HTML",reply_markup=InlineKeyboardMarkup(inline_keyboard=rows2))
        return

    if user_id not in temp_user:
        temp_user[user_id] = {}

    if user_id not in user_state:
        user_state[user_id] = None

    if user_id not in user_locks:
        user_locks[user_id] = asyncio.Lock()

    # SOZLAMALAR HOLATLARI
    if user_state.get(user_id) == "change_name":
        conn2 = _get_db_conn()
        cur2 = conn2.cursor()
        cur2.execute("UPDATE users SET full_name=%s WHERE user_id=%s", (message.text, user_id))
        conn2.commit(); cur2.close(); conn2.close()
        user_state[user_id] = None
        await message.answer(f"✅ Ism o'zgartirildi: {message.text}")
        return

    if user_state.get(user_id) == "change_bdate":
        try:
            from datetime import datetime
            bdate = datetime.strptime(message.text.strip(), "%d.%m.%Y").date()
            conn2 = _get_db_conn()
            cur2 = conn2.cursor()
            cur2.execute("UPDATE users SET birth_date=%s WHERE user_id=%s", (bdate, user_id))
            conn2.commit(); cur2.close(); conn2.close()
            user_state[user_id] = None
            await message.answer(f"✅ Tug'ilgan kun saqlandi: {message.text}")
        except Exception:
            await message.answer("❌ Format xato! Masalan: 15.03.2015")
        return

    if user_state.get(user_id) == "change_school_settings":
        conn2 = _get_db_conn()
        cur2 = conn2.cursor()
        cur2.execute("UPDATE users SET school=%s WHERE user_id=%s", (message.text, user_id))
        conn2.commit(); cur2.close(); conn2.close()
        user_state[user_id] = None
        await message.answer(f"✅ Maktab saqlandi: {message.text}")
        return

    # REGISTRATSIYA
    if user_state.get(user_id) and user_state.get(user_id) not in ("text_answer",):

        await register_handler(message)
        return

    try:
        conn = _get_db_conn()
        cur = conn.cursor()

        cur.execute("""
        UPDATE users
        SET last_seen=NOW()
        WHERE user_id=%s
        """, (user_id,))

        conn.commit()
        conn.close()

    except:
        pass

    # ── Kitob yuklash matn handler ──
    if (str(admin_state.get(user_id) or "").startswith("ai_rasm_custom") and message.text):
        parts_s = admin_state[user_id].split(":")
        style = parts_s[1] if len(parts_s)>1 else "multik"
        tavsif = message.text.strip()
        admin_state.pop(user_id, None)
        status_r = await message.answer(f"🤔 Tushunmoqda...\n«{tavsif[:60]}»")

        async def do_smart_rasm():
            try:
                from rasim_generator import generate_smart
                await status_r.edit_text(f"🎨 Chizilmoqda (HD)...\n«{tavsif[:60]}»")
                img, prompt = await generate_smart(tavsif, "ta'lim", "1", style, is_admin=True)
                if img:
                    from aiogram.types import BufferedInputFile
                    sent = await message.answer_photo(
                        BufferedInputFile(img, "rasm.png"),
                        caption=f"🎨 {tavsif[:60]}\n\n📝 <i>{(prompt or '')[:800]}</i>",
                        parse_mode="HTML"
                    )
                    await status_r.edit_text(
                        f"✅ Tayyor!\n\nDB ga saqlash uchun nom yozing:\n"
                        f"Masalan: <code>1-01-1-01-01-01-001-1</code>",
                        parse_mode="HTML"
                    )
                    fid = sent.photo[-1].file_id
                    admin_state[user_id] = f"save_rasm:{fid}"
                else:
                    await status_r.edit_text("❌ Rasm yaratilmadi. Qayta urinib ko'ring.")
            except Exception as e:
                await status_r.edit_text(f"❌ {e}")

        asyncio.create_task(do_smart_rasm())
        return

    if admin_state.get(user_id) == "ai_rasm_custom" and message.text:
        tavsif = message.text.strip()
        admin_state.pop(user_id, None)
        status_r = await message.answer(f"⏳ Rasm yaratilmoqda...\n🖌 {tavsif[:60]}")
        async def do_custom_rasm():
            try:
                from rasim_generator import generate_dalle
                img_bytes = await generate_dalle(tavsif)
                if img_bytes:
                    from aiogram.types import BufferedInputFile
                    sent = await message.answer_photo(
                        BufferedInputFile(img_bytes, "rasm.png"),
                        caption=f"🎨 {tavsif[:100]}"
                    )
                    fid = sent.photo[-1].file_id
                    # Nomi so'raladi
                    admin_state[user_id] = f"save_rasm:{fid}"
                    await status_r.edit_text(
                        "✅ Rasm tayyor!\n\nDB ga saqlash uchun nom yozing:\n"
                        "Masalan: <code>1-01-1-01-01-01-001-1</code>",
                        parse_mode="HTML"
                    )
                else:
                    await status_r.edit_text("❌ Rasm yaratilmadi. API kalitni tekshiring.")
            except Exception as e:
                await status_r.edit_text(f"❌ {e}")
        asyncio.create_task(do_custom_rasm())
        return

    if (str(admin_state.get(user_id) or "").startswith("save_rasm:") and message.text):
        fid = admin_state[user_id].split(":",1)[1]
        name = message.text.strip()
        admin_state.pop(user_id, None)
        conn2 = _get_db_conn(); cur2 = conn2.cursor()
        cur2.execute("""INSERT INTO images(name,file_id) VALUES(%s,%s)
                        ON CONFLICT(name) DO UPDATE SET file_id=EXCLUDED.file_id""",
                    (name, fid))
        conn2.commit(); cur2.close(); conn2.close()
        await message.answer(f"✅ Saqlandi! Kod: <code>{name}</code>", parse_mode="HTML")
        return

    if admin_state.get(user_id) == "mtt_sinf_fan" and message.text:
        txt_sf = message.text.strip()
        if "|" not in txt_sf:
            await message.answer("❌ Format: <code>sinf | fan</code>\nMasalan: <code>2 | Ingliz tili</code>", parse_mode="HTML")
            return
        parts_sf = txt_sf.split("|", 1)
        gr_sf  = parts_sf[0].strip()
        fan_sf = parts_sf[1].strip()
        admin_state[user_id] = f"mtt_mavzu:{gr_sf}:{fan_sf}"
        await message.answer(
            f"✅ {gr_sf}-sinf | {fan_sf}\n\n"
            f"Endi chorak/mavzularni yozing:\n\n"
            f"1/ Alphabet review\n"
            f"1/ Hello Greetings\n"
            f"2/ My family\n"
            f"2/ My house\n\n"
            f"Format: chorak_raqami/ mavzu_nomi"
        )
        return

    if str(admin_state.get(user_id) or "").startswith("mtt_fan_input:") and message.text:
        gr_fi = admin_state[user_id].split(":")[1]
        fan_fi = message.text.strip()
        admin_state[user_id] = f"mtt_mavzu:{gr_fi}:{fan_fi}"
        await message.answer(
            f"✅ Fan: {fan_fi}\n\n"
            f"Endi mavzularni yozing:\n\n"
            f"1/ Alphabet review\n"
            f"1/ Hello Greetings\n"
            f"2/ My family\n\n"
            f"Format: chorak_raqami/ mavzu_nomi"
        )
        return

    if str(admin_state.get(user_id) or "").startswith("mtt_mavzu:") and message.text:
        parts3 = admin_state[user_id].split(":", 2)
        gr3, fan3 = parts3[1], parts3[2]
        admin_state.pop(user_id, None)
        text3 = message.text.strip()

        # Mavzularni parse qilish
        mavzular3 = []
        for line in text3.split("\n"):
            line = line.strip()
            if not line: continue
            if "/" in line:
                parts_l = line.split("/", 1)
                chorak_n = parts_l[0].strip()
                mavzu_n  = parts_l[1].strip()
                if mavzu_n:
                    mavzular3.append((chorak_n, mavzu_n))

        if not mavzular3:
            await message.answer("❌ Format noto'g'ri!\n\nMisol:\n1/ Alphabet review\n1/ Greetings\n2/ My family")
            admin_state[user_id] = f"mtt_mavzu:{gr3}:{fan3}"
            return

        await message.answer(f"⏳ DTS shablon yaratilmoqda...\n{len(mavzular3)} ta mavzu")

        # 1-QADAM: DTS shablon yaratish
        from shablon_yaratish import _create_shablon
        from aiogram.types import BufferedInputFile
        buf3 = await _create_shablon(gr3, fan3, mavzular3)
        fname3 = f"DTS_{gr3}sinf_{fan3[:15].replace(' ','_')}.xlsx"
        await message.answer_document(
            BufferedInputFile(buf3.read(), filename=fname3),
            caption=(
                f"1️⃣ DTS Shablon tayyor!\n"
                f"🏫 {gr3}-sinf | 📚 {fan3}\n"
                f"📝 {len(mavzular3)} ta mavzu x 2 qator\n\n"
                f"Bu faylni to'ldirib botga yuboring:\n"
                f"Bob, Bo'lim, Kichik mavzu ustunlarini\n\n"
                f"To'ldirib yuborilsa — DTS import qilinadi\n"
                f"va test shabloni tayyorlanadi!"
            )
        )
        # Test sozlamalari — DTS dan keyin
        admin_state[user_id] = f"mtt_test_sozlama:{gr3}:{fan3}"
        from ai_generatori import _gen_settings_kb, gen_state
        state4 = gen_state.get(user_id, {})
        state4["grade"] = gr3; state4["subject"] = fan3
        state4["mavzular_text"] = text3
        state4.setdefault("gen_groups",[
            {"diff":"oson",    "type":"single_choice","count":5},
            {"diff":"orta",    "type":"single_choice","count":5},
            {"diff":"qiyin",   "type":"single_choice","count":5},
            {"diff":"murakkab","type":"single_choice","count":5},
        ])
        gen_state[user_id] = state4
        await message.answer(
            f"2️⃣ Test shablon sozlamalari\n"
            f"🏫 {gr3}-sinf | 📚 {fan3}\n"
            f"📝 {len(mavzular3)} ta mavzu\n\n"
            f"Har qiyinlikdan nechta savol?",
            reply_markup=_gen_settings_kb(state4["gen_groups"])
        )
        return

    if admin_state.get(user_id) == "kitob_yuklash" and message.text:
        txt = message.text.strip().strip("`").strip()
        parts = [x.strip() for x in txt.split("|")]
        if len(parts) < 2:
            await message.answer(
                "❌ Format:\n<code>Kitob nomi | Fan | Sinf | Muallif</code>",
                parse_mode="HTML"
            )
            return
        admin_state[f"{user_id}_kitob_info"] = {
            "title":   parts[0],
            "fan":     parts[1] if len(parts)>1 else "",
            "sinf":    parts[2] if len(parts)>2 else "",
            "muallif": " | ".join(parts[3:]) if len(parts)>3 else "",
        }
        admin_state[user_id] = "kitob_yuklash_pdf"
        info = admin_state[f"{user_id}_kitob_info"]
        await message.answer(
            f"✅ Saqlandi:\n📖 {info['title']}\n"
            f"📚 {info['fan']} | 🏫 {info['sinf']}-sinf\n"
            f"✍️ {info['muallif']}\n\nPDF faylni yuboring 👇"
        )
        return

    if (
        admin_state.get(user_id) == "dts_import"
        and message.document
    ):
        from dts_import_handlers import DTSImportState
        await state.set_state(DTSImportState.waiting_excel)
        await dts_excel_import(
            message,
            state
        )
        admin_state[user_id] = None
        return

    # Shablon import
    if message.document:
        from shablon_yaratish import handle_shablon_document, shablon_state
        if shablon_state.get(user_id, {}).get("step") == "import_wait":
            await handle_shablon_document(message, user_id, bot)
            return

    # Shablon mavzu matni
    from shablon_yaratish import shablon_state as sh_state, handle_shablon_message
    # Tugma bosilgan bo'lsa shablon_state ni o'tkazib yuboramiz
    _menu_btns = {
        "📋 Shablonlar","🧪 Test shablon","📥 Test import",
        "📋 Topik shablon","📥 Topik import",
        "📚 Dars shablon","📥 Dars import",
        "🔙 Admin menyu","📊 Test statistikasi",
        "🤖 AI Generator","📚 Shablon yaratish",
        "🔙 Ortga",
    }
    if (
        message.text not in _menu_btns
        and sh_state.get(user_id, {}).get("step") in ("sinf_fan", "mavzular")
    ):
        await handle_shablon_message(message, user_id)
        return

    # ── Excel RASMLAR varog'i bo'lsa — avtomatik rasm yaratish ──
    if (message.document and _is_admin(user_id) and
            message.document.file_name and
            (message.document.file_name or "").endswith(".xlsx")):
        # RASMLAR varog'i borligini tekshiramiz
        from io import BytesIO
        buf_check = BytesIO()
        await message.bot.download(message.document.file_id, destination=buf_check)
        buf_check.seek(0)
        try:
            import openpyxl as _opx
            _wb = _opx.load_workbook(buf_check, data_only=True)
            if "RASMLAR" in _wb.sheetnames:
                _wb.close()
                buf_check.seek(0)
                xl_bytes2 = buf_check.read()
                admin_state[f"{user_id}_rasm_xl"] = xl_bytes2
                await message.answer(
                    "🖼 RASMLAR varog\'i topildi!\n\nQaysi uslubda chizilsin?",
                    reply_markup=InlineKeyboardMarkup(inline_keyboard=[
                        [InlineKeyboardButton(text="🎠 Multik — bolalarga mos (tavsiya)", callback_data="xl_style:multik")],
                        [InlineKeyboardButton(text="📸 Hayotiy — realistik",              callback_data="xl_style:hayotiy")],
                        [InlineKeyboardButton(text="✏️ Chizma — darslik uslubi",          callback_data="xl_style:chizma")],
                        [InlineKeyboardButton(text="📚 Darslik — ta\'lim diagramma",     callback_data="xl_style:darslik")],
                        [InlineKeyboardButton(text="🎮 3D — hajmli",                      callback_data="xl_style:3d")],
                        [InlineKeyboardButton(text="⚡ Avto (multik) — hoziroq",          callback_data="xl_style:multik_auto")],
                        [InlineKeyboardButton(text="🔄 Qayta yaratish (eskilarni o\'chir)", callback_data="xl_style:multik_force")],
                    ])
                )
                return
            _wb.close()
        except: pass

    # ── Kitob PDF yuklash ──
    if (message.document and _is_admin(user_id) and
            message.document.file_name and
            message.document.file_name.lower().endswith(".pdf")):
        # State bo'lmasa ham qabul qilamiz
        info = admin_state.get(f"{user_id}_kitob_info", {})
        if not info:
            # Fayl nomidan ma'lumot olish
            fname = message.document.file_name.replace(".pdf","")
            info = {"title": fname, "fan": "Matematika", "sinf": "7", "muallif": ""}
        title   = info.get("title", "Kitob")
        fan     = info.get("fan", "Matematika")
        sinf    = info.get("sinf", "7")
        muallif = info.get("muallif", "")
        fid     = message.document.file_id
        admin_state[user_id] = None
        admin_state.pop(f"{user_id}_kitob_info", None)
        status_k = await message.answer(
            f"📖 {title}\n⏳ PDF qabul qilindi, yuklanmoqda..."
        )

        async def do_process_kitob():
            try:
                import kitob_bazasi as _kb
                import tempfile, os as _os, aiohttp as _ah

                await status_k.edit_text("⏳ Fayl ma'lumoti olinmoqda...")
                print(f"[KITOB] get_file boshlandi")

                try:
                    file_info = await message.bot.get_file(fid)
                    file_path = file_info.file_path
                    print(f"[KITOB] file_path: {file_path}")
                except Exception as fe:
                    print(f"[KITOB] get_file xato: {fe}")
                    await status_k.edit_text(
                        f"❌ Fayl yuklab bo'lmadi!\n"
                        f"Sabab: {fe}\n\n"
                        f"Telegram 20MB dan katta fayllarni bermir.\n"
                        f"Faylni siqib (zip) yoki bo'lib yuboring."
                    )
                    return

                token = message.bot.token
                url = f"https://api.telegram.org/file/bot{token}/{file_path}"

                tmp_k = tempfile.NamedTemporaryFile(suffix=".pdf", delete=False)
                async with _ah.ClientSession() as sess:
                    async with sess.get(url, timeout=_ah.ClientTimeout(total=120)) as resp:
                        chunk_size = 65536
                        downloaded = 0
                        while True:
                            chunk = await resp.content.read(chunk_size)
                            if not chunk: break
                            tmp_k.write(chunk)
                            downloaded += len(chunk)
                            if downloaded % (1024*1024) == 0:
                                mb = downloaded // (1024*1024)
                                print(f"[KITOB] {mb}MB yuklanди")
                tmp_k.close()
                print(f"[KITOB] PDF saqlandi: {tmp_k.name}")

                async def upd(msg):
                    print(f"[KITOB] {msg}")
                    try: await status_k.edit_text(msg)
                    except: pass

                await upd("⏳ Boshlandi...")
                result = await _kb.load_book_to_db(
                    file_path=tmp_k.name,
                    sinf=sinf, fan=fan, muallif=muallif,
                    progress_cb=upd
                )
                _os.remove(tmp_k.name)
                await message.answer(
                    f"✅ Kitob saqlandi!\n📖 {title}\n"
                    f"📄 {result['pages']} bet | 📐 {result['exercises']} misol\n"
                    f"🔑 Book ID: {result['book_id']}"
                )
            except Exception as e:
                import traceback
                print(f"[KITOB ERROR] {e}\n{traceback.format_exc()}")
                await message.answer(f"❌ Xato: {e}")
            finally:
                admin_state.pop(user_id, None)
                admin_state.pop(f"{user_id}_kitob_info", None)

        asyncio.create_task(do_process_kitob())
        return

    # ── Excel shablon yuborilsa — faqat "excel_merge" state da ──
    if (message.document and _is_admin(user_id) and
            message.document.file_name and
            (message.document.file_name or "").endswith(".xlsx") and
            admin_state.get(user_id) == "excel_merge"):

        first_key = f"merge_file1_{user_id}"

        if first_key in admin_state:
            # Ikkinchi fayl keldi — birlashtirish
            file1_bytes = admin_state.pop(first_key)
            buf2 = BytesIO()
            await message.bot.download(message.document.file_id, destination=buf2)
            st = await message.answer("⏳ Birlashtirilyapti...")
            try:
                from excel_merger import merge_excel
                res_bytes, info = merge_excel(file1_bytes, buf2.getvalue())
                if res_bytes:
                    from aiogram.types import BufferedInputFile
                    await message.answer_document(
                        BufferedInputFile(res_bytes, "birlashtirilgan.xlsx"),
                        caption=info
                    )
                    await st.delete()
                else:
                    await st.edit_text(info)
            except Exception as e:
                await st.edit_text(f"❌ Xato: {e}")
        else:
            # Birinchi fayl — DB dan to'ldirishga urinish
            buf1 = BytesIO()
            await message.bot.download(message.document.file_id, destination=buf1)
            file1_bytes = buf1.getvalue()
            st = await message.answer("⏳ DB dan savollar qidirilmoqda...")
            try:
                from excel_merger import fill_from_db
                res_bytes, info = fill_from_db(file1_bytes)
                if res_bytes and "To'ldirildi" in info:
                    from aiogram.types import BufferedInputFile
                    fname_out = "toldiriilgan_" + (message.document.file_name or "shablon.xlsx")
                    await message.answer_document(
                        BufferedInputFile(res_bytes, fname_out),
                        caption=info
                    )
                    await st.delete()
                else:
                    admin_state[first_key] = file1_bytes
                    await st.edit_text(
                        f"📎 {message.document.file_name or 'fayl'} saqlandi.\n"
                        f"DB da savol topilmadi.\n\n"
                        f"2-fayl yuboring (savol fayli) — birlashtiraman!"
                    )
            except Exception as e:
                admin_state[first_key] = file1_bytes
                await st.edit_text(
                    f"📎 {message.document.file_name or 'fayl'} saqlandi.\n"
                    f"2-fayl yuboring — birlashtiraman!"
                )
        return

    if (
        admin_state.get(user_id) == "test_import"
        and message.document
    ):

        await prepare_test_import(message, user_id)

        return

    if user_state.get(message.from_user.id) == "in_test":
        # Test paytida matn yozsa — yumshoq eslatma
        try:
            await message.answer(
                "🧪 Test davom etyapti!\n"
                "Javob berish uchun tugmalardan foydalaning.\n\n"
                "Chiqish: /menu"
            )
        except: pass
        return

    if user_state.get(message.from_user.id) == "in_lesson":
        # Dars paytida matn yozsa — yumshoq eslatma
        try:
            await message.answer(
                "📖 Dars davom etyapti!\n"
                "Oldinga o'tish uchun tugmalardan foydalaning.\n\n"
                "Chiqish: /menu"
            )
        except: pass
        return

    if user_state.get(user_id) == "text_answer":

        _javob = message.text
        _chat  = message.chat.id
        _mid   = message.message_id

        await check_text_answer(
            user_id,
            _javob,
            message
        )

        # Foydalanuvchi yozgan javobni o'chiramiz — ekran to'lmasin
        try:
            await bot.delete_message(_chat, _mid)
        except Exception:
            pass

        return

    if message.text == "👥 Foydalanuvchilar statistikasi":

        if message.from_user.id not in ADMINS:
            return

        conn = _get_db_conn()
        cur = conn.cursor()

        cur.execute("SELECT COUNT(*) FROM users")
        total = cur.fetchone()[0]

        cur.execute("""
        SELECT COUNT(*)
        FROM users
        WHERE DATE(last_seen)=CURRENT_DATE
        """)
        today = cur.fetchone()[0]

        cur.execute("""
        SELECT COUNT(*)
        FROM users
        WHERE last_seen >= NOW() - INTERVAL '30 day'
        """)
        month = cur.fetchone()[0]

        conn.close()

        await message.answer(
            f"👥 Jami foydalanuvchilar: {total}\n\n"
            f"📅 Bugun kirganlar: {today}\n"
            f"🗓 Oxirgi 30 kun: {month}"
        )

        return

    elif message.text == "🧪 Test sinovi":
        if user_id not in ADMINS:
            return
        from test_sinovi import show_sinov_start
        await show_sinov_start(message, user_id)
        return

    elif message.text == "👥 Foydalanuvchilar":
        if user_id not in ADMINS:
            return
        conn2 = _get_db_conn()
        cur2 = conn2.cursor()
        cur2.execute("SELECT COUNT(*) FROM users")
        total = cur2.fetchone()[0]
        cur2.execute("SELECT role, COUNT(*) FROM users GROUP BY role ORDER BY COUNT(*) DESC")
        roles = cur2.fetchall()
        cur2.close(); conn2.close()
        text = f"👥 Foydalanuvchilar: {total} ta\n\n"
        for role, cnt in roles:
            text += f"• {role}: {cnt} ta\n"
        await message.answer(text)
        return

    elif message.text == "🖼 Rasmlar boshqaruvi":
        if user_id not in ADMINS: return
        # 2 xil rejim
        await message.answer(
            "🖼 Rasmlar boshqaruvi",
            reply_markup=InlineKeyboardMarkup(inline_keyboard=[
                [InlineKeyboardButton(text="📤 Rasm yuklash (sinf→fan→mavzu)", callback_data="img_upload_start")],
                [InlineKeyboardButton(text="📋 Barcha rasmlar", callback_data="img_panel")],
            ])
        )
        return

    elif message.text == "🖼 Rasmlar boshqaruvi":
        if user_id not in ADMINS:
            return
        from image_admin import show_image_panel
        await show_image_panel(message)
        return

    elif message.text == "📚 Shablon yaratish":
        if user_id not in ADMINS:
            return
        from shablon_yaratish import show_shablon_menu
        await show_shablon_menu(message, user_id)
        return

    elif message.text in ("🤖 AI Generator", "🧪 Test shablon"):
        if user_id not in ADMINS:
            return
        from ai_generatori import show_gen_start
        await show_gen_start(message, user_id)
        return

    elif message.text in ("📚 DTS boshqaruvi", "🧭 DTS topik boshqaruvi"):
        # Shablon state ni tozalamiz
        from shablon_yaratish import shablon_state
        shablon_state.pop(user_id, None)
        await dts_menu(message)
        return

    elif message.text == "📊 Test statistikasi":

        await message.answer(
            "📊 Test statistikasi / Generator",
            reply_markup=ReplyKeyboardMarkup(
                keyboard=[
                    [KeyboardButton(text="📚 Mavzular statistikasi"),
                    KeyboardButton(text="📊 Generator statistikasi")],
                    [KeyboardButton(text="▶️ Generatorni boshlash"),
                    KeyboardButton(text="⏹ Generatorni to‘xtatish")],
                    [KeyboardButton(text="🔙 Ortga")]
                ],
                  resize_keyboard=True
            )
        )

        return

    elif message.text == "🧪 Bilimni sinash":
        # Admin uchun test
        if user_id not in ADMINS:
            pass  # o'quvchi uchun allaqachon handled
        else:
            # Admin uchun ham xuddi o'quvchi kabi
            conn2 = _get_db_conn(); cur2 = conn2.cursor()
            cur2.execute("""
                SELECT grade FROM (SELECT DISTINCT grade FROM dts_tree WHERE is_deleted=FALSE) _g
                ORDER BY CASE WHEN grade ~ '^[0-9]+$' THEN grade::int ELSE 9999 END, grade
            """)
            _grades = [r[0] for r in cur2.fetchall()]
            cur2.close(); conn2.close()
            rows = [[InlineKeyboardButton(
                text=f"{g}-sinf" if str(g).isdigit() else str(g),
                callback_data=f"stnav_grade:{g}"
            )] for g in _grades]
            rows.append([InlineKeyboardButton(text="⚡ Tezkor 20ta", callback_data="tset_start_quick")])
            await message.answer(
                "🧪 Test — sinf tanlang:",
                reply_markup=InlineKeyboardMarkup(inline_keyboard=rows)
            )
            return

    elif message.text == "📦 Kitob Word":
        if user_id not in ADMINS: return
        # Kitoblar ro'yxati
        conn2 = _get_db_conn(); cur2 = conn2.cursor()
        cur2.execute("SELECT id, title, fan, sinf FROM books ORDER BY id DESC LIMIT 10")
        books = cur2.fetchall(); cur2.close(); conn2.close()
        if not books:
            await message.answer("❌ Hali kitob yuklanmagan.\n«📖 Kitob yuklash» dan boshlang.")
            return
        rows = []
        for b in books:
            rows.append([InlineKeyboardButton(
                text=f"📖 {b[1][:25]} ({b[2]}, {b[3]}-sinf)",
                callback_data=f"book_make:{b[0]}"
            )])
        rows.append([InlineKeyboardButton(text="📦 To'liq paket (ZIP)", callback_data="book_full_pkg")])
        await message.answer(
            "📖 Kitob yaratish\n\nQaysi kitobdan Word fayl yasaymiz?\n"
            "(15 betlik bo'laklarga bo'linadi, ZIP arxivda yuboriladi)",
            reply_markup=InlineKeyboardMarkup(inline_keyboard=rows)
        )
        return

    elif message.text and message.text.startswith("📊 Hisobotlar"):
        if user_id not in ADMINS: return
        conn2 = _get_db_conn(); cur2 = conn2.cursor()
        # O'qilmagan xatolar
        cur2.execute("SELECT COUNT(*) FROM error_log WHERE is_read=FALSE")
        unread = cur2.fetchone()[0] or 0
        # O'qilgan xatolar
        cur2.execute("SELECT COUNT(*) FROM error_log WHERE is_read=TRUE")
        read_cnt = cur2.fetchone()[0] or 0
        cur2.close(); conn2.close()

        rows = [
            [InlineKeyboardButton(text="📊 Test natijalari (Excel)", callback_data="rep_test")],
            [InlineKeyboardButton(text="📈 O'quvchi taraqqiyoti (Excel)", callback_data="rep_prog")],
            [InlineKeyboardButton(text="📅 Dars rejasi", callback_data="rep_plan")],
        ]
        # Xatolar — aniq ajratilgan
        if unread > 0:
            rows.append([InlineKeyboardButton(
                text=f"🔴 O'qilmagan xatolar: {unread} ta — KO'RISH",
                callback_data="err_unread"
            )])
        rows.append([InlineKeyboardButton(
            text=f"✅ O'qilgan xatolar: {read_cnt} ta",
            callback_data="err_read"
        )])
        rows.append([InlineKeyboardButton(
            text="🗑 Barcha xatolarni tozalash",
            callback_data="err_clear"
        )])

        txt = (
            f"📊 Hisobotlar & Xatolar\n\n"
            f"🔴 O'qilmagan: {unread} ta\n"
            f"✅ O'qilgan: {read_cnt} ta"
        )
        await message.answer(txt, reply_markup=InlineKeyboardMarkup(inline_keyboard=rows))
        return

    elif message.text == "📊 Hisobotlar":
        if user_id not in ADMINS: return
        from jadval_generator import test_results_text, weak_analysis_text
        text2  = test_results_text(days=30)
        text3  = weak_analysis_text()
        await message.answer(text2)
        await message.answer(text3)
        # Excel ham yuboramiz
        from jadval_generator import test_results_excel
        from aiogram.types import BufferedInputFile
        try:
            buf = test_results_excel(days=30)
            await message.answer_document(
                BufferedInputFile(buf.read(), "test_natijalari.xlsx"),
                caption="📊 Test natijalari (Excel)"
            )
        except Exception as e:
            await message.answer(f"Excel xato: {e}")
        return

    elif message.text == "📅 Dars rejasi":
        if user_id not in ADMINS: return
        admin_state[user_id] = "dars_rejasi"
        await message.answer(
            "📅 Dars rejasi\n\nSinf va fanni yozing:\n<code>1 | Ingliz tili</code>",
            parse_mode="HTML"
        )
        return

    elif message.text == "📈 Taraqqiyot":
        if user_id not in ADMINS: return
        from jadval_generator import student_progress_text, student_progress_excel
        from aiogram.types import BufferedInputFile
        text2 = student_progress_text()
        await message.answer(text2)
        try:
            buf = student_progress_excel()
            await message.answer_document(
                BufferedInputFile(buf.read(), "taraqqiyot.xlsx"),
                caption="📈 O'quvchi taraqqiyoti (Excel)"
            )
        except Exception as e:
            await message.answer(f"Excel xato: {e}")
        return

    elif message.text == "📝 Shablon to'ldirish":
        if user_id not in ADMINS: return
        admin_state[user_id] = "excel_merge"
        await message.answer(
            "📝 Shablon to'ldirish\n\n"
            "Bo'sh shablon Excel faylini yuboring.\n"
            "Bot DB dagi savollardan avtomatik to'ldiradi.\n\n"
            "Agar DB da savol yo'q bo'lsa — 2-fayl (savol fayli) ham yuborasiz."
        )
        return

    elif message.text == "📚 Kitoblar ▾":
        if user_id not in ADMINS: return
        conn2=_get_db_conn();cur2=conn2.cursor()
        try:
            cur2.execute("SELECT id,title,sinf,total_pages FROM books ORDER BY id DESC LIMIT 10")
            books=cur2.fetchall()
        except: books=[]
        cur2.close();conn2.close()
        rows2=[[InlineKeyboardButton(text=f"📖 {b[1]} ({b[3]} bet)",callback_data=f"kitob_info:{b[0]}")] for b in books]
        rows2.append([
            InlineKeyboardButton(text="📤 PDF yuklash",callback_data="kitob_upload"),

        ])
        await message.answer("📚 Kitoblar:", reply_markup=InlineKeyboardMarkup(inline_keyboard=rows2))
        return

    elif message.text == "📚 Kitoblar ▾":
        if user_id not in ADMINS: return
        await message.answer(
            "📚 Kitoblar bo'limi:",
            reply_markup=InlineKeyboardMarkup(inline_keyboard=[
                [InlineKeyboardButton(text="📖 Kitob yuklash",    callback_data="menu_kitob_yuklash")],
                [InlineKeyboardButton(text="🎓 Kitobni o'qit (AI)", callback_data="menu_kitob_oqit")],
                [InlineKeyboardButton(text="📦 Kitobdan Word",    callback_data="menu_kitob_word")],
            ])
        )
        return

    elif message.text == "🧠 Bilimlar ▾":
        if user_id not in ADMINS: return
        await message.answer(
            "🧠 Bilimlar bo'limi:",
            reply_markup=InlineKeyboardMarkup(inline_keyboard=[
                [InlineKeyboardButton(text="🔍 Bilim qidirish",          callback_data="menu_bilim_qidir")],
                [InlineKeyboardButton(text="📚 Bilimni mustahkamlash",   callback_data="menu_bilim_must")],
                [InlineKeyboardButton(text="🧪 Bilimni sinash (test)",   callback_data="menu_bilim_sin")],
            ])
        )
        return

    elif message.text == "🚀 Mavzu tayyorla":
        if user_id not in ADMINS: return
        admin_state[user_id] = "mtt_sinf_fan"
        await message.answer(
            "🚀 Mavzu tayyorlash\n\n"
            "Sinf va fanni yozing:\n"
            "Format: <code>sinf | fan</code>\n\n"
            "Masalan:\n"
            "<code>2 | Ingliz tili</code>\n"
            "<code>3 | Matematika</code>",
            parse_mode="HTML"
        )
        return

    elif message.text == "🎨 AI Rasm yaratish":
        if user_id not in ADMINS: return
        conn2 = _get_db_conn(); cur2 = conn2.cursor()
        cur2.execute("""
            SELECT grade FROM (
                SELECT DISTINCT d.grade FROM dts_tree d
                JOIN generated_tests g ON g.topic_code=d.topic_code
                WHERE d.is_deleted=FALSE
            ) _g
            ORDER BY CASE WHEN grade ~ '^[0-9]+$' THEN grade::int ELSE 99 END
        """)
        grades = [r[0] for r in cur2.fetchall()]
        cur2.close(); conn2.close()
        rows = [[InlineKeyboardButton(
            text=f"🏫 {gr}-sinf" if str(gr).isdigit() else f"📚 {gr}",
            callback_data=f"rasm_grade:{gr}"
        )] for gr in grades]
        rows.append([InlineKeyboardButton(text="✏️ O'zim tavsif beraman", callback_data="ai_rasm_custom")])
        await message.answer(
            "🎨 AI Rasm yaratish (BEPUL)\n\nSinf tanlang:",
            reply_markup=InlineKeyboardMarkup(inline_keyboard=rows)
        )
        return
    elif message.text == "📖 Kitob yuklash":
        if user_id not in ADMINS: return
        admin_state[user_id] = "kitob_yuklash"
        await message.answer(
            "📖 Kitob yuklash\n\n"
            "Quyidagi formatda yozing:\n"
            "<code>Kitob nomi | Fan | Sinf | Muallif</code>\n\n"
            "Masalan:\n"
            "<code>Matematika 5 | Matematika | 5 | Mirzayev A.</code>\n\n"
            "Keyin PDF faylni yuboring.",
            parse_mode="HTML"
        )
        return

    elif message.text == "🎓 Kitob o'qit":
        if user_id not in ADMINS: return
        # Oxirgi yuklangan kitobni o'qitamiz
        conn2 = _get_db_conn(); cur2 = conn2.cursor()
        cur2.execute("SELECT id,title,fan,sinf FROM books ORDER BY id DESC LIMIT 5")
        books = cur2.fetchall(); cur2.close(); conn2.close()
        if not books:
            await message.answer("❌ Hali kitob yuklanmagan."); return
        rows = [[InlineKeyboardButton(
            text=f"📖 {b[1]} ({b[2]}, {b[3]}-sinf)",
            callback_data=f"train_book:{b[0]}:{b[2]}:{b[3]}"
        )] for b in books]
        await message.answer(
            "🎓 Qaysi kitobni o'qitamiz?\n(AI tahlil qilib bilim bazasiga saqlaydi)",
            reply_markup=InlineKeyboardMarkup(inline_keyboard=rows)
        )
        return

    elif message.text == "🔍 Bilim qidirish":
        if user_id not in ADMINS: return
        admin_state[user_id] = "kitob_qidirish"
        await message.answer("🔍 Qidiruv so'zini yozing:")
        return

    elif message.text == "🔧 Test tuzatmalar":
        if user_id not in ADMINS: return
        conn2 = _get_db_conn(); cur2 = conn2.cursor()
        cur2.execute("""
            SELECT id, test_id, topic_code, question, comment, created_at
            FROM test_corrections WHERE status='new'
            ORDER BY created_at DESC LIMIT 15
        """)
        rows2 = cur2.fetchall(); cur2.close(); conn2.close()
        if not rows2:
            await message.answer("✅ Tuzatish so'rovlari yo'q!"); return
        lines = [f"🔧 Tuzatish so'rovlari ({len(rows2)} ta):\n"]
        kbs = []
        for cid, tid, tc, q, comm, cat in rows2:
            lines.append(f"📝 {q[:60]}...")
            if tid:
                kbs.append([InlineKeyboardButton(
                    text=f"✏️ #{tid} — ko'rish",
                    callback_data=f"admin_fix_test:{tid}"
                )])
        await message.answer("\n".join(lines[:10]),
                             reply_markup=InlineKeyboardMarkup(inline_keyboard=kbs[:10]))
        return

    elif message.text and ("🆘 Xatolar" in message.text):
        if user_id not in ADMINS:
            return
        conn2 = _get_db_conn()
        cur2  = conn2.cursor()
        cur2.execute("""
            SELECT id, user_id, username, error_text, created_at
            FROM error_log
            ORDER BY created_at DESC
            LIMIT 20
        """)
        errors = cur2.fetchall()
        # O'qilgan deb belgilash
        cur2.execute("UPDATE error_log SET is_read=TRUE WHERE is_read=FALSE")
        conn2.commit(); cur2.close(); conn2.close()

        if not errors:
            await message.answer(
                "✅ Xatolar yo'q!",
                reply_markup=get_main_keyboard("Admin", unread_errors=0)
            )
            return

        lines = [f"🆘 So'nggi xatolar ({len(errors)} ta):\n"]
        for i, (eid, uid, uname, etxt, eat) in enumerate(errors, 1):
            t = eat.strftime("%d.%m %H:%M") if eat else "?"
            short = (etxt or "")[:150].replace("\n", " ")
            lines.append(f"{i}. [{t}] {uname}({uid})\n   {short}\n")

        # Uzun bo'lsa bo'lib yuborish
        text = "\n".join(lines)
        while text:
            await message.answer(text[:4096])
            text = text[4096:]

        await message.answer(
            "✅ Hammasi o'qilgan deb belgilandi.",
            reply_markup=get_main_keyboard("Admin", unread_errors=0)
        )
        return

    elif message.text == "📖 Darslar holati":
        if user_id not in ADMINS:
            return
        conn2 = _get_db_conn()
        cur2  = conn2.cursor()
        cur2.execute("""
            SELECT
                d.grade,
                d.subject_name,
                COUNT(DISTINCT d.topic_code)           AS jami,
                COUNT(DISTINCT l.topic_code)           AS bor,
                COUNT(DISTINCT d.topic_code) -
                COUNT(DISTINCT l.topic_code)           AS yoq
            FROM dts_tree d
            LEFT JOIN teacher_lessons l ON l.topic_code = d.topic_code
            WHERE d.is_deleted = FALSE
            GROUP BY d.grade, d.subject_name
            ORDER BY
                CASE WHEN d.grade ~ '^[0-9]+$' THEN d.grade::int ELSE 99 END,
                d.subject_name
        """)
        rows2 = cur2.fetchall()
        cur2.close(); conn2.close()

        if not rows2:
            await message.answer("📭 DTS daraxt bo'sh.")
            return

        lines = ["📖 Darslar holati\n"]
        cur_grade = None
        t_jami = t_bor = t_yoq = 0
        for grade, subj, jami, bor, yoq in rows2:
            if grade != cur_grade:
                cur_grade = grade
                lines.append(f"\n🎓 {grade}-sinf:")
            pct = round(bor*100/jami) if jami else 0
            bar = "🟩" if pct==100 else ("🟨" if pct>=50 else "🟥")
            lines.append(
                f"  {bar} {subj}\n"
                f"     ✅ {bor} dars bor  |  ❌ {yoq} yo'q  |  📚 {jami} jami ({pct}%)"
            )
            t_jami += jami; t_bor += bor; t_yoq += yoq

        pct_t = round(t_bor*100/t_jami) if t_jami else 0
        lines.append(
            f"\n━━━━━━━━━━━━━━\n"
            f"📊 Jami: ✅ {t_bor} | ❌ {t_yoq} | 📚 {t_jami} ({pct_t}%)"
        )
        await message.answer("\n".join(lines))
        return

    elif message.text == "📚 Mavzular statistikasi":
        if user_id not in ADMINS:
            return
        conn2 = _get_db_conn()
        cur2  = conn2.cursor()
        cur2.execute("""
            SELECT
                t.grade,
                t.subject_name,
                COUNT(DISTINCT t.topic_code)                              AS mavzu_soni,
                COUNT(g.id)                                               AS test_soni,
                COUNT(DISTINCT CASE WHEN g.id IS NULL THEN t.topic_code END) AS bosh_mavzu
            FROM dts_tree t
            LEFT JOIN generated_tests g ON g.topic_code = t.topic_code
            WHERE t.is_deleted = FALSE
            GROUP BY t.grade, t.subject_name
            ORDER BY t.grade, t.subject_name
        """)
        rows = cur2.fetchall()
        cur2.close(); conn2.close()

        if not rows:
            await message.answer("📭 Hozircha mavzu ma'lumoti yo'q.")
            return

        # Sinf bo'yicha guruhlash
        from collections import defaultdict
        by_grade = defaultdict(list)
        for grade, subj, mavzu, test, bosh in rows:
            by_grade[grade].append((subj, mavzu, test, bosh))

        lines = ["📊 Mavzular statistikasi\n"]
        total_m = total_t = total_b = 0
        for grade in sorted(by_grade.keys(), key=lambda x: int(x) if str(x).isdigit() else 99):
            lines.append(f"\n🎓 {grade}-sinf:")
            for subj, mavzu, test, bosh in by_grade[grade]:
                avg = round(test/mavzu, 1) if mavzu else 0
                bar = "🟩" if avg >= 5 else ("🟨" if avg >= 2 else "🟥")
                lines.append(
                    f"  {bar} {subj}\n"
                    f"     📚 {mavzu} mavzu | 🧪 {test} test | ⚠️ {bosh} bo'sh"
                )
                total_m += mavzu; total_t += test; total_b += bosh

        lines.append(
            f"\n━━━━━━━━━━━━━━\n"
            f"📊 Jami: {total_m} mavzu | {total_t} test | {total_b} bo'sh mavzu"
        )
        await message.answer("\n".join(lines))
        return

    elif message.text and message.text.endswith("-sinf"):

        grade = message.text.replace("-sinf", "")

        topic_stats_state[user_id] = {
            "grade": grade,
            "level": "subjects"
        }

        subjects = get_subjects_by_grade(
            grade
        )

        keyboard = []

        for subject in subjects:

            keyboard.append([
                KeyboardButton(text=subject)
            ])

        keyboard.append([
            KeyboardButton(text="🔙 Ortga")
        ])

        await message.answer(
            "📖 Fanni tanlang",
            reply_markup=ReplyKeyboardMarkup(
                keyboard=keyboard,
                resize_keyboard=True
            )
        )

        return

    elif (
        user_id in topic_stats_state
        and f"topic_{message.text}" in topic_stats_state[user_id]
    ):
        topic_stats_state[user_id]["level"] = "topic_stats"
        topic_code = topic_stats_state[user_id][
            f"topic_{message.text}"
        ]

        info = get_topic_name(topic_code)
        stats = get_topic_statistics(topic_code)

        topic_stats_state[user_id]["selected_topic"] = topic_code

        total = stats[0] or 0
        oson = stats[1] or 0
        orta = stats[2] or 0
        qiyin = stats[3] or 0
        murakkab = stats[4] or 0

        await message.answer(
            f"🔑 {topic_code}\n\n"
            f"🎓 Sinf: {info[0]}\n"
            f"📚 Fan: {info[1]}\n"
            f"🗓 Chorak: {info[2]}\n"
            f"📖 Bob: {info[3]}\n"
            f"📂 Bo'lim: {info[4]}\n"
            f"📘 Mavzu: {info[5]}\n"
            f"📌 Kichik mavzu: {info[6]}\n\n"
            f"📊 Jami test: {total}\n"
            f"🟢 Oson: {oson}\n"
            f"🟡 O'rta: {orta}\n"
            f"🟠 Qiyin: {qiyin}\n"
            f"🔴 Murakkab: {murakkab}",
            reply_markup=InlineKeyboardMarkup(inline_keyboard=[
                [
                    InlineKeyboardButton(text="📄 Excel shablon", callback_data=f"ts_excel:{topic_code}"),
                    InlineKeyboardButton(text="📥 Import", callback_data=f"ts_import:{topic_code}"),
                ],
                [
                    InlineKeyboardButton(text="🤖 AI Yaratish", callback_data=f"ts_gen:{topic_code}"),
                    InlineKeyboardButton(text="🗑 O'chirish", callback_data=f"ts_del:{topic_code}"),
                ],
                [
                    InlineKeyboardButton(text="👁 Testlarni ko'rish", callback_data=f"ts_view:{topic_code}:0"),
                ],
            ])
        )

        return

    elif message.text == "📥 Test import qilish":

        admin_state[user_id] = "test_import"

        await message.answer(
            "Excel fayl yuboring"
        )

        return

    elif message.text == "📥 DTS import":

        await dts_import_menu(
            message,
            admin_state,
            user_id
        )

        return

    # ===== 📋 SHABLONLAR — yagona shablon/import markazi =====
    elif message.text == "📋 Shablonlar":
        if user_id not in ADMINS:
            return
        await message.answer(
            "📋 Shablonlar va Import\n\n"
            "Chap — bo'sh shablon olish, o'ng — to'ldirilganni import qilish:",
            reply_markup=ReplyKeyboardMarkup(
                keyboard=[
                    [KeyboardButton(text="📋 Topik shablon"), KeyboardButton(text="📥 Topik import")],
                    [KeyboardButton(text="🧪 Test shablon"),  KeyboardButton(text="📥 Test import")],
                    [KeyboardButton(text="📚 Dars shablon"),  KeyboardButton(text="📥 Dars import")],
                    [KeyboardButton(text="🔙 Admin menyu")],
                ],
                resize_keyboard=True
            )
        )
        return

    elif message.text == "📋 Topik shablon":
        if user_id not in ADMINS:
            return
        from shablon_yaratish import shablon_state
        shablon_state[user_id] = {"step": "sinf_fan"}
        await message.answer(
            "📋 Topik kod uchun shablon\n\n"
            "Sinf va fanni yozing:\nMasalan: 1 Ingliz tili"
        )
        return

    elif message.text == "📥 Topik import":
        if user_id not in ADMINS:
            return
        from dts_import_handlers import DTSImportState
        admin_state[user_id] = None
        await state.set_state(DTSImportState.waiting_excel)
        await message.answer(
            "📄 DTS Excel faylini yuboring\n\n"
            "Bekor qilish: /menu",
            reply_markup=InlineKeyboardMarkup(inline_keyboard=[[
                InlineKeyboardButton(text="❌ Bekor qilish", callback_data="cancel_import")
            ]])
        )
        return

    elif message.text == "📥 Test import":
        if user_id not in ADMINS:
            return
        admin_state[user_id] = "test_import"
        await message.answer("📥 Test import\n\nTo'ldirilgan Excel faylni yuboring:")
        return

    elif message.text == "✅ Ha, import qil" and admin_state.get(user_id) == "test_import_confirm":
        if user_id not in ADMINS:
            return
        path = test_import_files.get(user_id)
        admin_state[user_id] = None
        if not path or not os.path.exists(path):
            await message.answer("❌ Fayl topilmadi, qaytadan yuboring.",
                                 reply_markup=get_main_keyboard("Admin"))
            return
        await message.answer("⏳ Import qilinmoqda...", reply_markup=get_main_keyboard("Admin"))
        await import_tests_excel(message, path, user_id)

        # RASM_MALUMOTI varaqidagi kodlarni olish
        try:
            import openpyxl as _oxr
            _wbr = _oxr.load_workbook(path, data_only=True)
            rasm_kodlar = []
            if "RASM_MALUMOTI" in _wbr.sheetnames:
                _wsr = _wbr["RASM_MALUMOTI"]
                for _rr in range(2, _wsr.max_row+1):
                    _id = _wsr.cell(_rr,1).value
                    if _id: rasm_kodlar.append(str(_id))
                _wbr.close()
            if rasm_kodlar:
                # DB ga saqlaymiz (redeploy bo'lsa ham saqlanadi)
                conn_rq=_get_db_conn();cur_rq=conn_rq.cursor()
                cur_rq.execute("""CREATE TABLE IF NOT EXISTS rasm_queue
                    (id SERIAL PRIMARY KEY, user_id BIGINT, kod TEXT, done BOOLEAN DEFAULT FALSE, created_at TIMESTAMP DEFAULT NOW())""")
                cur_rq.execute("DELETE FROM rasm_queue WHERE user_id=%s",(user_id,))
                for kk in rasm_kodlar:
                    cur_rq.execute("INSERT INTO rasm_queue(user_id,kod) VALUES(%s,%s)",(user_id,kk))
                conn_rq.commit();cur_rq.close();conn_rq.close()
                await message.answer(
                    f"🖼 Rasmlar kutilmoqda!\n\n"
                    f"📊 Jami: {len(rasm_kodlar)} ta rasm kodi saqlandi\n\n"
                    f"Endi collage rasmlarni yuboring:\n"
                    f"Har collage = 30 ta rasm (5 qator × 6 ustun)\n\n"
                    f"1-collage yuboring 👇"
                )
        except Exception as _er:
            pass

        try:
            os.remove(path)
        except Exception:
            pass
        test_import_files.pop(user_id, None)
        return

    elif message.text == "❌ Bekor" and admin_state.get(user_id) == "test_import_confirm":
        admin_state[user_id] = None
        path = test_import_files.pop(user_id, None)
        if path and os.path.exists(path):
            try:
                os.remove(path)
            except Exception:
                pass
        await message.answer("❌ Import bekor qilindi.",
                             reply_markup=get_main_keyboard("Admin"))
        return

    elif message.text == "📚 Dars shablon":
        if user_id not in ADMINS:
            return
        await lesson_admin.la_show_grades(message)
        return

    elif message.text == "📥 Dars import":
        if user_id not in ADMINS:
            return
        await state.set_state(lesson_admin.LessonAdminState.waiting_excel)
        await message.answer("📥 Dars import\n\nTo'ldirilgan Excel faylini yuboring:")
        return

    elif message.text == "🔙 Admin menyu":
        if user_id not in ADMINS:
            return
        await message.answer("⚙️ Admin menyusi", reply_markup=get_main_keyboard("Admin"))
        return
    # ===== Shablonlar markazi tugadi =====

    elif (
        user_id in topic_stats_state
        and "grade" in topic_stats_state[user_id]
        and "topics" not in topic_stats_state[user_id]
    ):

        grade = topic_stats_state[user_id]["grade"]

        subject_name = message.text

        topics = get_topics_by_grade_subject(
            grade,
            subject_name
        )

        topic_stats_state[user_id]["subject"] = subject_name
        topic_stats_state[user_id]["topics"] = topics
        topic_stats_state[user_id]["page"] = 0
        topic_stats_state[user_id]["level"] = "topics"

        await show_topics_page(
            message,
            user_id
        )

        return

    elif message.text == "⬅️ Oldingi":

        if user_id not in topic_stats_state:
            return

        if topic_stats_state[user_id]["page"] > 0:

            topic_stats_state[user_id]["page"] -= 1

        await show_topics_page(
            message,
            user_id
        )

        return

    elif message.text == "➡️ Keyingi":

        if user_id not in topic_stats_state:
            return

        total = len(
            topic_stats_state[user_id]["topics"]
        )

        current_page = topic_stats_state[user_id]["page"]

        max_page = (total - 1) // 10

        if current_page < max_page:

            topic_stats_state[user_id]["page"] += 1

        await show_topics_page(
            message,
            user_id
        )

        return

    elif message.text == "🔙 Ortga":

        if user_id not in topic_stats_state:
            return

        level = topic_stats_state[user_id].get("level")

        if level == "topic_stats":

            topic_stats_state[user_id]["level"] = "topics"

            await show_topics_page(
                message,
                user_id
            )

            return

        elif level == "topics":

            grade = topic_stats_state[user_id]["grade"]

            keyboards = []

            subjects = get_subjects_by_grade(
                grade
            )

            for subject in subjects:
                keyboards.append([
                    KeyboardButton(
                        text=subject
                    )
                ])

            keyboards.append([
                KeyboardButton(
                    text="🔙 Ortga"
                )
            ])

            topic_stats_state[user_id]["level"] = "subjects"

            await message.answer(
                "📖 Fanni tanlang",
                reply_markup=ReplyKeyboardMarkup(
                    keyboard=keyboards,
                    resize_keyboard=True
                )
            )

            return

        elif level == "subjects":

            grade = topic_stats_state[user_id]["grade"]

            subjects = get_subjects_by_grade(
                grade
            )

            keyboard = []

            for subject in subjects:

                keyboard.append([
                    KeyboardButton(
                        text=subject
                    )
                ])

            keyboard.append([
                KeyboardButton(
                    text="🔙 Ortga"
                )
            ])

            await message.answer(
                "📖 Fanni tanlang",
                reply_markup=ReplyKeyboardMarkup(
                    keyboard=keyboard,
                    resize_keyboard=True
                )
            )

            return

    elif message.text and "-" in message.text:

        topic_code = message.text.strip()

        info = get_topic_name(
            topic_code
        )

        if not info:
            return

        stats = get_topic_statistics(
            topic_code
        )

        mavzu = info[0]
        kichik = info[1]

        await message.answer(

            f"🔑 {topic_code}\n\n"

            f"📖 Mavzu:\n"
            f"{mavzu}\n\n"

            f"📌 Kichik mavzu:\n"
            f"{kichik}\n\n"

            f"📊 Jami test: {stats[0]}\n\n"

            f"🟢 Oson: {stats[1] or 0}\n"
            f"🟡 O'rta: {stats[2] or 0}\n"
            f"🟠 Qiyin: {stats[3] or 0}\n"
            f"🔴 Murakkab: {stats[4] or 0}\n\n"

            f"📝 Single choice: {stats[5] or 0}\n"
            f"☑️ Multiple choice: {stats[6] or 0}\n"
            f"✅ True/False: {stats[7] or 0}\n"
            f"✍️ Write answer: {stats[8] or 0}"
        )

        return

    elif message.text == "📄 Excel shablon":
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl import Workbook
        import io

        # Tanlangan mavzu kodini olish
        topic_code = ""
        grade = ""
        if user_id in topic_stats_state:
            topic_code = topic_stats_state[user_id].get("selected_topic", "")
            grade = str(topic_stats_state[user_id].get("grade", ""))

        # Sinf bo'yicha age_group
        age_map = {
            "1": "6-7", "2": "7-8", "3": "8-9", "4": "9-10",
            "5": "10-11", "6": "11-12", "7": "12-13", "8": "13-14",
            "9": "14-15", "10": "15-16", "11": "16-17"
        }
        age_group = age_map.get(grade, "10-11")

        wb = Workbook()
        ws = wb.active
        ws.title = "TESTLAR"

        headers = [
            "topic_code", "difficulty", "situation", "question",
            "option_a", "option_b", "option_c", "option_d",
            "correct_answer", "explanation", "question_type",
            "is_latex", "image_url", "audio_text", "language",
            "life_level", "age_group", "time_limit"
        ]

        # Header qatori
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="2E86AB")
            cell.alignment = Alignment(horizontal="center")

        # 40 ta qator — har biri uchun difficulty va nom
        def get_difficulty(n):
            if n <= 10: return "oson"
            elif n <= 20: return "o'rta"
            elif n <= 30: return "qiyin"
            else: return "murakkab"

        for i in range(1, 41):
            diff = get_difficulty(i)
            row = [
                topic_code,              # topic_code — o'zi
                diff,                    # difficulty
                "oddiy",                 # situation
                "",                      # question
                "",                      # option_a
                "",                      # option_b
                "",                      # option_c
                "",                      # option_d
                "",                      # correct_answer
                "",                      # explanation
                "single_choice",         # question_type
                False,                   # is_latex
                f"{topic_code}-{i}",     # image_url — kod-1, kod-2...
                "",                      # audio_text
                "uz",                    # language
                1,                       # life_level
                age_group,               # age_group
                60,                      # time_limit
            ]
            ws.append(row)
            # Difficulty rangini ko'rsatish
            colors = {"oson": "C8F7C5", "o'rta": "FFF3CD", "qiyin": "FFD7C4", "murakkab": "F5C6CB"}
            ws.cell(row=i+1, column=2).fill = PatternFill("solid", fgColor=colors.get(diff, "FFFFFF"))

        # Ustun kengliklari
        ws.column_dimensions['A'].width = 35
        ws.column_dimensions['B'].width = 12
        ws.column_dimensions['C'].width = 10
        ws.column_dimensions['D'].width = 50
        for col in ['E','F','G','H']:
            ws.column_dimensions[col].width = 20
        ws.column_dimensions['I'].width = 15
        ws.column_dimensions['J'].width = 30

        # Ma'lumot varag'i
        ws2 = wb.create_sheet("MA'LUMOT")
        ws2.append(["Maydon", "Qiymatlar"])
        ws2.append(["difficulty", "oson | o'rta | qiyin | murakkab"])
        ws2.append(["situation", "oddiy | hayotiy | laboratoriya"])
        ws2.append(["question_type", "single_choice | write_answer | true_false"])
        ws2.append(["correct_answer", "A | B | C | D  (yoki matn)"])
        ws2.append(["is_latex", "True | False"])
        ws2.append(["language", "uz | ru | en"])
        ws2.append(["age_group", f"{age_group} (avtomatik)"])
        ws2.append(["time_limit", "60 (sekund)"])
        ws2.append(["", ""])
        ws2.append(["Mavzu kodi", topic_code or "—"])
        ws2.append(["Sinf", f"{grade}-sinf"])
        ws2.append(["", ""])
        ws2.append(["1-10 qator", "OSON 🟢"])
        ws2.append(["11-20 qator", "O'RTA 🟡"])
        ws2.append(["21-30 qator", "QIYIN 🟠"])
        ws2.append(["31-40 qator", "MURAKKAB 🔴"])

        # Xotiraga saqlash
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        fname = f"shablon_{topic_code or 'mavzu'}.xlsx"
        from aiogram.types import BufferedInputFile
        await message.answer_document(
            BufferedInputFile(buf.read(), filename=fname),
            caption=(
                f"📋 Test shabloni tayyor!\n\n"
                f"🔑 Mavzu kodi: {topic_code or '—'}\n"
                f"🎓 Sinf: {grade or '—'}\n"
                f"👶 Yosh guruhi: {age_group}\n\n"
                f"📝 40 ta qator:\n"
                f"🟢 1-10: Oson\n"
                f"🟡 11-20: O'rta\n"
                f"🟠 21-30: Qiyin\n"
                f"🔴 31-40: Murakkab\n\n"
                f"Faqat bo'sh ustunlarni to'ldiring!"
            )
        )
        return
    
    elif message.text == "▶️ Generatorni boshlash":

        global generator_process

        if generator_process and generator_process.poll() is None:

            await message.answer(
                "Generator ishlayapti"
            )

            return

        generator_process = subprocess.Popen(
            ["python", "test_generator.py"]
        )

        await message.answer(
            "Generator ishga tushdi"
        )

        return

    elif message.text == "⏹ Generatorni to‘xtatish":

        if generator_process and generator_process.poll() is None:

            generator_process.terminate()

            await message.answer(
                "Generator to‘xtatildi"
            )

        else:

            await message.answer(
                "Generator ishlamayapti"
            )

        return
    elif message.text == "📊 Generator statistikasi":

        conn = _get_db_conn()
        cur = conn.cursor()

        cur.execute("""
            SELECT COUNT(*)
            FROM dts_tree
        """)
        total_topics = cur.fetchone()[0]

        cur.execute("""
            SELECT COUNT(DISTINCT topic_code)
            FROM generated_tests
        """)
        completed_topics = cur.fetchone()[0]

        cur.execute("""
            SELECT COUNT(*)
            FROM generated_tests
        """)
        total_tests = cur.fetchone()[0]

        remaining_topics = total_topics - completed_topics

        progress = round(
            completed_topics * 100 / total_topics,
            1
        ) if total_topics else 0

        cur.close()
        conn.close()

        await message.answer(
            f"📚 Jami mavzular: {total_topics}\n"
            f"✅ Test yaratilgan mavzular: {completed_topics}\n"
            f"❌ Qolgan mavzular: {remaining_topics}\n"
            f"📝 Jami testlar: {total_tests}\n"
            f"📈 Progress: {progress}%"
        )

    elif message.text == "⬅ Ortga":

        await admin_main_menu(
            message
        )

        return

    elif message.text == "📊 DTS statistika":

        await dts_statistics(
            message
        )

        return

    elif message.text == "📤 DTS export":

        await dts_export_menu(
            message
        )

        return

    elif message.text == "📤 Hammasini export":

        await dts_export_all(
            message
        )

        return
        
    # parallel message bloklash
    
    if user_id not in user_locks:
        user_locks[user_id] = asyncio.Lock()
        
    async with user_locks[user_id]:

      #  action = TEXT_TO_ID.get(message.text)

        # 🔙 ORTGA
        if message.text == BACK:

            user_id = message.from_user.id

            # history bo‘lsa
            if user_id in state_history and len(state_history[user_id]) > 1:

                # hozirgi state ni olib tashlash
                state_history[user_id].pop()

                # oldingi state
                prev_state = state_history[user_id][-1]

                user_state[user_id] = prev_state

                # CLASS ga qaytish
                if prev_state == "class":

                    await message.answer(
                        "Sinf tanlang:",
                        reply_markup=base_keyboard(CLASSES)
                    )
                    return

                # SUBJECT ga qaytish
                elif prev_state == "subject":

                    selected_class = temp_user[user_id]["class"]

                    subjects = SUBJECTS_BY_CLASS.get(selected_class)

                    flat_subjects = []

                    for row in subjects:
                        flat_subjects.extend(row)

                    await message.answer(
                        "Fan tanlang:",
                        reply_markup=base_keyboard(flat_subjects)
                    )
                    return

                elif prev_state == "db_school":

                    await message.answer(
                        "🏫 Maktab turini tanlang:",
                        reply_markup=base_keyboard([
                            "all",
                            "🏫 Oddiy",
                            "⭐ IDUM",
                            "🏆 Prezident",
                            "🏢 Xususiy"
                        ])
                    )
                    return

        # ===== SURVEY RESULTS =====
        elif message.text == "📋 So‘rovnoma natijalari":

            if message.from_user.id not in ADMINS:
                return

            conn = _get_db_conn()
            cur = conn.cursor()

            cur.execute("""
            SELECT surveys.question,
            survey_answers.answer,
            COUNT(*)
            FROM survey_answers
            JOIN surveys
            ON surveys.id = survey_answers.survey_id
            GROUP BY surveys.question, survey_answers.answer
            """)

            rows = cur.fetchall()

            conn.close()

            if not rows:

                await message.answer(
                    "❌ Natijalar yo‘q"
                )
                return

            text = "📋 So‘rovnoma natijalari\n\n"

            current_question = ""

            for question, answer, count in rows:

                if question != current_question:

                    current_question = question

                    text += f"\n📝 {question}\n"

                text += f"• {answer} — {count} ta\n"

            await message.answer(text)

            return

        elif user_state.get(message.from_user.id) == "admin_region":

            temp_user[message.from_user.id]["admin_region"] = message.text

            districts = REGIONS.get(message.text, [])

            flat = []

            for row in districts:
                flat.extend(row)

            user_state[message.from_user.id] = "admin_district"

            await message.answer(
                "Tuman tanlang:",
                reply_markup=base_keyboard(flat)
            )

            return

        elif user_state.get(message.from_user.id) == "admin_district":

            temp_user[message.from_user.id]["admin_district"] = message.text

            conn = _get_db_conn()
            cur = conn.cursor()

            cur.execute("""
            SELECT DISTINCT school
            FROM users
            WHERE district=%s
            """, (message.text,))

            rows = cur.fetchall()

            conn.close()

            schools = [r[0] for r in rows if r[0]]

            user_state[message.from_user.id] = "admin_school"

            await message.answer(
                "Maktab tanlang:",
                reply_markup=base_keyboard(schools)
            )

            return

        elif user_state.get(message.from_user.id) == "admin_school":

            school = message.text

            conn = _get_db_conn()
            cur = conn.cursor()

            cur.execute("""
            SELECT AVG(score * 100.0 / total)
            FROM results
            WHERE school=%s
            """, (school,))

            avg = cur.fetchone()[0]

            conn.close()

            if avg is None:

                await message.answer(
                    "❌ Ma’lumot yo‘q"
                )
                return

            avg = round(avg, 1)

            bar = "█" * int(avg // 10)
            empty = "░" * (10 - int(avg // 10))

            text = (
                f"🏫 {school}\n\n"
                f"{bar}{empty}\n"
                f"📊 O‘rtacha: {avg}%"
            )

            await message.answer(text)

            user_state[message.from_user.id] = None

            return

        elif user_state.get(message.from_user.id) == "teacher_level":

            temp_user[message.from_user.id]["teacher_level"] = message.text

            subjects = SUBJECTS_BY_LEVEL.get(message.text, [])

            user_state[message.from_user.id] = "teacher_subject"

            await message.answer(
                "Fan tanlang:",
                reply_markup=base_keyboard(subjects)
            )

            return

        elif user_state.get(message.from_user.id) == "teacher_subject":

            temp_user[message.from_user.id]["subject"] = message.text

            user_state[message.from_user.id] = "test_type"

            await message.answer(
                "Test turini tanlang:",
                reply_markup=base_keyboard(TEST_TYPES)
            )

            return

        # 🏠 HOME
        elif message.text in (HOME, HOME2):

            conn = _get_db_conn()
            cur = conn.cursor()
            cur.execute("SELECT role FROM users WHERE user_id=%s", (message.from_user.id,))
            user = cur.fetchone()
            conn.close()

            role = user[0] if user else None
            user_state[message.from_user.id] = None

            # Oxirgi 10 xabarni o'chirish
            try:
                for i in range(message.message_id - 1, message.message_id - 15, -1):
                    try:
                        await message.bot.delete_message(message.chat.id, i)
                    except Exception:
                        pass
            except Exception:
                pass

            from student_dashboard import build_dashboard
            try:
                text, kb = await build_dashboard(message.from_user.id)
                await message.answer(text, reply_markup=kb)
            except Exception:
                pass
            await message.answer("👇 Menyu:", reply_markup=get_main_keyboard(role))
            return

        elif message.text == "⚙️ Akkaunt sozlamalari":

            await message.answer(
                "Sozlamalar:",
                reply_markup=ReplyKeyboardMarkup(
                    keyboard=[
                        [KeyboardButton(text="🔄 Rolni almashtirish")],
                        [KeyboardButton(text="🌍 Hududni almashtirish")],
                        [KeyboardButton(text="🏫 Maktabni almashtirish")],
                        [KeyboardButton(text="🎓 Sinfni almashtirish")],
                        [KeyboardButton(text=BACK)]
                    ],
                    resize_keyboard=True
                )
            )
        elif message.text == "🔄 Rolni almashtirish":

            user_state[message.from_user.id] = "change_role"

            await message.answer(
                "Yangi rolni tanlang:",
                reply_markup=make_keyboard(["🧒 O‘quvchi", "👨‍🏫 O‘qituvchi"])
            )

        elif user_state.get(message.from_user.id) == "change_role":

            conn = _get_db_conn()
            cur = conn.cursor()

            cur.execute("""
            UPDATE users
            SET role=%s
            WHERE user_id=%s
            """, (
                message.text,
                message.from_user.id
            ))

            conn.commit()
            conn.close()

            user_state[message.from_user.id] = None

            await message.answer(
                f"✅ Rol o‘zgartirildi: {message.text}",
                reply_markup=get_main_keyboard(message.text)
            )

        elif message.text == "🌍 Hududni almashtirish":

            user_state[message.from_user.id] = "change_region"

            await message.answer(
                "Viloyatni tanlang:",
                reply_markup=make_keyboard(REGIONS.keys())
            )

            return

        elif user_state.get(message.from_user.id) == "change_region":

            temp_user[message.from_user.id] = {
                "new_region": message.text
            }

            districts = REGIONS.get(message.text, [])

            flat = []

            for row in districts:
                flat.extend(row)

            user_state[message.from_user.id] = "change_district"

            await message.answer(
                "Tumanni tanlang:",
                reply_markup=base_keyboard(flat)
            )

            return

        elif user_state.get(message.from_user.id) == "change_district":

            conn = _get_db_conn()
            cur = conn.cursor()

            cur.execute("""
            UPDATE users
            SET region=%s, district=%s
            WHERE user_id=%s
            """, (
                temp_user[message.from_user.id]["new_region"],
                message.text,
                message.from_user.id
            ))

            conn.commit()

            cur.execute(
                "SELECT role FROM users WHERE user_id=%s",
                (message.from_user.id,)
            )

            role = cur.fetchone()[0]

            conn.close()

            user_state[message.from_user.id] = None

            await message.answer(
                "✅ Hudud o‘zgartirildi",
                reply_markup=get_main_keyboard(role)
            )

            return

# ===== MAKTABNI ALMASHTIRISH =====

        elif message.text == "🏫 Maktabni almashtirish":

            user_state[user_id] = "change_school_type"

            await message.answer(
                "Maktab turini tanlang:",
                reply_markup=make_keyboard(SCHOOL_TYPES)
            )

            return

        elif user_state.get(user_id) == "change_school_type":

            temp_user[user_id]["new_school_type"] = message.text

            user_state[user_id] = "change_school"

            await message.answer(
                "Maktab raqami kiriting:"
            )

            return

        elif user_state.get(user_id) == "change_school":

            temp_user[user_id]["new_school"] = (
                f"{temp_user[user_id]['new_school_type']} - {message.text}"
            )

            school_type = temp_user[user_id]["new_school_type"]

            if school_type == "🏫 Oddiy davlat maktabi":
                classes = [c for c in CLASSES if "🏫 Oddiy" in c]

            elif school_type == "⭐ Ixtisoslashgan (IDUM)":
                classes = [c for c in CLASSES if "⭐ IDUM" in c]

            elif school_type == "🏆 Prezident maktabi":
                classes = [c for c in CLASSES if "🏆 Prezident" in c]

            else:
                classes = [c for c in CLASSES if "🏢 Xususiy" in c]

            user_state[user_id] = "change_school_class"

            await message.answer(
                "Sinfni tanlang:",
                reply_markup=base_keyboard(classes)
            )

            return

        elif user_state.get(user_id) == "change_school_class":

            conn = _get_db_conn()
            cur = conn.cursor()

            cur.execute("""
            UPDATE users
            SET school=%s, class=%s
            WHERE user_id=%s
            """, (
                temp_user[user_id]["new_school"],
                message.text,
                user_id
            ))

            conn.commit()

            cur.execute("""
            SELECT role
            FROM users
            WHERE user_id=%s
            """, (user_id,))

            row = cur.fetchone()

            conn.close()

            role = row[0] if row else "🧒 O‘quvchi"

            user_state[user_id] = None

            await message.answer(
                "✅ Maktab va sinf o‘zgartirildi",
                reply_markup=get_main_keyboard(role)
            )

            return

        # ===== SINFNI ALMASHTIRISH =====
        elif message.text == "🎓 Sinfni almashtirish":

            conn = _get_db_conn()
            cur = conn.cursor()

            cur.execute("""
            SELECT school
            FROM users
            WHERE user_id=%s
            """, (message.from_user.id,))

            row = cur.fetchone()

            conn.close()

            school = row[0] if row else ""

            if "🏫 Oddiy" in school:
                classes = [c for c in CLASSES if "🏫 Oddiy" in c]

            elif "⭐ IDUM" in school:
                classes = [c for c in CLASSES if "⭐ IDUM" in c]

            elif "🏆 Prezident" in school:
                classes = [c for c in CLASSES if "🏆 Prezident" in c]

            else:
                classes = [c for c in CLASSES if "🏢 Xususiy" in c]

            user_state[message.from_user.id] = "change_class"

            await message.answer(
                "Yangi sinfni tanlang:",
                reply_markup=base_keyboard(classes)
            )

            return

        elif user_state.get(message.from_user.id) == "change_class":

            conn = _get_db_conn()
            cur = conn.cursor()

            cur.execute("""
            UPDATE users
            SET class=%s
            WHERE user_id=%s
            """, (
                message.text,
                message.from_user.id
            ))

            conn.commit()

            cur.execute("""
            SELECT role FROM users
            WHERE user_id=%s
            """, (message.from_user.id,))

            row = cur.fetchone()

            conn.close()

            role = row[0] if row else "🧒 O‘quvchi"

            user_state[message.from_user.id] = None

            await message.answer(
                f"✅ Sinf o‘zgartirildi: {message.text}",
                reply_markup=get_main_keyboard(role)
            )

            return
        
        elif user_state.get(message.from_user.id) == "survey_work":

            data = user_test[message.from_user.id]

            data["answers"][data["index"]] = message.text.upper()

            # oxiri
            if data["index"] >= len(data["surveys"]) - 1:

                conn = _get_db_conn()
                cur = conn.cursor()

                cur.execute("""
                UPDATE users
                SET survey_done=1
                WHERE user_id=%s
                """, (message.from_user.id,))

                conn.commit()
                conn.close()

                user_state[message.from_user.id] = None

                await message.answer(
                    "✅ So‘rovnoma tugadi",
                    reply_markup=get_main_keyboard(
                        temp_user[message.from_user.id]["role"]
                    )
                )

                return

            data["index"] += 1

            q = data["surveys"][data["index"]]

            text = (
                f"{data['index']+1}/{len(data['surveys'])}\n\n"
                f"{q[2]}\n\n"
                f"A) {q[4]}\n"
                f"B) {q[5]}\n"
                f"C) {q[6]}\n"
                f"D) {q[7]}"
            )

            await message.answer(text)

            return

def _mk_ts_kb(st2, cnt_total):
    """ts_start settings uchun ✅ li keyboard."""
    def c(cond): return "✅ " if cond else ""
    cnt   = st2.get("ts_count", 20)
    diff  = st2.get("ts_diff", "all")
    timed = st2.get("ts_timed", "mix")   # True/False/"mix"
    write = st2.get("ts_write", False)   # True/False/"mix"
    img   = st2.get("ts_img", "mix")     # True/False/"mix"

    # Son tugmalari — mavjuddan oshmasin, maksimal 50 ta
    MAX_SAVOL = 50
    barcha = min(cnt_total, MAX_SAVOL)
    son_qatori = []
    for n in (10, 20, 40):
        if n < barcha:
            son_qatori.append(InlineKeyboardButton(
                text=f"{c(cnt==n)}{n} ta", callback_data=f"ts_cnt_{n}"))
    _yorliq = f"Barchasi ({barcha})" if cnt_total <= MAX_SAVOL else f"Maks {barcha} ta"
    son_qatori.append(InlineKeyboardButton(
        text=f"{c(cnt==barcha)}{_yorliq}",
        callback_data=f"ts_cnt_{barcha}"))

    return InlineKeyboardMarkup(inline_keyboard=[
        son_qatori,
        [InlineKeyboardButton(text=f"{c(diff=='oson')}🟢 Oson",   callback_data="ts_dif_oson"),
         InlineKeyboardButton(text=f"{c(diff=='orta')}🟡 O'rta",  callback_data="ts_dif_orta"),
         InlineKeyboardButton(text=f"{c(diff=='qiyin')}🔴 Qiyin", callback_data="ts_dif_qiyin"),
         InlineKeyboardButton(text=f"{c(diff=='all')}🌈 Aralash",  callback_data="ts_dif_all")],
        [InlineKeyboardButton(text=f"{c(timed==True)}⏱ Vaqtli",    callback_data="ts_time_1"),
         InlineKeyboardButton(text=f"{c(timed==False)}∞ Vaqtsiz",  callback_data="ts_time_0"),
         InlineKeyboardButton(text=f"{c(timed=='mix')}⚡ Aralash",  callback_data="ts_time_mix")],
        [InlineKeyboardButton(text=f"{c(write==False)}🔘 Tugmali",   callback_data="ts_wr_0"),
         InlineKeyboardButton(text=f"{c(write==True)}✍️ Yozuvli",   callback_data="ts_wr_1"),
         InlineKeyboardButton(text=f"{c(write=='mix')}🔀 Aralash",   callback_data="ts_wr_mix")],
        [InlineKeyboardButton(text=f"{c(img==True)}🖼 Rasmli",       callback_data="ts_img_1"),
         InlineKeyboardButton(text=f"{c(img==False)}📝 Rasmsiz",     callback_data="ts_img_0"),
         InlineKeyboardButton(text=f"{c(img=='mix')}🔀 Aralash",     callback_data="ts_img_mix")],
        [InlineKeyboardButton(text="▶️ Boshlash", callback_data="ts_go")],
    ])

@dp.callback_query()
async def test_buttons(call: CallbackQuery, state: FSMContext):
    user_id = _eff_uid(call.from_user.id)
    # BIRINCHI call.answer() — Telegram "yuklanyapti" ni darhol to'xtatadi
    try:
        await call.answer()
    except Exception:
        pass
    try:
        await _test_buttons_inner(call, state, user_id)
    except Exception as _e:
        await _error_and_home(call, user_id, _e, "Xatolik")

async def _test_buttons_inner(call: CallbackQuery, state: FSMContext, user_id: int):

    # ═══ DTS NAVIGATOR ═══

    # ═══ BARCHA LA_ CALLBACKLAR (lesson_admin) ═══
    # ── DELEGATE ──
    try:
        if call.data.startswith("tg_") or call.data == "tg_back":
            from cb_togarak import handle_tg
            if await handle_tg(call,user_id,admin_state,user_state,temp_user,bot): return
        if call.data.startswith("stg_"):
            from cb_student_tg import handle_stg
            if await handle_stg(call,user_id,admin_state,user_state,temp_user,bot): return
        if (call.data.startswith("kb_") or call.data.startswith("parent_") or
            call.data.startswith("nacc_") or
            call.data.startswith("lesson_") or call.data.startswith("tset_") or
            call.data.startswith("reg_") or call.data.startswith("reg:") or
            call.data.startswith("rq_") or call.data.startswith("speak_") or
            call.data.startswith("restart_lesson:") or call.data.startswith("resume_lesson:") or
            call.data in ("ai_stop","cancel_import","noop_timer","speak_all","speak_question",
                          "test_settings","test_skip","test_next_from_result",
                          "lesson_exit","lesson_finish_confirm","lesson_help","lesson_help_close",
                          "lesson_help_next","lesson_help_prev","lesson_next","lesson_prev","lesson_speak")):
            from cb_kabinet import handle_kb
            if await handle_kb(call,user_id,admin_state,user_state,temp_user,bot): return
        if (call.data.startswith("kitob_") or call.data.startswith("stnav_") or
            call.data.startswith("sin_fan:") or call.data.startswith("sin_gr:") or
            call.data.startswith("sin_mavzu:") or
            call.data in ("mustah_back","stnav_back_grade","kitob_qolda","kitob_upload")):
            from cb_kitob import handle_kitob
            if await handle_kitob(call,user_id,admin_state,user_state,temp_user,bot): return
        if (call.data.startswith("sin_") or
            call.data.startswith("menu_bilim_") or call.data.startswith("menu_kitob_") or
            call.data.startswith("menu_ai_") or call.data.startswith("mtt_") or
            call.data.startswith("mustah_") or call.data.startswith("sinash_") or
            call.data.startswith("rasm_") or call.data.startswith("ai_") or
            call.data.startswith("xl_") or call.data.startswith("rep_") or
            call.data.startswith("err_") or
            call.data in ("ai_rasm_auto","ai_rasm_custom","rasm_back","mtt_back",
                          "sinash_back","menu_ai_train","rep_plan","rep_prog","rep_test",
                          "err_clear","err_read","err_unread")):
            from cb_test_nav import handle_test_nav
            if await handle_test_nav(call,user_id,admin_state,user_state,temp_user,bot): return
    except Exception as _de:
        print(f"delegate xato: {_de}")
        import traceback; traceback.print_exc()
        try: await call.answer("⚠️ Xatolik yuz berdi, qayta urinib ko'ring", show_alert=True)
        except: pass
        return
    if call.data.startswith("la_") or call.data in (
        "la_sel_all","la_dl_sel","la_tmpl","la_imp",
        "la_back_grades","la_home"
    ):
        import lesson_admin as _la
        d = call.data
        if d.startswith("la_gs|p|"):      await _la.la_grades_page(call); return
        if d.startswith("la_g|"):         await _la.la_grade(call); return
        if d.startswith("la_s|"):         await _la.la_subject(call); return
        if d.startswith("la_q|"):         await _la.la_quarter(call); return
        if d.startswith("la_b|"):         await _la.la_bob(call); return
        if d.startswith("la_bl|"):        await _la.la_bolim(call); return
        if d.startswith("la_m|"):         await _la.la_mavzu(call); return
        if d.startswith("la_sel|"):       await _la.la_toggle_select(call); return
        if d.startswith("la_tp|"):        await _la.la_topics_page(call); return
        if d == "la_sel_all":             await _la.la_select_all(call); return
        if d == "la_dl_sel":              await _la.la_download_selected(call); return
        if d == "la_tmpl":                await _la.la_template(call); return
        if d == "la_imp":                 await _la.la_import_prompt(call); return
        if d == "la_back_grades":         await _la.la_back_grades(call); return
        if d == "la_home":                await _la.la_home(call); return
        if d.startswith("la_ld|"):        await _la.la_lesson_detail(call); return
        if d.startswith("la_prev|"):      await _la.la_preview_lesson(call); return
        if d.startswith("la_edit|"):      await _la.la_edit_lesson(call); return
        if d.startswith("la_delc|"):      await _la.la_delete_confirm(call); return
        if d.startswith("la_dely|"):      await _la.la_delete_yes(call); return
        return

    # ════════════════════════════

    # ═══ BARCHA DTS_ CALLBACKLAR (to'liq) ═══
    if call.data.startswith("dts_"):
        import dts_import_handlers as _dts
        d = call.data

        # Navigator
        if d == "dts_navigator":      await _dts.dts_navigator(call); return
        if d.startswith("dts_nav_"):
            page = int(d[8:])
            await _dts.dts_navigator(call, page=page); return
        if d.startswith("dts_del_topic:"):
            tc = d[14:]
            conn2=_get_db_conn();cur2=conn2.cursor()
            cur2.execute("UPDATE dts_tree SET is_deleted=TRUE WHERE topic_code=%s",(tc,))
            # Agar sinf bo'sh qolsa — sinf ham "yo'qoladi" (barcha topiklar o'chirilgan)
            conn2.commit();cur2.close();conn2.close()
            await call.answer("✅ Topik o'chirildi",show_alert=True)
            # Sahifani yangilash
            await _dts.dts_navigator(call)
            return
        if d == "dts_menu":           await _dts.dts_menu(call); return
        if d == "dts_search":         await _dts.dts_search(call); return
        if d == "dts_fast_search":    await _dts.dts_fast_search(call); return
        if d == "dts_adv_search":     await _dts.dts_adv_search(call); return
        if d == "dts_import":         await _dts.dts_import(call); return
        if d == "dts_confirm_import": await _dts.dts_confirm_import(call); return
        if d == "dts_problems":       await _dts.dts_problems(call); return
        if d == "dts_download_errors":await _dts.dts_download_errors(call); return
        if d == "dts_cancel_import":  await _dts.dts_cancel_import(call); return
        if d == "dts_export":
            try: await _dts.dts_export(call)
            except: pass
            return

        if d.startswith("dts_grade_"):   await _dts.dts_grade(call); return
        if d.startswith("dts_subject_"): await _dts.dts_subject(call); return
        if d.startswith("dts_quarter_"): await _dts.dts_quarter(call); return
        if d.startswith("dts_bob_"):     await _dts.dts_bob(call); return
        if d.startswith("dts_bolim_"):   await _dts.dts_bolim(call); return
        if d.startswith("dts_mavzu_"):   await _dts.dts_mavzu(call); return
        if d.startswith("dts_small_"):   await _dts.dts_small(call); return
        if d.startswith("dts_adv_"):     await _dts.dts_adv_grade(call); return
        return  # Noma'lum dts_ callback — jimgina o'tkazib yubor
    # ════════════════════════════


    if call.data.startswith("ts_settings:"):
        # ts_start ga yo'naltirish
        tc_set = call.data.split(":")[1]
        await call.answer()
        # ts_start handler ga pass qilamiz
        call.data = f"ts_start:{tc_set}"

    # ═══ MAVZU TANLANDI (aniq topic kodlar bilan) ═══
    if call.data.startswith("ts_sel:"):
        import ts_cache
        sel = ts_cache.ol(call.data[7:])
        if not sel or not sel["topic_codes"]:
            await call.message.answer(
                "⚠️ Bu ro'yxat eskirgan.\n\nMavzularni qaytadan oching."
            )
            return

        topic_codes = sel["topic_codes"]
        mname   = sel["mavzu_name"] or "Mavzu"
        gr_sel  = sel["grade"]
        subj_sel = sel["subject"]

        conn2 = _get_db_conn(); cur2 = conn2.cursor()
        cur2.execute("SELECT COUNT(*) FROM generated_tests WHERE topic_code = ANY(%s)",
                    (topic_codes,))
        cnt = cur2.fetchone()[0]
        cur2.close(); conn2.close()

        print(f"[ts_sel] '{mname[:30]}' sinf={gr_sel} -> "
              f"{len(topic_codes)} topic, {cnt} test | {topic_codes[:3]}")

        if cnt == 0:
            await call.message.answer(f"❌ '{mname}' mavzusi uchun test yo'q!"); return

        from storage import user_state as _us
        if not isinstance(_us.get(user_id), dict): _us[user_id] = {}
        _us[user_id].update({
            "ts_topic": topic_codes[0],
            "ts_topic_codes": topic_codes,
            "ts_mavzu_name": mname,
            "ts_grade": gr_sel, "ts_subject": subj_sel,
            "ts_count": 20, "ts_diff": "all",
            "ts_timed": True, "ts_write": False, "ts_img": "mix",
            "ts_ovoz": False,
            "_ts_cnt_total": cnt
        })
        _sarlavha = f"🧪 {mname[:45]}"
        if gr_sel:   _sarlavha += f"\n🎓 {gr_sel}-sinf"
        if subj_sel: _sarlavha += f" · 📚 {subj_sel[:25]}"
        await call.message.answer(
            f"{_sarlavha}\n📊 Jami: {cnt} ta savol\n\nSozlamalarni tanlang:",
            reply_markup=_mk_ts_kb(_us[user_id], cnt)
        )
        return

    if call.data.startswith("ts_mavzu:"):
        # Format: ts_mavzu:{mavzu_code}|{grade}|{nom_hash}
        _p = call.data[9:].split("|")
        mavzu_code = _p[0]
        gr_sel   = _p[1] if len(_p) > 1 and _p[1] else None
        nom_hash = _p[2] if len(_p) > 2 and _p[2] else None

        # ESKI TUGMA (sinf yoki hash yo'q) — noto'g'ri testlar chiqadi
        if not gr_sel or not nom_hash:
            print(f"[ts_mavzu] ❌ eski tugma: {call.data}")
            await call.message.answer(
                "⚠️ Bu eski ro'yxat.\n\n"
                "Mavzular ro'yxatini qaytadan oching:\n"
                "🧪 Bilimni sinash → sinf → fan"
            )
            return

        conn2 = _get_db_conn(); cur2 = conn2.cursor()
        cur2.execute("""SELECT topic_code, mavzu_name, subject_name
            FROM dts_tree
            WHERE mavzu_code=%s AND grade::TEXT=%s AND is_deleted=FALSE""",
            (mavzu_code, str(gr_sel)))
        qatorlar = cur2.fetchall()

        # Mavzu NOMI bo'yicha ajratamiz — AYNAN bosilgan mavzu
        _mos = [r for r in qatorlar if _mavzu_hash(r[1] or "") == nom_hash]
        if not _mos:
            cur2.close(); conn2.close()
            print(f"[ts_mavzu] ❌ nom_hash={nom_hash} mos kelmadi "
                  f"(kod={mavzu_code} sinf={gr_sel}, {len(qatorlar)} qator)")
            await call.message.answer(
                "⚠️ Mavzu topilmadi.\n\nRo'yxatni qaytadan oching."
            )
            return
        qatorlar = _mos

        topic_codes = sorted({r[0] for r in qatorlar})
        mname    = (qatorlar[0][1] or mavzu_code)
        subj_sel = qatorlar[0][2]

        cur2.execute("SELECT COUNT(*) FROM generated_tests WHERE topic_code = ANY(%s)",
                    (topic_codes,))
        cnt = cur2.fetchone()[0]
        cur2.close(); conn2.close()

        print(f"[ts_mavzu] '{mname[:30]}' sinf={gr_sel} -> "
              f"{len(topic_codes)} topic, {cnt} test | kodlar: {topic_codes[:3]}")

        if cnt == 0:
            await call.message.answer(f"❌ '{mname}' mavzusi uchun test yo'q!"); return
        from storage import user_state as _us
        if not isinstance(_us.get(user_id), dict): _us[user_id] = {}
        _us[user_id].update({
            "ts_topic": topic_codes[0],
            "ts_topic_codes": topic_codes,
            "ts_mavzu_name": mname,
            "ts_grade": gr_sel, "ts_subject": subj_sel,
            "ts_count": 20, "ts_diff": "all",
            "ts_timed": True, "ts_write": False, "ts_img": "mix",
            "ts_ovoz": False,
            "_ts_cnt_total": cnt
        })
        _sarlavha = f"🧪 {mname[:45]}"
        _sarlavha += f"\n🎓 {gr_sel}-sinf"
        if subj_sel: _sarlavha += f" · 📚 {subj_sel[:25]}"
        await call.message.answer(
            f"{_sarlavha}\n📊 Jami: {cnt} ta savol ({len(topic_codes)} ta bo'lim)"
            f"\n\nSozlamalarni tanlang:",
            reply_markup=_mk_ts_kb(_us[user_id], cnt)
        )
        return

    if call.data.startswith("ts_start:"):
        topic_code = call.data.split(":")[1]
        # Test sonini tekshirish
        conn2 = _get_db_conn(); cur2 = conn2.cursor()
        cur2.execute("SELECT COUNT(*) FROM generated_tests WHERE topic_code=%s", (topic_code,))
        cnt = cur2.fetchone()[0]; cur2.close(); conn2.close()
        if cnt == 0:
            await call.message.answer("❌ Bu mavzu uchun test yo'q!")
            return
        # Sozlamalar ekrani
        from storage import user_state as _us
        if not isinstance(_us.get(user_id), dict): _us[user_id] = {}
        _us[user_id].update({"ts_topic": topic_code,
                              "ts_topic_codes": [topic_code],
                              "ts_count": 20,
                              "ts_diff": "all", "ts_timed": True,
                              "ts_write": False, "ts_img": "mix",
                              "ts_ovoz": False,
                              "_ts_cnt_total": cnt})
        await call.message.answer(
            f"🧪 Test: {topic_code}\n📊 Jami: {cnt} ta savol\n\nSozlamalarni tanlang:",
            reply_markup=_mk_ts_kb(_us[user_id], cnt)
        )
        return

    if (call.data.startswith("ts_cnt_") or call.data.startswith("ts_dif_")
            or call.data.startswith("ts_wr_") or call.data.startswith("ts_time_")
            or call.data.startswith("ts_img_") or call.data.startswith("ts_ovoz_")):
        from storage import user_state as _us
        if not isinstance(_us.get(user_id), dict): _us[user_id] = {}
        st2 = _us[user_id]
        if call.data.startswith("ts_cnt_"):   st2["ts_count"] = int(call.data.split("_")[-1])
        elif call.data.startswith("ts_dif_"): st2["ts_diff"]  = call.data.replace("ts_dif_","")
        elif call.data == "ts_wr_1":          st2["ts_write"] = True
        elif call.data == "ts_wr_0":          st2["ts_write"] = False
        elif call.data == "ts_wr_mix":        st2["ts_write"] = "mix"
        elif call.data == "ts_time_1":        st2["ts_timed"] = True
        elif call.data == "ts_time_0":        st2["ts_timed"] = False
        elif call.data == "ts_time_mix":      st2["ts_timed"] = "mix"
        elif call.data == "ts_img_1":         st2["ts_img"]   = True
        elif call.data == "ts_img_0":         st2["ts_img"]   = False
        elif call.data == "ts_img_mix":       st2["ts_img"]   = "mix"
        elif call.data == "ts_ovoz_1":        st2["ts_ovoz"]  = True
        elif call.data == "ts_ovoz_0":        st2["ts_ovoz"]  = False
        await call.answer("✅")
        # Filtrga mos test sonini qayta hisoblaymiz
        cnt_total = st2.get("_ts_cnt_total", 999)
        try:
            tcs = st2.get("ts_topic_codes") or ([st2["ts_topic"]] if st2.get("ts_topic") else [])
            if tcs:
                d2 = st2.get("ts_diff", "all")
                w2 = st2.get("ts_write", False)
                i2 = st2.get("ts_img", "mix")
                f1 = "" if d2 == "all" else f"AND difficulty='{d2}'"
                if w2 == "mix": f2 = ""
                elif w2 is True: f2 = "AND question_type = 'write_answer'"
                else: f2 = "AND question_type != 'write_answer'"
                if i2 == "mix": f3 = ""
                elif i2 is True: f3 = "AND image_file_id IS NOT NULL AND image_file_id <> ''"
                else: f3 = "AND (image_file_id IS NULL OR image_file_id = '')"
                c3 = _get_db_conn(); cr3 = c3.cursor()
                cr3.execute(f"""SELECT COUNT(*) FROM generated_tests
                    WHERE topic_code=ANY(%s) {f1} {f2} {f3}""", (tcs,))
                mavjud = (cr3.fetchone() or [0])[0]
                cr3.close(); c3.close()
                cnt_total = mavjud
                st2["_ts_mavjud"] = mavjud
                # Tanlangan son mavjuddan ko'p bo'lsa — kamaytiramiz
                _chek = min(mavjud, 50)
                if st2.get("ts_count", 20) > _chek and _chek > 0:
                    st2["ts_count"] = _chek
        except Exception as e:
            print(f"[ts_filter] {e}")

        try:
            await call.message.edit_reply_markup(reply_markup=_mk_ts_kb(st2, cnt_total))
        except Exception: pass
        return

    if call.data.startswith("book_make:"):
        book_id3 = int(call.data.split(":")[1])
        await call.answer()
        status3 = await call.message.answer("⏳ Word fayllar tayyorlanmoqda...")
        async def do_make():
            try:
                from hujjat_generator import create_book_archive
                zip_buf, info = create_book_archive(book_id3, pages_per_chunk=15)
                if not zip_buf:
                    await status3.edit_text("❌ " + info); return
                from aiogram.types import BufferedInputFile
                await call.message.answer_document(
                    BufferedInputFile(zip_buf.read(), f"kitob_{book_id3}.zip"),
                    caption=info
                )
                await status3.delete()
            except Exception as e:
                await status3.edit_text(f"❌ Xato: {e}")
        asyncio.create_task(do_make())
        return

    if call.data == "book_full_pkg":
        await call.answer()
        admin_state[user_id] = "full_pkg"
        await call.message.answer(
            "📦 To'liq paket\n\nSinf va fanni yozing:\n<code>1 | Ingliz tili</code>",
            parse_mode="HTML"
        )
        return

    if call.data.startswith("train_book:"):
        parts2 = call.data.split(":")
        book_id2 = int(parts2[1])
        fan2  = parts2[2] if len(parts2)>2 else ""
        sinf2 = parts2[3] if len(parts2)>3 else ""
        await call.answer()
        status2 = await call.message.answer(
            f"🎓 O'qitish boshlanmoqda...\n"
            f"📖 Kitob #{book_id2} | {fan2} | {sinf2}-sinf\n\n"
            f"⏳ Bu bir necha daqiqa davom etishi mumkin..."
        )
        async def do_train():
            try:
                from pedagog_trainer import train_from_book
                async def prog(msg):
                    try: await status2.edit_text(msg)
                    except: pass
                result = await train_from_book(book_id2, fan2, sinf2, prog)
                await call.message.answer(
                    f"✅ O'qitish yakunlandi!\n"
                    f"🧠 {result['facts']} ta bilim saqlandio'n"
                    f"🧪 {result['tests']} ta test yaratildi\n"
                    f"Endi Gemini/GPT siz ham javob bera olaman!"
                )
            except Exception as e:
                await call.message.answer(f"❌ Xato: {e}")
        asyncio.create_task(do_train())
        return

    if call.data.startswith("report_test:"):
        cur_idx = int(call.data.split(":")[1])
        await call.answer()
        # Timer ni pauza qilamiz
        from storage import test_sessions as _ts2
        s2 = _ts2.get(user_id,{})
        if s2.get("timer_task"):
            try: s2["timer_task"].cancel()
            except: pass
            s2["timer_task"] = None
        user_state[user_id] = f"report_comment:{cur_idx}"
        await call.message.answer(
            "✏️ Xato haqida yozing:\n\n"
            "• Javob noto'g'ri\n"
            "• Savol tushunarsiz\n"
            "• Rasm mos emas\n\n"
            "Yozib yuboring → test davom etadi"
        )
        return

    if call.data.startswith("admin_fix_test:"):
        # Admin tuzatish paneli
        tid = int(call.data.split(":")[1])
        conn2 = _get_db_conn(); cur2 = conn2.cursor()
        cur2.execute("""
            SELECT g.id, g.topic_code, g.question, g.option_a, g.option_b, g.option_c, g.option_d,
                   g.correct_answer, g.explanation, g.image_url, g.question_type,
                   c.comment
            FROM generated_tests g
            LEFT JOIN test_corrections c ON c.test_id=g.id
            WHERE g.id=%s LIMIT 1
        """, (tid,))
        row2 = cur2.fetchone(); cur2.close(); conn2.close()
        if not row2:
            await call.message.answer("❌ Test topilmadi"); return
        text2 = (
            f"✏️ Test #{tid} tuzatish\n\n"
            f"📝 Savol:\n{row2[2]}\n\n"
            f"A) {row2[3]}\nB) {row2[4]}\nC) {row2[5]}\nD) {row2[6]}\n\n"
            f"✅ To'g'ri: {row2[7]}\n"
            f"💡 Izoh: {row2[8] or '-'}\n"
            f"🖼 Rasm: {row2[9] or '-'}\n"
            f"🔑 Tur: {row2[10]}\n\n"
            f"👤 Izoh: {row2[11] or '-'}"
        )
        await call.message.answer(
            text2,
            reply_markup=InlineKeyboardMarkup(inline_keyboard=[
                [InlineKeyboardButton(text="✏️ Savolni o'zgartir", callback_data=f"edit_q:{tid}"),
                 InlineKeyboardButton(text="✅ To'g'rini o'zgartir", callback_data=f"edit_ans:{tid}")],
                [InlineKeyboardButton(text="💡 Izohni o'zgartir", callback_data=f"edit_expl:{tid}"),
                 InlineKeyboardButton(text="🖼 Rasmni o'zgartir", callback_data=f"edit_img:{tid}")],
                [InlineKeyboardButton(text="🗑 O'chir", callback_data=f"del_test:{tid}"),
                 InlineKeyboardButton(text="✅ OK (saqla)", callback_data=f"fix_ok:{tid}")],
            ])
        )
        return

    if call.data.startswith("edit_q:"):
        tid=int(call.data.split(":")[1])
        await call.answer()
        admin_state[user_id]=f"edit_test_field:{tid}:question"
        conn2=_get_db_conn();cur2=conn2.cursor()
        cur2.execute("SELECT question FROM generated_tests WHERE id=%s",(tid,))
        row2=cur2.fetchone();cur2.close();conn2.close()
        await call.message.answer(f"✏️ Yangi savolni yozing:\n\nHozirgi:\n{row2[0] if row2 else ''}")
        return

    if call.data.startswith("edit_ans:"):
        tid=int(call.data.split(":")[1])
        await call.answer()
        conn2=_get_db_conn();cur2=conn2.cursor()
        cur2.execute("SELECT option_a,option_b,option_c,option_d,correct_answer FROM generated_tests WHERE id=%s",(tid,))
        row2=cur2.fetchone();cur2.close();conn2.close()
        if row2:
            await call.message.answer(
                f"✅ To'g'ri javobni tanlang:\n\nA) {row2[0]}\nB) {row2[1]}\nC) {row2[2]}\nD) {row2[3]}\n\nHozir: {row2[4]}",
                reply_markup=InlineKeyboardMarkup(inline_keyboard=[
                    [InlineKeyboardButton(text=f"A) {row2[0][:30]}",callback_data=f"set_ans:{tid}:A")],
                    [InlineKeyboardButton(text=f"B) {row2[1][:30]}",callback_data=f"set_ans:{tid}:B")],
                    [InlineKeyboardButton(text=f"C) {row2[2][:30]}",callback_data=f"set_ans:{tid}:C")],
                    [InlineKeyboardButton(text=f"D) {row2[3][:30]}",callback_data=f"set_ans:{tid}:D")],
                ])
            )
        return

    if call.data.startswith("set_ans:"):
        parts2=call.data.split(":");tid=int(parts2[1]);ans=parts2[2]
        conn2=_get_db_conn();cur2=conn2.cursor()
        cur2.execute("UPDATE generated_tests SET correct_answer=%s WHERE id=%s",(ans,tid))
        conn2.commit();cur2.close();conn2.close()
        await call.answer(f"✅ To'g'ri javob {ans} ga o'zgartirildi",show_alert=True)
        return

    if call.data.startswith("edit_expl:"):
        tid=int(call.data.split(":")[1])
        await call.answer()
        admin_state[user_id]=f"edit_test_field:{tid}:explanation"
        conn2=_get_db_conn();cur2=conn2.cursor()
        cur2.execute("SELECT explanation FROM generated_tests WHERE id=%s",(tid,))
        row2=cur2.fetchone();cur2.close();conn2.close()
        await call.message.answer(f"💡 Yangi izohni yozing:\n\nHozirgi:\n{row2[0] if row2 else '-'}")
        return

    if call.data.startswith("edit_img:"):
        tid=int(call.data.split(":")[1])
        await call.answer()
        admin_state[user_id]=f"edit_test_field:{tid}:image_url"
        conn2=_get_db_conn();cur2=conn2.cursor()
        cur2.execute("SELECT image_url FROM generated_tests WHERE id=%s",(tid,))
        row2=cur2.fetchone();cur2.close();conn2.close()
        await call.message.answer(f"🖼 Yangi rasm kodini yozing:\n\nHozirgi: {row2[0] if row2 else '-'}\n\nMasalan: 3-02-1-001-1")
        return

    if call.data.startswith("fix_ok:"):
        tid = int(call.data.split(":")[1])
        conn2 = _get_db_conn(); cur2 = conn2.cursor()
        cur2.execute("UPDATE test_corrections SET status='resolved' WHERE test_id=%s", (tid,))
        conn2.commit(); cur2.close(); conn2.close()
        await call.answer("✅ Belgilandi")
        await call.message.edit_reply_markup(reply_markup=None)
        return

    if call.data.startswith("del_test:"):
        if user_id not in ADMINS: return
        tid = int(call.data.split(":")[1])
        conn2 = _get_db_conn(); cur2 = conn2.cursor()
        cur2.execute("DELETE FROM generated_tests WHERE id=%s", (tid,))
        cur2.execute("UPDATE test_corrections SET status='deleted' WHERE test_id=%s", (tid,))
        conn2.commit(); cur2.close(); conn2.close()
        await call.answer("🗑 O'chirildi", show_alert=True)
        await call.message.edit_text("🗑 Test o'chirildi")
        return

    if call.data == "ts_go":
        from storage import user_state as _us
        st2 = _us.get(user_id) if isinstance(_us.get(user_id), dict) else {}
        tc   = st2.get("ts_topic","")
        topic_codes = st2.get("ts_topic_codes") or ([tc] if tc else [])
        cnt2 = min(st2.get("ts_count", 20), 50)   # maksimal 50 ta savol
        diff = st2.get("ts_diff", "all")
        write= st2.get("ts_write", False)
        if not topic_codes:
            await call.answer("❌ Mavzu tanlanmagan — qayta tanlang", show_alert=True)
            return
        conn2 = _get_db_conn(); cur2 = conn2.cursor()
        diff_f = "" if diff=="all" else f"AND difficulty='{diff}'"
        # write: False=faqat test, True=faqat yozma, "mix"=aralash
        if write == "mix": type_f = ""
        elif write == True: type_f = "AND question_type = 'write_answer'"
        else: type_f = "AND question_type != 'write_answer'"
        img    = st2.get("ts_img", "mix")
        # "Rasmli" = rasm CHIZILGAN (file_id bor). Faqat kod bo'lsa rasm ko'rinmaydi.
        if img == "mix":
            img_f = ""
        elif img is True:
            img_f = "AND image_file_id IS NOT NULL AND image_file_id <> ''"
        else:
            img_f = "AND (image_file_id IS NULL OR image_file_id = '')"
        cur2.execute(f"""
            SELECT question,option_a,option_b,option_c,option_d,
                   correct_answer,explanation,question_type,is_latex,
                   COALESCE(NULLIF(image_file_id,''), image_url) AS image_url,audio_text,language,time_limit,topic_code
            FROM generated_tests WHERE topic_code=ANY(%s) {diff_f} {type_f} {img_f}
            ORDER BY RANDOM() LIMIT %s
        """, (topic_codes, cnt2))
        _xom = cur2.fetchall(); cur2.close(); conn2.close()

        # Nazorat: har savol tanlangan kodlardan ekanini tekshiramiz
        _kod_set = set(topic_codes)
        _begona = [r[13] for r in _xom if r[13] not in _kod_set]
        if _begona:
            print(f"[ts_go] ⚠️ BEGONA KODLAR: {set(_begona)}")
        tests = [r[:13] for r in _xom if r[13] in _kod_set]
        print(f"[ts_go] topics={len(topic_codes)} soralgan={cnt2} topildi={len(tests)} "
              f"img={img} write={write} diff={diff}")

        # ADMIN uchun tashxis — qaysi kod, qaysi savol
        if _is_admin(user_id):
            try:
                _d = [f"🔍 <b>Tashxis</b>", ""]
                _d.append(f"📚 Mavzu: {st2.get('ts_mavzu_name','?')}")
                _d.append(f"🎓 Sinf: {st2.get('ts_grade','?')}")
                _d.append(f"🔢 Topic kodlar ({len(topic_codes)}):")
                for _k in topic_codes[:6]:
                    _d.append(f"   <code>{_k}</code>")
                if len(topic_codes) > 6:
                    _d.append(f"   ... yana {len(topic_codes)-6} ta")
                _d.append(f"\n📝 Olingan savollar ({len(tests)}):")
                for _r in _xom[:4]:
                    _d.append(f"   <code>{_r[13]}</code>")
                    _d.append(f"   • {str(_r[0])[:50]}")
                await call.message.answer("\n".join(_d)[:3800], parse_mode="HTML")
            except Exception as _de:
                print(f"[ts_go] tashxis: {_de}")
        if not tests:
            await call.answer("❌ Bu filtr bo'yicha test topilmadi!", show_alert=True)
            return
        await call.answer()
        # Test paytida menyu klaviaturasini yashiramiz
        try:
            from aiogram.types import ReplyKeyboardRemove
            _hid = await call.message.answer("🧪 Test boshlandi", reply_markup=ReplyKeyboardRemove())
            try: await _hid.delete()
            except: pass
        except Exception: pass
        from test_engine import start_test
        timed_ = st2.get("ts_timed", True)
        ovoz_  = False   # ovoz FAQAT tugma bosilganda — hech qachon avtomatik
        # test_engine eski versiya bo'lsa avto_ovoz qabul qilmasligi mumkin
        try:
            await start_test(user_id, tests, call.message, timed=timed_, avto_ovoz=ovoz_)
        except TypeError:
            await start_test(user_id, tests, call.message, timed=timed_)
        # Sozlamani sessiyaga yozamiz — test_engine o'qishi mumkin
        from storage import test_sessions as _ts
        if user_id in _ts:
            _ts[user_id]["topic_code"] = tc
            _ts[user_id]["avto_ovoz"] = ovoz_
        return

    # ═══════════════════════

    if call.data.startswith("next_lesson_"):
        topic_code = call.data.replace("next_lesson_", "")
        await open_teacher_lesson(call.message, topic_code, _user_id=call.from_user.id)
        await call.answer()
        return

    if call.data == "force_dashboard":
        await call.answer()
        try:
            from student_dashboard import build_dashboard_full
            text, kb = await build_dashboard_full(call.from_user.id)
            await call.message.edit_text(text, reply_markup=kb)
        except Exception:
            pass
        return

    if call.data == "go_sleep":
        await call.answer()
        await call.message.edit_text(
            "😴 Yaxshi uxlang!\n🌙 Ertaga yanada kuchli bo'lasiz! 💪"
        )
        return

    if call.data == "go_rest":
        await call.answer()
        await call.message.edit_text(
            "🎮 Yaxshi dam oling!\n"
            "Ruhiy kuch to'plash ham o'rganish! 💚"
        )
        return

    if call.data.startswith("sh_"):
        from shablon_yaratish import handle_shablon_callback
        await handle_shablon_callback(call, user_id)
        return

    if call.data.startswith("sinov_"):
        from test_sinovi import handle_sinov_callback
        await handle_sinov_callback(call, user_id)
        return

    if call.data.startswith("ts_"):
        parts = call.data.split(":")
        action = parts[0][3:]  # excel, import, gen, del, view

        if action == "excel":
            topic_code = parts[1]
            # Excel shablon yuborish
            topic_stats_state[user_id]["selected_topic"] = topic_code
            await call.answer()
            # Excel shablon generatsiya
            import openpyxl, io
            from openpyxl.styles import Font, PatternFill
            from aiogram.types import BufferedInputFile
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "TESTLAR"
            headers = ["topic_code","difficulty","situation","question","option_a","option_b","option_c","option_d","correct_answer","explanation","question_type","is_latex","image_url","audio_text","language","life_level","age_group","time_limit"]
            for col, h in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=h)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill("solid", fgColor="2E86AB")
            conn_t = _get_db_conn()
            cur_t = conn_t.cursor()
            cur_t.execute("SELECT grade FROM dts_tree WHERE topic_code=%s LIMIT 1", (topic_code,))
            row_t = cur_t.fetchone()
            cur_t.close(); conn_t.close()
            grade = row_t[0] if row_t else "1"
            age_map = {"1":"6-7","2":"7-8","3":"8-9","4":"9-10","5":"10-11","6":"11-12","7":"12-13","8":"13-14","9":"14-15","10":"15-16","11":"16-17"}
            age_group = age_map.get(str(grade), "10-11")
            def get_diff(n):
                if n<=10: return "oson"
                elif n<=20: return "o'rta"
                elif n<=30: return "qiyin"
                else: return "murakkab"
            for i in range(1, 41):
                ws.append([topic_code, get_diff(i), "oddiy", "", "", "", "", "", "", "", "single_choice", False, f"{topic_code}-{i}", None, "uz", 1, age_group, 60])
            buf = io.BytesIO()
            wb.save(buf); buf.seek(0)
            await call.message.answer_document(
                BufferedInputFile(buf.read(), filename=f"shablon_{topic_code}.xlsx"),
                caption=f"📋 Shablon: {topic_code}"
            )

        elif action == "import":
            topic_code = parts[1]
            admin_state[user_id] = "test_import"
            await call.answer()
            await call.message.answer("📥 Excel fayl yuboring")

        elif action == "gen":
            topic_code = parts[1]
            await call.answer("🤖 AI yaratmoqda...", show_alert=True)
            # DTS dan ma'lumot olish
            conn_g = _get_db_conn()
            cur_g = conn_g.cursor()
            cur_g.execute("""
                SELECT grade, subject_name, mavzu_name, kichik_name
                FROM dts_tree WHERE topic_code=%s LIMIT 1
            """, (topic_code,))
            row_g = cur_g.fetchone()
            cur_g.close(); conn_g.close()
            if row_g:
                grade, subject, mavzu, kichik = row_g
                from ai_generatori import _generate_questions, _age_group
                msg = await call.message.answer(f"🤖 AI ishlamoqda...\n📌 {kichik}")
                try:
                    questions = await _generate_questions(grade, subject, mavzu, kichik, topic_code)
                    conn_s = _get_db_conn()
                    cur_s = conn_s.cursor()
                    for q in questions:
                        cur_s.execute("""
                            INSERT INTO generated_tests
                            (topic_code,question,option_a,option_b,option_c,option_d,
                             correct_answer,explanation,question_type,is_latex,
                             image_url,audio_text,language,life_level,age_group,
                             time_limit,difficulty,situation)
                            VALUES(%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
                        """, (topic_code, q.get("question",""), q.get("a",""), q.get("b",""),
                              q.get("c",""), q.get("d",""), q.get("correct",""),
                              q.get("explanation",""), q.get("question_type","single_choice"),
                              False, q.get("image_url"), None, "uz", 1,
                              _age_group(grade), q.get("time_limit",60),
                              q.get("difficulty","oson"), "oddiy"))
                    conn_s.commit(); cur_s.close(); conn_s.close()
                    await msg.edit_text(f"✅ {len(questions)} ta savol yaratildi!\n🔑 {topic_code}")
                except Exception as ex:
                    await msg.edit_text(f"❌ Xato: {ex}")

        elif action == "del":
            topic_code = parts[1]
            await call.message.edit_reply_markup(reply_markup=InlineKeyboardMarkup(inline_keyboard=[
                [
                    InlineKeyboardButton(text="✅ Ha, o'chir", callback_data=f"ts_del_yes:{topic_code}"),
                    InlineKeyboardButton(text="❌ Yo'q", callback_data=f"ts_del_no:{topic_code}"),
                ]
            ]))
            await call.answer("⚠️ Tasdiqlang!")

        elif action == "del_yes":
            topic_code = parts[1]
            conn_d = _get_db_conn()
            cur_d = conn_d.cursor()
            cur_d.execute("DELETE FROM generated_tests WHERE topic_code=%s", (topic_code,))
            cnt = cur_d.rowcount
            conn_d.commit(); cur_d.close(); conn_d.close()
            await call.answer(f"✅ {cnt} ta test o'chirildi!", show_alert=True)
            await call.message.edit_reply_markup(reply_markup=None)

        elif action == "del_no":
            await call.answer("❌ Bekor qilindi")
            await call.message.edit_reply_markup(reply_markup=InlineKeyboardMarkup(inline_keyboard=[
                [
                    InlineKeyboardButton(text="📄 Excel shablon", callback_data=f"ts_excel:{parts[1]}"),
                    InlineKeyboardButton(text="📥 Import", callback_data=f"ts_import:{parts[1]}"),
                ],
                [
                    InlineKeyboardButton(text="🤖 AI Yaratish", callback_data=f"ts_gen:{parts[1]}"),
                    InlineKeyboardButton(text="🗑 O'chirish", callback_data=f"ts_del:{parts[1]}"),
                ],
                [
                    InlineKeyboardButton(text="👁 Testlarni ko'rish", callback_data=f"ts_view:{parts[1]}:0"),
                ],
            ]))

        elif action == "view":
            topic_code = parts[1]
            offset = int(parts[2]) if len(parts) > 2 else 0
            conn_v = _get_db_conn()
            cur_v = conn_v.cursor()
            cur_v.execute("SELECT COUNT(*) FROM generated_tests WHERE topic_code=%s", (topic_code,))
            total_v = cur_v.fetchone()[0]
            cur_v.execute("""
                SELECT id, difficulty, question, correct_answer
                FROM generated_tests WHERE topic_code=%s
                ORDER BY difficulty, id LIMIT 5 OFFSET %s
            """, (topic_code, offset))
            tests = cur_v.fetchall()
            cur_v.close(); conn_v.close()

            text = f"👁 Testlar: {topic_code}\nJami: {total_v}\n\n"
            for tid, diff, q, correct in tests:
                text += f"[{diff}] {q[:60]}...\n✅ {correct[:30]}\n\n"

            nav = []
            if offset > 0:
                nav.append(InlineKeyboardButton(text="◀️", callback_data=f"ts_view:{topic_code}:{offset-5}"))
            if offset + 5 < total_v:
                nav.append(InlineKeyboardButton(text="▶️", callback_data=f"ts_view:{topic_code}:{offset+5}"))

            kb_rows = []
            if nav: kb_rows.append(nav)
            kb_rows.append([InlineKeyboardButton(text="🗑 Barchasini o'chir", callback_data=f"ts_del:{topic_code}")])

            await call.message.edit_text(text, reply_markup=InlineKeyboardMarkup(inline_keyboard=kb_rows))
            await call.answer()

        return

    if call.data.startswith("gen_") or call.data.startswith("gg_cnt_") or call.data.startswith("gg_tp_") or call.data in ("gen_go", "gen_template"):
        from ai_generatori import handle_gen_callback
        await handle_gen_callback(call, user_id)
        return

    if call.data.startswith("sts_"):
        from student_settings import handle_settings_callback
        await handle_settings_callback(call, user_id, user_state)
        return

    if call.data.startswith("img_"):
        from image_admin import handle_img_callback
        await handle_img_callback(call, user_id)
        return

    if call.data == "go_home_dashboard":
        await call.answer()
        conn2 = _get_db_conn()
        cur2  = conn2.cursor()
        cur2.execute("SELECT role FROM users WHERE user_id=%s", (call.from_user.id,))
        row = cur2.fetchone()
        cur2.close(); conn2.close()
        role = row[0] if row else "🧒 O'quvchi"

        # Eski xabarlarni o'chirish
        try:
            for i in range(call.message.message_id, call.message.message_id - 20, -1):
                try:
                    await bot.delete_message(call.message.chat.id, i)
                except Exception:
                    pass
        except Exception:
            pass

        try:
            from student_dashboard import build_dashboard
            text, kb = await build_dashboard(call.from_user.id)
            await bot.send_message(call.message.chat.id, text, reply_markup=kb)
        except Exception:
            pass
        await bot.send_message(
            call.message.chat.id,
            "🏠 Bosh menyu",
            reply_markup=get_main_keyboard(role)
        )
        return

    if call.data == "go_home":
        await call.answer()
        conn2 = _get_db_conn()
        cur2  = conn2.cursor()
        cur2.execute("SELECT role FROM users WHERE user_id=%s", (call.from_user.id,))
        row = cur2.fetchone()
        cur2.close(); conn2.close()
        role = row[0] if row else "🧒 O'quvchi"
        await call.message.answer(
            "🏠 Asosiy menyu",
            reply_markup=get_main_keyboard(role)
        )
        return

    if call.data in ("lesson_continue", "lesson_continue_force"):
        from datetime import datetime
        now        = datetime.now()
        hour       = now.hour
        weekday    = now.weekday()
        is_night   = hour >= 22 or hour < 6
        is_weekend = weekday >= 5
        forced     = call.data == "lesson_continue_force"

        if is_night and not forced:
            await call.answer("🌙 Tun vaqti! Uxlash sog'liq uchun muhim.", show_alert=True)
            return

        if is_weekend and not forced:
            await call.answer()
            await call.message.answer(
                "🏖 Bugun dam olish kuni!\n\nBaribir dars o'rganasizmi?",
                reply_markup=InlineKeyboardMarkup(inline_keyboard=[[
                    InlineKeyboardButton(text="✅ Ha, o'rganaman", callback_data="lesson_continue_force"),
                    InlineKeyboardButton(text="🎮 Yo'q, dam olaman", callback_data="go_rest"),
                ]])
            )
            return

        await call.answer()
        try:
            await call.message.edit_reply_markup(reply_markup=None)
        except Exception:
            pass
        await open_teacher_lesson(call.message, _user_id=call.from_user.id)
        return

    if call.data == "lesson_repeat":
        from progress import get_repeat_topics
        topics = get_repeat_topics(call.from_user.id)
        if topics:
            topic_code = topics[0][0]
            await open_teacher_lesson(call.message, topic_code, _user_id=call.from_user.id)
        else:
            await call.answer("Takrorlanadigan mavzu yo'q!", show_alert=True)
        await call.answer()
        return

    if call.data.startswith("fan_select|"):
        parts      = call.data.split("|")
        grade      = parts[1]
        subj       = parts[2]
        from progress import get_next_topic
        next_topic = get_next_topic(call.from_user.id, grade, None)
        if next_topic and next_topic[3] == subj:
            await open_teacher_lesson(call.message, next_topic[0], _user_id=call.from_user.id)
        else:
            conn2 = _get_db_conn()
            cur2  = conn2.cursor()
            cur2.execute("""
                SELECT t.topic_code FROM dts_tree t
                LEFT JOIN learned_topics lt
                    ON lt.topic_code=t.topic_code AND lt.user_id=%s
                WHERE t.grade=%s AND t.subject_name=%s
                  AND lt.topic_code IS NULL AND t.is_deleted=FALSE
                ORDER BY t.topic_code LIMIT 1
            """, (call.from_user.id, grade, subj))
            row = cur2.fetchone()
            cur2.close(); conn2.close()
            if row:
                await open_teacher_lesson(call.message, row[0], _user_id=call.from_user.id)
            else:
                await call.answer(f"✅ {subj} tugallangan!", show_alert=True)
        await call.answer()
        return

    if call.data == "exam_later":
        await call.answer("⏰ Keyinroq eslatiladi")
        conn = _get_db_conn()
        cur  = conn.cursor()
        cur.execute("SELECT role FROM users WHERE user_id=%s", (call.from_user.id,))
        row  = cur.fetchone()
        cur.close(); conn.close()
        role = row[0] if row else "🧒 O'quvchi"
        await call.message.delete()
        await call.message.answer("🏠 Asosiy menyu", reply_markup=get_main_keyboard(role))
        return

    if call.data.startswith("ans_"):

        answer = call.data.replace(
            "ans_",
            ""
        )

        await check_button_answer(
            call.from_user.id,
            answer,
            call.message
        )

        return

    # ═══ BOSH EKRAN (dashboard) ═══
    if call.data.startswith("dash_"):
        await call.answer()
        import student_dashboard as _sd
        uid_d = call.from_user.id
        try:
            if call.data == "dash_refresh":
                txt, kb = await _sd.build_dashboard(uid_d)
            elif call.data == "dash_togarak":
                txt, kb = await _sd.build_togarak_stat(uid_d)
            elif call.data == "dash_maktab":
                txt, kb = await _sd.build_maktab_stat(uid_d)
            elif call.data == "dash_bugun":
                txt, kb = await _sd.build_bugungi(uid_d)
            elif call.data == "dash_vazifa":
                txt, kb = await _sd.build_vazifalar(uid_d)
            elif call.data == "dash_stat":
                txt, kb = await _sd.build_maktab_stat(uid_d)
            else:
                return
            try:
                await call.message.edit_text(txt[:4000], parse_mode="HTML", reply_markup=kb)
            except Exception:
                await call.message.answer(txt[:4000], parse_mode="HTML", reply_markup=kb)
        except Exception as e:
            import traceback; traceback.print_exc()
            try: await call.message.answer(f"⚠️ Xatolik: {e}")
            except Exception: pass
        return

    # ═══════════════════════════════════════════
    # 👨‍👩‍👧 OTA-ONA PANELI (op_)
    # ═══════════════════════════════════════════
    if call.data.startswith("op_"):
        qism = call.data.split(":")
        bolim = qism[0][3:]           # asosiy / yoqlama / vazifa / imtihon / baho / nazorat
        child_id = int(qism[1])
        await _ota_ekran(call, user_id, child_id, bolim)
        return

    # ═══════════════════════════════════════════
    # 📲 AKKAUNT KO'CHIRISH (ak_)
    # ═══════════════════════════════════════════
    if call.data.startswith("ak_"):
        import akkaunt as _ak
        _ak.jadval()
        qism = call.data.split(":")
        amal = qism[0]
        tg = _tg_id(user_id)

        # ── Menyu ──
        if amal == "ak_menu":
            lst = _ak.akkauntlar(tg)
            t = [f"📱 <b>Akkauntlar ({len(lst)}/{_ak.MAX_AKKAUNT})</b>\n"]
            for uid2, idx, ism, rol2, sinf2, aktiv in lst:
                belgi = "🟢" if aktiv else "⚪"
                t.append(f"{belgi} {ism or '—'} · {rol2 or '—'} {sinf2 or ''}")
            rows = [
                [InlineKeyboardButton(text="📤 Boshqa telefonga ko'chirish",
                                      callback_data="ak_kod")],
                [InlineKeyboardButton(text="📥 Boshqa telefondagi akkauntga kirish",
                                      callback_data="ak_kirish")],
            ]
            if len(lst) > 1:
                rows.append([InlineKeyboardButton(text="🔓 Shu telefondan uzish",
                                                  callback_data="ak_uz")])
            await call.message.answer("\n".join(t), parse_mode="HTML",
                                      reply_markup=InlineKeyboardMarkup(inline_keyboard=rows))
            return

        # ── Ko'chirish kodi ──
        if amal == "ak_kod":
            kod = _ak.kochirish_kod(user_id, tg)
            if not kod:
                await call.answer("❌ Kod yaratilmadi", show_alert=True); return
            ism, _, _ = (lambda r: r)(_get_user_qisqa(user_id))
            await call.message.answer(
                f"📤 <b>Boshqa telefonga ko'chirish</b>\n\n"
                f"👤 {ism or 'Akkaunt'}\n\n"
                f"Kod:\n\n<code>{kod}</code>\n\n"
                f"⏳ {_ak.KOD_MUDDAT} daqiqa amal qiladi.\n\n"
                f"Yangi telefonda botga kiring:\n"
                f"«👤 Kabinet → 📱 Akkauntlar → 📥 Kirish»\n"
                f"va shu kodni yozing.\n\n"
                f"⚠️ Kodni hech kimga bermang — akkauntingizga kirib oladi.",
                parse_mode="HTML")
            return

        # ── Kod bilan kirish ──
        if amal == "ak_kirish":
            if not _ak.joy_bormi(tg):
                await call.answer(
                    f"❌ Bu telefonda {_ak.MAX_AKKAUNT} ta akkaunt bor. Avval bittasini uzing.",
                    show_alert=True)
                return
            user_state[user_id] = "ak_kod_kirit"
            await call.message.answer(
                "📥 <b>Akkauntga kirish</b>\n\n"
                "Eski telefoningizda olingan\n"
                "8 belgili kodni yozing:",
                parse_mode="HTML")
            return

        # ── Uzish ──
        if amal == "ak_uz":
            lst = _ak.akkauntlar(tg)
            rows = [[InlineKeyboardButton(
                text=f"🔓 {(a[2] or '—')[:26]} {a[4] or ''}",
                callback_data=f"ak_uzyes:{a[0]}")] for a in lst if a[0]]
            await call.message.answer(
                "🔓 Qaysi akkauntni shu telefondan uzamiz?\n\n"
                "<i>Ma'lumot o'chmaydi — boshqa telefonda qoladi.</i>",
                parse_mode="HTML",
                reply_markup=InlineKeyboardMarkup(inline_keyboard=rows))
            return

        if amal == "ak_uzyes":
            uid2 = int(qism[1])
            ok, xabar = _ak.uzib_qoy(tg, uid2)
            _eff_clear(tg)
            await call.answer(xabar[:190], show_alert=not ok)
            if ok:
                try: await call.message.edit_text(xabar)
                except Exception: pass
            return
        return

    # ═══════════════════════════════════════════
    # 👨‍👩‍👧 OTA-ONA ↔ FARZAND (fk_)
    # ═══════════════════════════════════════════
    if call.data.startswith("fk_"):
        import ota_ona as _oo
        qism = call.data.split(":")
        amal = qism[0]

        # ── Ota-ona kod so'raydi ──
        if amal == "fk_add":
            import akkaunt as _ak
            tg = _tg_id(user_id)
            # Shu telefondagi boshqa akkauntlar (kod kerak emas — egasi o'zi)
            boshqa = [a for a in _ak.akkauntlar(tg) if a[0] and int(a[0]) != user_id]

            rows = []
            for uid2, idx, ism, rol2, sinf2, aktiv in boshqa:
                if _oo.bogliqmi(user_id, int(uid2)):
                    continue
                rows.append([InlineKeyboardButton(
                    text=f"📱 {(ism or '—')[:24]} {sinf2 or ''}",
                    callback_data=f"fk_shu:{uid2}")])

            rows.append([InlineKeyboardButton(text="🔢 Kod bilan ulash",
                                              callback_data="fk_kodla")])
            matn = "➕ <b>Farzand qo'shish</b>\n\n"
            if rows[:-1]:
                matn += ("📱 Shu telefondagi akkauntlar — bir bosishda ulanadi:\n"
                         "(tasdiq so'ralmaydi, chunki telefon sizniki)\n\n")
            matn += "🔢 Boshqa telefondagi farzand — kod kerak."
            await call.message.answer(matn, parse_mode="HTML",
                                      reply_markup=InlineKeyboardMarkup(inline_keyboard=rows))
            return

        # ── Shu telefondagi akkauntni ulash ──
        if amal == "fk_shu":
            child = int(qism[1])
            if child == user_id:
                await call.answer("❌ O'zingizni ulay olmaysiz", show_alert=True); return
            import akkaunt as _ak
            tg = _tg_id(user_id)
            # Haqiqatan shu telefondami?
            egasi = {int(a[0]) for a in _ak.akkauntlar(tg) if a[0]}
            if child not in egasi:
                await call.answer("❌ Bu akkaunt shu telefonda emas", show_alert=True); return

            _oo.bogla(user_id, child)
            ism_c, _, sinf_c = _oo.kim(child)
            await call.answer("✅ Ulandi")
            await call.message.answer(
                f"✅ <b>{ism_c or 'Farzand'}</b> {sinf_c or ''} ulandi.\n\n"
                f"«👤 Kabinet → 👨‍👩‍👧 Farzandlarim» dan ko'rasiz.\n"
                f"Akkauntga kirish uchun «🔄 Akkaunt almashtirish».",
                parse_mode="HTML")
            return

        # ── Kod bilan ulash ──
        if amal == "fk_kodla":
            user_state[user_id] = "parent_link_id"
            await call.message.answer(
                "🔢 <b>Kod bilan ulash</b>\n\n"
                "Farzandingiz o'z telefonida:\n"
                "«👤 Kabinet → 🔗 Ota-onani ulash»\n"
                "dan 6 xonali kod oladi.\n\n"
                "Shu kodni yozing:",
                parse_mode="HTML")
            return

        # ── Farzand kod oladi ──
        if amal == "fk_kod":
            kod = _oo.kod_yarat(user_id)
            if not kod:
                await call.answer("❌ Kod yaratilmadi", show_alert=True); return
            await call.message.answer(
                f"🔗 <b>Ota-onani ulash</b>\n\n"
                f"Kodingiz:\n\n<code>{kod}</code>\n\n"
                f"⏳ {_oo.KOD_MUDDAT} daqiqa amal qiladi.\n\n"
                f"Ota-onangiz botda «👨‍👩‍👧 Farzand qo'shish» dan\n"
                f"shu kodni kiritsin. So'ng sizdan tasdiq so'raladi.",
                parse_mode="HTML")
            return

        # ── Farzand tasdiqladi ──
        if amal == "fk_ok":
            parent_id = int(qism[1]); kod = qism[2]
            tekshir = _oo.kod_tekshir(kod)
            if tekshir != user_id:
                await call.answer("⚠️ Kod muddati o'tgan", show_alert=True)
                try: await call.message.edit_text("⏳ Kod muddati o'tgan. Yangi kod oling.")
                except Exception: pass
                return

            ok = _oo.bogla(parent_id, user_id)
            _oo.kod_ochir(kod)
            ism_p, _, _ = _oo.kim(parent_id)
            ism_c, _, sinf_c = _oo.kim(user_id)

            await call.answer("✅ Ulandi")
            try:
                await call.message.edit_text(
                    f"✅ <b>{ism_p or 'Ota-ona'}</b> ulandi.\n\n"
                    f"Endi u natijalaringizni ko'radi.\n"
                    f"Istalgan payt uzib qo'yishingiz mumkin:\n"
                    f"«👤 Kabinet → 🔗 Ota-onani ulash»",
                    parse_mode="HTML")
            except Exception: pass

            try:
                await bot.send_message(
                    _tg_id(parent_id),
                    f"✅ <b>Bog'lanish tasdiqlandi</b>\n\n"
                    f"👤 {ism_c or 'Farzand'} {sinf_c or ''}\n\n"
                    f"Endi uning baholari, davomati va\n"
                    f"imtihon natijalarini ko'rasiz.",
                    parse_mode="HTML")
            except Exception: pass
            return

        # ── Farzand rad etdi ──
        if amal == "fk_no":
            parent_id = int(qism[1]); kod = qism[2]
            _oo.kod_ochir(kod)
            await call.answer("❌ Rad etildi")
            try:
                await call.message.edit_text(
                    "❌ So'rov rad etildi.\n\nHech qanday ma'lumot ulashilmadi.")
            except Exception: pass
            try:
                await bot.send_message(_tg_id(parent_id),
                    "❌ Farzandingiz ulanish so'rovini rad etdi.")
            except Exception: pass
            return

        # ── Farzandlar ro'yxati (ota-ona) ──
        if amal == "fk_list":
            lst = _oo.farzandlar(user_id)
            if not lst:
                await call.message.answer(
                    "👨‍👩‍👧 Hali farzand ulanmagan.\n\n"
                    "Farzandingizdan kod so'rang.")
                return
            rows = [[InlineKeyboardButton(text=f"🔓 {ism[:26]} {sinf} — uzish",
                     callback_data=f"fk_del:{cid}")] for cid, ism, sinf in lst]
            await call.message.answer(
                f"👨‍👩‍👧 <b>Farzandlarim ({len(lst)})</b>", parse_mode="HTML",
                reply_markup=InlineKeyboardMarkup(inline_keyboard=rows))
            return

        # ── Ota-onalar ro'yxati (farzand) ──
        if amal == "fk_mine":
            lst = _oo.otalar(user_id)
            rows = [[InlineKeyboardButton(text="🔗 Yangi kod olish", callback_data="fk_kod")]]
            if lst:
                qatorlar = [f"👨‍👩‍👧 <b>Ulangan ota-onalar ({len(lst)})</b>", ""]
                for pid, ism in lst:
                    qatorlar.append(f"• {ism}")
                    rows.append([InlineKeyboardButton(
                        text=f"🔓 {ism[:28]} — uzish", callback_data=f"fk_cut:{pid}")])
                t = "\n".join(qatorlar)
            else:
                t = ("👨‍👩‍👧 Hali ota-ona ulanmagan.\n\n"
                     "Kod oling va ota-onangizga bering.")
            await call.message.answer(t, parse_mode="HTML",
                                      reply_markup=InlineKeyboardMarkup(inline_keyboard=rows))
            return

        # ── Uzish ──
        if amal in ("fk_del", "fk_cut"):
            bosh = int(qism[1])
            if amal == "fk_del":   # ota-ona farzandni uzadi
                ok = _oo.uzib_qoy(user_id, bosh); kim2 = bosh
            else:                  # farzand ota-onani uzadi
                ok = _oo.uzib_qoy(bosh, user_id); kim2 = bosh
            await call.answer("✅ Uzildi" if ok else "⚠️ Topilmadi")
            if ok:
                try:
                    await call.message.edit_text("🔓 Bog'lanish uzildi.")
                except Exception: pass
                try:
                    await bot.send_message(_tg_id(kim2), "🔓 Bog'lanish uzildi.")
                except Exception: pass
            return
        return

    # ═══════════════════════════════════════════
    # 🎓 TO'GARAK TESTLARI (tt_)
    # ═══════════════════════════════════════════
    if call.data == "tt_lock":
        await call.answer("🔒 Bu mavzu hali o'tilmagan", show_alert=True)
        return

    if call.data.startswith(("tt_tg:", "tt_x:", "tt_all:", "tt_p:", "tt_go:")):
        import togarak_test as _tt
        qism = call.data.split(":")
        amal = qism[0]
        tgid = int(qism[1])

        kalit = f"tt_sel:{user_id}:{tgid}"
        sah_kalit = f"tt_sah:{user_id}:{tgid}"
        ro_yxat = _tt.mavzular(tgid)
        if not ro_yxat:
            await call.answer("❌ Bu to'garak fani bo'yicha test yo'q", show_alert=True)
            return

        tanlangan = set(temp_user.get(kalit, []))
        sahifa = int(temp_user.get(sah_kalit, 0))

        if amal == "tt_tg":
            tanlangan = set(); sahifa = 0
        elif amal == "tt_x":
            i = int(qism[2])
            if i in tanlangan: tanlangan.discard(i)
            else:              tanlangan.add(i)
        elif amal == "tt_all":
            ochiqlar = {i for i, m in enumerate(ro_yxat) if m[3]}
            tanlangan = set() if tanlangan >= ochiqlar else ochiqlar
        elif amal == "tt_p":
            sahifa = int(qism[2])
        elif amal == "tt_go":
            if not tanlangan:
                await call.answer("❌ Mavzu belgilanmagan", show_alert=True); return
            kodlar = _tt.kodlar(ro_yxat, tanlangan)
            nomlar = _tt.nomlar(ro_yxat, tanlangan)
            jami = _tt.test_soni(kodlar)
            if jami == 0:
                await call.answer("❌ Test topilmadi", show_alert=True); return

            fan, tg_nomi = _tt.togarak_fan(tgid)
            nom = nomlar[0] if len(nomlar) == 1 else f"{len(nomlar)} ta mavzu"

            import ts_cache
            sid = ts_cache.saqla(kodlar, nom, "", fan, jami)
            print(f"[tt_go] tgid={tgid} {len(kodlar)} topic, {jami} test -> ts_sel:{sid}")

            temp_user.pop(kalit, None); temp_user.pop(sah_kalit, None)

            chiqadi = min(jami, _tt.MAX_TEST)
            from storage import user_state as _us
            if not isinstance(_us.get(user_id), dict): _us[user_id] = {}
            _us[user_id].update({
                "ts_topic": kodlar[0], "ts_topic_codes": kodlar,
                "ts_mavzu_name": nom, "ts_grade": "", "ts_subject": fan,
                "ts_count": chiqadi, "ts_diff": "all",
                "ts_timed": True, "ts_write": False, "ts_img": "mix",
                "ts_ovoz": False, "_ts_cnt_total": jami,
            })
            await call.message.answer(
                f"🎓 {tg_nomi}\n📚 {nom}\n"
                f"📊 {jami} ta savol · {chiqadi} tasi chiqadi\n\n"
                f"Sozlamalarni tanlang:",
                reply_markup=_mk_ts_kb(_us[user_id], jami)
            )
            return

        temp_user[kalit] = list(tanlangan)
        temp_user[sah_kalit] = sahifa

        ochiq = sum(1 for m in ro_yxat if m[3])
        yopiq = len(ro_yxat) - ochiq
        _, tg_nomi = _tt.togarak_fan(tgid)
        sarlavha = f"🎓 {tg_nomi}\n📚 {ochiq} mavzu ochiq"
        if yopiq: sarlavha += f" · 🔒 {yopiq} yopiq"
        if tanlangan:
            n = _tt.test_soni(_tt.kodlar(ro_yxat, tanlangan))
            sarlavha += f"\n☑️ {len(tanlangan)} tanlandi · {n} ta savol"
        sarlavha += "\n\nMavzularni belgilang:"

        kb = _tt.mavzu_kb(tgid, ro_yxat, tanlangan, sahifa)
        try:
            await call.message.edit_text(sarlavha, reply_markup=kb)
        except Exception:
            await call.message.answer(sarlavha, reply_markup=kb)
        return

    # ═══════════════════════════════════════════
    # 📊 IMTIHON — TEST YURITUVCHISI (imq / imstop)
    # ═══════════════════════════════════════════
    if call.data.startswith("imstop:"):
        import baholash as _bh
        _bh.seans_tugat(user_id)
        await call.answer("🛑 To'xtatildi")
        try: await call.message.edit_text("🛑 Imtihon to'xtatildi.\nNatija saqlanmadi.")
        except Exception: pass
        return

    if call.data.startswith("imq:"):
        import baholash as _bh
        _, iid, idx, javob = call.data.split(":")
        iid = int(iid); idx = int(idx)

        st = _bh.seans(user_id)
        if not st or st["idx"] != idx:
            await call.answer("⚠️ Bu savol allaqachon javoblangan", show_alert=True)
            return

        togri, tugadi, foiz = _bh.javob_tekshir(user_id, javob)
        await call.answer("✅ To'g'ri" if togri else "❌ Noto'g'ri")

        try: await call.message.delete()
        except Exception: pass

        if not tugadi:
            s = st["savollar"][st["idx"]]
            matn = _bh.savol_matni(s, st["idx"], len(st["savollar"]))
            kb = _bh.savol_kb(iid, st["idx"])
            if s[6]:
                try:
                    await call.message.answer_photo(s[6], caption=matn, reply_markup=kb)
                    return
                except Exception: pass
            await call.message.answer(matn, reply_markup=kb)
            return

        # ── IMTIHON TUGADI ──
        st = _bh.seans_tugat(user_id)
        _bh.baho_qoy(iid, user_id, foiz, manba="test")
        imt = _bh.imtihon_ol(iid)
        d = _bh.daraja(foiz)

        await call.message.answer(
            f"🏁 <b>Imtihon tugadi</b>\n\n"
            f"📝 {imt['nomi'] if imt else '—'}\n"
            f"✅ To'g'ri: {st['togri']}/{len(st['savollar'])}\n"
            f"📊 Natija: <b>{foiz}%</b>\n"
            f"🎖 {d}\n\n"
            f"<i>Natija o'qituvchi va ota-onangizga ko'rinadi.</i>",
            parse_mode="HTML"
        )

        # O'qituvchi va ota-onaga xabar
        if imt:
            try:
                c = _get_db_conn(); cr = c.cursor()
                cr.execute("SELECT full_name FROM users WHERE user_id=%s", (user_id,))
                ism = (cr.fetchone() or ["O'quvchi"])[0]
                cr.execute("SELECT parent_id FROM parent_child WHERE child_id=%s", (user_id,))
                otalar = [r[0] for r in cr.fetchall()]
                cr.close(); c.close()
            except Exception:
                ism = "O'quvchi"; otalar = []

            xabar = f"📊 {ism}\n📝 {imt['nomi']}\n🎯 Natija: {foiz}% · {d}"
            for kim in [imt["teacher_id"]] + otalar:
                if kim:
                    try: await bot.send_message(_tg_id(kim), xabar)
                    except Exception: pass
        return

    # ═══════════════════════════════════════════
    # 👨‍🏫 IMTIHON BOSHQARUVI (im_)
    # ═══════════════════════════════════════════
    if call.data.startswith("im_"):
        import baholash as _bh
        _bh.jadval()
        qism = call.data.split(":")
        amal = qism[0]

        # ── Imtihonlar ro'yxati ──
        if amal == "im_menu":
            tgid = int(qism[1])
            lst = _bh.imtihonlar(tgid)
            t = ["📊 <b>Imtihonlar</b>\n"]
            rows = []
            for iid, nomi, turi, sana, n in lst:
                belgi = "✍️" if turi == "yozma" else "🧪"
                t.append(f"{belgi} {nomi} — {n} ta baho")
                rows.append([InlineKeyboardButton(
                    text=f"{belgi} {nomi[:28]} ({n})", callback_data=f"im_ko:{iid}")])
            if not lst:
                t.append("<i>Hali imtihon yo'q.</i>")
            rows.append([InlineKeyboardButton(text="➕ Yangi imtihon", callback_data=f"im_new:{tgid}")])
            rows.append([InlineKeyboardButton(text="🏆 Reyting", callback_data=f"im_reyt:{tgid}")])
            rows.append([InlineKeyboardButton(text="⬅️ Orqaga", callback_data=f"tg_info:{tgid}")])
            try:
                await call.message.edit_text("\n".join(t), parse_mode="HTML",
                                             reply_markup=InlineKeyboardMarkup(inline_keyboard=rows))
            except Exception:
                await call.message.answer("\n".join(t), parse_mode="HTML",
                                          reply_markup=InlineKeyboardMarkup(inline_keyboard=rows))
            return

        # ── Yangi imtihon: turini tanlash ──
        if amal == "im_new":
            tgid = int(qism[1])
            await call.message.answer(
                "➕ <b>Yangi imtihon</b>\n\nTurini tanlang:",
                parse_mode="HTML",
                reply_markup=InlineKeyboardMarkup(inline_keyboard=[
                    [InlineKeyboardButton(text="✍️ Yozma — o'zim baho qo'yaman",
                                          callback_data=f"im_tur:{tgid}:yozma")],
                    [InlineKeyboardButton(text="🧪 Test — bot baholaydi",
                                          callback_data=f"im_tur:{tgid}:test")],
                    [InlineKeyboardButton(text="⬅️ Orqaga", callback_data=f"im_menu:{tgid}")],
                ]))
            return

        if amal == "im_tur":
            tgid = int(qism[1]); turi = qism[2]
            admin_state[user_id] = f"im_nom:{tgid}:{turi}"
            await call.message.answer(
                f"{'✍️ Yozma' if turi=='yozma' else '🧪 Test'} imtihon\n\n"
                f"Imtihon nomini yozing:\nMasalan: <code>1-chorak yakuniy</code>",
                parse_mode="HTML")
            return

        # ── Test imtihoni uchun mavzu manbasi ──
        if amal == "im_manba":
            tgid = int(qism[1]); iid = int(qism[2]); manba = qism[3]
            import togarak_test as _tt
            ro_yxat = _tt.mavzular(tgid)
            ochiq = [m for m in ro_yxat if m[3]]
            if not ochiq:
                await call.answer("❌ Ochiq mavzu yo'q", show_alert=True); return

            if manba == "random":
                kodlar = sorted({k for m in ochiq for k in m[1]})
                c = _get_db_conn(); cr = c.cursor()
                cr.execute("UPDATE togarak_imtihon SET topic_codes=%s WHERE id=%s",
                           (kodlar, iid))
                c.commit(); cr.close(); c.close()
                await _im_elon(call, tgid, iid, bot)
                return

            # O'qituvchi tanlaydi — mavzu ro'yxati
            rows = [[InlineKeyboardButton(text=f"📗 {m[0][:34]} ({m[2]})",
                     callback_data=f"im_mv:{tgid}:{iid}:{i}")]
                    for i, m in enumerate(ro_yxat) if m[3]]
            rows.append([InlineKeyboardButton(text="⬅️ Orqaga", callback_data=f"im_menu:{tgid}")])
            await call.message.answer("Qaysi mavzudan?",
                                      reply_markup=InlineKeyboardMarkup(inline_keyboard=rows))
            return

        if amal == "im_mv":
            tgid = int(qism[1]); iid = int(qism[2]); i = int(qism[3])
            import togarak_test as _tt
            ro_yxat = _tt.mavzular(tgid)
            if i >= len(ro_yxat):
                await call.answer("❌ Topilmadi", show_alert=True); return
            kodlar = ro_yxat[i][1]
            c = _get_db_conn(); cr = c.cursor()
            cr.execute("UPDATE togarak_imtihon SET topic_codes=%s WHERE id=%s", (kodlar, iid))
            c.commit(); cr.close(); c.close()
            await _im_elon(call, tgid, iid, bot)
            return

        # ── Imtihonni ko'rish / baholash ──
        if amal == "im_ko":
            iid = int(qism[1])
            imt = _bh.imtihon_ol(iid)
            if not imt:
                await call.answer("❌ Topilmadi", show_alert=True); return
            nat = _bh.natijalar(iid)
            belgi = "✍️" if imt["turi"] == "yozma" else "🧪"
            t = [f"{belgi} <b>{imt['nomi']}</b>\n"]
            for i, (uid2, ism, foiz, manba) in enumerate(nat, 1):
                t.append(f"{i}. {ism or uid2} — <b>{float(foiz):.0f}%</b> {_bh.daraja(float(foiz))}")
            if not nat:
                t.append("<i>Hali baho qo'yilmagan.</i>")

            rows = []
            if imt["turi"] == "yozma":
                rows.append([InlineKeyboardButton(text="✍️ Baho qo'yish",
                                                  callback_data=f"im_bal:{iid}")])
            else:
                rows.append([InlineKeyboardButton(text="📢 Qayta e'lon qilish",
                                                  callback_data=f"im_elon:{iid}")])
            rows.append([InlineKeyboardButton(text="⬅️ Orqaga",
                                              callback_data=f"im_menu:{imt['togarak_id']}")])
            await call.message.answer("\n".join(t)[:3800], parse_mode="HTML",
                                      reply_markup=InlineKeyboardMarkup(inline_keyboard=rows))
            return

        # ── Yozma: o'quvchi tanlash ──
        if amal == "im_bal":
            iid = int(qism[1])
            imt = _bh.imtihon_ol(iid)
            if not imt: return
            c = _get_db_conn(); cr = c.cursor()
            cr.execute("""SELECT a.user_id, COALESCE(u.full_name,'—')
                FROM togarak_azolar a LEFT JOIN users u ON u.user_id=a.user_id
                WHERE a.togarak_id=%s AND a.aktiv=TRUE""", (imt["togarak_id"],))
            azolar = cr.fetchall(); cr.close(); c.close()
            mavjud = {r[0]: float(r[2]) for r in _bh.natijalar(iid)}
            rows = []
            for uid2, ism in azolar:
                b = f" — {mavjud[uid2]:.0f}%" if uid2 in mavjud else ""
                rows.append([InlineKeyboardButton(text=f"👤 {ism[:30]}{b}",
                                                  callback_data=f"im_u:{iid}:{uid2}")])
            rows.append([InlineKeyboardButton(text="⬅️ Orqaga", callback_data=f"im_ko:{iid}")])
            await call.message.answer("O'quvchini tanlang:",
                                      reply_markup=InlineKeyboardMarkup(inline_keyboard=rows))
            return

        if amal == "im_u":
            iid = int(qism[1]); uid2 = int(qism[2])
            admin_state[user_id] = f"im_foiz:{iid}:{uid2}"
            await call.message.answer(
                "Foizni yozing (1–100):\nMasalan: <code>85</code>", parse_mode="HTML")
            return

        if amal == "im_elon":
            iid = int(qism[1])
            imt = _bh.imtihon_ol(iid)
            if imt:
                await _im_elon(call, imt["togarak_id"], iid, bot)
            return

        # ── Reyting ──
        if amal == "im_reyt":
            tgid = int(qism[1])
            r = _bh.reyting(tgid)
            if not r:
                await call.answer("❌ A'zo yo'q", show_alert=True); return
            medal = ["🥇", "🥈", "🥉"]
            t = ["🏆 <b>To'garak reytingi</b>",
                 "<i>imtihon 80% + vazifa 10% + test 10%</i>\n"]
            for i, (uid2, ism, yak, imt, vaz, tst) in enumerate(r):
                m = medal[i] if i < 3 else f"{i+1}."
                t.append(f"{m} <b>{ism}</b> — {yak}%")
                t.append(f"     🎓 {imt if imt is not None else '—'}"
                         f" · 📝 {vaz}% · 🧪 {tst}%")
            t.append(f"\n<i>Jami {len(r)} o'quvchi</i>")
            await call.message.answer("\n".join(t)[:3800], parse_mode="HTML",
                reply_markup=InlineKeyboardMarkup(inline_keyboard=[[
                    InlineKeyboardButton(text="⬅️ Orqaga", callback_data=f"im_menu:{tgid}")]]))
            return

        # ── O'quvchi imtihonni boshlaydi ──
        if amal == "im_start":
            iid = int(qism[1])
            imt = _bh.imtihon_ol(iid)
            if not imt or not imt["aktiv"]:
                await call.answer("❌ Imtihon yopilgan", show_alert=True); return
            nat = {r[0] for r in _bh.natijalar(iid)}
            if user_id in nat:
                await call.answer("✅ Siz allaqachon topshirgansiz", show_alert=True); return
            savollar = _bh.savollar_ol(imt["topic_codes"], imt["savol_soni"])
            if not savollar:
                await call.answer("❌ Savol topilmadi", show_alert=True); return
            _bh.seans_boshla(user_id, iid, savollar)
            try: await call.message.edit_reply_markup(reply_markup=None)
            except Exception: pass
            s = savollar[0]
            matn = _bh.savol_matni(s, 0, len(savollar))
            kb = _bh.savol_kb(iid, 0)
            if s[6]:
                try:
                    await call.message.answer_photo(s[6], caption=matn, reply_markup=kb); return
                except Exception: pass
            await call.message.answer(matn, reply_markup=kb)
            return
        return

    # ═══ TEST O'CHIRISH TASDIG'I ═══
    if call.data == "ochir_no":
        _ochir_soqi.pop(user_id, None)
        await call.answer("❌ Bekor qilindi")
        try: await call.message.edit_text("❌ O'chirish bekor qilindi.")
        except Exception: pass
        return

    if call.data in ("ochir_yes", "ochir_yes_rasm"):
        if not _is_admin(user_id):
            await call.answer("❌ Ruxsat yo'q", show_alert=True); return
        so = _ochir_soqi.pop(user_id, None)
        if not so:
            await call.answer("⚠️ So'rov eskirgan, qaytadan yozing", show_alert=True); return
        shart, args, izoh, kutilgan = so
        rasm_ham = (call.data == "ochir_yes_rasm")
        await call.answer("🗑 O'chirilmoqda...")
        try:
            c = _get_db_conn(); cr = c.cursor()

            nrasm = 0
            if rasm_ham:
                # Avval rasm kodlarini yig'amiz (testlar o'chishidan oldin)
                cr.execute(f"""SELECT DISTINCT image_url FROM generated_tests
                    WHERE {shart} AND image_url IS NOT NULL AND image_url<>''""", tuple(args))
                kodlar = [r[0] for r in cr.fetchall()]
                if kodlar:
                    cr.execute("DELETE FROM images WHERE name = ANY(%s)", (kodlar,))
                    nrasm = cr.rowcount
                    cr.execute("DELETE FROM rasm_tavsif WHERE image_id = ANY(%s)", (kodlar,))

            cr.execute(f"DELETE FROM generated_tests WHERE {shart}", tuple(args))
            n = cr.rowcount
            c.commit(); cr.close(); c.close()
            print(f"[ochir] {izoh} -> {n} test, {nrasm} rasm o'chirildi")

            xabar = f"✅ <b>{n} ta test o'chirildi</b>\n📍 {izoh}\n"
            if rasm_ham:
                xabar += f"🖼 {nrasm} ta rasm ham o'chirildi\n"
                xabar += "\n<i>Yangi Excel yuklasangiz rasmlar qaytadan chiziladi.</i>"
            else:
                xabar += "\n<i>Topic kodlar va rasmlar joyida.</i>"
            await call.message.edit_text(xabar, parse_mode="HTML")
        except Exception as e:
            await call.message.edit_text(f"❌ Xato: {e}")
        return

    if call.data == "ochrasm_no":
        await call.answer("❌ Bekor qilindi")
        try: await call.message.edit_text("❌ Bekor qilindi.")
        except Exception: pass
        return

    if call.data == "ochrasm_yes":
        if not _is_admin(user_id):
            await call.answer("❌ Ruxsat yo'q", show_alert=True); return
        await call.answer("🗑 O'chirilmoqda...")
        try:
            c = _get_db_conn(); cr = c.cursor()
            cr.execute("""UPDATE generated_tests SET image_file_id=NULL
                WHERE image_file_id IS NOT NULL AND image_file_id<>''""")
            n = cr.rowcount
            cr.execute("DELETE FROM images")
            c.commit(); cr.close(); c.close()
            await call.message.edit_text(
                f"✅ <b>{n} ta rasm o'chirildi</b>\nTestlar joyida.\n\n"
                f"<i>Excel qayta yuklasangiz rasmlar qayta chiziladi.</i>",
                parse_mode="HTML")
        except Exception as e:
            await call.message.edit_text(f"❌ Xato: {e}")
        return

    if call.data == "rsmlat":
        await call.answer("⏸ Keyinroq")
        try: await call.message.edit_text("⏸ Rasm chizish kechiktirildi.\nErtaga yana so'rayman.")
        except Exception: pass
        return

    if call.data == "rsmres":
        await call.answer("▶️ Davom etmoqda...")
        try: await call.message.edit_reply_markup(reply_markup=None)
        except Exception: pass
        asyncio.create_task(_auto_generate_images(call.from_user.id, resume=True))
        return

    if call.data == "test_stop":
        await call.answer()
        await call.message.answer(
            "🛑 Testni to'xtatmoqchimisiz?",
            reply_markup=InlineKeyboardMarkup(inline_keyboard=[[
                InlineKeyboardButton(text="✅ Ha, to'xtat", callback_data="test_stop_yes"),
                InlineKeyboardButton(text="❌ Yo'q, davom", callback_data="test_stop_no"),
            ]])
        )
        return

    if call.data == "test_stop_yes":
        await call.answer("✅ To'xtatildi")
        try: await call.message.delete()
        except: pass
        await stop_test(call.from_user.id, call.message)
        # Menyuni qaytaramiz
        try:
            conn9=_get_db_conn();cur9=conn9.cursor()
            cur9.execute("SELECT role FROM users WHERE user_id=%s",(call.from_user.id,))
            r9=cur9.fetchone();cur9.close();conn9.close()
            await call.message.answer("🏠 Bosh menyu", reply_markup=get_main_keyboard(r9[0] if r9 else ""))
        except Exception: pass
        return

    if call.data == "test_stop_no":
        await call.answer("▶️ Davom etilmoqda!")
        try: await call.message.delete()
        except: pass
        return


async def _ota_ekran(source, parent_id, child_id, bolim):
    """Ota-ona uchun farzand ekranlari."""
    import ota_panel as _op
    import ota_ona as _oo

    if not _oo.bogliqmi(parent_id, child_id):
        await _javob(source, "❌ Bu farzand sizga ulanmagan.")
        return

    info = _op.farzand_info(child_id)
    bosh = f"👤 <b>{info['ism']}</b> {info['sinf']}\n"

    if bolim == "asosiy":
        naz = _op.nazorat(child_id)
        t = [bosh]
        if naz:
            for y in naz:
                t.append(f"\n📚 <b>{y['togarak']}</b>")
                t.append(f"   🎯 Yakuniy: <b>{y['yakuniy']}%</b> {y['daraja']}")
                if y["orin"]:
                    t.append(f"   🏆 O'rin: {y['orin']}/{y['jami']}")
        else:
            t.append("\n<i>Hali to'garakka a'zo emas.</i>")
        kb = InlineKeyboardMarkup(inline_keyboard=[
            [InlineKeyboardButton(text="📋 Yo'qlama", callback_data=f"op_yoqlama:{child_id}"),
             InlineKeyboardButton(text="📝 Uy vazifasi", callback_data=f"op_vazifa:{child_id}")],
            [InlineKeyboardButton(text="🎓 Imtihonlar", callback_data=f"op_imtihon:{child_id}"),
             InlineKeyboardButton(text="⭐ Baholar", callback_data=f"op_baho:{child_id}")],
            [InlineKeyboardButton(text="📊 Nazorat", callback_data=f"op_nazorat:{child_id}")],
        ])
        await _javob(source, "\n".join(t)[:3800], kb)
        return

    if bolim == "yoqlama":
        y = _op.yoqlama(child_id)
        t = [bosh, "📋 <b>Davomat</b>\n"]
        if y:
            for nomi, keldi, jami, foiz in y:
                belgi = "🟢" if foiz >= 80 else ("🟡" if foiz >= 60 else "🔴")
                t.append(f"{belgi} {nomi}")
                t.append(f"   {keldi}/{jami} dars · {foiz}%")
        else:
            t.append("<i>Yo'qlama ma'lumoti yo'q.</i>")
        await _javob(source, "\n".join(t)[:3800], _ota_orqaga(child_id))
        return

    if bolim == "vazifa":
        v = _op.vazifalar(child_id)
        t = [bosh, "📝 <b>Uy vazifalari</b>\n"]
        if v:
            bajardi = sum(1 for x in v if x[3])
            t.append(f"✅ Topshirdi: {bajardi}/{len(v)}\n")
            for nomi, mavzu, dl, ok in v[:12]:
                belgi = "✅" if ok else "⏳"
                sana = f" · {dl}" if dl else ""
                t.append(f"{belgi} {mavzu[:38]}{sana}")
        else:
            t.append("<i>Vazifa yo'q.</i>")
        await _javob(source, "\n".join(t)[:3800], _ota_orqaga(child_id))
        return

    if bolim == "imtihon":
        im = _op.imtihonlar(child_id)
        t = [bosh, "🎓 <b>Imtihon natijalari</b>\n"]
        if im:
            for nomi, foiz, turi, sana in im:
                f = float(foiz or 0)
                belgi = "🏆" if f >= 90 else ("⭐" if f >= 75 else ("👍" if f >= 60 else "📖"))
                tur = "✍️" if turi == "yozma" else "🧪"
                s = sana.strftime("%d.%m") if sana else ""
                t.append(f"{belgi} {tur} {nomi[:28]} — <b>{f:.0f}%</b>  {s}")
        else:
            t.append("<i>Hali imtihon topshirmagan.</i>")
        await _javob(source, "\n".join(t)[:3800], _ota_orqaga(child_id))
        return

    if bolim == "baho":
        b = _op.baholar(child_id)
        tn = _op.test_natijalari(child_id)
        t = [bosh, "⭐ <b>Baholar</b>\n"]
        if b:
            t.append("👨‍🏫 O'qituvchi bahosi:")
            for nomi, baho, sana in b[:10]:
                s = sana.strftime("%d.%m") if hasattr(sana, "strftime") else ""
                t.append(f"   {baho} · {nomi[:24]} {s}")
        if tn:
            t.append("\n🧪 Mustaqil testlar:")
            for fan, foiz, _ in tn[:8]:
                t.append(f"   {fan[:24]} — {int(foiz or 0)}%")
        if not b and not tn:
            t.append("<i>Baho yo'q.</i>")
        await _javob(source, "\n".join(t)[:3800], _ota_orqaga(child_id))
        return

    if bolim == "nazorat":
        naz = _op.nazorat(child_id)
        t = [bosh, "📊 <b>Umumiy nazorat</b>\n"]
        if naz:
            for y in naz:
                t.append(f"📚 <b>{y['togarak']}</b>")
                t.append(f"   🎯 Yakuniy: <b>{y['yakuniy']}%</b> {y['daraja']}")
                imt = y["imtihon"] if y["imtihon"] is not None else "—"
                t.append(f"   🎓 Imtihon: {imt} <i>(80%)</i>")
                t.append(f"   📝 Vazifa: {y['vazifa']}% <i>(10%)</i>")
                t.append(f"   🧪 Test: {y['test']}% <i>(10%)</i>")
                if y["orin"]:
                    t.append(f"   🏆 Reyting: {y['orin']}/{y['jami']}")
                t.append("")
        else:
            t.append("<i>To'garak yo'q.</i>")
        await _javob(source, "\n".join(t)[:3800], _ota_orqaga(child_id))
        return


def _ota_orqaga(child_id):
    return InlineKeyboardMarkup(inline_keyboard=[[
        InlineKeyboardButton(text="⬅️ Orqaga", callback_data=f"op_asosiy:{child_id}")]])


async def _javob(source, matn, kb=None):
    """Message yoki CallbackQuery ga javob."""
    try:
        if hasattr(source, "message") and source.message is not None:
            try:
                await source.message.edit_text(matn, parse_mode="HTML", reply_markup=kb)
            except Exception:
                await source.message.answer(matn, parse_mode="HTML", reply_markup=kb)
        else:
            await source.answer(matn, parse_mode="HTML", reply_markup=kb)
    except Exception as e:
        print(f"[ota_ekran] {e}")


async def _im_elon(call, tgid, iid, bot):
    """Test imtihonini a'zolarga e'lon qiladi."""
    import baholash as _bh
    imt = _bh.imtihon_ol(iid)
    if not imt:
        return
    try:
        c = _get_db_conn(); cr = c.cursor()
        cr.execute("SELECT user_id FROM togarak_azolar WHERE togarak_id=%s AND aktiv=TRUE",
                   (tgid,))
        azolar = [r[0] for r in cr.fetchall()]
        cr.close(); c.close()
    except Exception as e:
        print(f"[im_elon] {e}"); azolar = []

    kb = InlineKeyboardMarkup(inline_keyboard=[[
        InlineKeyboardButton(text="▶️ Imtihonni boshlash", callback_data=f"im_start:{iid}")]])
    n = 0
    for uid2 in azolar:
        try:
            await bot.send_message(
                _tg_id(uid2),
                f"📢 <b>Yangi imtihon</b>\n\n"
                f"📝 {imt['nomi']}\n"
                f"🧪 {imt['savol_soni']} ta savol\n\n"
                f"Tayyor bo'lsangiz boshlang:",
                parse_mode="HTML", reply_markup=kb)
            n += 1
        except Exception:
            pass
    await call.message.answer(f"📢 Imtihon {n} ta o'quvchiga yuborildi.")


async def _health_server():
    """Railway health check uchun oddiy HTTP server."""
    from aiohttp import web
    async def health(request):
        return web.Response(text="OK", status=200)
    app = web.Application()
    app.router.add_get("/health", health)
    app.router.add_get("/", health)
    runner = web.AppRunner(app)
    await runner.setup()
    site = web.TCPSite(runner, "0.0.0.0", 8080)
    await site.start()
    print("✅ Health server port 8080 da ishga tushdi")

# ═══ BRAIN message handler (AI yordamchi) ═══
@dp.message()
async def brain_handler(message: Message, state: FSMContext):
    uid = message.from_user.id if message.from_user else 0
    if uid in ADMINS: return
    if not message.text: return
    if user_state.get(uid) in ("text_answer", "in_test"): return
    if user_state.get(uid) != "ai_mode": return
    if message.text.startswith("/"): return
    menu_buttons = {
        "🎯 Bugungi reja","📚 Bilimni mustahkamlash","🧪 Bilimni sinash",
        "📈 Rivojlanishim","🌍 Hamjamiyat","👤 Kabinet","📚 To'garaklar",
        "🎨 Rasm chizdir","🤖 Yordamchi",
    }
    if message.text in menu_buttons: return
    try:
        from brain import process_message as _brain
        conn_ = _get_db_conn(); cur_ = conn_.cursor()
        cur_.execute("SELECT class, role FROM users WHERE user_id=%s", (uid,))
        row_ = cur_.fetchone(); cur_.close(); conn_.close()
        grade_ = str(row_[0]) if row_ and row_[0] else None
        role_ = str(row_[1]) if row_ and row_[1] else ""
        res = await _brain(message.text, uid, grade_, role=role_)
        if res.get("message"):
            await message.answer(res["message"])
        if res.get("action") == "START_TEST" and res.get("topic"):
            conn2 = _get_db_conn(); cur2 = conn2.cursor()
            cur2.execute("""
                SELECT question,option_a,option_b,option_c,option_d,
                       correct_answer,explanation,question_type,is_latex,
                       COALESCE(NULLIF(image_file_id,''), image_url) AS image_url,audio_text,language,time_limit
                FROM generated_tests WHERE topic_code=%s ORDER BY RANDOM() LIMIT 20
            """, (res["topic"]["topic_code"],))
            tests_ = cur2.fetchall(); cur2.close(); conn2.close()
            if tests_:
                from test_engine import start_test
                await start_test(uid, tests_, message)
    except Exception as e:
        print(f"brain xato: {e}")

async def main():
    print("BOT ISHGA TUSHDI 🚀")
    init_db()
    # Barcha state larni tozalash
    user_state.clear()
    from storage import lesson_state, temp_user, registration_message, reg_kbd_message
    lesson_state.clear()
    temp_user.clear()
    registration_message.clear()
    reg_kbd_message.clear()
    # Health server
    try:
        asyncio.create_task(_health_server())
        asyncio.create_task(_daily_image_resume())
        try:
            import ts_cache
            _n = ts_cache.tozala(7)
            if _n: print(f"[ts_cache] {_n} ta eski yozuv o'chirildi")
        except Exception as _e:
            print(f"[ts_cache] {_e}")
        try:
            import baholash
            baholash.jadval()
            print("[baholash] jadvallar tayyor")
        except Exception as _e:
            print(f"[baholash] {_e}")
        try:
            import ota_ona
            ota_ona.jadval()
            print("[ota_ona] jadvallar tayyor")
        except Exception as _e:
            print(f"[ota_ona] {_e}")
        try:
            import akkaunt
            akkaunt.jadval()      # uid ustuni + migratsiya
            print("[akkaunt] jadvallar tayyor")
        except Exception as _e:
            print(f"[akkaunt] {_e}")
    except Exception as _he:
        print(f"Health server xato: {_he}")
    # Foydalanuvchilar jimgina davom etaveradi (xabar yuborilmaydi)
    # Haftalik hisobot — har dushanba
    async def weekly_report_task():
        while True:
            import asyncio
            from datetime import datetime
            now = datetime.now()
            # Har dushanba 08:00 da
            if now.weekday() == 0 and now.hour == 8 and now.minute == 0:
                try:
                    from features import get_weekly_report
                    conn_ = _get_db_conn(); cur_ = conn_.cursor()
                    # Barcha ota-onalarga hisobot
                    cur_.execute("""SELECT p.parent_id, p.child_id, u.full_name,
                        a.togarak_id, t.nomi FROM parent_child p
                        JOIN users u ON u.user_id=p.child_id
                        JOIN togarak_azolar a ON a.user_id=p.child_id AND a.aktiv=TRUE
                        JOIN togaraklar t ON t.id=a.togarak_id AND t.aktiv=TRUE""")
                    rows_ = cur_.fetchall(); cur_.close(); conn_.close()
                    for r in rows_:
                        rep = get_weekly_report(r[3], r[1])
                        txt = (f"📊 Haftalik hisobot\n👤 {r[2]}\n📚 {r[4]}\n\n"
                               f"📋 Davomat: {rep['keldi']} kun keldi\n"
                               f"⭐ O'rt.baho: {rep['avg_baho']}\n"
                               f"📝 Vazifa: {rep['hw_done']} ta topshirdi")
                        try: await bot.send_message(r[0], txt)
                        except: pass
                except Exception as e:
                    print(f"weekly_report: {e}")
            await asyncio.sleep(60)
    # Extra routerlar ulash
    try:
        from router_extra import router as extra_router
        dp.include_router(extra_router)
    except Exception as e:
        print(f'router_extra: {e}')
    try:
        from handler_parent import router as parent_router
        dp.include_router(parent_router)
    except Exception as e:
        print(f'handler_parent: {e}')
    asyncio.create_task(weekly_report_task())
    try:
        from router_extra import tolov_eslatma_task
        asyncio.create_task(tolov_eslatma_task(bot))
    except: pass

    # Webhook yoki polling
    WEBHOOK_URL = os.getenv("WEBHOOK_URL","")
    if WEBHOOK_URL:
        from aiogram.webhook.aiohttp_server import SimpleRequestHandler, setup_application
        from aiohttp import web as _web2
        WEBHOOK_PATH = f"/webhook/{bot.token}"
        await bot.set_webhook(f"{WEBHOOK_URL}{WEBHOOK_PATH}")
        print(f"✅ Webhook: {WEBHOOK_URL}{WEBHOOK_PATH}")
        app2 = _web2.Application()
        SimpleRequestHandler(dispatcher=dp, bot=bot).register(app2, path=WEBHOOK_PATH)
        setup_application(app2, dp, bot=bot)
        runner2 = _web2.AppRunner(app2)
        await runner2.setup()
        site2 = _web2.TCPSite(runner2, "0.0.0.0", 8080)
        await site2.start()
        await asyncio.Event().wait()
    else:
        await dp.start_polling(bot)

if __name__ == "__main__":
    asyncio.run(main())
