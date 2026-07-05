"""rasim_generator.py — HF Space + Pollinations zaxira"""
import os, asyncio, aiohttp, psycopg2, re, urllib.parse

DATABASE_URL = os.getenv("DATABASE_URL","")
OPENAI_KEY   = os.getenv("OPENAI_API_KEY","")
HF_SPACE_URL = "https://azizbekahmat-rasm2.hf.space"

STYLES = {
    "multik":   "cute cartoon style, Disney Pixar, vibrant colors, child-friendly",
    "hayotiy":  "realistic photo style, photorealistic, natural lighting",
    "chizma":   "hand-drawn pencil sketch, educational diagram",
    "akvarell": "watercolor painting, soft pastel colors",
    "darslik":  "textbook scientific illustration, educational diagram",
    "3d":       "3D render, CGI, vibrant colors",
    "komiks":   "comic book style, bold outlines, bright colors",
}

def get_quality_suffix(style="multik"):
    return f"{STYLES.get(style,STYLES['multik'])}, white background, no text, high quality"

UZ_EN = {
    "och ko'k":"light blue","to'q ko'k":"dark blue","och yashil":"light green",
    "yorqin":"bright","och":"light","kulrang":"gray","jigarrang":"brown",
    "binafsha":"purple","qizil":"red","ko'k":"blue","yashil":"green",
    "sariq":"yellow","oq":"white","qora":"black","pushti":"pink","oltin":"golden",
    "fon":"background","maydon":"field","osmon":"sky","o'rtada":"in center",
    "pastda":"below","yuqorida":"above","quyosh":"sun","oy":"moon",
    "yulduz":"star","bulut":"cloud","yomg'ir":"rain","qor":"snow",
    "tog'":"mountain","daryo":"river","dengiz":"sea","daraxt":"tree",
    "gul":"flower","plyaj":"beach","yoz":"summer","qish":"winter",
    "bahor":"spring","kuz":"autumn","bola":"child","bolalar":"children",
    "qiz":"girl","o'g'il":"boy","o'quvchi":"student","o'qituvchi":"teacher",
    "ona":"mother","ota":"father","oila":"family",
    "o'tirmoqda":"sitting","turmoqda":"standing","yugurmoqda":"running",
    "o'ynamoqda":"playing","o'qimoqda":"reading","yozmoqda":"writing",
    "charaqlab":"shining","mushuk":"cat","it":"dog","baliq":"fish",
    "qush":"bird","ot":"horse","sigir":"cow","kapalak":"butterfly",
    "olma":"apple","banan":"banana","uzum":"grapes","nok":"pear",
    "tarvuz":"watermelon","sabzi":"carrot","pomidor":"tomato",
    "maktab":"school","sinf":"classroom","kitob":"book","qalam":"pencil",
    "uy":"house","shahar":"city","mashina":"car","avtobus":"bus",
    "futbol":"football","suzish":"swimming","katta":"big","kichik":"small",
}

def uz_to_en(tavsif):
    t = tavsif.lower()
    for uz, en in sorted(UZ_EN.items(), key=lambda x: -len(x[0])):
        t = t.replace(uz, en)
    t = re.sub(r"[^\x00-\x7F]", " ", t)
    t = re.sub(r"\s+", " ", t).strip()
    return t

async def _tavsif_to_prompt(tavsif, fan="ta'lim", sinf="1", style="multik"):
    gkey = os.getenv("GEMINI_API_KEY","")
    if gkey:
        try:
            url = f"https://generativelanguage.googleapis.com/v1beta/models/gemini-1.5-flash:generateContent?key={gkey}"
            body = {"contents":[{"parts":[{"text":
                f"Translate to English for AI image generation. "
                f"Grade {sinf} {fan} educational, child-friendly cartoon. "
                f"Description: {tavsif}\n"
                f"Return ONLY the English prompt, max 20 words."}]}],
                "generationConfig":{"maxOutputTokens":60}}
            async with aiohttp.ClientSession() as s:
                async with s.post(url,json=body,timeout=aiohttp.ClientTimeout(total=8)) as r:
                    if r.status==200:
                        d=await r.json()
                        eng=d["candidates"][0]["content"]["parts"][0]["text"].strip()
                        return f"{eng}, {get_quality_suffix(style)}"
        except: pass
    eng = uz_to_en(tavsif)
    return f"{eng}, grade {sinf} educational, {get_quality_suffix(style)}"

async def generate_hf_space(prompt):
    """HF Space Gradio API."""
    try:
        url = f"{HF_SPACE_URL}/api/predict"
        async with aiohttp.ClientSession() as s:
            async with s.post(url, json={"data":[prompt]},
                             timeout=aiohttp.ClientTimeout(total=90)) as r:
                if r.status==200:
                    data=await r.json()
                    result=data.get("data",[{}])
                    if result:
                        val=result[0]
                        if isinstance(val,dict):
                            img_url=val.get("url") or val.get("path","")
                            if img_url:
                                if not img_url.startswith("http"):
                                    img_url=f"{HF_SPACE_URL}/file={img_url}"
                                async with s.get(img_url) as ir:
                                    if ir.status==200: return await ir.read()
                        elif isinstance(val,str) and val.startswith("http"):
                            async with s.get(val) as ir:
                                if ir.status==200: return await ir.read()
    except Exception as e:
        print(f"HF Space: {e}")
    return None

async def generate_hf(prompt, style="multik"):
    """HF Space → Pollinations zaxira."""
    clean = re.sub(r"[^\x00-\x7F]","",prompt)
    clean = re.sub(r"\s+"," ",clean).strip()[:250]

    img = await generate_hf_space(clean)
    if img and len(img)>1000: return img

    url=(f"https://image.pollinations.ai/prompt/{urllib.parse.quote(clean)}"
         f"?width=768&height=512&nologo=true&model=flux")
    for _ in range(3):
        try:
            async with aiohttp.ClientSession() as s:
                async with s.get(url,timeout=aiohttp.ClientTimeout(total=45)) as r:
                    if r.status==200:
                        data=await r.read()
                        if len(data)>1000: return data
                    elif r.status==429: await asyncio.sleep(10)
        except: await asyncio.sleep(3)
    return None

async def generate_dalle(prompt):
    if not OPENAI_KEY: return None
    try:
        async with aiohttp.ClientSession() as s:
            async with s.post("https://api.openai.com/v1/images/generations",
                json={"model":"dall-e-3","prompt":prompt,"n":1,"size":"1024x1024"},
                headers={"Authorization":f"Bearer {OPENAI_KEY}"},
                timeout=aiohttp.ClientTimeout(total=60)) as r:
                if r.status==200:
                    d=await r.json()
                    async with s.get(d["data"][0]["url"]) as ir:
                        return await ir.read() if ir.status==200 else None
    except: pass
    return None

async def generate_image(question, fan="ta'lim", sinf="1", style="multik"):
    prompt=await _tavsif_to_prompt(question,fan,sinf,style)
    img=await generate_hf(prompt,style)
    if not img: img=await generate_dalle(prompt)
    return img

async def generate_and_save(topic_code,question,fan,sinf,img_num,bot,chat_id,style="multik"):
    name=f"{topic_code}-{img_num}"
    try:
        conn=psycopg2.connect(DATABASE_URL);cur=conn.cursor()
        cur.execute("SELECT file_id FROM images WHERE name=%s",(name,))
        row=cur.fetchone();cur.close();conn.close()
        if row: return row[0]
    except: pass
    img_bytes=await generate_image(question,fan,sinf,style)
    if not img_bytes: return None
    try:
        from aiogram.types import BufferedInputFile
        sent=await bot.send_photo(chat_id,BufferedInputFile(img_bytes,f"{name}.png"),caption=f"🖼 {name}")
        fid=sent.photo[-1].file_id
        conn=psycopg2.connect(DATABASE_URL);cur=conn.cursor()
        cur.execute("INSERT INTO images(name,file_id) VALUES(%s,%s) ON CONFLICT(name) DO UPDATE SET file_id=EXCLUDED.file_id",(name,fid))
        conn.commit();cur.close();conn.close()
        return fid
    except Exception as e:
        print(f"Saqlash: {e}")
    return None

async def generate_from_excel(excel_bytes,bot,chat_id,progress_cb=None,style="multik",force=False):
    import openpyxl
    from io import BytesIO
    async def p(msg):
        if progress_cb: await progress_cb(msg)
    wb=openpyxl.load_workbook(BytesIO(excel_bytes),data_only=True)
    if "RASMLAR" not in wb.sheetnames:
        return {"error":"RASMLAR varog'i topilmadi!"}
    ws_r=wb["RASMLAR"]
    topic_info={}
    if "MALUMOT" in wb.sheetnames:
        ws_m=wb["MALUMOT"]
        for r in range(2,ws_m.max_row+1):
            tc=ws_m.cell(r,2).value;sinf=ws_m.cell(r,3).value;fan=ws_m.cell(r,4).value
            if tc: topic_info[str(tc)]={"sinf":str(sinf or "1"),"fan":str(fan or "ta'lim")}
    items=[]
    for r in range(2,ws_r.max_row+1):
        kod=ws_r.cell(r,1).value;mav=ws_r.cell(r,2).value;tavs=ws_r.cell(r,3).value
        if kod and tavs:
            tc="-".join(str(kod).split("-")[:-1])
            info=topic_info.get(tc,{})
            items.append({"kod":str(kod),"tavsif":str(tavs),"sinf":info.get("sinf","1"),"fan":info.get("fan","ta'lim")})
    if not items: return {"error":"Ma'lumot yo'q!"}
    existing=set()
    if not force:
        try:
            kods=[it["kod"] for it in items]
            conn=psycopg2.connect(DATABASE_URL);cur=conn.cursor()
            ph=",".join(["%s"]*len(kods))
            cur.execute(f"SELECT name FROM images WHERE name IN ({ph})",kods)
            existing={r[0] for r in cur.fetchall()};cur.close();conn.close()
        except: pass
    todo=[it for it in items if it["kod"] not in existing]
    skipped=len(existing)
    await p(f"📊 Jami: {len(items)} | Yangi: {len(todo)} | O'tkazildi: {skipped}")
    created=errors=0
    for i,item in enumerate(todo,1):
        try:
            prompt=await _tavsif_to_prompt(item["tavsif"],item["fan"],item["sinf"],style)
            img=await generate_hf(prompt,style)
            if not img: img=await generate_dalle(prompt)
            if img:
                from aiogram.types import BufferedInputFile
                sent=await bot.send_photo(chat_id,BufferedInputFile(img,f"{item['kod']}.png"),caption=f"🖼 {item['kod']}")
                fid=sent.photo[-1].file_id
                c2=psycopg2.connect(DATABASE_URL);cu2=c2.cursor()
                cu2.execute("INSERT INTO images(name,file_id) VALUES(%s,%s) ON CONFLICT(name) DO UPDATE SET file_id=EXCLUDED.file_id",(item["kod"],fid))
                c2.commit();cu2.close();c2.close();created+=1
            else: errors+=1
        except Exception as e:
            errors+=1;print(f"rasm: {e}")
        if i%5==0 or i==len(todo):
            await p(f"⏳ {round(i*100/len(todo))}% ({i}/{len(todo)}) | ✅{created} | ❌{errors}")
        await asyncio.sleep(3)
    await p(f"🎉 Tayyor!\n✅ {created} | ⏭ {skipped} | ❌ {errors}")
    for attempt in range(2):
        if errors==0: break
        try:
            all_kods=[it["kod"] for it in todo]
            conn=psycopg2.connect(DATABASE_URL);cur=conn.cursor()
            ph=",".join(["%s"]*len(all_kods))
            cur.execute(f"SELECT name FROM images WHERE name IN ({ph})",all_kods)
            done_kods={r[0] for r in cur.fetchall()};cur.close();conn.close()
            retry=[it for it in todo if it["kod"] not in done_kods]
        except: break
        if not retry: break
        await p(f"🔄 {attempt+2}-urinish: {len(retry)} ta...")
        await asyncio.sleep(5)
        r_c=r_e=0
        for i,item in enumerate(retry,1):
            try:
                prompt=await _tavsif_to_prompt(item["tavsif"],item["fan"],item["sinf"],style)
                img=await generate_hf(prompt,style)
                if not img: img=await generate_dalle(prompt)
                if img:
                    from aiogram.types import BufferedInputFile
                    sent=await bot.send_photo(chat_id,BufferedInputFile(img,f"{item['kod']}.png"),caption=f"🖼 {item['kod']}")
                    fid=sent.photo[-1].file_id
                    c2=psycopg2.connect(DATABASE_URL);cu2=c2.cursor()
                    cu2.execute("INSERT INTO images(name,file_id) VALUES(%s,%s) ON CONFLICT(name) DO UPDATE SET file_id=EXCLUDED.file_id",(item["kod"],fid))
                    c2.commit();cu2.close();c2.close();r_c+=1;created+=1
                else: r_e+=1
            except: r_e+=1
            if i%5==0 or i==len(retry):
                await p(f"🔄{attempt+2}: {round(i*100/len(retry))}% | ✅{r_c} | ❌{r_e}")
            await asyncio.sleep(3)
        errors=r_e
    await p(f"🏁 Hammasi tugadi!\n✅ {created} ta | ❌ {errors} ta")
    return {"created":created,"skipped":skipped,"errors":errors}

async def generate_topic_images(topic_code,fan,sinf,count=20,bot=None,chat_id=0,progress_cb=None,style="multik"):
    async def p(msg):
        if progress_cb: await progress_cb(msg)
    conn=psycopg2.connect(DATABASE_URL);cur=conn.cursor()
    cur.execute("SELECT id,question FROM generated_tests WHERE topic_code=%s ORDER BY id LIMIT %s",(topic_code,count))
    tests=cur.fetchall();cur.close();conn.close()
    created=skipped=errors=0
    for i,(tid,question) in enumerate(tests,1):
        await p(f"🖼 {i}/{len(tests)}...")
        try:
            fid=await generate_and_save(topic_code,question,fan,sinf,i,bot,chat_id,style)
            if fid:
                c2=psycopg2.connect(DATABASE_URL);cu2=c2.cursor()
                cu2.execute("UPDATE generated_tests SET image_url=%s WHERE id=%s",(f"{topic_code}-{i}",tid))
                c2.commit();cu2.close();c2.close();created+=1
            else: skipped+=1
        except: errors+=1
        await asyncio.sleep(3)
    await p(f"✅{created} | ⏭{skipped} | ❌{errors}")
    return {"created":created,"skipped":skipped,"errors":errors}

async def auto_generate_subject_images(fan,bot,chat_id,progress_cb=None):
    return 0
