"""
jadval_generator.py — Jadval va hisobot tizimi
Excel + Telegram inline ko'rinish
"""
import os, io, psycopg2
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import (Font, PatternFill, Alignment,
                              Border, Side)
from openpyxl.utils import get_column_letter

DATABASE_URL = os.getenv("DATABASE_URL")

def db():
    return psycopg2.connect(DATABASE_URL)

# ── Stil yordamchilar ──
def hcell(ws, row, col, val, color="2E86AB", bold=True, width=None):
    c = ws.cell(row, col, val)
    c.font      = Font(bold=bold, color="FFFFFF", size=11)
    c.fill      = PatternFill("solid", fgColor=color)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border    = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"),  bottom=Side(style="thin")
    )
    if width and hasattr(ws, 'column_dimensions'):
        ws.column_dimensions[get_column_letter(col)].width = width
    return c

def dcell(ws, row, col, val, bg=None, center=False):
    c = ws.cell(row, col, val)
    if bg:
        c.fill = PatternFill("solid", fgColor=bg)
    c.alignment = Alignment(
        horizontal="center" if center else "left",
        vertical="center", wrap_text=True
    )
    c.border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"),  bottom=Side(style="thin")
    )
    return c

def pct_color(pct):
    if pct >= 90: return "C6EFCE"   # Yashil
    if pct >= 70: return "FFEB9C"   # Sariq
    if pct >= 50: return "FFCC99"   # To'q sariq
    return "FFC7CE"                  # Qizil

# ══════════════════════════════════════
# 1. TEST NATIJALARI JADVALI
# ══════════════════════════════════════
def test_results_excel(fan=None, sinf=None, days=30) -> io.BytesIO:
    conn = db(); cur = conn.cursor()
    date_from = datetime.now() - timedelta(days=days)
    fan_f  = "AND d.subject_name=%s" if fan else ""
    sinf_f = "AND u.class=%s" if sinf else ""
    params = [date_from]
    if fan:  params.append(fan)
    if sinf: params.append(sinf)

    cur.execute(f"""
        SELECT u.full_name, u.class,
               d.subject_name, d.kichik_name,
               COUNT(*) as jami,
               SUM(CASE WHEN ts.correct THEN 1 ELSE 0 END) as togri,
               ROUND(100.0*SUM(CASE WHEN ts.correct THEN 1 ELSE 0 END)/COUNT(*)) as pct,
               MIN(ts.created_at)::date as sana
        FROM test_history ts
        JOIN users u ON u.user_id = ts.user_id
        LEFT JOIN dts_tree d ON d.topic_code = ts.topic_code
        WHERE ts.created_at >= %s {fan_f} {sinf_f}
        GROUP BY u.full_name, u.class, d.subject_name, d.kichik_name, ts.created_at::date
        ORDER BY u.class, u.full_name, ts.created_at::date DESC
    """, params)
    rows = cur.fetchall()
    cur.close(); conn.close()

    wb = Workbook(); ws = wb.active
    ws.title = "Test natijalari"
    ws.row_dimensions[1].height = 40

    headers = ["O'quvchi", "Sinf", "Fan", "Mavzu",
               "Jami savol", "To'g'ri", "Natija %", "Sana"]
    widths   = [25, 8, 20, 30, 12, 10, 12, 12]
    for i, (h, w) in enumerate(zip(headers, widths), 1):
        hcell(ws, 1, i, h, width=w)

    for r, row in enumerate(rows, 2):
        pct  = row[6] or 0
        bg   = pct_color(pct) if r % 2 == 0 else None
        for c, val in enumerate(row, 1):
            center = c in (2, 5, 6, 7, 8)
            if c == 7:
                dcell(ws, r, c, f"{val}%", pct_color(pct), True)
            else:
                dcell(ws, r, c, val, bg, center)

    # Umumiy statistika
    if rows:
        ws.append([])
        total_row = len(rows) + 3
        hcell(ws, total_row, 1, "JAMI STATISTIKA", "1B4F72", width=25)
        avg_pct = sum(r[6] or 0 for r in rows) / len(rows)
        dcell(ws, total_row, 7, f"{avg_pct:.1f}%", pct_color(avg_pct), True)

    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return buf

def test_results_text(fan=None, sinf=None, days=30) -> str:
    """Bot ichida ko'rsatish uchun matn."""
    conn = db(); cur = conn.cursor()
    date_from = datetime.now() - timedelta(days=days)
    params = [date_from]
    if fan:  params.append(fan)
    if sinf: params.append(sinf)
    fan_f  = "AND d.subject_name=%s" if fan else ""
    sinf_f = "AND u.class=%s" if sinf else ""

    cur.execute(f"""
        SELECT u.full_name, u.class,
               ROUND(100.0*SUM(CASE WHEN ts.correct THEN 1 ELSE 0 END)/COUNT(*)) as pct,
               COUNT(*) as jami
        FROM test_history ts
        JOIN users u ON u.user_id = ts.user_id
        LEFT JOIN dts_tree d ON d.topic_code = ts.topic_code
        WHERE ts.created_at >= %s {fan_f} {sinf_f}
        GROUP BY u.full_name, u.class
        ORDER BY pct DESC LIMIT 20
    """, params)
    rows = cur.fetchall()
    cur.close(); conn.close()

    if not rows:
        return "📊 Hali test ma'lumotlari yo'q."

    lines = [f"📊 Test natijalari (oxirgi {days} kun)\n"]
    for i, (name, cls, pct, total) in enumerate(rows, 1):
        icon = "🥇" if i==1 else "🥈" if i==2 else "🥉" if i==3 else "📌"
        bar  = "█"*int((pct or 0)//10) + "░"*(10-int((pct or 0)//10))
        lines.append(f"{icon} {name} ({cls}-sinf)\n"
                     f"   {bar} {pct}% | {total} ta test")
    return "\n".join(lines)

# ══════════════════════════════════════
# 2. DARS REJASI JADVALI
# ══════════════════════════════════════
def lesson_plan_excel(sinf: str, fan: str, weeks: int = 4) -> io.BytesIO:
    conn = db(); cur = conn.cursor()
    cur.execute("""
        SELECT DISTINCT d.mavzu_name, d.kichik_name, d.topic_code,
               EXISTS(SELECT 1 FROM teacher_lessons tl WHERE tl.topic_code=d.topic_code) as has_lesson,
               EXISTS(SELECT 1 FROM generated_tests gt WHERE gt.topic_code=d.topic_code) as has_test
        FROM dts_tree d
        WHERE d.grade=%s AND d.subject_name=%s AND d.is_deleted=FALSE
        ORDER BY d.topic_code
        LIMIT %s
    """, (sinf, fan, weeks*5))
    topics = cur.fetchall()
    cur.close(); conn.close()

    wb = Workbook(); ws = wb.active
    ws.title = f"{sinf}-sinf {fan} rejasi"

    # Sarlavha
    ws.merge_cells("A1:H1")
    c = ws.cell(1, 1, f"{sinf}-sinf | {fan} | {weeks} haftalik dars rejasi")
    c.font      = Font(bold=True, size=14, color="FFFFFF")
    c.fill      = PatternFill("solid", fgColor="1A5276")
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 35

    headers = ["Hafta", "Kun", "Mavzu", "Kichik mavzu",
               "Dars", "Test", "Topik kodi", "Izoh"]
    widths   = [8, 8, 25, 30, 8, 8, 28, 20]
    for i, (h, w) in enumerate(zip(headers, widths), 1):
        hcell(ws, 2, i, h, "2E86AB", width=w)

    days_uz = ["Dushanba","Seshanba","Chorshanba","Payshanba","Juma"]
    week_colors = ["EBF5FB","E9F7EF","FEF9E7","FDEDEC","F4ECF7"]

    row = 3
    for w_num in range(1, weeks+1):
        for d_num, day in enumerate(days_uz):
            idx  = (w_num-1)*5 + d_num
            if idx >= len(topics): break
            t = topics[idx]
            bg = week_colors[(w_num-1) % 5]
            dcell(ws, row, 1, f"{w_num}-hafta", bg, True)
            dcell(ws, row, 2, day, bg, True)
            dcell(ws, row, 3, t[0], bg)                    # mavzu
            dcell(ws, row, 4, t[1], bg)                    # kichik_mavzu
            dcell(ws, row, 5, "✅" if t[3] else "❌", bg, True)
            dcell(ws, row, 6, "✅" if t[4] else "❌", bg, True)
            dcell(ws, row, 7, t[2], bg)                    # topic_code
            dcell(ws, row, 8, "", bg)
            ws.row_dimensions[row].height = 20
            row += 1

    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return buf

# ══════════════════════════════════════
# 3. O'QUVCHI TARAQQIYOT JADVALI
# ══════════════════════════════════════
def student_progress_excel(sinf: str = None) -> io.BytesIO:
    conn = db(); cur = conn.cursor()
    sinf_f = "WHERE u.class=%s" if sinf else ""
    cur.execute(f"""
        SELECT u.user_id, u.full_name, u.class,
               COUNT(DISTINCT lp.topic_code) as darslar,
               COALESCE(ts_stats.togri, 0) as togri,
               COALESCE(ts_stats.jami, 0) as jami,
               COALESCE(ROUND(100.0*ts_stats.togri/NULLIF(ts_stats.jami,0)),0) as pct
        FROM users u
        LEFT JOIN lesson_progress lp ON lp.user_id = u.user_id
        LEFT JOIN (
            SELECT user_id,
                   SUM(CASE WHEN correct THEN 1 ELSE 0 END) togri,
                   COUNT(*) jami
            FROM test_history GROUP BY user_id
        ) ts_stats ON ts_stats.user_id = u.user_id
        {sinf_f}
        GROUP BY u.user_id, u.full_name, u.class,
                 ts_stats.togri, ts_stats.jami
        ORDER BY u.class, pct DESC NULLS LAST
    """, ([sinf] if sinf else []))
    rows = cur.fetchall()
    cur.close(); conn.close()

    wb = Workbook(); ws = wb.active
    ws.title = "O'quvchi taraqqiyoti"

    headers = ["#", "O'quvchi", "Sinf", "O'tilgan darslar",
               "To'g'ri javob", "Jami test", "Natija %", "Daraja"]
    widths   = [5, 25, 8, 18, 14, 12, 12, 15]
    for i, (h, w) in enumerate(zip(headers, widths), 1):
        hcell(ws, 1, i, h, width=w)

    for r, row in enumerate(rows, 2):
        pct   = row[6] or 0
        bg    = pct_color(pct)
        grade = "🏆 A'lo" if pct>=90 else "👍 Yaxshi" if pct>=70 else "📈 O'rta" if pct>=50 else "📚 Past"
        dcell(ws, r, 1, r-1,    center=True)
        dcell(ws, r, 2, row[1])
        dcell(ws, r, 3, row[2], center=True)
        dcell(ws, r, 4, row[3], center=True)
        dcell(ws, r, 5, row[4], center=True)
        dcell(ws, r, 6, row[5], center=True)
        dcell(ws, r, 7, f"{pct}%", bg, True)
        dcell(ws, r, 8, grade,  bg)

    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return buf

def student_progress_text(sinf: str = None) -> str:
    conn = db(); cur = conn.cursor()
    sinf_f = "WHERE u.class=%s" if sinf else ""
    cur.execute(f"""
        SELECT u.full_name, u.class,
               COUNT(DISTINCT lp.topic_code) as darslar,
               COALESCE(ROUND(100.0*SUM(CASE WHEN th.correct THEN 1 ELSE 0 END)/
                              NULLIF(COUNT(th.id),0)),0) as pct
        FROM users u
        LEFT JOIN lesson_progress lp ON lp.user_id=u.user_id
        LEFT JOIN test_history th ON th.user_id=u.user_id
        {sinf_f}
        GROUP BY u.full_name, u.class
        ORDER BY pct DESC LIMIT 15
    """, ([sinf] if sinf else []))
    rows = cur.fetchall()
    cur.close(); conn.close()
    if not rows:
        return "📈 Taraqqiyot ma'lumotlari yo'q."
    lines = ["📈 O'quvchi taraqqiyoti\n"]
    medals = ["🥇","🥈","🥉"]
    for i, (name, cls, darslar, pct) in enumerate(rows, 1):
        icon = medals[i-1] if i<=3 else "📌"
        bar  = "█"*int((pct or 0)//10) + "░"*(10-int((pct or 0)//10))
        lines.append(f"{icon} {name} ({cls}-sinf)\n"
                     f"   {bar} {pct}% | {darslar} dars")
    return "\n".join(lines)

# ══════════════════════════════════════
# 4. TAHLIL — ZAIF MAVZULAR
# ══════════════════════════════════════
def weak_analysis_text(fan: str = None) -> str:
    try:
        conn = db(); cur = conn.cursor()
        fan_f = "WHERE fan=%s" if fan else ""
        cur.execute(f"""
            SELECT mavzu, fan, error_count
            FROM weak_topics {fan_f}
            ORDER BY error_count DESC LIMIT 10
        """, ([fan] if fan else []))
        rows = cur.fetchall()
        cur.close(); conn.close()
    except: return "📊 Ma'lumot yo'q."

    if not rows:
        return "✅ Zaif mavzular yo'q — hammasi yaxshi!"
    lines = ["⚠️ Eng ko'p xato qilingan mavzular:\n"]
    for mavzu, fan_, cnt in rows:
        bar = "🔴"*min(cnt//2,5) + "⚪"*(5-min(cnt//2,5))
        lines.append(f"{bar} {mavzu} ({fan_}) — {cnt} ta xato")
    return "\n".join(lines)
