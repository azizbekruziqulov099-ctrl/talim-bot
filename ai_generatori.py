"""
ai_generator.py — AI bilan test generatsiya qilish
OpenAI GPT-4o-mini orqali har mavzu uchun 20 ta savol
"""
import psycopg2, os, json, asyncio
from aiogram.types import InlineKeyboardMarkup, InlineKeyboardButton
from openai_client import client

DATABASE_URL = os.getenv("DATABASE_URL")

def db(): return psycopg2.connect(DATABASE_URL)

# Generator holati
gen_state = {}

def _gen_settings_kb(groups):
    DIFFS = [("oson","🟢"),("orta","🟡"),("qiyin","🔴"),("murakkab","⚫")]
    rows = []
    for g in groups:
        d = g["diff"]; icon = next(em for df,em in DIFFS if df==d)
        cnt = g["count"]; tp = g["type"]
        rows.append([
            InlineKeyboardButton(text=f"{icon} {d.capitalize()}", callback_data="noop"),
            InlineKeyboardButton(text="✅0" if cnt==0  else "0",  callback_data=f"gg_cnt_{d}_0"),
            InlineKeyboardButton(text="✅5" if cnt==5  else "5",  callback_data=f"gg_cnt_{d}_5"),
            InlineKeyboardButton(text="✅10" if cnt==10 else "10", callback_data=f"gg_cnt_{d}_10"),
            InlineKeyboardButton(text="✅15" if cnt==15 else "15", callback_data=f"gg_cnt_{d}_15"),
            InlineKeyboardButton(text="✅20" if cnt==20 else "20", callback_data=f"gg_cnt_{d}_20"),
        ])
        rows.append([
            InlineKeyboardButton(
                text="✅ 🔘 Tugmali" if tp=="single_choice" else "🔘 Tugmali",
                callback_data=f"gg_tp_{d}_choice"
            ),
            InlineKeyboardButton(
                text="✅ ✍️ Yozuvli" if tp=="write_answer" else "✍️ Yozuvli",
                callback_data=f"gg_tp_{d}_write"
            ),
        ])
    total = sum(g["count"] for g in groups)
    rows.append([InlineKeyboardButton(
        text=f"🤖 AI bilan yaratish (jami: {total} ta mavzuga)",
        callback_data="gen_go"
    )])
    return InlineKeyboardMarkup(inline_keyboard=rows)

  # user_id -> {grade, subject, selected_topics, ...}


# ─────────────────────────────────────────
# 1. SINF TANLASH
# ─────────────────────────────────────────
async def show_gen_start(message, user_id):
    gen_state[user_id] = {}
    # DB dan barcha sinflarni olamiz (CEFR va boshqalar ham)
    conn = db(); cur = conn.cursor()
    cur.execute("""
        SELECT grade FROM (
            SELECT DISTINCT grade FROM dts_tree WHERE is_deleted=FALSE
        ) _g
        ORDER BY
            CASE WHEN grade ~ '^[0-9]+$' THEN grade::int ELSE 9999 END,
            grade
    """)
    grades = [r[0] for r in cur.fetchall()]
    cur.close(); conn.close()

    def btn_label(g):
        return f"{g}-sinf" if str(g).isdigit() else str(g)

    kb = InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text=btn_label(g), callback_data=f"gen_grade:{g}")
         for g in grades[i:i+4]]
        for i in range(0, len(grades), 4)
    ])
    await message.answer("🤖 AI Test Generator\n\nSinfni tanlang:", reply_markup=kb)


# ─────────────────────────────────────────
# 2. FAN TANLASH
# ─────────────────────────────────────────
async def show_subjects(call, user_id, grade):
    gen_state[user_id] = {"grade": grade, "selected": []}
    conn = db(); cur = conn.cursor()
    cur.execute("""
        SELECT DISTINCT subject_name FROM dts_tree
        WHERE grade=%s AND is_deleted=FALSE
        ORDER BY subject_name
    """, (grade,))
    subjects = [r[0] for r in cur.fetchall()]
    cur.close(); conn.close()

    kb = InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text=s, callback_data=f"gen_subj:{s[:40]}")]
        for s in subjects
    ] + [[InlineKeyboardButton(text="◀️ Orqaga", callback_data="gen_start")]])

    await call.message.edit_text(
        f"🤖 AI Generator\n🎓 {grade + ('-sinf' if str(grade).isdigit() else '')}\n\nFanni tanlang:",
        reply_markup=kb
    )


# ─────────────────────────────────────────
# 3. KICHIK MAVZULAR RO'YXATI (checkbox)
# ─────────────────────────────────────────
async def show_topics(call, user_id, subject):
    state = gen_state.get(user_id, {})
    grade = state.get("grade", "1")
    state["subject"] = subject
    gen_state[user_id] = state

    conn = db(); cur = conn.cursor()
    cur.execute("""
        SELECT topic_code, kichik_name, mavzu_name,
               (SELECT COUNT(*) FROM generated_tests WHERE topic_code=t.topic_code) as cnt
        FROM dts_tree t
        WHERE grade=%s AND subject_name=%s AND is_deleted=FALSE
        ORDER BY topic_code
    """, (grade, subject))
    topics = cur.fetchall()
    cur.close(); conn.close()

    state["topics"] = {t[0]: {"kichik": t[1], "mavzu": t[2], "cnt": t[3]} for t in topics}
    selected = state.get("selected", [])

    state["filter"] = "all"  # Default: barcha mavzular
    gen_state[user_id] = state
    await _render_topics(call.message, user_id, selected, grade, subject, edit=True)


async def _render_topics(message, user_id, selected, grade, subject, edit=False):
    state = gen_state.get(user_id, {})
    topics_list = state.get("topics", {})
    filt = state.get("filter", "empty")
    page = state.get("page", 0)
    PAGE_SIZE = 15

    # Filterlash
    if filt == "empty":
        filtered = {k: v for k, v in topics_list.items() if v["cnt"] == 0}
    else:
        filtered = topics_list

    items = list(filtered.items())
    total = len(items)
    page_items = items[page*PAGE_SIZE:(page+1)*PAGE_SIZE]

    empty_cnt = sum(1 for v in topics_list.values() if v["cnt"] == 0)
    all_cnt = len(topics_list)

    text = (
        f"🤖 AI Generator\n"
        f"🎓 {grade}-sinf | 📚 {subject}\n\n"
        f"🔴 Bo'sh: {empty_cnt} ta | 📊 Jami: {all_cnt} ta\n"
        f"✅ Tanlangan: {len(selected)} ta\n"
        f""
    )
    filter_label = "🔴 Faqat bo'sh mavzular" if filt == "empty" else "📋 Barcha mavzular"
    text += f"{filter_label} ({page*PAGE_SIZE+1}-{min((page+1)*PAGE_SIZE, total)}/{total}):"

    rows = []
    for code, info in page_items:
        is_sel = code in selected
        cnt = info["cnt"]
        name = info["kichik"][:22] if info["kichik"] else code[-15:]
        check = "✅" if is_sel else "⬜"
        status = f"✓{cnt}" if cnt > 0 else "✗0"
        rows.append([InlineKeyboardButton(
            text=f"{check} {name} [{status}]",
            callback_data=f"gen_toggle:{code}"
        )])

    # Navigatsiya
    nav = []
    if page > 0:
        nav.append(InlineKeyboardButton(text="◀️", callback_data=f"gen_page:{page-1}"))
    _btn = "📋 Barchasi" if filt == "empty" else "🔴 Bo'shlar"
    nav.append(InlineKeyboardButton(
        text=_btn,
        callback_data="gen_filter_toggle"
    ))
    if (page+1)*PAGE_SIZE < total:
        nav.append(InlineKeyboardButton(text="▶️", callback_data=f"gen_page:{page+1}"))
    rows.append(nav)

    bottom = [InlineKeyboardButton(text="◀️ Fan", callback_data="gen_subj_back")]
    if selected:
        bottom.append(InlineKeyboardButton(text="❌ Tozalash", callback_data="gen_clear"))
    rows.append(bottom)

    # Barchasini tanlash + Sozlamalar + Boshlash
    action_row = []
    if filt == "empty" and empty_cnt > 0:
        action_row.append(InlineKeyboardButton(
            text=f"☑️ Barcha bo'shlarni ({empty_cnt})",
            callback_data="gen_select_all"
        ))
    elif filt != "empty":
        action_row.append(InlineKeyboardButton(
            text=f"☑️ Barchasini ({all_cnt})",
            callback_data="gen_select_all"
        ))
    if action_row:
        rows.append(action_row)

    if selected:
        rows.append([InlineKeyboardButton(
            text=f"⚙️ Sozlamalar → 🚀 Boshlash ({len(selected)} mavzu)",
            callback_data="gen_settings"
        )])

    kb = InlineKeyboardMarkup(inline_keyboard=rows)

    if edit:
        try:
            await message.edit_text(text, reply_markup=kb)
        except:
            await message.answer(text, reply_markup=kb)
    else:
        await message.answer(text, reply_markup=kb)


# ─────────────────────────────────────────
# 4. TOGGLE — mavzu belgilash/olib tashlash
# ─────────────────────────────────────────
async def toggle_topic(call, user_id, code):
    state = gen_state.get(user_id, {})
    selected = state.get("selected", [])

    if code in selected:
        selected.remove(code)
    else:
        if False:  # limit olib tashlandi
            await call.answer("✅")
            return
        selected.append(code)

    state["selected"] = selected
    gen_state[user_id] = state
    grade = state.get("grade", "1")
    subject = state.get("subject", "")
    await _render_topics(call.message, user_id, selected, grade, subject, edit=True)
    await call.answer()


# ─────────────────────────────────────────
# 5. AI BILAN GENERATSIYA
# ─────────────────────────────────────────
async def run_generator(call, user_id):
    state = gen_state.get(user_id, {})
    selected = state.get("selected", [])
    topics_list = state.get("topics", {})
    grade = state.get("grade", "1")
    subject = state.get("subject", "")

    if not selected:
        await call.answer("❌ Mavzu tanlanmagan!", show_alert=True)
        return

    await call.answer()
    # Shablon rejimi — AI emas, bo'sh shablon
    await _generate_template(call, user_id, selected, topics_list, grade, subject)

    conn = db(); cur = conn.cursor()
    total_saved = 0
    errors = []

    # Status xabari
    status_msg = await call.message.answer(
        f"🤖 AI ishlamoqda...\n📚 {len(selected)} ta mavzu\n\n⏳ Tayyorlanmoqda..."
    )

    for idx, code in enumerate(selected):
        info = topics_list.get(code, {})
        kichik = info.get("kichik", code)
        mavzu = info.get("mavzu", "")

        try:
            # Status yangilash
            await status_msg.edit_text(
                f"🤖 AI ishlamoqda...\n"
                f"📚 {len(selected)} ta mavzu\n\n"
                f"{idx}/{len(selected)} ✅\n"
                f"⏳ Hozir: {kichik[:40]}"
            )

            # AI dan savol olish
            questions = await _generate_questions(grade, subject, mavzu, kichik, code)

            if questions:
                # DBga saqlash
                for q in questions:
                    cur.execute("""
                        INSERT INTO generated_tests
                        (topic_code, question, option_a, option_b, option_c, option_d,
                         correct_answer, explanation, question_type, is_latex,
                         image_url, audio_text, language, life_level, age_group,
                         time_limit, difficulty, situation)
                        VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
                        ON CONFLICT DO NOTHING
                    """, (
                        code,
                        q.get("question",""),
                        q.get("a",""), q.get("b",""), q.get("c",""), q.get("d",""),
                        q.get("correct",""),
                        q.get("explanation",""),
                        q.get("question_type","single_choice"),
                        False,
                        q.get("image_url"),
                        None, "uz", 1,
                        _age_group(grade),
                        q.get("time_limit", 60),
                        q.get("difficulty","oson"),
                        "oddiy"
                    ))
                    total_saved += 1
                conn.commit()

        except Exception as ex:
            errors.append(f"{kichik[:30]}: {str(ex)[:50]}")

        await asyncio.sleep(1)  # Rate limit

    cur.close(); conn.close()

    # Yakuniy hisobot
    err_text = ""
    if errors:
        err_text = f"\n\n⚠️ Xatolar ({len(errors)}):\n" + "\n".join(errors[:5])

    await status_msg.edit_text(
        f"✅ Generatsiya tugadi!\n\n"
        f"📊 Saqlangan: {total_saved} ta savol\n"
        f"📚 Mavzular: {len(selected)} ta{err_text}\n\n"
        f"📥 Excel yaratilmoqda..."
    )

    # Excel eksport — rasmli savollar ko'rinib tursin
    try:
        excel_buf = await _export_to_excel(selected, topics_list, grade, subject)
        from aiogram.types import BufferedInputFile
        await call.message.answer_document(
            BufferedInputFile(excel_buf, filename=f"savol_{grade}sinf_{subject[:10]}.xlsx"),
            caption=(
                "📋 Shablon tayyor!\n\n"
                + f"📚 {len(selected)} ta mavzu × 20 qator\n"
                + f"📊 Jami: {len(selected)*20} ta qator\n\n"
                + "✏️ Savol, javob, izohlarni to'ldiring\n"
                + "📥 Keyin: Test import qilish"
            )
        )
    except Exception as ex:
        await call.message.answer(f"Excel xato: {ex}")


async def _generate_template(call, user_id, selected, topics_list, grade, subject):
    import openpyxl, io
    from openpyxl.styles import Font, PatternFill, Alignment
    from aiogram.types import BufferedInputFile

    wb = openpyxl.Workbook()
    ws1 = wb.active
    ws1.title = "TESTLAR"

    headers = [
        "topic_code","difficulty","situation","question",
        "option_a","option_b","option_c","option_d",
        "correct_answer","explanation","question_type",
        "is_latex","image_url","audio_text","language",
        "life_level","age_group","time_limit"
    ]
    for col, h in enumerate(headers, 1):
        cell = ws1.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="2E86AB")
        cell.alignment = Alignment(horizontal="center")

    diff_colors = {
        "oson": "C8F7C5", "o'rta": "FFF3CD",
        "qiyin": "FFD7C4", "murakkab": "F5C6CB"
    }
    diff_times = {"oson": 60, "o'rta": 55, "qiyin": 50, "murakkab": 45}
    difficulties = ["oson"]*5 + ["o'rta"]*5 + ["qiyin"]*5 + ["murakkab"]*5

    conn = db(); cur = conn.cursor()
    row_num = 2

    for code in selected:
        cur.execute("SELECT grade FROM dts_tree WHERE topic_code=%s LIMIT 1", (code,))
        r = cur.fetchone()
        g = r[0] if r else grade
        age = _age_group(g)

        for i, diff in enumerate(difficulties):
            n = i + 1
            img_url = f"{code}-{n}"  # Hammaga rasm

            ws1.append([
                code, diff, "oddiy", "", "", "", "", "",
                "", "", "single_choice", False,
                img_url, "", "uz", 1, age, diff_times[diff]
            ])
            color = diff_colors[diff]
            for col in range(1, 19):
                ws1.cell(row_num, col).fill = PatternFill("solid", fgColor=color)
            row_num += 1

        ws1.append([""] * 18)
        row_num += 1

    cur.close(); conn.close()

    ws1.column_dimensions["A"].width = 28
    ws1.column_dimensions["D"].width = 55
    for c in ["E","F","G","H"]: ws1.column_dimensions[c].width = 22
    ws1.column_dimensions["I"].width = 22
    ws1.column_dimensions["J"].width = 35
    ws1.column_dimensions["M"].width = 28

    ws2 = wb.create_sheet("MALUMOT")
    ws2.append(["#","Topic code","Sinf","Fan","Chorak","Bob","Bolim","Mavzu","Kichik mavzu","Test soni"])
    for cell in ws2[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="2E86AB")

    conn2 = db(); cur2 = conn2.cursor()
    for idx, code in enumerate(selected, 1):
        cur2.execute("""
            SELECT grade,subject_name,quarter,bob_name,bolim_name,mavzu_name,kichik_name
            FROM dts_tree WHERE topic_code=%s LIMIT 1
        """, (code,))
        r = cur2.fetchone()
        cur2.execute("SELECT COUNT(*) FROM generated_tests WHERE topic_code=%s", (code,))
        cnt = cur2.fetchone()[0]
        if r:
            ws2.append([idx, code, r[0], r[1], r[2], r[3], r[4], r[5], r[6], cnt])
        else:
            ws2.append([idx, code, "","","","","","","", cnt])
    cur2.close(); conn2.close()

    for col, w in zip(["A","B","C","D","E","F","G","H","I","J"],
                      [4, 28, 6, 18, 8, 35, 35, 35, 35, 10]):
        ws2.column_dimensions[col].width = w

    buf = io.BytesIO()
    wb.save(buf); buf.seek(0)

    fname = f"shablon_{grade}sinf_{len(selected)}mavzu.xlsx"
    await call.message.answer_document(
        BufferedInputFile(buf.read(), filename=fname),
        caption=(
            "✅ Shablon tayyor!\n"
            + f"📚 {len(selected)} mavzu x 20 qator\n"
            + f"📊 Jami: {len(selected)*20} qator\n\n"
            + "Bo\'sh: savol, javoblar, izoh\n"
            + "To\'ldirib import qiling!"
        )
    )


def _age_group(grade):
    m = {"1":"6-7","2":"7-8","3":"8-9","4":"9-10","5":"10-11",
         "6":"11-12","7":"12-13","8":"13-14","9":"14-15","10":"15-16","11":"16-17"}
    return m.get(str(grade), "10-11")


async def _export_to_excel(selected, topics_list, grade, subject):
    """Saqlangan savollarni Excel ga chiqarish"""
    import openpyxl, io
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "SAVOLLAR"

    headers = [
        "topic_code", "kichik_mavzu", "difficulty", "question_type",
        "question", "option_a", "option_b", "option_c", "option_d",
        "correct_answer", "explanation", "image_url", "time_limit"
    ]

    # Header
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="2E86AB")
        cell.alignment = Alignment(horizontal="center")

    conn = db(); cur = conn.cursor()
    row_num = 2

    diff_colors = {
        "oson": "C8F7C5", "o'rta": "FFF3CD",
        "qiyin": "FFD7C4", "murakkab": "F5C6CB"
    }

    for code in selected:
        info = topics_list.get(code, {})
        kichik = info.get("kichik", code)

        cur.execute("""
            SELECT question, option_a, option_b, option_c, option_d,
                   correct_answer, explanation, image_url, time_limit,
                   difficulty, question_type
            FROM generated_tests
            WHERE topic_code=%s
            ORDER BY difficulty, id
        """, (code,))
        rows = cur.fetchall()

        for r in rows:
            q, a, b, c, d, correct, expl, img, tl, diff, qtype = r

            ws.cell(row_num, 1).value = code
            ws.cell(row_num, 2).value = kichik
            ws.cell(row_num, 3).value = diff
            ws.cell(row_num, 4).value = qtype
            ws.cell(row_num, 5).value = q
            ws.cell(row_num, 6).value = a
            ws.cell(row_num, 7).value = b
            ws.cell(row_num, 8).value = c
            ws.cell(row_num, 9).value = d
            ws.cell(row_num, 10).value = correct
            ws.cell(row_num, 11).value = expl
            ws.cell(row_num, 12).value = img
            ws.cell(row_num, 13).value = tl

            # Rang — qiyinlikka qarab
            color = diff_colors.get(diff, "FFFFFF")
            for col in range(1, 14):
                ws.cell(row_num, col).fill = PatternFill("solid", fgColor=color)

            # Rasmli savolni belgilash
            if img:
                ws.cell(row_num, 12).font = Font(bold=True, color="FF0000")

            row_num += 1

        # Mavzular orasiga bo'sh qator
        row_num += 1

    cur.close(); conn.close()

    # Ustun kengliklari
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['E'].width = 50
    for col in ['F','G','H','I','J']:
        ws.column_dimensions[col].width = 20
    ws.column_dimensions['K'].width = 40
    ws.column_dimensions['L'].width = 25

    # Rasmlar varag'i — faqat rasmli savollar
    ws2 = wb.create_sheet("RASMLI SAVOLLAR")
    ws2.append(["image_url", "kichik_mavzu", "savol"])

    conn2 = db(); cur2 = conn2.cursor()
    for code in selected:
        info = topics_list.get(code, {})
        kichik = info.get("kichik", code)
        cur2.execute("""
            SELECT image_url, question FROM generated_tests
            WHERE topic_code=%s AND image_url IS NOT NULL
            ORDER BY id
        """, (code,))
        for img_url, q in cur2.fetchall():
            ws2.append([img_url, kichik, q])
    cur2.close(); conn2.close()

    ws2['A1'].font = Font(bold=True)
    ws2['B1'].font = Font(bold=True)
    ws2['C1'].font = Font(bold=True)
    ws2.column_dimensions['A'].width = 30
    ws2.column_dimensions['B'].width = 30
    ws2.column_dimensions['C'].width = 60

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


async def _generate_questions(grade, subject, mavzu, kichik, topic_code):
    """OpenAI API orqali 20 ta savol yaratish"""

    # Yosh guruhi
    age = _age_group(grade)

    # Pedagogik yo'riqnoma
    from pedagogy import get_pedagogy, get_tag_examples
    pedagogy_note = get_pedagogy(grade, subject)
    tag_examples = get_tag_examples(grade, subject)

    # Darajaga mos qiyinlik tavsifi
    grade_int = int(grade) if str(grade).isdigit() else 1

    if grade_int <= 2:
        level_desc = f"""DARAJALAR ({grade}-sinf, {age} yosh — boshlang'ich):
- oson (5 ta, 60s): Rasmga qarab bitta so'z tanish. Eng oddiy. Masalan: "Qaysi bola baland?"
- o'rta (5 ta, 55s): Oddiy jumla to'ldirish yoki tarjima. Masalan: "Tom ... bo'yli." 
- qiyin (5 ta, 50s): 2-3 so'zli savol, tushunish kerak. Masalan: "Who has long hair?"
- murakkab (5 ta, 50s): Qisqa gap tuzish yoki tanlash. Masalan: "What does she look like?"
ESLATMA: {grade}-sinf uchun MURAKKAB ham oddiy bo'lsin — faqat 1-2 ta tushuncha birlashtirilsin!"""
    elif grade_int <= 5:
        level_desc = f"""DARAJALAR ({grade}-sinf, {age} yosh — o'rta bosqich):
- oson (5 ta, 60s): Asosiy tushunchani aniqlash
- o'rta (5 ta, 55s): Qo'llash, to'ldirish, oddiy tahlil
- qiyin (5 ta, 50s): Qoidani qo'llash, tushunish
- murakkab (5 ta, 50s): Amaliy vaziyat, sintez"""
    else:
        level_desc = f"""DARAJALAR ({grade}-sinf, {age} yosh — yuqori bosqich):
- oson (5 ta, 60s): Asosiy bilim
- o'rta (5 ta, 55s): Tushunish va qo'llash
- qiyin (5 ta, 50s): Tahlil va baholash
- murakkab (5 ta, 50s): Sintez, tanqidiy fikrlash"""

    prompt = f"""Siz {grade}-sinf {subject} fani o\'qituvchisisiz.
Fan: {subject} | Mavzu: {mavzu} | Kichik mavzu: {kichik}

{pedagogy_note}

{tag_examples}

QATIY QOIDALAR:
1. Savol ichida to\'g\'ri javob bo\'lmasin (qavsda ham emas!)
2. correct = to\'g\'ri javobning AYNAN matni (A/B/C/D emas!)
3. Faqat "{kichik}" mavzusiga oid savol yoz!
4. 4 javob bir-biriga o\'xshash bo\'lsin, faqat 1 tasi to\'g\'ri
5. 6 tasida has_image: true (oson va o\'rta darajada)
6. 2 tasida write_answer (oson 1 ta, o\'rta 1 ta)
7. write_answer da: a="", b="", c="", d=""

TARTIB: Avval 5 ta oson, keyin 5 ta o'rta, keyin 5 ta qiyin, keyin 5 ta murakkab!
Har daraja ichida TAKRORLANMASIN — har savol o'ziga xos bo'lsin!

FAQAT JSON (markdown, izoh, boshqa matn yozma):
[
  {{"question":"oson savol 1","a":"...","b":"...","c":"...","d":"...","correct":"...","explanation":"...","difficulty":"oson","time_limit":60,"question_type":"single_choice","has_image":true}},
  {{"question":"oson savol 2","a":"","b":"","c":"","d":"","correct":"...","explanation":"...","difficulty":"oson","time_limit":60,"question_type":"write_answer","has_image":false}},
  {{"question":"oson savol 3","a":"...","b":"...","c":"...","d":"...","correct":"...","explanation":"...","difficulty":"oson","time_limit":60,"question_type":"single_choice","has_image":false}},
  {{"question":"oson savol 4","a":"...","b":"...","c":"...","d":"...","correct":"...","explanation":"...","difficulty":"oson","time_limit":60,"question_type":"single_choice","has_image":true}},
  {{"question":"oson savol 5","a":"...","b":"...","c":"...","d":"...","correct":"...","explanation":"...","difficulty":"oson","time_limit":60,"question_type":"single_choice","has_image":true}},
  ... (keyin 5 ta o'rta, 5 ta qiyin, 5 ta murakkab)
]"""


    # OpenAI API
    import re
    from concurrent.futures import ThreadPoolExecutor

    def call_openai():
        response = client.chat.completions.create(
            model="gpt-4o",
            messages=[
                {"role": "system", "content": "Faqat to'g'ri JSON array qaytargin. AYNAN 20 ta element bo'lsin — kam bo'lmasin! Boshqa hech narsa yozma. Markdown, izoh, ```json kabi belgilar ishlatma."},
                {"role": "user", "content": prompt}
            ],
            max_tokens=8000,
            temperature=0.5
        )
        return response.choices[0].message.content

    loop = asyncio.get_event_loop()
    with ThreadPoolExecutor() as pool:
        text = await loop.run_in_executor(pool, call_openai)

    # JSON ni tozalab olish
    import re
    # Markdown code block larni olib tashlash
    text = re.sub(r'```json\s*', '', text)
    text = re.sub(r'```\s*', '', text)
    text = text.strip()

    # JSON array topish
    match = re.search(r'\[.*\]', text, re.DOTALL)
    if not match:
        raise ValueError(f"JSON topilmadi: {text[:300]}")

    json_str = match.group()
    # Nazorat belgilarini tozalash
    json_str = re.sub(r'[\x00-\x08\x0b\x0c\x0e-\x1f\x7f]', '', json_str)

    questions = json.loads(json_str)

    # Fieldlarni to'ldirish
    img_counter = 0
    for i, q in enumerate(questions):
        q.setdefault("difficulty", "oson")
        q.setdefault("time_limit", 60)
        q.setdefault("question_type", "single_choice")
        q.setdefault("has_image", False)

        # Rasmli savol uchun image_url
        if q.get("has_image"):
            img_counter += 1
            q["image_url"] = f"{topic_code}-{img_counter}"
        else:
            q["image_url"] = None

        # write_answer uchun javoblarni tozalash
        if q.get("question_type") == "write_answer":
            q["a"] = q["b"] = q["c"] = q["d"] = ""

    return questions


# ─────────────────────────────────────────
# CALLBACK HANDLER
# ─────────────────────────────────────────
async def handle_gen_callback(call, user_id):
    data = call.data

    if data == "gen_start":
        await call.answer()
        await show_gen_start(call.message, user_id)

    elif data.startswith("gen_grade:"):
        grade = data[10:]
        await call.answer()
        await show_subjects(call, user_id, grade)

    elif data.startswith("gen_subj:"):
        subject = data[9:]
        await call.answer()
        await show_topics(call, user_id, subject)

    elif data == "gen_subj_back":
        state = gen_state.get(user_id, {})
        grade = state.get("grade", "1")
        await call.answer()
        await show_subjects(call, user_id, grade)

    elif data.startswith("gen_toggle:"):
        code = data[11:]
        await toggle_topic(call, user_id, code)

    elif data == "gen_clear":
        state = gen_state.get(user_id, {})
        state["selected"] = []
        gen_state[user_id] = state
        grade = state.get("grade", "1")
        subject = state.get("subject", "")
        await _render_topics(call.message, user_id, [], grade, subject, edit=True)
        await call.answer("✅ Tozalandi")

    elif data.startswith("gen_page:"):
        page = int(data[9:])
        state = gen_state.get(user_id, {})
        state["page"] = page
        gen_state[user_id] = state
        selected = state.get("selected", [])
        grade = state.get("grade", "1")
        subject = state.get("subject", "")
        await _render_topics(call.message, user_id, selected, grade, subject, edit=True)
        await call.answer()

    elif data == "gen_filter_toggle":
        state = gen_state.get(user_id, {})
        state["filter"] = "all" if state.get("filter") == "empty" else "empty"
        state["page"] = 0
        gen_state[user_id] = state
        selected = state.get("selected", [])
        grade = state.get("grade", "1")
        subject = state.get("subject", "")
        await _render_topics(call.message, user_id, selected, grade, subject, edit=True)
        await call.answer()

    elif data == "gen_run":
        await run_generator(call, user_id)
    elif data == "gen_select_all":
        state = gen_state.get(user_id, {})
        grade   = state.get("grade", "1")
        subject = state.get("subject", "")
        filt    = state.get("filter", "empty")
        # DB dan barcha mavzularni olamiz
        conn2 = psycopg2.connect(DATABASE_URL); cur2 = conn2.cursor()
        cur2.execute("""
            SELECT d.topic_code,
                COUNT(g.id) as cnt
            FROM dts_tree d
            LEFT JOIN generated_tests g ON g.topic_code=d.topic_code
            WHERE d.grade=%s AND d.subject_name=%s AND d.is_deleted=FALSE
            GROUP BY d.topic_code
        """, (grade, subject))
        rows2 = cur2.fetchall(); cur2.close(); conn2.close()
        if filt == "empty":
            all_codes = [r[0] for r in rows2 if r[1] == 0]
        else:
            all_codes = [r[0] for r in rows2]
        state["selected"] = all_codes
        gen_state[user_id] = state
        await _render_topics(call.message, user_id, all_codes, grade, subject, edit=True)
        await call.answer(f"✅ {len(all_codes)} ta tanlandi")

    elif data == "gen_settings":
        state = gen_state.get(user_id, {})
        selected = state.get("selected", [])
        if not selected:
            await call.answer("❌ Mavzu tanlanmagan!", show_alert=True); return
        state.setdefault("gen_groups", [
            {"diff":"oson",    "type":"single_choice","count":5},
            {"diff":"orta",    "type":"single_choice","count":5},
            {"diff":"qiyin",   "type":"single_choice","count":5},
            {"diff":"murakkab","type":"single_choice","count":5},
        ])
        gen_state[user_id] = state
        await call.answer()
        await call.message.answer(
            f"⚙️ Sozlamalar\n📚 {len(selected)} ta mavzu tanlangan\n\n"
            f"Har qiyinlikdan nechta va qanday turdagi savol?",
            reply_markup=_gen_settings_kb(state["gen_groups"])
        )

    elif data.startswith("gg_cnt_") or data.startswith("gg_tp_"):
        state = gen_state.setdefault(user_id, {})
        groups = state.setdefault("gen_groups", [
            {"diff":"oson",    "type":"single_choice","count":5},
            {"diff":"orta",    "type":"single_choice","count":5},
            {"diff":"qiyin",   "type":"single_choice","count":5},
            {"diff":"murakkab","type":"single_choice","count":5},
        ])
        if data.startswith("gg_cnt_"):
            parts2 = data[7:].rsplit("_",1)
            d2, cnt2 = parts2[0], int(parts2[1])
            for g in groups:
                if g["diff"] == d2: g["count"] = cnt2
        elif data.startswith("gg_tp_"):
            parts2 = data[6:].rsplit("_",1)
            d2, tp2 = parts2[0], parts2[1]
            for g in groups:
                if g["diff"] == d2:
                    g["type"] = "single_choice" if tp2=="choice" else "write_answer"
        await call.answer("✅")
        try: await call.message.edit_reply_markup(reply_markup=_gen_settings_kb(groups))
        except: pass

    elif data == "gen_go":
        state = gen_state.get(user_id, {})
        selected = state.get("selected", [])
        groups   = state.get("gen_groups", [
            {"diff":"oson","type":"single_choice","count":5},
            {"diff":"orta","type":"single_choice","count":5},
            {"diff":"qiyin","type":"single_choice","count":5},
            {"diff":"murakkab","type":"single_choice","count":5},
        ])
        total_per = sum(g["count"] for g in groups)
        if not selected or total_per == 0:
            await call.answer("❌ Mavzu yoki son tanlanmagan!", show_alert=True); return
        await call.answer()
        # gen_groups ni gen_state ga saqlash
        state["groups"] = groups
        gen_state[user_id] = state
        await run_generator(call, user_id)
