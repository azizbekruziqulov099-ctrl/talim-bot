"""handler_excel.py — Excel shablon handlerlari"""
import psycopg2, asyncio, os, re
from io import BytesIO
from keyboards import get_main_keyboard
from aiogram.types import InlineKeyboardMarkup, InlineKeyboardButton, BufferedInputFile, FSInputFile
from aiogram.fsm.context import FSMContext
from storage import user_state, admin_state, temp_user
DATABASE_URL = os.getenv("DATABASE_URL","")
ADMINS = list(map(int, os.getenv("ADMINS","0").split(",")))
def _get_db_conn(): return psycopg2.connect(DATABASE_URL)

async def handle_excel_msg(message, user_id, admin_state, user_state, temp_user, bot):
    """True qaytarsa ishladi"""
    # ── Excel shablon yuborilsa — faqat "excel_merge" state da ──
    if (message.document and user_id in ADMINS and
            message.document.file_name and
            (message.document.file_name or "").endswith(".xlsx") and
            admin_state.get(user_id) == "excel_merge"):

        first_key = f"merge_file1_{user_id}"

        if first_key in admin_state:
            # Ikkinchi fayl keldi — birlashtirish
            file1_bytes = admin_state.pop(first_key)
            buf2 = BytesIO()
            await message.bot.download(message.document.file_id, destination=buf2)
            st = await message.answer("⏳ Birlashtirilyapti...")
            try:
                from excel_merger import merge_excel
                res_bytes, info = merge_excel(file1_bytes, buf2.getvalue())
                if res_bytes:
                    from aiogram.types import BufferedInputFile
                    await message.answer_document(
                        BufferedInputFile(res_bytes, "birlashtirilgan.xlsx"),
                        caption=info
                    )
                    await st.delete()
                else:
                    await st.edit_text(info)
            except Exception as e:
                await st.edit_text(f"❌ Xato: {e}")
        else:
            # Birinchi fayl — DB dan to'ldirishga urinish
            buf1 = BytesIO()
            await message.bot.download(message.document.file_id, destination=buf1)
            file1_bytes = buf1.getvalue()
            st = await message.answer("⏳ DB dan savollar qidirilmoqda...")
            try:
                from excel_merger import fill_from_db
                res_bytes, info = fill_from_db(file1_bytes)
                if res_bytes and "To'ldirildi" in info:
                    from aiogram.types import BufferedInputFile
                    fname_out = "toldiriilgan_" + (message.document.file_name or "shablon.xlsx")
                    await message.answer_document(
                        BufferedInputFile(res_bytes, fname_out),
                        caption=info
                    )
                    await st.delete()
                else:
                    admin_state[first_key] = file1_bytes
                    await st.edit_text(
                        f"📎 {message.document.file_name or 'fayl'} saqlandi.\n"
                        f"DB da savol topilmadi.\n\n"
                        f"2-fayl yuboring (savol fayli) — birlashtiraman!"
                    )
            except Exception as e:
                admin_state[first_key] = file1_bytes
                await st.edit_text(
                    f"📎 {message.document.file_name or 'fayl'} saqlandi.\n"
                    f"2-fayl yuboring — birlashtiraman!"
                )
        return

    if (
        admin_state.get(user_id) == "test_import"
        and message.document
    ):

        await prepare_test_import(message, user_id)

        return

    if user_state.get(message.from_user.id) == "in_test":
        # Test paytida matn yozsa — yumshoq eslatma
        try:
            await message.answer(
                "🧪 Test davom etyapti!\n"
                "Javob berish uchun tugmalardan foydalaning.\n\n"
                "Chiqish: /menu"
            )
        except: pass
        return

    if user_state.get(message.from_user.id) == "in_lesson":
        # Dars paytida matn yozsa — yumshoq eslatma
        try:
            await message.answer(
                "📖 Dars davom etyapti!\n"
                "Oldinga o'tish uchun tugmalardan foydalaning.\n\n"
                "Chiqish: /menu"
            )
        except: pass
        return

    if user_state.get(message.from_user.id) == "text_answer":

        await check_text_answer(
            message.from_user.id,
            message.text,
            message
        )

        return

    if message.text == "👥 Foydalanuvchilar statistikasi":

        if message.from_user.id not in ADMINS:
            return

        conn = _get_db_conn()
        cur = conn.cursor()

        cur.execute("SELECT COUNT(*) FROM users")
        total = cur.fetchone()[0]

        cur.execute("""
        SELECT COUNT(*)
        FROM users
        WHERE DATE(last_seen)=CURRENT_DATE
        """)
        today = cur.fetchone()[0]

        cur.execute("""
        SELECT COUNT(*)
        FROM users
        WHERE last_seen >= NOW() - INTERVAL '30 day'
        """)
        month = cur.fetchone()[0]

        conn.close()

        await message.answer(
            f"👥 Jami foydalanuvchilar: {total}\n\n"
            f"📅 Bugun kirganlar: {today}\n"
            f"🗓 Oxirgi 30 kun: {month}"
        )

        return

    elif message.text == "🧪 Test sinovi":
        if user_id not in ADMINS:
            return
        from test_sinovi import show_sinov_start
        await show_sinov_start(message, user_id)
        return

    elif message.text == "👥 Foydalanuvchilar":
        if user_id not in ADMINS:
            return
        conn2 = _get_db_conn()
        cur2 = conn2.cursor()
        cur2.execute("SELECT COUNT(*) FROM users")
        total = cur2.fetchone()[0]
        cur2.execute("SELECT role, COUNT(*) FROM users GROUP BY role ORDER BY COUNT(*) DESC")
        roles = cur2.fetchall()
        cur2.close(); conn2.close()
        text = f"👥 Foydalanuvchilar: {total} ta\n\n"
        for role, cnt in roles:
            text += f"• {role}: {cnt} ta\n"
        await message.answer(text)
        return

    elif message.text == "🖼 Rasmlar boshqaruvi":
        if user_id not in ADMINS: return
        # 2 xil rejim
        await message.answer(
            "🖼 Rasmlar boshqaruvi",
            reply_markup=InlineKeyboardMarkup(inline_keyboard=[
                [InlineKeyboardButton(text="📤 Rasm yuklash (sinf→fan→mavzu)", callback_data="img_upload_start")],
                [InlineKeyboardButton(text="📋 Barcha rasmlar", callback_data="img_panel")],
            ])
        )
        return

    elif message.text == "🖼 Rasmlar boshqaruvi":
        if user_id not in ADMINS:
            return
        from image_admin import show_image_panel
        await show_image_panel(message)
        return

    elif message.text == "📚 Shablon yaratish":
        if user_id not in ADMINS:
            return
        from shablon_yaratish import show_shablon_menu
        await show_shablon_menu(message, user_id)
        return

    elif message.text in ("🤖 AI Generator", "🧪 Test shablon"):
        if user_id not in ADMINS:
            return
        from ai_generatori import show_gen_start
        await show_gen_start(message, user_id)
        return

    elif message.text in ("📚 DTS boshqaruvi", "🧭 DTS topik boshqaruvi"):
        # Shablon state ni tozalamiz
        from shablon_yaratish import shablon_state
        shablon_state.pop(user_id, None)
        await dts_menu(message)
        return

    elif message.text == "📊 Test statistikasi":

        await message.answer(
            "📊 Test statistikasi / Generator",
            reply_markup=ReplyKeyboardMarkup(
                keyboard=[
                    [KeyboardButton(text="📚 Mavzular statistikasi"),
                    KeyboardButton(text="📊 Generator statistikasi")],
                    [KeyboardButton(text="▶️ Generatorni boshlash"),
                    KeyboardButton(text="⏹ Generatorni to‘xtatish")],
                    [KeyboardButton(text="🔙 Ortga")]
                ],
                  resize_keyboard=True
            )
        )

        return

    elif message.text == "🧪 Bilimni sinash":
        # Admin uchun test
        if user_id not in ADMINS:
            pass  # o'quvchi uchun allaqachon handled
        else:
            # Admin uchun ham xuddi o'quvchi kabi
            conn2 = _get_db_conn(); cur2 = conn2.cursor()
            cur2.execute("""
                SELECT grade FROM (SELECT DISTINCT grade FROM dts_tree WHERE is_deleted=FALSE) _g
                ORDER BY CASE WHEN grade ~ '^[0-9]+$' THEN grade::int ELSE 9999 END, grade
            """)
            _grades = [r[0] for r in cur2.fetchall()]
            cur2.close(); conn2.close()
            rows = [[InlineKeyboardButton(
                text=f"{g}-sinf" if str(g).isdigit() else str(g),
                callback_data=f"stnav_grade:{g}"
            )] for g in _grades]
            rows.append([InlineKeyboardButton(text="⚡ Tezkor 20ta", callback_data="tset_start_quick")])
            await message.answer(
                "🧪 Test — sinf tanlang:",
                reply_markup=InlineKeyboardMarkup(inline_keyboard=rows)
            )
            return

    elif message.text == "📦 Kitob Word":
        if user_id not in ADMINS: return
        # Kitoblar ro'yxati
        conn2 = _get_db_conn(); cur2 = conn2.cursor()
        cur2.execute("SELECT id, title, fan, sinf FROM books ORDER BY id DESC LIMIT 10")
        books = cur2.fetchall(); cur2.close(); conn2.close()
        if not books:
            await message.answer("❌ Hali kitob yuklanmagan.\n«📖 Kitob yuklash» dan boshlang.")
            return
        rows = []
        for b in books:
            rows.append([InlineKeyboardButton(
                text=f"📖 {b[1][:25]} ({b[2]}, {b[3]}-sinf)",
                callback_data=f"book_make:{b[0]}"
            )])
        rows.append([InlineKeyboardButton(text="📦 To'liq paket (ZIP)", callback_data="book_full_pkg")])
        await message.answer(
            "📖 Kitob yaratish\n\nQaysi kitobdan Word fayl yasaymiz?\n"
            "(15 betlik bo'laklarga bo'linadi, ZIP arxivda yuboriladi)",
            reply_markup=InlineKeyboardMarkup(inline_keyboard=rows)
        )
        return

    elif message.text and message.text.startswith("📊 Hisobotlar"):
        if user_id not in ADMINS: return
        conn2 = _get_db_conn(); cur2 = conn2.cursor()
        # O'qilmagan xatolar
        cur2.execute("SELECT COUNT(*) FROM error_log WHERE is_read=FALSE")
        unread = cur2.fetchone()[0] or 0
        # O'qilgan xatolar
        cur2.execute("SELECT COUNT(*) FROM error_log WHERE is_read=TRUE")
        read_cnt = cur2.fetchone()[0] or 0
        cur2.close(); conn2.close()

        rows = [
            [InlineKeyboardButton(text="📊 Test natijalari (Excel)", callback_data="rep_test")],
            [InlineKeyboardButton(text="📈 O'quvchi taraqqiyoti (Excel)", callback_data="rep_prog")],
            [InlineKeyboardButton(text="📅 Dars rejasi", callback_data="rep_plan")],
        ]
        # Xatolar — aniq ajratilgan
        if unread > 0:
            rows.append([InlineKeyboardButton(
                text=f"🔴 O'qilmagan xatolar: {unread} ta — KO'RISH",
                callback_data="err_unread"
            )])
        rows.append([InlineKeyboardButton(
            text=f"✅ O'qilgan xatolar: {read_cnt} ta",
            callback_data="err_read"
        )])
        rows.append([InlineKeyboardButton(
            text="🗑 Barcha xatolarni tozalash",
            callback_data="err_clear"
        )])

        txt = (
            f"📊 Hisobotlar & Xatolar\n\n"
            f"🔴 O'qilmagan: {unread} ta\n"
            f"✅ O'qilgan: {read_cnt} ta"
        )
        await message.answer(txt, reply_markup=InlineKeyboardMarkup(inline_keyboard=rows))
        return

    elif message.text == "📊 Hisobotlar":
        if user_id not in ADMINS: return
        from jadval_generator import test_results_text, weak_analysis_text
        text2  = test_results_text(days=30)
        text3  = weak_analysis_text()
        await message.answer(text2)
        await message.answer(text3)
        # Excel ham yuboramiz
        from jadval_generator import test_results_excel
        from aiogram.types import BufferedInputFile
        try:
            buf = test_results_excel(days=30)
            await message.answer_document(
                BufferedInputFile(buf.read(), "test_natijalari.xlsx"),
                caption="📊 Test natijalari (Excel)"
            )
        except Exception as e:
            await message.answer(f"Excel xato: {e}")
        return

    elif message.text == "📅 Dars rejasi":
        if user_id not in ADMINS: return
        admin_state[user_id] = "dars_rejasi"
        await message.answer(
            "📅 Dars rejasi\n\nSinf va fanni yozing:\n<code>1 | Ingliz tili</code>",
            parse_mode="HTML"
        )
        return

    elif message.text == "📈 Taraqqiyot":
        if user_id not in ADMINS: return
        from jadval_generator import student_progress_text, student_progress_excel
        from aiogram.types import BufferedInputFile
        text2 = student_progress_text()
        await message.answer(text2)
        try:
            buf = student_progress_excel()
            await message.answer_document(
                BufferedInputFile(buf.read(), "taraqqiyot.xlsx"),
                caption="📈 O'quvchi taraqqiyoti (Excel)"
            )
        except Exception as e:
            await message.answer(f"Excel xato: {e}")
        return

    elif message.text == "📝 Shablon to'ldirish":
        if user_id not in ADMINS: return
        admin_state[user_id] = "excel_merge"
        await message.answer(
            "📝 Shablon to'ldirish\n\n"
            "Bo'sh shablon Excel faylini yuboring.\n"
            "Bot DB dagi savollardan avtomatik to'ldiradi.\n\n"
            "Agar DB da savol yo'q bo'lsa — 2-fayl (savol fayli) ham yuborasiz."
        )
        return

    elif message.text == "📚 Kitoblar ▾":
        if user_id not in ADMINS: return
        conn2=_get_db_conn();cur2=conn2.cursor()
        try:
            cur2.execute("SELECT id,title,sinf,total_pages FROM books ORDER BY id DESC LIMIT 10")
            books=cur2.fetchall()
        except: books=[]
        cur2.close();conn2.close()
        rows2=[[InlineKeyboardButton(text=f"📖 {b[1]} ({b[3]} bet)",callback_data=f"kitob_info:{b[0]}")] for b in books]
        rows2.append([
            InlineKeyboardButton(text="📤 PDF yuklash",callback_data="kitob_upload"),

        ])
        await message.answer("📚 Kitoblar:", reply_markup=InlineKeyboardMarkup(inline_keyboard=rows2))
        return

    elif message.text == "📚 Kitoblar ▾":
        if user_id not in ADMINS: return
        await message.answer(
            "📚 Kitoblar bo'limi:",
            reply_markup=InlineKeyboardMarkup(inline_keyboard=[
                [InlineKeyboardButton(text="📖 Kitob yuklash",    callback_data="menu_kitob_yuklash")],
                [InlineKeyboardButton(text="🎓 Kitobni o'qit (AI)", callback_data="menu_kitob_oqit")],
                [InlineKeyboardButton(text="📦 Kitobdan Word",    callback_data="menu_kitob_word")],
            ])
        )
        return

    elif message.text == "🧠 Bilimlar ▾":
        if user_id not in ADMINS: return
        await message.answer(
            "🧠 Bilimlar bo'limi:",
            reply_markup=InlineKeyboardMarkup(inline_keyboard=[
                [InlineKeyboardButton(text="🔍 Bilim qidirish",          callback_data="menu_bilim_qidir")],
                [InlineKeyboardButton(text="📚 Bilimni mustahkamlash",   callback_data="menu_bilim_must")],
                [InlineKeyboardButton(text="🧪 Bilimni sinash (test)",   callback_data="menu_bilim_sin")],
            ])
        )
        return

    elif message.text == "🚀 Mavzu tayyorla":
        if user_id not in ADMINS: return
        admin_state[user_id] = "mtt_sinf_fan"
        await message.answer(
            "🚀 Mavzu tayyorlash\n\n"
            "Sinf va fanni yozing:\n"
            "Format: <code>sinf | fan</code>\n\n"
            "Masalan:\n"
            "<code>2 | Ingliz tili</code>\n"
            "<code>3 | Matematika</code>",
            parse_mode="HTML"
        )
        return

    elif message.text == "🎨 AI Rasm yaratish":
        if user_id not in ADMINS: return
        conn2 = _get_db_conn(); cur2 = conn2.cursor()
        cur2.execute("""
            SELECT grade FROM (
                SELECT DISTINCT d.grade FROM dts_tree d
                JOIN generated_tests g ON g.topic_code=d.topic_code
                WHERE d.is_deleted=FALSE
            ) _g
            ORDER BY CASE WHEN grade ~ '^[0-9]+$' THEN grade::int ELSE 99 END
        """)
        grades = [r[0] for r in cur2.fetchall()]
        cur2.close(); conn2.close()
        rows = [[InlineKeyboardButton(
            text=f"🏫 {gr}-sinf" if str(gr).isdigit() else f"📚 {gr}",
            callback_data=f"rasm_grade:{gr}"
        )] for gr in grades]
        rows.append([InlineKeyboardButton(text="✏️ O'zim tavsif beraman", callback_data="ai_rasm_custom")])
        await message.answer(
            "🎨 AI Rasm yaratish (BEPUL)\n\nSinf tanlang:",
            reply_markup=InlineKeyboardMarkup(inline_keyboard=rows)
        )
        return
    elif message.text == "📖 Kitob yuklash":
        if user_id not in ADMINS: return
        admin_state[user_id] = "kitob_yuklash"
        await message.answer(
            "📖 Kitob yuklash\n\n"
            "Quyidagi formatda yozing:\n"
            "<code>Kitob nomi | Fan | Sinf | Muallif</code>\n\n"
            "Masalan:\n"
            "<code>Matematika 5 | Matematika | 5 | Mirzayev A.</code>\n\n"
            "Keyin PDF faylni yuboring.",
            parse_mode="HTML"
        )
        return

    elif message.text == "🎓 Kitob o'qit":
        if user_id not in ADMINS: return
        # Oxirgi yuklangan kitobni o'qitamiz
        conn2 = _get_db_conn(); cur2 = conn2.cursor()
        cur2.execute("SELECT id,title,fan,sinf FROM books ORDER BY id DESC LIMIT 5")
        books = cur2.fetchall(); cur2.close(); conn2.close()
        if not books:
            await message.answer("❌ Hali kitob yuklanmagan."); return
        rows = [[InlineKeyboardButton(
            text=f"📖 {b[1]} ({b[2]}, {b[3]}-sinf)",
            callback_data=f"train_book:{b[0]}:{b[2]}:{b[3]}"
        )] for b in books]
        await message.answer(
            "🎓 Qaysi kitobni o'qitamiz?\n(AI tahlil qilib bilim bazasiga saqlaydi)",
            reply_markup=InlineKeyboardMarkup(inline_keyboard=rows)
        )
        return

    elif message.text == "🔍 Bilim qidirish":
        if user_id not in ADMINS: return
        admin_state[user_id] = "kitob_qidirish"
        await message.answer("🔍 Qidiruv so'zini yozing:")
        return

    elif message.text == "🔧 Test tuzatmalar":
        if user_id not in ADMINS: return
        conn2 = _get_db_conn(); cur2 = conn2.cursor()
        cur2.execute("""
            SELECT id, test_id, topic_code, question, comment, created_at
            FROM test_corrections WHERE status='new'
            ORDER BY created_at DESC LIMIT 15
        """)
        rows2 = cur2.fetchall(); cur2.close(); conn2.close()
        if not rows2:
            await message.answer("✅ Tuzatish so'rovlari yo'q!"); return
        lines = [f"🔧 Tuzatish so'rovlari ({len(rows2)} ta):\n"]
        kbs = []
        for cid, tid, tc, q, comm, cat in rows2:
            lines.append(f"📝 {q[:60]}...")
            if tid:
                kbs.append([InlineKeyboardButton(
                    text=f"✏️ #{tid} — ko'rish",
                    callback_data=f"admin_fix_test:{tid}"
                )])
        await message.answer("\n".join(lines[:10]),
                             reply_markup=InlineKeyboardMarkup(inline_keyboard=kbs[:10]))
        return

    elif message.text and ("🆘 Xatolar" in message.text):
        if user_id not in ADMINS:
            return
        conn2 = _get_db_conn()
        cur2  = conn2.cursor()
        cur2.execute("""
            SELECT id, user_id, username, error_text, created_at
            FROM error_log
            ORDER BY created_at DESC
            LIMIT 20
        """)
        errors = cur2.fetchall()
        # O'qilgan deb belgilash
        cur2.execute("UPDATE error_log SET is_read=TRUE WHERE is_read=FALSE")
        conn2.commit(); cur2.close(); conn2.close()

        if not errors:
            await message.answer(
                "✅ Xatolar yo'q!",
                reply_markup=get_main_keyboard("Admin", unread_errors=0)
            )
            return

        lines = [f"🆘 So'nggi xatolar ({len(errors)} ta):\n"]
        for i, (eid, uid, uname, etxt, eat) in enumerate(errors, 1):
            t = eat.strftime("%d.%m %H:%M") if eat else "?"
            short = (etxt or "")[:150].replace("\n", " ")
            lines.append(f"{i}. [{t}] {uname}({uid})\n   {short}\n")

        # Uzun bo'lsa bo'lib yuborish
        text = "\n".join(lines)
        while text:
            await message.answer(text[:4096])
            text = text[4096:]

        await message.answer(
            "✅ Hammasi o'qilgan deb belgilandi.",
            reply_markup=get_main_keyboard("Admin", unread_errors=0)
        )
        return

    elif message.text == "📖 Darslar holati":
        if user_id not in ADMINS:
            return
        conn2 = _get_db_conn()
        cur2  = conn2.cursor()
        cur2.execute("""
            SELECT
                d.grade,
                d.subject_name,
                COUNT(DISTINCT d.topic_code)           AS jami,
                COUNT(DISTINCT l.topic_code)           AS bor,
                COUNT(DISTINCT d.topic_code) -
                COUNT(DISTINCT l.topic_code)           AS yoq
            FROM dts_tree d
            LEFT JOIN teacher_lessons l ON l.topic_code = d.topic_code
            WHERE d.is_deleted = FALSE
            GROUP BY d.grade, d.subject_name
            ORDER BY
                CASE WHEN d.grade ~ '^[0-9]+$' THEN d.grade::int ELSE 99 END,
                d.subject_name
        """)
        rows2 = cur2.fetchall()
        cur2.close(); conn2.close()

        if not rows2:
            await message.answer("📭 DTS daraxt bo'sh.")
            return

        lines = ["📖 Darslar holati\n"]
        cur_grade = None
        t_jami = t_bor = t_yoq = 0
        for grade, subj, jami, bor, yoq in rows2:
            if grade != cur_grade:
                cur_grade = grade
                lines.append(f"\n🎓 {grade}-sinf:")
            pct = round(bor*100/jami) if jami else 0
            bar = "🟩" if pct==100 else ("🟨" if pct>=50 else "🟥")
            lines.append(
                f"  {bar} {subj}\n"
                f"     ✅ {bor} dars bor  |  ❌ {yoq} yo'q  |  📚 {jami} jami ({pct}%)"
            )
            t_jami += jami; t_bor += bor; t_yoq += yoq

        pct_t = round(t_bor*100/t_jami) if t_jami else 0
        lines.append(
            f"\n━━━━━━━━━━━━━━\n"
            f"📊 Jami: ✅ {t_bor} | ❌ {t_yoq} | 📚 {t_jami} ({pct_t}%)"
        )
        await message.answer("\n".join(lines))
        return

    elif message.text == "📚 Mavzular statistikasi":
        if user_id not in ADMINS:
            return
        conn2 = _get_db_conn()
        cur2  = conn2.cursor()
        cur2.execute("""
            SELECT
                t.grade,
                t.subject_name,
                COUNT(DISTINCT t.topic_code)                              AS mavzu_soni,
                COUNT(g.id)                                               AS test_soni,
                COUNT(DISTINCT CASE WHEN g.id IS NULL THEN t.topic_code END) AS bosh_mavzu
            FROM dts_tree t
            LEFT JOIN generated_tests g ON g.topic_code = t.topic_code
            WHERE t.is_deleted = FALSE
            GROUP BY t.grade, t.subject_name
            ORDER BY t.grade, t.subject_name
        """)
        rows = cur2.fetchall()
        cur2.close(); conn2.close()

        if not rows:
            await message.answer("📭 Hozircha mavzu ma'lumoti yo'q.")
            return

        # Sinf bo'yicha guruhlash
        from collections import defaultdict
        by_grade = defaultdict(list)
        for grade, subj, mavzu, test, bosh in rows:
            by_grade[grade].append((subj, mavzu, test, bosh))

        lines = ["📊 Mavzular statistikasi\n"]
        total_m = total_t = total_b = 0
        for grade in sorted(by_grade.keys(), key=lambda x: int(x) if str(x).isdigit() else 99):
            lines.append(f"\n🎓 {grade}-sinf:")
            for subj, mavzu, test, bosh in by_grade[grade]:
                avg = round(test/mavzu, 1) if mavzu else 0
                bar = "🟩" if avg >= 5 else ("🟨" if avg >= 2 else "🟥")
                lines.append(
                    f"  {bar} {subj}\n"
                    f"     📚 {mavzu} mavzu | 🧪 {test} test | ⚠️ {bosh} bo'sh"
                )
                total_m += mavzu; total_t += test; total_b += bosh

        lines.append(
            f"\n━━━━━━━━━━━━━━\n"
            f"📊 Jami: {total_m} mavzu | {total_t} test | {total_b} bo'sh mavzu"
        )
        await message.answer("\n".join(lines))
        return

    elif message.text and message.text.endswith("-sinf"):

        grade = message.text.replace("-sinf", "")

        topic_stats_state[user_id] = {
            "grade": grade,
            "level": "subjects"
        }

        subjects = get_subjects_by_grade(
            grade
        )

        keyboard = []

        for subject in subjects:

            keyboard.append([
                KeyboardButton(text=subject)
            ])

        keyboard.append([
            KeyboardButton(text="🔙 Ortga")
        ])

        await message.answer(
            "📖 Fanni tanlang",
            reply_markup=ReplyKeyboardMarkup(
                keyboard=keyboard,
                resize_keyboard=True
            )
        )

        return

    elif (
        user_id in topic_stats_state
        and f"topic_{message.text}" in topic_stats_state[user_id]
    ):
        topic_stats_state[user_id]["level"] = "topic_stats"
        topic_code = topic_stats_state[user_id][
            f"topic_{message.text}"
        ]

        info = get_topic_name(topic_code)
        stats = get_topic_statistics(topic_code)

        topic_stats_state[user_id]["selected_topic"] = topic_code

        total = stats[0] or 0
        oson = stats[1] or 0
        orta = stats[2] or 0
        qiyin = stats[3] or 0
        murakkab = stats[4] or 0

        await message.answer(
            f"🔑 {topic_code}\n\n"
            f"🎓 Sinf: {info[0]}\n"
            f"📚 Fan: {info[1]}\n"
            f"🗓 Chorak: {info[2]}\n"
            f"📖 Bob: {info[3]}\n"
            f"📂 Bo'lim: {info[4]}\n"
            f"📘 Mavzu: {info[5]}\n"
            f"📌 Kichik mavzu: {info[6]}\n\n"
            f"📊 Jami test: {total}\n"
            f"🟢 Oson: {oson}\n"
            f"🟡 O'rta: {orta}\n"
            f"🟠 Qiyin: {qiyin}\n"
            f"🔴 Murakkab: {murakkab}",
            reply_markup=InlineKeyboardMarkup(inline_keyboard=[
                [
                    InlineKeyboardButton(text="📄 Excel shablon", callback_data=f"ts_excel:{topic_code}"),
                    InlineKeyboardButton(text="📥 Import", callback_data=f"ts_import:{topic_code}"),
                ],
                [
                    InlineKeyboardButton(text="🤖 AI Yaratish", callback_data=f"ts_gen:{topic_code}"),
                    InlineKeyboardButton(text="🗑 O'chirish", callback_data=f"ts_del:{topic_code}"),
                ],
                [
                    InlineKeyboardButton(text="👁 Testlarni ko'rish", callback_data=f"ts_view:{topic_code}:0"),
                ],
            ])
        )

        return

    elif message.text == "📥 Test import qilish":

        admin_state[user_id] = "test_import"

        await message.answer(
            "Excel fayl yuboring"
        )

        return

    elif message.text == "📥 DTS import":

        await dts_import_menu(
            message,
            admin_state,
            user_id
        )

        return

    # ===== 📋 SHABLONLAR — yagona shablon/import markazi =====
    elif message.text == "📋 Shablonlar":
        if user_id not in ADMINS:
            return
        await message.answer(
            "📋 Shablonlar va Import\n\n"
            "Chap — bo'sh shablon olish, o'ng — to'ldirilganni import qilish:",
            reply_markup=ReplyKeyboardMarkup(
                keyboard=[
                    [KeyboardButton(text="📋 Topik shablon"), KeyboardButton(text="📥 Topik import")],
                    [KeyboardButton(text="🧪 Test shablon"),  KeyboardButton(text="📥 Test import")],
                    [KeyboardButton(text="📚 Dars shablon"),  KeyboardButton(text="📥 Dars import")],
                    [KeyboardButton(text="🔙 Admin menyu")],
                ],
                resize_keyboard=True
            )
        )
        return

    elif message.text == "📋 Topik shablon":
        if user_id not in ADMINS:
            return
        from shablon_yaratish import shablon_state
        shablon_state[user_id] = {"step": "sinf_fan"}
        await message.answer(
            "📋 Topik kod uchun shablon\n\n"
            "Sinf va fanni yozing:\nMasalan: 1 Ingliz tili"
        )
        return

    elif message.text == "📥 Topik import":
        if user_id not in ADMINS:
            return
        from dts_import_handlers import DTSImportState
        admin_state[user_id] = None
        await state.set_state(DTSImportState.waiting_excel)
        await message.answer(
            "📄 DTS Excel faylini yuboring\n\n"
            "Bekor qilish: /menu",
            reply_markup=InlineKeyboardMarkup(inline_keyboard=[[
                InlineKeyboardButton(text="❌ Bekor qilish", callback_data="cancel_import")
            ]])
        )
        return

    elif message.text == "📥 Test import":
        if user_id not in ADMINS:
            return
        admin_state[user_id] = "test_import"
        await message.answer("📥 Test import\n\nTo'ldirilgan Excel faylni yuboring:")
        return

    elif message.text == "✅ Ha, import qil" and admin_state.get(user_id) == "test_import_confirm":
        if user_id not in ADMINS:
            return
        path = test_import_files.get(user_id)
        admin_state[user_id] = None
        if not path or not os.path.exists(path):
            await message.answer("❌ Fayl topilmadi, qaytadan yuboring.",
                                 reply_markup=get_main_keyboard("Admin"))
            return
        await message.answer("⏳ Import qilinmoqda...", reply_markup=get_main_keyboard("Admin"))
        await import_tests_excel(message, path, user_id)

        # RASM_MALUMOTI varaqidagi kodlarni olish
        try:
            import openpyxl as _oxr
            _wbr = _oxr.load_workbook(path, data_only=True)
            rasm_kodlar = []
            if "RASM_MALUMOTI" in _wbr.sheetnames:
                _wsr = _wbr["RASM_MALUMOTI"]
                for _rr in range(2, _wsr.max_row+1):
                    _id = _wsr.cell(_rr,1).value
                    if _id: rasm_kodlar.append(str(_id))
                _wbr.close()
            if rasm_kodlar:
                # DB ga saqlaymiz (redeploy bo'lsa ham saqlanadi)
                conn_rq=_get_db_conn();cur_rq=conn_rq.cursor()
                cur_rq.execute("""CREATE TABLE IF NOT EXISTS rasm_queue
                    (id SERIAL PRIMARY KEY, user_id BIGINT, kod TEXT, done BOOLEAN DEFAULT FALSE, created_at TIMESTAMP DEFAULT NOW())""")
                cur_rq.execute("DELETE FROM rasm_queue WHERE user_id=%s",(user_id,))
                for kk in rasm_kodlar:
                    cur_rq.execute("INSERT INTO rasm_queue(user_id,kod) VALUES(%s,%s)",(user_id,kk))
                conn_rq.commit();cur_rq.close();conn_rq.close()
                await message.answer(
                    f"🖼 Rasmlar kutilmoqda!\n\n"
                    f"📊 Jami: {len(rasm_kodlar)} ta rasm kodi saqlandi\n\n"
                    f"Endi collage rasmlarni yuboring:\n"
                    f"Har collage = 30 ta rasm (5 qator × 6 ustun)\n\n"
                    f"1-collage yuboring 👇"
                )
        except Exception as _er:
            pass

        try:
            os.remove(path)
        except Exception:
            pass
        test_import_files.pop(user_id, None)
        return

    elif message.text == "❌ Bekor" and admin_state.get(user_id) == "test_import_confirm":
        admin_state[user_id] = None
        path = test_import_files.pop(user_id, None)
        if path and os.path.exists(path):
            try:
                os.remove(path)
            except Exception:
                pass
        await message.answer("❌ Import bekor qilindi.",
                             reply_markup=get_main_keyboard("Admin"))
        return

    elif message.text == "📚 Dars shablon":
        if user_id not in ADMINS:
            return
        await lesson_admin.la_show_grades(message)
        return

    elif message.text == "📥 Dars import":
        if user_id not in ADMINS:
            return
        await state.set_state(lesson_admin.LessonAdminState.waiting_excel)
        await message.answer("📥 Dars import\n\nTo'ldirilgan Excel faylini yuboring:")
        return

    elif message.text == "🔙 Admin menyu":
        if user_id not in ADMINS:
            return
        await message.answer("⚙️ Admin menyusi", reply_markup=get_main_keyboard("Admin"))
        return
    # ===== Shablonlar markazi tugadi =====

    elif (
        user_id in topic_stats_state
        and "grade" in topic_stats_state[user_id]
        and "topics" not in topic_stats_state[user_id]
    ):

        grade = topic_stats_state[user_id]["grade"]

        subject_name = message.text

        topics = get_topics_by_grade_subject(
            grade,
            subject_name
        )

        topic_stats_state[user_id]["subject"] = subject_name
        topic_stats_state[user_id]["topics"] = topics
        topic_stats_state[user_id]["page"] = 0
        topic_stats_state[user_id]["level"] = "topics"

        await show_topics_page(
            message,
            user_id
        )

        return

    elif message.text == "⬅️ Oldingi":

        if user_id not in topic_stats_state:
            return

        if topic_stats_state[user_id]["page"] > 0:

            topic_stats_state[user_id]["page"] -= 1

        await show_topics_page(
            message,
            user_id
        )

        return

    elif message.text == "➡️ Keyingi":

        if user_id not in topic_stats_state:
            return

        total = len(
            topic_stats_state[user_id]["topics"]
        )

        current_page = topic_stats_state[user_id]["page"]

        max_page = (total - 1) // 10

        if current_page < max_page:

            topic_stats_state[user_id]["page"] += 1

        await show_topics_page(
            message,
            user_id
        )

        return

    elif message.text == "🔙 Ortga":

        if user_id not in topic_stats_state:
            return

        level = topic_stats_state[user_id].get("level")

        if level == "topic_stats":

            topic_stats_state[user_id]["level"] = "topics"

            await show_topics_page(
                message,
                user_id
            )

            return

        elif level == "topics":

            grade = topic_stats_state[user_id]["grade"]

            keyboards = []

            subjects = get_subjects_by_grade(
                grade
            )

            for subject in subjects:
                keyboards.append([
                    KeyboardButton(
                        text=subject
                    )
                ])

            keyboards.append([
                KeyboardButton(
                    text="🔙 Ortga"
                )
            ])

            topic_stats_state[user_id]["level"] = "subjects"

            await message.answer(
                "📖 Fanni tanlang",
                reply_markup=ReplyKeyboardMarkup(
                    keyboard=keyboards,
                    resize_keyboard=True
                )
            )

            return

        elif level == "subjects":

            grade = topic_stats_state[user_id]["grade"]

            subjects = get_subjects_by_grade(
                grade
            )

            keyboard = []

            for subject in subjects:

                keyboard.append([
                    KeyboardButton(
                        text=subject
                    )
                ])

            keyboard.append([
                KeyboardButton(
                    text="🔙 Ortga"
                )
            ])

            await message.answer(
                "📖 Fanni tanlang",
                reply_markup=ReplyKeyboardMarkup(
                    keyboard=keyboard,
                    resize_keyboard=True
                )
            )

            return

    elif message.text and "-" in message.text:

        topic_code = message.text.strip()

        info = get_topic_name(
            topic_code
        )

        if not info:
            return

        stats = get_topic_statistics(
            topic_code
        )

        mavzu = info[0]
        kichik = info[1]

        await message.answer(

            f"🔑 {topic_code}\n\n"

            f"📖 Mavzu:\n"
            f"{mavzu}\n\n"

            f"📌 Kichik mavzu:\n"
            f"{kichik}\n\n"

            f"📊 Jami test: {stats[0]}\n\n"

            f"🟢 Oson: {stats[1] or 0}\n"
            f"🟡 O'rta: {stats[2] or 0}\n"
            f"🟠 Qiyin: {stats[3] or 0}\n"
            f"🔴 Murakkab: {stats[4] or 0}\n\n"

            f"📝 Single choice: {stats[5] or 0}\n"
            f"☑️ Multiple choice: {stats[6] or 0}\n"
            f"✅ True/False: {stats[7] or 0}\n"
            f"✍️ Write answer: {stats[8] or 0}"
        )

        return

    elif message.text == "📄 Excel shablon":
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl import Workbook
        import io

        # Tanlangan mavzu kodini olish
        topic_code = ""
        grade = ""
        if user_id in topic_stats_state:
            topic_code = topic_stats_state[user_id].get("selected_topic", "")
            grade = str(topic_stats_state[user_id].get("grade", ""))

        # Sinf bo'yicha age_group
        age_map = {
            "1": "6-7", "2": "7-8", "3": "8-9", "4": "9-10",
            "5": "10-11", "6": "11-12", "7": "12-13", "8": "13-14",
            "9": "14-15", "10": "15-16", "11": "16-17"
        }
        age_group = age_map.get(grade, "10-11")

        wb = Workbook()
        ws = wb.active
        ws.title = "TESTLAR"

        headers = [
            "topic_code", "difficulty", "situation", "question",
            "option_a", "option_b", "option_c", "option_d",
            "correct_answer", "explanation", "question_type",
            "is_latex", "image_url", "audio_text", "language",
            "life_level", "age_group", "time_limit"
        ]

        # Header qatori
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="2E86AB")
            cell.alignment = Alignment(horizontal="center")

        # 40 ta qator — har biri uchun difficulty va nom
        def get_difficulty(n):
            if n <= 10: return "oson"
            elif n <= 20: return "o'rta"
            elif n <= 30: return "qiyin"
            else: return "murakkab"

        for i in range(1, 41):
            diff = get_difficulty(i)
            row = [
                topic_code,              # topic_code — o'zi
                diff,                    # difficulty
                "oddiy",                 # situation
                "",                      # question
                "",                      # option_a
                "",                      # option_b
                "",                      # option_c
                "",                      # option_d
                "",                      # correct_answer
                "",                      # explanation
                "single_choice",         # question_type
                False,                   # is_latex
                f"{topic_code}-{i}",     # image_url — kod-1, kod-2...
                "",                      # audio_text
                "uz",                    # language
                1,                       # life_level
                age_group,               # age_group
                60,                      # time_limit
            ]
            ws.append(row)
            # Difficulty rangini ko'rsatish
            colors = {"oson": "C8F7C5", "o'rta": "FFF3CD", "qiyin": "FFD7C4", "murakkab": "F5C6CB"}
            ws.cell(row=i+1, column=2).fill = PatternFill("solid", fgColor=colors.get(diff, "FFFFFF"))

        # Ustun kengliklari
        ws.column_dimensions['A'].width = 35
        ws.column_dimensions['B'].width = 12
        ws.column_dimensions['C'].width = 10
        ws.column_dimensions['D'].width = 50
        for col in ['E','F','G','H']:
            ws.column_dimensions[col].width = 20
        ws.column_dimensions['I'].width = 15
        ws.column_dimensions['J'].width = 30

        # Ma'lumot varag'i
        ws2 = wb.create_sheet("MA'LUMOT")
        ws2.append(["Maydon", "Qiymatlar"])
        ws2.append(["difficulty", "oson | o'rta | qiyin | murakkab"])
        ws2.append(["situation", "oddiy | hayotiy | laboratoriya"])
        ws2.append(["question_type", "single_choice | write_answer | true_false"])
        ws2.append(["correct_answer", "A | B | C | D  (yoki matn)"])
        ws2.append(["is_latex", "True | False"])
        ws2.append(["language", "uz | ru | en"])
        ws2.append(["age_group", f"{age_group} (avtomatik)"])
        ws2.append(["time_limit", "60 (sekund)"])
        ws2.append(["", ""])
        ws2.append(["Mavzu kodi", topic_code or "—"])
        ws2.append(["Sinf", f"{grade}-sinf"])
        ws2.append(["", ""])
        ws2.append(["1-10 qator", "OSON 🟢"])
        ws2.append(["11-20 qator", "O'RTA 🟡"])
        ws2.append(["21-30 qator", "QIYIN 🟠"])
        ws2.append(["31-40 qator", "MURAKKAB 🔴"])

        # Xotiraga saqlash
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        fname = f"shablon_{topic_code or 'mavzu'}.xlsx"
        from aiogram.types import BufferedInputFile
        await message.answer_document(
            BufferedInputFile(buf.read(), filename=fname),
            caption=(
                f"📋 Test shabloni tayyor!\n\n"
                f"🔑 Mavzu kodi: {topic_code or '—'}\n"
                f"🎓 Sinf: {grade or '—'}\n"
                f"👶 Yosh guruhi: {age_group}\n\n"
                f"📝 40 ta qator:\n"
                f"🟢 1-10: Oson\n"
                f"🟡 11-20: O'rta\n"
                f"🟠 21-30: Qiyin\n"
                f"🔴 31-40: Murakkab\n\n"
                f"Faqat bo'sh ustunlarni to'ldiring!"
            )
        )
        return
    
    elif message.text == "▶️ Generatorni boshlash":

        global generator_process

        if generator_process and generator_process.poll() is None:

            await message.answer(
                "Generator ishlayapti"
            )

            return

        generator_process = subprocess.Popen(
            ["python", "test_generator.py"]
        )

        await message.answer(
            "Generator ishga tushdi"
        )

        return

    elif message.text == "⏹ Generatorni to‘xtatish":

        if generator_process and generator_process.poll() is None:

            generator_process.terminate()

            await message.answer(
                "Generator to‘xtatildi"
            )

        else:

            await message.answer(
                "Generator ishlamayapti"
            )

        return
    elif message.text == "📊 Generator statistikasi":

        conn = _get_db_conn()
        cur = conn.cursor()

        cur.execute("""
            SELECT COUNT(*)
            FROM dts_tree
        """)
        total_topics = cur.fetchone()[0]

        cur.execute("""
            SELECT COUNT(DISTINCT topic_code)
            FROM generated_tests
        """)
        completed_topics = cur.fetchone()[0]

        cur.execute("""
            SELECT COUNT(*)
            FROM generated_tests
        """)
        total_tests = cur.fetchone()[0]

        remaining_topics = total_topics - completed_topics

        progress = round(
            completed_topics * 100 / total_topics,
            1
        ) if total_topics else 0

        cur.close()
        conn.close()

        await message.answer(
            f"📚 Jami mavzular: {total_topics}\n"
            f"✅ Test yaratilgan mavzular: {completed_topics}\n"
            f"❌ Qolgan mavzular: {remaining_topics}\n"
            f"📝 Jami testlar: {total_tests}\n"
            f"📈 Progress: {progress}%"
        )

    elif message.text == "⬅ Ortga":

        await admin_main_menu(
            message
        )

        return

    elif message.text == "📊 DTS statistika":

        await dts_statistics(
            message
        )

        return

    elif message.text == "📤 DTS export":

        await dts_export_menu(
            message
        )

        return

    elif message.text == "📤 Hammasini export":

        await dts_export_all(
            message
        )

        return
        
    # parallel message bloklash
    
    if user_id not in user_locks:
        user_locks[user_id] = asyncio.Lock()
        
    async with user_locks[user_id]:

      #  action = TEXT_TO_ID.get(message.text)

        # 🔙 ORTGA
        if message.text == BACK:

            user_id = message.from_user.id

            # history bo‘lsa
            if user_id in state_history and len(state_history[user_id]) > 1:

                # hozirgi state ni olib tashlash
                state_history[user_id].pop()

                # oldingi state
                prev_state = state_history[user_id][-1]

                user_state[user_id] = prev_state

                # CLASS ga qaytish
                if prev_state == "class":

                    await message.answer(
                        "Sinf tanlang:",
                        reply_markup=base_keyboard(CLASSES)
                    )
                    return

                # SUBJECT ga qaytish
                elif prev_state == "subject":

                    selected_class = temp_user[user_id]["class"]

                    subjects = SUBJECTS_BY_CLASS.get(selected_class)

                    flat_subjects = []

                    for row in subjects:
                        flat_subjects.extend(row)

                    await message.answer(
                        "Fan tanlang:",
                        reply_markup=base_keyboard(flat_subjects)
                    )
                    return

                elif prev_state == "db_school":

                    await message.answer(
                        "🏫 Maktab turini tanlang:",
                        reply_markup=base_keyboard([
                            "all",
                            "🏫 Oddiy",
                            "⭐ IDUM",
                            "🏆 Prezident",
                            "🏢 Xususiy"
                        ])
                    )
                    return

        # ===== SURVEY RESULTS =====
        elif message.text == "📋 So‘rovnoma natijalari":

            if message.from_user.id not in ADMINS:
                return

            conn = _get_db_conn()
            cur = conn.cursor()

            cur.execute("""
            SELECT surveys.question,
            survey_answers.answer,
            COUNT(*)
            FROM survey_answers
            JOIN surveys
            ON surveys.id = survey_answers.survey_id
            GROUP BY surveys.question, survey_answers.answer
            """)

            rows = cur.fetchall()

            conn.close()

            if not rows:

                await message.answer(
                    "❌ Natijalar yo‘q"
                )
                return

            text = "📋 So‘rovnoma natijalari\n\n"

            current_question = ""

            for question, answer, count in rows:

                if question != current_question:

                    current_question = question

                    text += f"\n📝 {question}\n"

                text += f"• {answer} — {count} ta\n"

            await message.answer(text)

            return

        elif user_state.get(message.from_user.id) == "admin_region":

            temp_user[message.from_user.id]["admin_region"] = message.text

            districts = REGIONS.get(message.text, [])

            flat = []

            for row in districts:
                flat.extend(row)

            user_state[message.from_user.id] = "admin_district"

            await message.answer(
                "Tuman tanlang:",
                reply_markup=base_keyboard(flat)
            )

            return

        elif user_state.get(message.from_user.id) == "admin_district":

            temp_user[message.from_user.id]["admin_district"] = message.text

            conn = _get_db_conn()
            cur = conn.cursor()

            cur.execute("""
            SELECT DISTINCT school
            FROM users
            WHERE district=%s
            """, (message.text,))

            rows = cur.fetchall()

            conn.close()

            schools = [r[0] for r in rows if r[0]]

            user_state[message.from_user.id] = "admin_school"

            await message.answer(
                "Maktab tanlang:",
                reply_markup=base_keyboard(schools)
            )

            return

        elif user_state.get(message.from_user.id) == "admin_school":

            school = message.text

            conn = _get_db_conn()
            cur = conn.cursor()

            cur.execute("""
            SELECT AVG(score * 100.0 / total)
            FROM results
            WHERE school=%s
            """, (school,))

            avg = cur.fetchone()[0]

            conn.close()

            if avg is None:

                await message.answer(
                    "❌ Ma’lumot yo‘q"
                )
                return

            avg = round(avg, 1)

            bar = "█" * int(avg // 10)
            empty = "░" * (10 - int(avg // 10))

            text = (
                f"🏫 {school}\n\n"
                f"{bar}{empty}\n"
                f"📊 O‘rtacha: {avg}%"
            )

            await message.answer(text)

            user_state[message.from_user.id] = None

            return

        elif user_state.get(message.from_user.id) == "teacher_level":

            temp_user[message.from_user.id]["teacher_level"] = message.text

            subjects = SUBJECTS_BY_LEVEL.get(message.text, [])

            user_state[message.from_user.id] = "teacher_subject"

            await message.answer(
                "Fan tanlang:",
                reply_markup=base_keyboard(subjects)
            )

            return

        elif user_state.get(message.from_user.id) == "teacher_subject":

            temp_user[message.from_user.id]["subject"] = message.text

            user_state[message.from_user.id] = "test_type"

            await message.answer(
                "Test turini tanlang:",
                reply_markup=base_keyboard(TEST_TYPES)
            )

            return

        # 🏠 HOME
        elif message.text in (HOME, HOME2):

            conn = _get_db_conn()
            cur = conn.cursor()
            cur.execute("SELECT role FROM users WHERE user_id=%s", (message.from_user.id,))
            user = cur.fetchone()
            conn.close()

            role = user[0] if user else None
            user_state[message.from_user.id] = None

            # Oxirgi 10 xabarni o'chirish
            try:
                for i in range(message.message_id - 1, message.message_id - 15, -1):
                    try:
                        await message.bot.delete_message(message.chat.id, i)
                    except Exception:
                        pass
            except Exception:
                pass

            from student_dashboard import build_dashboard
            try:
                text, kb = await build_dashboard(message.from_user.id)
                await message.answer(text, reply_markup=kb)
            except Exception:
                pass
            await message.answer("👇 Menyu:", reply_markup=get_main_keyboard(role))
            return

        elif message.text == "⚙️ Akkaunt sozlamalari":

            await message.answer(
                "Sozlamalar:",
                reply_markup=ReplyKeyboardMarkup(
                    keyboard=[
                        [KeyboardButton(text="🔄 Rolni almashtirish")],
                        [KeyboardButton(text="🌍 Hududni almashtirish")],
                        [KeyboardButton(text="🏫 Maktabni almashtirish")],
                        [KeyboardButton(text="🎓 Sinfni almashtirish")],
                        [KeyboardButton(text=BACK)]
                    ],
                    resize_keyboard=True
                )
            )
        elif message.text == "🔄 Rolni almashtirish":

            user_state[message.from_user.id] = "change_role"

            await message.answer(
                "Yangi rolni tanlang:",
                reply_markup=make_keyboard(["🧒 O‘quvchi", "👨‍🏫 O‘qituvchi"])
            )

        elif user_state.get(message.from_user.id) == "change_role":

            conn = _get_db_conn()
            cur = conn.cursor()

            cur.execute("""
            UPDATE users
            SET role=%s
            WHERE user_id=%s
            """, (
                message.text,
                message.from_user.id
            ))

            conn.commit()
            conn.close()

            user_state[message.from_user.id] = None

            await message.answer(
                f"✅ Rol o‘zgartirildi: {message.text}",
                reply_markup=get_main_keyboard(message.text)
            )

        elif message.text == "🌍 Hududni almashtirish":

            user_state[message.from_user.id] = "change_region"

            await message.answer(
                "Viloyatni tanlang:",
                reply_markup=make_keyboard(REGIONS.keys())
            )

            return

        elif user_state.get(message.from_user.id) == "change_region":

            temp_user[message.from_user.id] = {
                "new_region": message.text
            }

            districts = REGIONS.get(message.text, [])

            flat = []

            for row in districts:
                flat.extend(row)

            user_state[message.from_user.id] = "change_district"

            await message.answer(
                "Tumanni tanlang:",
                reply_markup=base_keyboard(flat)
            )

            return

        elif user_state.get(message.from_user.id) == "change_district":

            conn = _get_db_conn()
            cur = conn.cursor()

            cur.execute("""
            UPDATE users
            SET region=%s, district=%s
            WHERE user_id=%s
            """, (
                temp_user[message.from_user.id]["new_region"],
                message.text,
                message.from_user.id
            ))

            conn.commit()

            cur.execute(
                "SELECT role FROM users WHERE user_id=%s",
                (message.from_user.id,)
            )

            role = cur.fetchone()[0]

            conn.close()

            user_state[message.from_user.id] = None

            await message.answer(
                "✅ Hudud o‘zgartirildi",
                reply_markup=get_main_keyboard(role)
            )

            return

# ===== MAKTABNI ALMASHTIRISH =====

        elif message.text == "🏫 Maktabni almashtirish":

            user_state[user_id] = "change_school_type"

            await message.answer(
                "Maktab turini tanlang:",
                reply_markup=make_keyboard(SCHOOL_TYPES)
            )

            return

        elif user_state.get(user_id) == "change_school_type":

            temp_user[user_id]["new_school_type"] = message.text

            user_state[user_id] = "change_school"

            await message.answer(
                "Maktab raqami kiriting:"
            )

            return

        elif user_state.get(user_id) == "change_school":

            temp_user[user_id]["new_school"] = (
                f"{temp_user[user_id]['new_school_type']} - {message.text}"
            )

            school_type = temp_user[user_id]["new_school_type"]

            if school_type == "🏫 Oddiy davlat maktabi":
                classes = [c for c in CLASSES if "🏫 Oddiy" in c]

            elif school_type == "⭐ Ixtisoslashgan (IDUM)":
                classes = [c for c in CLASSES if "⭐ IDUM" in c]

            elif school_type == "🏆 Prezident maktabi":
                classes = [c for c in CLASSES if "🏆 Prezident" in c]

            else:
                classes = [c for c in CLASSES if "🏢 Xususiy" in c]

            user_state[user_id] = "change_school_class"

            await message.answer(
                "Sinfni tanlang:",
                reply_markup=base_keyboard(classes)
            )

            return

        elif user_state.get(user_id) == "change_school_class":

            conn = _get_db_conn()
            cur = conn.cursor()

            cur.execute("""
            UPDATE users
            SET school=%s, class=%s
            WHERE user_id=%s
            """, (
                temp_user[user_id]["new_school"],
                message.text,
                user_id
            ))

            conn.commit()

            cur.execute("""
            SELECT role
            FROM users
            WHERE user_id=%s
            """, (user_id,))

            row = cur.fetchone()

            conn.close()

            role = row[0] if row else "🧒 O‘quvchi"

            user_state[user_id] = None

            await message.answer(
                "✅ Maktab va sinf o‘zgartirildi",
                reply_markup=get_main_keyboard(role)
            )

            return

        # ===== SINFNI ALMASHTIRISH =====
        elif message.text == "🎓 Sinfni almashtirish":

            conn = _get_db_conn()
            cur = conn.cursor()

            cur.execute("""
            SELECT school
            FROM users
            WHERE user_id=%s
            """, (message.from_user.id,))

            row = cur.fetchone()

            conn.close()

            school = row[0] if row else ""

            if "🏫 Oddiy" in school:
                classes = [c for c in CLASSES if "🏫 Oddiy" in c]

            elif "⭐ IDUM" in school:
                classes = [c for c in CLASSES if "⭐ IDUM" in c]

            elif "🏆 Prezident" in school:
                classes = [c for c in CLASSES if "🏆 Prezident" in c]

            else:
                classes = [c for c in CLASSES if "🏢 Xususiy" in c]

            user_state[message.from_user.id] = "change_class"

            await message.answer(
                "Yangi sinfni tanlang:",
                reply_markup=base_keyboard(classes)
            )

            return

        elif user_state.get(message.from_user.id) == "change_class":

            conn = _get_db_conn()
            cur = conn.cursor()

            cur.execute("""
            UPDATE users
            SET class=%s
            WHERE user_id=%s
            """, (
                message.text,
                message.from_user.id
            ))

            conn.commit()

            cur.execute("""
            SELECT role FROM users
            WHERE user_id=%s
            """, (message.from_user.id,))

            row = cur.fetchone()

            conn.close()

            role = row[0] if row else "🧒 O‘quvchi"

            user_state[message.from_user.id] = None

            await message.answer(
                f"✅ Sinf o‘zgartirildi: {message.text}",
                reply_markup=get_main_keyboard(role)
            )

            return
        
        elif user_state.get(message.from_user.id) == "survey_work":

            data = user_test[message.from_user.id]

            data["answers"][data["index"]] = message.text.upper()

            # oxiri
            if data["index"] >= len(data["surveys"]) - 1:

                conn = _get_db_conn()
                cur = conn.cursor()

                cur.execute("""
                UPDATE users
                SET survey_done=1
                WHERE user_id=%s
                """, (message.from_user.id,))

                conn.commit()
                conn.close()

                user_state[message.from_user.id] = None

                await message.answer(
                    "✅ So‘rovnoma tugadi",
                    reply_markup=get_main_keyboard(
                        temp_user[message.from_user.id]["role"]
                    )
                )

                return

            data["index"] += 1

            q = data["surveys"][data["index"]]

            text = (
                f"{data['index']+1}/{len(data['surveys'])}\n\n"
                f"{q[2]}\n\n"
                f"A) {q[4]}\n"
                f"B) {q[5]}\n"
                f"C) {q[6]}\n"
                f"D) {q[7]}"
            )

            await message.answer(text)

            return

def _mk_ts_kb(st2, cnt_total):
    """ts_start settings uchun ✅ li keyboard."""
    def c(cond): return "✅ " if cond else ""
    cnt   = st2.get("ts_count", 20)
    diff  = st2.get("ts_diff", "all")
    timed = st2.get("ts_timed", "mix")   # True/False/"mix"
    write = st2.get("ts_write", False)   # True/False/"mix"
    img   = st2.get("ts_img", "mix")     # True/False/"mix"
    return InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text=f"{c(cnt==20)}20 ta",    callback_data="ts_cnt_20"),
         InlineKeyboardButton(text=f"{c(cnt==40)}40 ta",    callback_data="ts_cnt_40"),
         InlineKeyboardButton(text=f"{c(cnt==cnt_total)}Barchasi ({cnt_total})", callback_data=f"ts_cnt_{cnt_total}")],
        [InlineKeyboardButton(text=f"{c(diff=='oson')}🟢 Oson",   callback_data="ts_dif_oson"),
         InlineKeyboardButton(text=f"{c(diff=='orta')}🟡 O'rta",  callback_data="ts_dif_orta"),
         InlineKeyboardButton(text=f"{c(diff=='qiyin')}🔴 Qiyin", callback_data="ts_dif_qiyin"),
         InlineKeyboardButton(text=f"{c(diff=='all')}🌈 Aralash",  callback_data="ts_dif_all")],
        [InlineKeyboardButton(text=f"{c(timed==True)}⏱ Vaqtli",    callback_data="ts_time_1"),
         InlineKeyboardButton(text=f"{c(timed==False)}∞ Vaqtsiz",  callback_data="ts_time_0"),
         InlineKeyboardButton(text=f"{c(timed=='mix')}⚡ Aralash",  callback_data="ts_time_mix")],
        [InlineKeyboardButton(text=f"{c(write==False)}🔘 Tugmali",   callback_data="ts_wr_0"),
         InlineKeyboardButton(text=f"{c(write==True)}✍️ Yozuvli",   callback_data="ts_wr_1"),
         InlineKeyboardButton(text=f"{c(write=='mix')}🔀 Aralash",   callback_data="ts_wr_mix")],
        [InlineKeyboardButton(text=f"{c(img==True)}🖼 Rasimli",      callback_data="ts_img_1"),
         InlineKeyboardButton(text=f"{c(img==False)}📝 Rasmsiz",     callback_data="ts_img_0"),
         InlineKeyboardButton(text=f"{c(img=='mix')}🔀 Aralash",     callback_data="ts_img_mix")],
        [InlineKeyboardButton(text="▶️ Boshlash", callback_data="ts_go")],
    ])

@dp.callback_query()

async def test_buttons(call: CallbackQuery, state: FSMContext):
    user_id = call.from_user.id
    # BIRINCHI call.answer() — Telegram "yuklanyapti" ni darhol to'xtatadi
    try:
        await call.answer()
    except Exception:
        pass
    try:
        await _test_buttons_inner(call, state, user_id)
    except Exception as _e:
        await _error_and_home(call, user_id, _e, "Xatolik")

async def _test_buttons_inner(call: CallbackQuery, state: FSMContext, user_id: int):

    # ═══ DTS NAVIGATOR ═══

    # ═══ BARCHA LA_ CALLBACKLAR (lesson_admin) ═══
    # ── DELEGATE ──
    try:
        if call.data.startswith("tg_") or call.data == "tg_back":
            from cb_togarak import handle_tg
            if await handle_tg(call,user_id,admin_state,user_state,temp_user,bot): return
        if call.data.startswith("stg_"):
            from cb_student_tg import handle_stg
            if await handle_stg(call,user_id,admin_state,user_state,temp_user,bot): return
        if call.data.startswith("kb_") or call.data.startswith("parent_"):
            from cb_kabinet import handle_kb
            if await handle_kb(call,user_id,admin_state,user_state,temp_user,bot): return
        if call.data.startswith("kitob_"):
            from cb_kitob import handle_kitob
            if await handle_kitob(call,user_id,admin_state,user_state,temp_user,bot): return
        if call.data.startswith("ts_") or call.data.startswith("sin_"):
            from cb_test_nav import handle_test_nav
            if await handle_test_nav(call,user_id,admin_state,user_state,temp_user,bot): return
    except Exception as _de: print(f"delegate: {_de}")
    if call.data.startswith("la_") or call.data in (
        "la_sel_all","la_dl_sel","la_tmpl","la_imp",
        "la_back_grades","la_home"
    ):
        import lesson_admin as _la
        d = call.data
        if d.startswith("la_gs|p|"):      await _la.la_grades_page(call); return
        if d.startswith("la_g|"):         await _la.la_grade(call); return
        if d.startswith("la_s|"):         await _la.la_subject(call); return
        if d.startswith("la_q|"):         await _la.la_quarter(call); return
        if d.startswith("la_b|"):         await _la.la_bob(call); return
        if d.startswith("la_bl|"):        await _la.la_bolim(call); return
        if d.startswith("la_m|"):         await _la.la_mavzu(call); return
        if d.startswith("la_sel|"):       await _la.la_toggle_select(call); return
        if d.startswith("la_tp|"):        await _la.la_topics_page(call); return
        if d == "la_sel_all":             await _la.la_select_all(call); return
        if d == "la_dl_sel":              await _la.la_download_selected(call); return
        if d == "la_tmpl":                await _la.la_template(call); return
        if d == "la_imp":                 await _la.la_import_prompt(call); return
        if d == "la_back_grades":         await _la.la_back_grades(call); return
        if d == "la_home":                await _la.la_home(call); return
        if d.startswith("la_ld|"):        await _la.la_lesson_detail(call); return
        if d.startswith("la_prev|"):      await _la.la_preview_lesson(call); return
        if d.startswith("la_edit|"):      await _la.la_edit_lesson(call); return
        if d.startswith("la_delc|"):      await _la.la_delete_confirm(call); return
        if d.startswith("la_dely|"):      await _la.la_delete_yes(call); return
        return

    return False
